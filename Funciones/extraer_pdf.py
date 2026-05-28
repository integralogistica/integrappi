import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# Columnas definidas por posición X (basado en extractos de Bancolombia)
COLUMNS = [
    ("FECHA", 23, 71),
    ("DESCRIPCION", 72, 239),
    ("SUCURSAL_CANAL", 240, 328),
    ("REFERENCIA_1", 329, 397),
    ("REFERENCIA_2", 398, 466),
    ("DOCUMENTO", 467, 525),
    ("VALOR", 526, 600),
]

# Tolerancia para agrupar palabras en la misma fila
Y_TOLERANCE = 5


def extract_transactions_by_position(pdf_path):
    """Extrae transacciones usando posiciones X/Y de las palabras en el PDF."""
    transactions = []
    header_info = {
        "empresa": "",
        "numero_cuenta": "",
        "nit": "",
        "tipo_cuenta": "",
        "saldo_efectivo": "",
        "saldo_canje": "",
        "saldo_total": "",
    }

    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            words = page.extract_words(keep_blank_chars=True)

            # Extraer info del encabezado (solo de la primera página)
            if page_idx == 0:
                header_info = _extract_header(words)
                # En la primera página, filtrar solo datos debajo del header
                fecha_headers = [w for w in words if w["text"] == "FECHA"]
                if fecha_headers:
                    data_start_y = fecha_headers[0]["top"] + 10
                    data_words = [w for w in words if w["top"] > data_start_y]
                else:
                    data_words = words
            else:
                # Páginas siguientes: tomar todas las palabras
                data_words = words

            # Filtrar palabras que son claramente no-data (pie de página)
            data_words = [w for w in data_words
                          if not w["text"].startswith("Página")
                          and not w["text"].startswith("Página")]

            # Agrupar palabras por fila
            rows = _group_words_by_row(data_words)

            for row_words in rows:
                parsed = _parse_row(row_words)
                if not parsed:
                    continue

                if parsed["tipo"] == "transaccion":
                    transactions.append(parsed)
                elif parsed["tipo"] == "orfana" and transactions:
                    prev = transactions[-1]
                    if parsed["referencia2"]:
                        if prev["referencia2"]:
                            prev["referencia2"] += " " + parsed["referencia2"]
                        else:
                            prev["referencia2"] = parsed["referencia2"]

    # Limpiar campo 'tipo' de las transacciones finales
    for txn in transactions:
        txn.pop("tipo", None)

    return header_info, transactions


def _extract_header(words):
    """Extrae información del encabezado del PDF."""
    info = {
        "empresa": "",
        "numero_cuenta": "",
        "nit": "",
        "tipo_cuenta": "",
        "saldo_efectivo": "",
        "saldo_canje": "",
        "saldo_total": "",
    }

    full_text = " ".join(w["text"] for w in words)

    # Empresa: primera línea del PDF (top más bajo)
    if words:
        min_top = min(w["top"] for w in words)
        empresa_words = [w for w in words if abs(w["top"] - min_top) < 2]
        info["empresa"] = " ".join(w["text"] for w in sorted(empresa_words, key=lambda w: w["x0"]))

    m = re.search(r"Cuenta:\s*(\d+)", full_text)
    if m:
        info["numero_cuenta"] = m.group(1)

    m = re.search(r"NIT:\s*(\d+)", full_text)
    if m:
        info["nit"] = m.group(1)

    m = re.search(r"Tipo de cuenta:\s*(\w+)", full_text)
    if m:
        info["tipo_cuenta"] = m.group(1)

    m = re.search(r"Saldo Efectivo Actual:\s*\$?([\d,]+\.\d{2})", full_text)
    if m:
        info["saldo_efectivo"] = m.group(1)

    m = re.search(r"Saldo en Canje Actual:\s*\$?([\d,]+\.\d{2})", full_text)
    if m:
        info["saldo_canje"] = m.group(1)

    m = re.search(r"Saldo Total Actual:\s*\$?([\d,]+\.\d{2})", full_text)
    if m:
        info["saldo_total"] = m.group(1)

    return info


def _group_words_by_row(words):
    """Agrupa palabras en filas basándose en su posición Y."""
    if not words:
        return []

    # Ordenar por posición Y
    sorted_words = sorted(words, key=lambda w: (w["top"], w["x0"]))

    rows = []
    current_row = [sorted_words[0]]

    for word in sorted_words[1:]:
        if abs(word["top"] - current_row[0]["top"]) <= Y_TOLERANCE:
            current_row.append(word)
        else:
            rows.append(current_row)
            current_row = [word]
    rows.append(current_row)

    return rows


def _parse_row(row_words):
    """Parsea una fila de palabras en una transacción usando posiciones de columna."""
    col_data = {col[0]: [] for col in COLUMNS}

    for word in row_words:
        assigned = False
        for col_name, x_start, x_end in COLUMNS:
            word_center = (word["x0"] + word["x1"]) / 2
            if x_start - 5 <= word_center <= x_end + 5:
                col_data[col_name].append(word["text"])
                assigned = True
                break
        if not assigned:
            # Si la palabra cae fuera de las columnas, asignarla a la más cercana
            word_center = (word["x0"] + word["x1"]) / 2
            best_col = min(COLUMNS, key=lambda c: abs(word_center - (c[1] + c[2]) / 2))
            col_data[best_col[0]].append(word["text"])

    fecha_str = " ".join(col_data["FECHA"])
    valor_str = " ".join(col_data["VALOR"])
    valor_str = valor_str.replace(",", "").replace(".", ",")
    desc_str = " ".join(col_data["DESCRIPCION"])
    suc_str = " ".join(col_data["SUCURSAL_CANAL"])
    ref1_str = " ".join(col_data["REFERENCIA_1"])
    ref2_str = " ".join(col_data["REFERENCIA_2"])
    doc_str = " ".join(col_data["DOCUMENTO"])

    # Fila de transacción válida: tiene fecha y valor
    if re.match(r"\d{4}/\d{2}/\d{2}", fecha_str) and re.match(r"-?[\d]+,\d{2}", valor_str):
        return {
            "tipo": "transaccion",
            "fecha": fecha_str,
            "descripcion": desc_str,
            "sucursal": suc_str,
            "referencia1": ref1_str,
            "referencia2": ref2_str,
            "documento": doc_str,
            "valor": valor_str,
        }

    # Fila huérfana: no tiene fecha ni valor, pero tiene datos en columnas de referencia/descripción
    orphan_data = ref1_str + ref2_str + doc_str + desc_str + suc_str
    if orphan_data.strip() and not fecha_str.strip() and not valor_str.strip():
        return {
            "tipo": "orfana",
            "descripcion": desc_str,
            "sucursal": suc_str,
            "referencia1": ref1_str,
            "referencia2": ref2_str,
            "documento": doc_str,
        }

    return None


def create_excel(header_info, transactions, output_path):
    """Crea el archivo Excel con formato profesional."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracto Bancario"

    # Estilos
    dark_blue = "003366"
    light_blue = "D6E4F0"
    green = "548235"
    red = "C00000"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Calibri", bold=True, color=dark_blue, size=14)
    subtitle_font = Font(name="Calibri", bold=True, color=dark_blue, size=11)
    normal_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", bold=True, size=10)
    header_fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    alt_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    # Encabezados de tabla directamente en la fila 1
    columns = ["FECHA", "DESCRIPCION", "SUCURSAL/CANAL", "REFERENCIA 1", "REFERENCIA 2", "DOCUMENTO", "VALOR"]
    col_widths = [14, 42, 18, 18, 18, 18, 20]

    for col_idx, (col_name, width) in enumerate(zip(columns, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    header_row = 1
    row = 2

    # Datos
    for i, txn in enumerate(transactions):
        data = [
            txn["fecha"],
            txn["descripcion"],
            txn["sucursal"],
            txn["referencia1"],
            txn["referencia2"],
            txn["documento"],
            txn["valor"],
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border

            if col_idx == 2:  # Descripción
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 7:  # Valor
                cell.alignment = Alignment(horizontal="right", vertical="center")
                val_str = str(value).replace(",", "").replace("$", "")
                try:
                    num = float(val_str)
                    if num < 0:
                        cell.font = Font(name="Calibri", size=10, color=red)
                    else:
                        cell.font = Font(name="Calibri", size=10, color=green)
                except ValueError:
                    pass
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if i % 2 == 1:
                cell.fill = alt_fill

        row += 1

    # Resumen
    row += 1
    ws.cell(row=row, column=1, value=f"Total transacciones: {len(transactions)}").font = subtitle_font

    # Congelar paneles y autofiltro
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:G{header_row + len(transactions)}"

    wb.save(output_path)


def process_pdf(pdf_path):
    """Procesa un PDF y genera el Excel correspondiente."""
    header_info, transactions = extract_transactions_by_position(pdf_path)

    if not transactions:
        return False, "No se encontraron transacciones en el PDF."

    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    default_name = f"{base_name}_extracto.xlsx"

    root = tk.Tk()
    root.withdraw()
    save_path = filedialog.asksaveasfilename(
        title="Guardar Excel como...",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Archivos Excel", "*.xlsx")],
    )
    root.destroy()

    if not save_path:
        return False, "Operación cancelada por el usuario."

    create_excel(header_info, transactions, save_path)
    return True, f"Excel generado: {save_path}\nTransacciones encontradas: {len(transactions)}"


def select_and_process():
    """Abre diálogo para seleccionar PDF(s) y procesarlos."""
    root = tk.Tk()
    root.withdraw()

    pdf_files = filedialog.askopenfilenames(
        title="Seleccionar archivo(s) PDF de extracto bancario",
        filetypes=[("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*")],
        initialdir=os.path.expanduser("~/Downloads"),
    )

    if not pdf_files:
        messagebox.showinfo("Cancelado", "No se seleccionó ningún archivo.")
        return

    results = []
    for pdf_path in pdf_files:
        try:
            success, msg = process_pdf(pdf_path)
            icon = "OK" if success else "ERROR"
            results.append(f"[{icon}] {os.path.basename(pdf_path)}:\n{msg}")
        except Exception as e:
            results.append(f"[ERROR] {os.path.basename(pdf_path)}: {str(e)}")

    messagebox.showinfo("Resultado", "\n\n".join(results))


if __name__ == "__main__":
    select_and_process()
