from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from .errors import RNDCBusinessError, RNDCNoDataError
from .models import ExploracionRutaRequest
from .service import normalizar

ENTRADAS = [
    "periodo", "configuracion", "origen", "destino", "condicion_carga",
    "unidad_transporte_nombre", "tipo_carga_nombre", "horas_totales_cargue",
    "horas_totales_descargue", "limit",
]
SALIDAS = [
    "fila_entrada", "estado", "mensaje", "periodo_resultado", "origen_codigo",
    "origen_nombre", "destino_codigo", "destino_nombre", "configuracion_resultado",
    "condicion_carga_resultado", "tipo_carga_codigo", "tipo_carga_nombre_resultado",
    "unidad_transporte_codigo", "unidad_transporte_nombre_resultado", "ruta_id", "via",
    "kilometros", "horas_recorrido", "valor_moviliza", "valor_hora",
    "horas_logisticas_total", "costo_total_calculado",
]
MAX_FILAS = 200


def _ajustar_hoja(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, column in enumerate(ws.columns, 1):
        width = min(max((len(str(cell.value or "")) for cell in column), default=10) + 2, 55)
        ws.column_dimensions[get_column_letter(index)].width = width


def crear_plantilla() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "consultas"
    ws.append(ENTRADAS)
    ws.append(["202608", "2L3", "08296000", "76892000", "1", "FURGON", "General", 3, 3, 20])
    ws.append(["202608", "3S3", "11001000", "05001000", "1", "FURGON", "General", 3, 3, 20])
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).number_format = "@"
        ws.cell(row, 3).number_format = "00000000"
        ws.cell(row, 4).number_format = "00000000"
    _ajustar_hoja(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _texto_codigo(value, longitud=None):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text.zfill(longitud) if longitud and text.isdigit() else text


def _numero_excel(value):
    """Convierte números RNDC a celdas numéricas sin alterar textos inválidos."""
    if value is None or value == "":
        return None
    try:
        number = Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, ValueError):
        return value
    return int(number) if number == number.to_integral_value() else float(number)


def _formatear_columnas_numericas(ws):
    headers = {cell.value: cell.column for cell in ws[1]}
    formatos = {
        "kilometros": "#,##0.00",
        "horas_recorrido": "#,##0.00",
        "valor_moviliza": "$#,##0",
        "valor_hora": "$#,##0",
        "horas_logisticas_total": "#,##0.00",
        "costo_total_calculado": "$#,##0",
    }
    for header, number_format in formatos.items():
        column = headers.get(header)
        if not column:
            continue
        for row in range(2, ws.max_row + 1):
            ws.cell(row, column).number_format = number_format


def leer_consultas_excel(content: bytes) -> list[tuple[int, ExploracionRutaRequest]]:
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("El archivo no es un Excel .xlsx válido") from exc
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [normalizar(x).lower().replace(" ", "_") for x in next(rows)]
    except StopIteration as exc:
        raise ValueError("El Excel está vacío") from exc
    faltantes = [x for x in ("periodo", "configuracion", "origen", "destino") if x not in headers]
    if faltantes:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(faltantes))
    consultas = []
    errores = []
    for row_number, values in enumerate(rows, 2):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        raw = dict(zip(headers, values))
        raw["periodo"] = _texto_codigo(raw.get("periodo"), 6)
        raw["configuracion"] = _texto_codigo(raw.get("configuracion"))
        raw["origen"] = _texto_codigo(raw.get("origen"), 8)
        raw["destino"] = _texto_codigo(raw.get("destino"), 8)
        if raw.get("condicion_carga") is not None:
            raw["condicion_carga"] = _texto_codigo(raw.get("condicion_carga"))
        for optional in ("condicion_carga", "unidad_transporte_nombre", "tipo_carga_nombre", "horas_totales_cargue", "horas_totales_descargue", "limit"):
            if raw.get(optional) is None or str(raw.get(optional)).strip() == "":
                raw.pop(optional, None)
        # El Excel permite revisar varias vías sin generar respuestas excesivas.
        # El endpoint JSON mantiene su default general de 200.
        raw.setdefault("limit", 20)
        try:
            consultas.append((row_number, ExploracionRutaRequest.model_validate(raw)))
        except ValidationError as exc:
            mensaje = "; ".join(f"{'.'.join(map(str, e['loc']))}: {e['msg']}" for e in exc.errors())
            errores.append((row_number, raw, mensaje))
    if len(consultas) + len(errores) > MAX_FILAS:
        raise ValueError(f"El archivo supera el máximo de {MAX_FILAS} filas")
    if not consultas and not errores:
        raise ValueError("El Excel no contiene filas de consulta")
    return consultas, errores


def _fila_base(payload):
    return [
        payload.periodo, payload.configuracion, payload.origen, payload.destino,
        payload.condicion_carga, payload.unidad_transporte_nombre,
        payload.tipo_carga_nombre, str(payload.horas_totales_cargue),
        str(payload.horas_totales_descargue), payload.limit,
    ]


def procesar_excel(content: bytes, client, resumir_rutas) -> BytesIO:
    consultas, errores_validacion = leer_consultas_excel(content)
    wb = Workbook()
    ws = wb.active
    ws.title = "resultados"
    ws.append(ENTRADAS + SALIDAS)

    for row_number, raw, mensaje in errores_validacion:
        base = [raw.get(x) for x in ENTRADAS]
        ws.append(base + [row_number, "ERROR_VALIDACION", mensaje] + [None] * (len(SALIDAS) - 3))

    for row_number, payload in consultas:
        base = _fila_base(payload)
        try:
            documentos = client.explorar(
                payload.periodo, payload.configuracion, payload.origen,
                payload.destino, payload.condicion_carga
            )
            rutas, _ = resumir_rutas(
                documentos, payload.limit, payload.unidad_transporte_nombre,
                payload.tipo_carga_nombre, payload.horas_totales_cargue,
                payload.horas_totales_descargue
            )
            if not rutas:
                ws.append(base + [row_number, "SIN_RESULTADO", "RNDC no devolvió coincidencias para los filtros"] + [None] * (len(SALIDAS) - 3))
                continue
            for ruta in rutas:
                ws.append(base + [
                    row_number, "OK", "", ruta.get("periodo"), ruta.get("origen_codigo"),
                    ruta.get("origen_nombre"), ruta.get("destino_codigo"), ruta.get("destino_nombre"),
                    ruta.get("configuracion"), ruta.get("condicion_carga"), ruta.get("tipo_carga_codigo"),
                    ruta.get("tipo_carga_nombre"), ruta.get("unidad_transporte_codigo"),
                    ruta.get("unidad_transporte_nombre"), ruta.get("ruta_id"), ruta.get("via"),
                    _numero_excel(ruta.get("kilometros")), _numero_excel(ruta.get("horas_recorrido")),
                    _numero_excel(ruta.get("valor_moviliza")), _numero_excel(ruta.get("valor_hora")),
                    _numero_excel(ruta.get("horas_logisticas_total")),
                    _numero_excel(ruta.get("costo_total_calculado")),
                ])
        except RNDCNoDataError as exc:
            ws.append(base + [row_number, "SIN_RESULTADO", str(exc)] + [None] * (len(SALIDAS) - 3))
        except RNDCBusinessError as exc:
            estado = "SIN_RESULTADO" if "RNDC13" in str(exc).upper() else "ERROR_RNDC"
            ws.append(base + [row_number, estado, str(exc)] + [None] * (len(SALIDAS) - 3))
        except Exception as exc:
            ws.append(base + [row_number, "ERROR", f"{type(exc).__name__}: {exc}"] + [None] * (len(SALIDAS) - 3))

    _formatear_columnas_numericas(ws)
    _ajustar_hoja(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
