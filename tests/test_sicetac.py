import unittest
import xml.etree.ElementTree as ET
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

import httpx
from openpyxl import Workbook, load_workbook

from rutas.sicetac import _resumir_rutas
from sicetac.config import COMBINACIONES, validar_combinaciones
from sicetac.excel_service import crear_plantilla, leer_consultas_excel, procesar_excel
from sicetac.errors import RNDCBusinessError, RNDCCredentialsError, RNDCNoDataError, RNDCSoapFaultError
from sicetac.models import ExploracionRutaRequest, calcular_costo_total
from sicetac.rndc_client import RNDCClient, SOAP_ACTION, construir_envolvente_soap, construir_xml_exploracion, construir_xml_rndc, interpretar_respuesta_soap, sanitizar
from sicetac.service import SicetacService, _coincide, consulta_id, periodo_anterior


def soap(inner, prefix="soapenv"):
    escaped = inner.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return f'<{prefix}:Envelope xmlns:{prefix}="http://schemas.xmlsoap.org/soap/envelope/"><{prefix}:Body><m:AtenderMensajeRNDCResponse xmlns:m="urn:x"><return>{escaped}</return></m:AtenderMensajeRNDCResponse></{prefix}:Body></{prefix}:Envelope>'.encode()


class XmlTests(unittest.TestCase):
    def test_xml_actual_comillas_y_escape(self):
        combo = dict(COMBINACIONES[0]); combo["origen_codigo"] = "12345678"
        data = construir_xml_rndc("u<&", "p&", "202608", combo)
        root = ET.fromstring(data)
        self.assertEqual(root.findtext("solicitud/tipo"), "6")
        self.assertEqual(root.findtext("solicitud/procesoid"), "26")
        self.assertEqual(root.findtext("documento/PERIODO"), "'202608'")
        self.assertEqual(root.findtext("acceso/username"), "u<&")

    def test_envolvente_y_action(self):
        body = construir_envolvente_soap(construir_xml_rndc("u", "p", "202608", COMBINACIONES[0]))
        ET.fromstring(body)  # La envoltura debe ser XML válido, sin namespaces duplicados.
        self.assertIn(b"AtenderMensajeRNDC", body)
        self.assertIn(b"Request", body)
        self.assertEqual(body.count(b"xmlns:xsi="), 1)
        self.assertEqual(SOAP_ACTION, "urn:BPMServicesIntf-IBPMServices#AtenderMensajeRNDC")

    def test_return_namespaces_y_multiples_documentos(self):
        inner = "<root><documento><periodo>202608</periodo><valormoviliza>1</valormoviliza></documento><documento><periodo>202608</periodo><valormoviliza>2</valormoviliza></documento></root>"
        self.assertEqual(len(interpretar_respuesta_soap(soap(inner, "s"))), 2)

    def test_error_credentials(self):
        with self.assertRaises(RNDCCredentialsError):
            interpretar_respuesta_soap(soap("<root><ErrorMSG>Usuario o contraseña inválida</ErrorMSG></root>"))

    def test_rndc11_con_sql_mal_escapado_es_sin_datos(self):
        malformed = "<ErrorMSG>Error RNDC11: Documento no encontrado. SELECT * WHERE ROWNUM <= 10000</ErrorMSG>"
        with self.assertRaises(RNDCNoDataError):
            interpretar_respuesta_soap(soap(malformed))

    def test_fault(self):
        data = b'<s:Envelope xmlns:s="http://schemas.xmlsoap.org/soap/envelope/"><s:Body><s:Fault><faultcode>x</faultcode><faultstring>boom</faultstring></s:Fault></s:Body></s:Envelope>'
        with self.assertRaises(RNDCSoapFaultError): interpretar_respuesta_soap(data)

    def test_sanitizacion(self):
        self.assertEqual(sanitizar("user=juan password=secreto", "juan", "secreto"), "user=*** password=***")

    def test_xml_amplio_omite_destino_y_condicion(self):
        data = construir_xml_rndc("u", "p", "202608", COMBINACIONES[0], amplia=True)
        root = ET.fromstring(data)
        self.assertEqual(root.findtext("documento/ORIGEN"), "'08296000'")
        self.assertIsNone(root.find("documento/DESTINO"))
        self.assertIsNone(root.find("documento/CONDICIONCARGAID"))

    def test_xml_exploracion_sin_origen_solo_filtra_periodo_y_configuracion(self):
        data = construir_xml_exploracion("u", "p", "202608", "3S3")
        root = ET.fromstring(data)
        self.assertEqual(root.findtext("documento/PERIODO"), "'202608'")
        self.assertEqual(root.findtext("documento/CONFIGURACIONESID"), "'3S3'")
        self.assertIsNone(root.find("documento/ORIGEN"))
        self.assertIsNone(root.find("documento/DESTINO"))

    def test_xml_exploracion_de_ruta_incluye_origen_y_destino(self):
        data = construir_xml_exploracion(
            "u", "p", "202608", "3S3", "11001000", "05001000", "1"
        )
        root = ET.fromstring(data)
        self.assertEqual(root.findtext("documento/ORIGEN"), "'11001000'")
        self.assertEqual(root.findtext("documento/DESTINO"), "'05001000'")
        self.assertEqual(root.findtext("documento/CONDICIONCARGAID"), "'1'")

    def test_alias_liviano_se_traduce_al_id_interno_rndc(self):
        data = construir_xml_exploracion(
            "u", "p", "202608", "2L3", "08296000", "76892000", "1"
        )
        root = ET.fromstring(data)
        self.assertEqual(root.findtext("documento/CONFIGURACIONESID"), "'2_7_8'")

    def test_cliente_reintenta_rndc13(self):
        respuestas = [
            httpx.Response(200, content=soap("<root><ErrorMSG>Error RNDC13: nodo incompatible</ErrorMSG></root>")),
            httpx.Response(200, content=soap("<root><documento><periodo>202608</periodo></documento></root>")),
        ]
        llamadas = []

        def handler(request):
            llamadas.append(request)
            return respuestas.pop(0)

        client = RNDCClient("http://rndc.test", "u", "p", transport=httpx.MockTransport(handler))
        client._min_request_interval = 0
        try:
            docs = client.explorar("202608", "3S3", "11001000", "05001000", "1")
        finally:
            client.close()
        self.assertEqual(len(docs), 1)
        self.assertEqual(len(llamadas), 2)


class DomainTests(unittest.TestCase):
    def test_payload_exploracion_normaliza_configuracion(self):
        payload = ExploracionRutaRequest(
            periodo="202608", configuracion="3s3", origen="11001000",
            destino="05001000"
        )
        self.assertEqual(payload.configuracion, "3S3")
        self.assertEqual(payload.condicion_carga, "1")
        self.assertEqual(payload.horas_totales_cargue, Decimal("3"))
        self.assertEqual(payload.horas_totales_descargue, Decimal("3"))

    def test_exploracion_filtra_furgon_antes_del_limite(self):
        documentos = [
            {"rutasid": "1", "nombreunidadtransporte": "PLATAFORMA"},
            {"rutasid": "2", "nombreunidadtransporte": "FURGON"},
            {"rutasid": "3", "nombreunidadtransporte": "Furgón"},
        ]
        rutas, total = _resumir_rutas(documentos, 1, unidad_transporte_nombre="furgon")
        self.assertEqual(total, 2)
        self.assertEqual(len(rutas), 1)
        self.assertEqual(rutas[0]["ruta_id"], "2")

    def test_exploracion_calcula_costo_con_horas_del_json(self):
        documentos = [{"rutasid": "1", "valormoviliza": "1000", "valorhora": "100"}]
        rutas, _ = _resumir_rutas(
            documentos, 10, horas_totales_cargue=Decimal("2.5"),
            horas_totales_descargue=Decimal("1.5")
        )
        self.assertEqual(rutas[0]["horas_logisticas_total"], "4.0")
        self.assertEqual(rutas[0]["costo_total_calculado"], "1400.0")


class ExcelTests(unittest.TestCase):
    def test_plantilla_contiene_dos_ejemplos_validos(self):
        consultas, errores = leer_consultas_excel(crear_plantilla().getvalue())
        self.assertEqual(len(consultas), 2)
        self.assertEqual(errores, [])
        self.assertEqual(consultas[0][1].origen, "08296000")
        self.assertEqual(consultas[0][1].limit, 20)

    def test_resultado_excel_incluye_una_fila_por_ruta(self):
        wb = Workbook(); ws = wb.active
        ws.append(["periodo", "configuracion", "origen", "destino", "condicion_carga"])
        ws.append([202608, "3S3", 11001000, 5001000, 1])
        source = BytesIO(); wb.save(source)

        class ExcelClient:
            def explorar(self, *args):
                return [{
                    "periodo": "202608", "origen": "11001000", "destino": "5001000",
                    "configuracion": "3S3", "condicioncarga": "CARGADO",
                    "rutasid": "106", "valormoviliza": "1000", "valorhora": "100",
                }]

        output = procesar_excel(source.getvalue(), ExcelClient(), _resumir_rutas)
        result = load_workbook(output, data_only=True).active
        headers = [cell.value for cell in result[1]]
        self.assertNotIn("consulta_id_usuario", headers)
        self.assertNotIn("fila_original", headers)
        self.assertIn("fila_entrada", headers)
        row = dict(zip(headers, [cell.value for cell in result[2]]))
        self.assertEqual(row["estado"], "OK")
        self.assertEqual(row["destino"], "05001000")
        self.assertEqual(row["limit"], 20)
        self.assertEqual(row["costo_total_calculado"], 1600)
        costo_cell = result.cell(2, headers.index("costo_total_calculado") + 1)
        self.assertEqual(costo_cell.data_type, "n")
        self.assertEqual(costo_cell.number_format, "$#,##0")

    def test_decimal_y_formula(self):
        value = calcular_costo_total("3873858", "101509", 3, 3)
        self.assertIsInstance(value, Decimal)
        self.assertEqual(value, Decimal("4482912"))

    def test_id_deterministico(self):
        doc = {"periodo_aplicado": "202608", "origen": "1", "destino": "2", "configuracion": "3S3", "condicion_carga": "1", "tipo_carga": "General", "unidad_transporte": "Furgón", "rutasid": "9"}
        self.assertEqual(consulta_id(doc), consulta_id(dict(doc)))
        changed = dict(doc, rutasid="10")
        self.assertNotEqual(consulta_id(doc), consulta_id(changed))

    def test_configuracion_completa(self):
        validar_combinaciones()
        self.assertEqual(len(COMBINACIONES), 5)
        self.assertEqual({x["configuracion_codigo"] for x in COMBINACIONES}, {"3S3", "2", "2L3"})

    def test_periodo_anterior(self):
        self.assertEqual(periodo_anterior("202601"), "202512")

    def test_condicion_cargado_equivale_al_codigo_uno(self):
        combo = COMBINACIONES[0]
        document = {
            "origen": combo["origen_codigo"],
            "destino": combo["destino_codigo"],
            "configuracion": combo["configuracion_codigo"],
            "condicioncarga": "CARGADO",
            "nombretipocarga": combo["tipo_carga"],
            "nombreunidadtransporte": combo["unidad_transporte"],
        }
        self.assertTrue(_coincide(document, combo))

    def test_divipola_rndc_sin_cero_inicial_coincide(self):
        combo = dict(COMBINACIONES[0], origen_codigo="08296000")
        document = {
            "origen": "8296000", "destino": combo["destino_codigo"],
            "configuracion": combo["configuracion_codigo"],
        }
        self.assertTrue(_coincide(document, combo))

    def test_configuracion_liviana_descriptiva_de_rndc_coincide(self):
        combo = COMBINACIONES[0]
        document = {
            "origen": combo["origen_codigo"], "destino": combo["destino_codigo"],
            "configuracion": "2L3 Liviano entre 7.5 y 8 Tonel.",
        }
        self.assertTrue(_coincide(document, combo))


class FakeRepo:
    def comprobar(self): pass
    def upsert_many(self, docs): return len(docs), 0


class FakeClient:
    def __init__(self, answers, broad_answers=None):
        self.answers, self.broad_answers, self.periods, self.broad_periods = answers, broad_answers or [], [], []
    def consultar(self, periodo, combination):
        self.periods.append(periodo)
        answer = self.answers.pop(0) if self.answers else []
        if isinstance(answer, Exception): raise answer
        return answer
    def consultar_amplia(self, periodo, combination):
        self.broad_periods.append(periodo)
        answer = self.broad_answers.pop(0) if self.broad_answers else []
        if isinstance(answer, Exception): raise answer
        return answer


class ServiceTests(unittest.TestCase):
    @patch("sicetac.service.COMBINACIONES", [COMBINACIONES[0]])
    def test_rndc13_activa_consulta_amplia_y_filtra_localmente(self):
        combo = COMBINACIONES[0]
        document = {
            "periodo": "202608", "origen": combo["origen_codigo"], "destino": combo["destino_codigo"],
            "configuracion": combo["configuracion_codigo"], "condicioncarga": "CARGADO",
            "nombretipocarga": combo["tipo_carga"], "nombreunidadtransporte": combo["unidad_transporte"],
            "valormoviliza": "100", "valorhora": "10",
        }
        client = FakeClient([RNDCBusinessError("Error RNDC13")], [[document]])
        result = SicetacService(client, FakeRepo()).ejecutar("202608", dry_run=False)
        self.assertEqual(result["documentos_insertados"], 1)
        self.assertEqual(client.broad_periods, ["202608"])

    def test_retrocede_solo_sin_documentos(self):
        client = FakeClient([[], [{"periodo":"202607", "origen":"08296000", "destino":"76892000", "configuracion":"2L3", "condicioncarga":"1", "valormoviliza":"1", "valorhora":"1"}]])
        SicetacService(client, FakeRepo()).ejecutar("202608", dry_run=True)
        self.assertEqual(client.periods[:2], ["202608", "202607"])

    def test_no_retrocede_autenticacion(self):
        client = FakeClient([RNDCCredentialsError("rechazada")])
        with self.assertRaises(RNDCCredentialsError): SicetacService(client, FakeRepo()).ejecutar("202608", dry_run=True)
        self.assertEqual(client.periods, ["202608"])


if __name__ == "__main__": unittest.main()
