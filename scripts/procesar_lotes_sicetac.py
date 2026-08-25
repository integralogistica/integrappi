from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from sicetac.config import Settings
from sicetac.excel_service import (_ajustar_hoja, _formatear_columnas_numericas,
                                    procesar_excel, resumir_rutas)
from sicetac.rndc_client import RNDCClient


def consolidar(resultados: Path, destino: Path):
    archivos = sorted(resultados.glob("resultado_lote_*.xlsx"))
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "resultados"
    headers = None
    estados = Counter()
    total = 0
    reemplazos = {}
    for retry in sorted(resultados.glob("reintento_*.xlsx")):
        wr = load_workbook(retry, read_only=True, data_only=True)
        rr = wr.active.iter_rows(values_only=True)
        rh = list(next(rr))
        id_index = rh.index("consulta_id_usuario")
        estado_index = rh.index("estado")
        for row in rr:
            if row[estado_index] == "OK":
                reemplazos.setdefault(row[id_index], []).append(list(row))
    reemplazados = set()
    for archivo in archivos:
        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        current_headers = list(next(rows))
        if headers is None:
            headers = current_headers
            ws_out.append(headers)
        estado_index = current_headers.index("estado")
        id_index = current_headers.index("consulta_id_usuario")
        for row in rows:
            consulta_id = row[id_index]
            nuevas = reemplazos.get(consulta_id)
            if nuevas and consulta_id not in reemplazados:
                for nueva in nuevas:
                    ws_out.append(nueva)
                    estados[str(nueva[estado_index])] += 1
                    total += 1
                reemplazados.add(consulta_id)
            elif not nuevas:
                ws_out.append(list(row))
                estados[str(row[estado_index])] += 1
                total += 1
    _formatear_columnas_numericas(ws_out)
    _ajustar_hoja(ws_out)
    wb_out.save(destino)
    return total, estados


def ejecutar(lotes: Path, resultados: Path):
    resultados.mkdir(parents=True, exist_ok=True)
    archivos = sorted(lotes.glob("sicetac_lote_*.xlsx"))
    settings = Settings.from_env()
    client = RNDCClient(settings.soap_url, settings.username, settings.password)
    try:
        for index, archivo in enumerate(archivos, 1):
            destino = resultados / f"resultado_lote_{index:03d}.xlsx"
            if destino.exists():
                print(f"[{index}/{len(archivos)}] omitido (ya existe): {destino.name}", flush=True)
                continue
            print(f"[{index}/{len(archivos)}] procesando {archivo.name}", flush=True)
            output = procesar_excel(archivo.read_bytes(), client, resumir_rutas)
            destino.write_bytes(output.getvalue())
            print(f"[{index}/{len(archivos)}] guardado {destino.name}", flush=True)
    finally:
        client.close()
    consolidado = resultados / "resultados_sicetac_consolidado.xlsx"
    total, estados = consolidar(resultados, consolidado)
    print(f"Consolidado: {consolidado} | filas={total} | estados={dict(estados)}", flush=True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Procesa y consolida lotes SICE-TAC por SOAP")
    parser.add_argument("lotes", type=Path)
    parser.add_argument("resultados", type=Path)
    args = parser.parse_args()
    load_dotenv()
    ejecutar(args.lotes, args.resultados)
