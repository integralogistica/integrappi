from __future__ import annotations

import argparse
import math
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def ajustar(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, column in enumerate(ws.columns, 1):
        width = min(max((len(str(c.value or "")) for c in column), default=10) + 2, 45)
        ws.column_dimensions[get_column_letter(index)].width = width


def dividir(origen: Path, salida: Path, tamano: int):
    wb = load_workbook(origen, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    data = [list(row) for row in rows if any(v is not None and str(v).strip() for v in row)]
    headers_salida = ["consulta_id_usuario", "fila_original"] + headers

    salida.mkdir(parents=True, exist_ok=False)
    manifiesto = Workbook()
    wm = manifiesto.active
    wm.title = "lotes"
    wm.append(["lote", "archivo", "filas", "fila_original_desde", "fila_original_hasta"])

    for lote, start in enumerate(range(0, len(data), tamano), 1):
        chunk = data[start:start + tamano]
        nombre = f"sicetac_lote_{lote:03d}.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "consultas"
        sheet.append(headers_salida)
        for offset, row in enumerate(chunk):
            fila_original = start + offset + 2
            sheet.append([f"C{fila_original - 1:06d}", fila_original] + row)
        ajustar(sheet)
        book.save(salida / nombre)
        wm.append([lote, nombre, len(chunk), start + 2, start + len(chunk) + 1])

    ajustar(wm)
    manifiesto.save(salida / "manifiesto_lotes.xlsx")
    return len(data), math.ceil(len(data) / tamano)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Divide consultas SICE-TAC conservando trazabilidad")
    parser.add_argument("origen", type=Path)
    parser.add_argument("salida", type=Path)
    parser.add_argument("--tamano", type=int, default=30)
    args = parser.parse_args()
    filas, lotes = dividir(args.origen, args.salida, args.tamano)
    print(f"{filas} consultas divididas en {lotes} lotes de hasta {args.tamano} filas")
