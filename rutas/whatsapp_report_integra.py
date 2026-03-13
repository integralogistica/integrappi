import os
from datetime import datetime
from typing import Optional, Dict, Any, List

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse, JSONResponse
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

mongo_uri = os.getenv("MONGO_URI")
client = MongoClient(mongo_uri)
db = client["integra"]

ruta_whatsapp_report = APIRouter(prefix="/whatsapp-report", tags=["whatsapp-report"])

coleccion_uso = db["uso_whatsapp"]

MAPA_INICIAL = {
    "TRANSPORTADOR_MENU": "🚚 Transportador",
    "EMPLOYEE_MENU": "🧑‍💼 Empleado",
    "CLIENTE_MENU": "🧾 Cliente",
}

def _parse_date_yyyy_mm_dd(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    return datetime.fromisoformat(s)  # "2026-01-26"

@ruta_whatsapp_report.get("/resumen")
def resumen_uso(
    desde: Optional[str] = Query(None, description="Fecha inicio YYYY-MM-DD"),
    hasta: Optional[str] = Query(None, description="Fecha fin YYYY-MM-DD"),
):
    """
    Retorna un resumen JSON del uso del bot en el rango de fechas dado:
    - Total de números únicos
    - Desglose por rol (transportador / empleado / cliente)
    - Actividad diaria (mensajes entrantes por día)
    """
    dt_desde = _parse_date_yyyy_mm_dd(desde)
    dt_hasta = _parse_date_yyyy_mm_dd(hasta)

    date_filter: Dict[str, Any] = {}
    if dt_desde or dt_hasta:
        date_filter["created_at"] = {}
        if dt_desde:
            date_filter["created_at"]["$gte"] = dt_desde
        if dt_hasta:
            date_filter["created_at"]["$lte"] = datetime(
                dt_hasta.year, dt_hasta.month, dt_hasta.day, 23, 59, 59
            )

    # ── Pipeline: resumen por teléfono ──────────────────────────────────
    pipeline_phones = [
        {"$match": date_filter if date_filter else {}},
        {
            "$addFields": {
                "initial_choice": {
                    "$cond": [
                        {
                            "$and": [
                                {"$eq": ["$event", "STATE_CHANGED"]},
                                {"$in": ["$state", ["TRANSPORTADOR_MENU", "EMPLOYEE_MENU", "CLIENTE_MENU"]]},
                            ]
                        },
                        "$state",
                        None,
                    ]
                }
            }
        },
        {
            "$group": {
                "_id": "$phone",
                "cnt_transportador": {
                    "$sum": {"$cond": [{"$eq": ["$initial_choice", "TRANSPORTADOR_MENU"]}, 1, 0]}
                },
                "cnt_empleado": {
                    "$sum": {"$cond": [{"$eq": ["$initial_choice", "EMPLOYEE_MENU"]}, 1, 0]}
                },
                "cnt_cliente": {
                    "$sum": {"$cond": [{"$eq": ["$initial_choice", "CLIENTE_MENU"]}, 1, 0]}
                },
            }
        },
    ]

    phones_data: List[Dict[str, Any]] = list(coleccion_uso.aggregate(pipeline_phones))

    total_numeros = len(phones_data)

    # Un número se clasifica en el rol que más usó; si usó varios, se cuenta en todos
    numeros_transportador = sum(1 for p in phones_data if p["cnt_transportador"] > 0)
    numeros_empleado      = sum(1 for p in phones_data if p["cnt_empleado"] > 0)
    numeros_cliente       = sum(1 for p in phones_data if p["cnt_cliente"] > 0)
    numeros_sin_rol       = sum(
        1 for p in phones_data
        if p["cnt_transportador"] == 0 and p["cnt_empleado"] == 0 and p["cnt_cliente"] == 0
    )

    # ── Pipeline: actividad diaria (mensajes IN por día) ─────────────────
    pipeline_diario = [
        {
            "$match": {
                **( date_filter if date_filter else {}),
                "direction": "IN",
                "event": "MESSAGE_RECEIVED",
            }
        },
        {
            "$group": {
                "_id": {
                    "year":  {"$year": "$created_at"},
                    "month": {"$month": "$created_at"},
                    "day":   {"$dayOfMonth": "$created_at"},
                },
                "mensajes": {"$sum": 1},
                "numeros_unicos": {"$addToSet": "$phone"},
            }
        },
        {"$sort": {"_id.year": 1, "_id.month": 1, "_id.day": 1}},
        {
            "$project": {
                "_id": 0,
                "fecha": {
                    "$dateToString": {
                        "format": "%Y-%m-%d",
                        "date": {
                            "$dateFromParts": {
                                "year": "$_id.year",
                                "month": "$_id.month",
                                "day": "$_id.day",
                            }
                        },
                    }
                },
                "mensajes": 1,
                "numeros_unicos": {"$size": "$numeros_unicos"},
            }
        },
    ]

    actividad_diaria: List[Dict[str, Any]] = list(coleccion_uso.aggregate(pipeline_diario))

    return JSONResponse({
        "rango": {
            "desde": desde or "sin límite",
            "hasta": hasta or "sin límite",
        },
        "totales": {
            "numeros_unicos": total_numeros,
            "como_transportador": numeros_transportador,
            "como_empleado": numeros_empleado,
            "como_cliente": numeros_cliente,
            "sin_rol_definido": numeros_sin_rol,
        },
        "actividad_diaria": actividad_diaria,
    })


@ruta_whatsapp_report.get("/numeros-por-estado")
def numeros_por_estado(
    desde: Optional[str] = Query(None, description="Fecha inicio YYYY-MM-DD"),
    hasta: Optional[str] = Query(None, description="Fecha fin YYYY-MM-DD"),
):
    """
    Retorna, por día, cuántos números únicos consultaron cada estado (transportador / empleado / cliente).
    Ejemplo: { "2026-03-13": { "transportador": 3, "empleado": 2, "cliente": 3 } }
    """
    dt_desde = _parse_date_yyyy_mm_dd(desde)
    dt_hasta = _parse_date_yyyy_mm_dd(hasta)

    match: Dict[str, Any] = {
        "event": "STATE_CHANGED",
        "state": {"$in": ["TRANSPORTADOR_MENU", "EMPLOYEE_MENU", "CLIENTE_MENU"]},
    }
    if dt_desde or dt_hasta:
        match["created_at"] = {}
        if dt_desde:
            match["created_at"]["$gte"] = dt_desde
        if dt_hasta:
            match["created_at"]["$lte"] = datetime(
                dt_hasta.year, dt_hasta.month, dt_hasta.day, 23, 59, 59
            )

    pipeline = [
        {"$match": match},
        {
            "$group": {
                "_id": {
                    "year":  {"$year": "$created_at"},
                    "month": {"$month": "$created_at"},
                    "day":   {"$dayOfMonth": "$created_at"},
                    "state": "$state",
                },
                "numeros_unicos": {"$addToSet": "$phone"},
            }
        },
        {
            "$project": {
                "_id": 0,
                "fecha": {
                    "$dateToString": {
                        "format": "%Y-%m-%d",
                        "date": {
                            "$dateFromParts": {
                                "year":  "$_id.year",
                                "month": "$_id.month",
                                "day":   "$_id.day",
                            }
                        },
                    }
                },
                "estado": "$_id.state",
                "cantidad": {"$size": "$numeros_unicos"},
            }
        },
        {"$sort": {"fecha": 1, "estado": 1}},
    ]

    rows: List[Dict[str, Any]] = list(coleccion_uso.aggregate(pipeline))

    MAPA_ESTADO = {
        "TRANSPORTADOR_MENU": "transportador",
        "EMPLOYEE_MENU":      "empleado",
        "CLIENTE_MENU":       "cliente",
    }

    resultado: Dict[str, Dict[str, int]] = {}
    for row in rows:
        fecha  = row["fecha"]
        estado = MAPA_ESTADO.get(row["estado"], row["estado"])
        if fecha not in resultado:
            resultado[fecha] = {"transportador": 0, "empleado": 0, "cliente": 0}
        resultado[fecha][estado] = row["cantidad"]

    return JSONResponse({
        "rango": {
            "desde": desde or "sin límite",
            "hasta": hasta or "sin límite",
        },
        "por_dia": resultado,
    })


def _pipeline_numeros_por_estado(match: Dict[str, Any]):
    return [
        {"$match": match},
        {
            "$group": {
                "_id": {
                    "year":  {"$year": "$created_at"},
                    "month": {"$month": "$created_at"},
                    "day":   {"$dayOfMonth": "$created_at"},
                    "state": "$state",
                },
                "numeros_unicos": {"$addToSet": "$phone"},
            }
        },
        {
            "$project": {
                "_id": 0,
                "fecha": {
                    "$dateToString": {
                        "format": "%Y-%m-%d",
                        "date": {
                            "$dateFromParts": {
                                "year":  "$_id.year",
                                "month": "$_id.month",
                                "day":   "$_id.day",
                            }
                        },
                    }
                },
                "estado": "$_id.state",
                "cantidad": {"$size": "$numeros_unicos"},
            }
        },
        {"$sort": {"fecha": 1, "estado": 1}},
    ]


@ruta_whatsapp_report.get("/numeros-por-estado/descargar-excel")
def descargar_excel_numeros_por_estado(
    desde: Optional[str] = Query(None, description="YYYY-MM-DD"),
    hasta: Optional[str] = Query(None, description="YYYY-MM-DD"),
):
    """
    Descarga un Excel con: Fecha | Transportadores | Empleados | Clientes
    """
    dt_desde = _parse_date_yyyy_mm_dd(desde)
    dt_hasta = _parse_date_yyyy_mm_dd(hasta)

    match: Dict[str, Any] = {
        "event": "STATE_CHANGED",
        "state": {"$in": ["TRANSPORTADOR_MENU", "EMPLOYEE_MENU", "CLIENTE_MENU"]},
    }
    if dt_desde or dt_hasta:
        match["created_at"] = {}
        if dt_desde:
            match["created_at"]["$gte"] = dt_desde
        if dt_hasta:
            match["created_at"]["$lte"] = datetime(
                dt_hasta.year, dt_hasta.month, dt_hasta.day, 23, 59, 59
            )

    rows: List[Dict[str, Any]] = list(coleccion_uso.aggregate(_pipeline_numeros_por_estado(match)))

    MAPA_ESTADO = {
        "TRANSPORTADOR_MENU": "transportador",
        "EMPLOYEE_MENU":      "empleado",
        "CLIENTE_MENU":       "cliente",
    }

    # Agrupar por fecha
    por_dia: Dict[str, Dict[str, int]] = {}
    for row in rows:
        fecha  = row["fecha"]
        estado = MAPA_ESTADO.get(row["estado"], row["estado"])
        if fecha not in por_dia:
            por_dia[fecha] = {"transportador": 0, "empleado": 0, "cliente": 0}
        por_dia[fecha][estado] = row["cantidad"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Números por Estado"

    ws.append(["Fecha", "Transportadores", "Empleados", "Clientes"])
    for fecha in sorted(por_dia.keys()):
        d = por_dia[fecha]
        ws.append([fecha, d["transportador"], d["empleado"], d["cliente"]])

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = max((len(str(cell.value or "")) for cell in ws[col_letter]), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="numeros_por_estado.xlsx"'},
    )


@ruta_whatsapp_report.get("/descargar-excel")
def descargar_excel(
    desde: Optional[str] = Query(None, description="YYYY-MM-DD"),
    hasta: Optional[str] = Query(None, description="YYYY-MM-DD"),
    limit: int = Query(5000, ge=1, le=50000),
):
    """
    Descarga un Excel con:
    phone | total_in | primera_interaccion | ultima_interaccion | opcion_inicial_mas_comun
    """
    dt_desde = _parse_date_yyyy_mm_dd(desde)
    dt_hasta = _parse_date_yyyy_mm_dd(hasta)

    match: Dict[str, Any] = {}
    if dt_desde or dt_hasta:
        match["created_at"] = {}
        if dt_desde:
            match["created_at"]["$gte"] = dt_desde
        if dt_hasta:
            # hasta inclusive (si guardas horas, esto incluye el día completo)
            match["created_at"]["$lte"] = datetime(dt_hasta.year, dt_hasta.month, dt_hasta.day, 23, 59, 59)

    pipeline = [
        {"$match": match} if match else {"$match": {}},

        # Marcamos "opción inicial" cuando el evento STATE_CHANGED cae en reminder de menús principales.
        {
            "$addFields": {
                "initial_choice": {
                    "$cond": [
                        {
                            "$and": [
                                {"$eq": ["$event", "STATE_CHANGED"]},
                                {"$in": ["$state", ["TRANSPORTADOR_MENU", "EMPLOYEE_MENU", "CLIENTE_MENU"]]},
                            ]
                        },
                        "$state",
                        None,
                    ]
                }
            }
        },

        # Resumen por phone:
        {
            "$group": {
                "_id": "$phone",
                "primera_interaccion": {"$min": "$created_at"},
                "ultima_interaccion": {"$max": "$created_at"},

                # total_in: cuenta SOLO los mensajes entrantes recibidos
                "total_in": {
                    "$sum": {
                        "$cond": [
                            {"$and": [
                                {"$eq": ["$direction", "IN"]},
                                {"$eq": ["$event", "MESSAGE_RECEIVED"]},
                            ]},
                            1,
                            0,
                        ]
                    }
                },

                # Contadores por opción inicial (todas las veces que eligió cada una)
                "cnt_transportador": {
                    "$sum": {"$cond": [{"$eq": ["$initial_choice", "TRANSPORTADOR_MENU"]}, 1, 0]}
                },
                "cnt_empleado": {
                    "$sum": {"$cond": [{"$eq": ["$initial_choice", "EMPLOYEE_MENU"]}, 1, 0]}
                },
                "cnt_cliente": {
                    "$sum": {"$cond": [{"$eq": ["$initial_choice", "CLIENTE_MENU"]}, 1, 0]}
                },
            }
        },

        # Determinar la opción más común
        {
            "$addFields": {
                "opcion_inicial_mas_comun": {
                    "$switch": {
                        "branches": [
                            {
                                "case": {"$and": [
                                    {"$gte": ["$cnt_transportador", "$cnt_empleado"]},
                                    {"$gte": ["$cnt_transportador", "$cnt_cliente"]},
                                    {"$gt": ["$cnt_transportador", 0]},
                                ]},
                                "then": "TRANSPORTADOR_MENU",
                            },
                            {
                                "case": {"$and": [
                                    {"$gte": ["$cnt_empleado", "$cnt_transportador"]},
                                    {"$gte": ["$cnt_empleado", "$cnt_cliente"]},
                                    {"$gt": ["$cnt_empleado", 0]},
                                ]},
                                "then": "EMPLOYEE_MENU",
                            },
                            {
                                "case": {"$and": [
                                    {"$gte": ["$cnt_cliente", "$cnt_transportador"]},
                                    {"$gte": ["$cnt_cliente", "$cnt_empleado"]},
                                    {"$gt": ["$cnt_cliente", 0]},
                                ]},
                                "then": "CLIENTE_MENU",
                            },
                        ],
                        "default": "SIN_DATOS",
                    }
                }
            }
        },

        {"$sort": {"total_in": -1}},
        {"$limit": limit},

        {
            "$project": {
                "_id": 0,
                "phone": "$_id",
                "total_in": 1,
                "primera_interaccion": 1,
                "ultima_interaccion": 1,
                "opcion_inicial_mas_comun": 1,
                "cnt_transportador": 1,
                "cnt_empleado": 1,
                "cnt_cliente": 1,
            }
        },
    ]

    rows: List[Dict[str, Any]] = list(coleccion_uso.aggregate(pipeline))

    # ---- Crear Excel ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen WhatsApp"

    headers = [
        "Número WhatsApp",
        "Total interacciones (IN)",
        "Primera interacción",
        "Última interacción",
        "Opción inicial más común",
        "Veces Transportador",
        "Veces Empleado",
        "Veces Cliente",
    ]
    ws.append(headers)

    for r in rows:
        opcion = r.get("opcion_inicial_mas_comun", "SIN_DATOS")
        opcion_legible = MAPA_INICIAL.get(opcion, "Sin datos")

        ws.append([
            r.get("phone"),
            r.get("total_in", 0),
            r.get("primera_interaccion"),
            r.get("ultima_interaccion"),
            opcion_legible,
            r.get("cnt_transportador", 0),
            r.get("cnt_empleado", 0),
            r.get("cnt_cliente", 0),
        ])

    # Auto-ajustar anchos
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    # Guardar a memoria y responder
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = "reporte_whatsapp_integra.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
