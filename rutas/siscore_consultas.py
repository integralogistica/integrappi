from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
from typing import List, Optional
import httpx
import logging
from datetime import datetime, timedelta, timezone

# Colombia = UTC-5 (sin horario de verano). El servidor corre en UTC y Mongo guarda
# los datetime como instantes UTC, por lo que los filtros por "día" deben alinearse
# al día Colombia sumando estas 5 h a los límites (que Mongo interpreta como UTC).
_OFFSET_COLOMBIA = timedelta(hours=5)
from dotenv import load_dotenv
import os
import pandas as pd
from pymongo import MongoClient
from Funciones.whatsapp_utils_integra import enviar_template_sync

load_dotenv()

router = APIRouter(prefix="/siscore", tags=["Siscore"])
logger = logging.getLogger(__name__)

# ── Hora confiable desde internet ────────────────────────────────────────────
# Hay PCs/servidores con el reloj desfasado. Para registrar fechas exactas (envío,
# aprobación, etc.) se obtiene la hora UTC real del header 'Date' de servidores
# confiables, cacheando el offset frente al reloj local (re-sync cada 10 min) y
# cayendo al reloj local si no hay internet. El sistema guarda fechas como UTC
# naive; la zona Colombia (UTC-5) se aplica al mostrarlas.
_HORA_CACHE = {'offset': None, 'ultima_sync': None}  # offset = hora_internet_utc - hora_local
_SERVIDORES_HORA = [
    'https://www.google.com/generate_204',
    'https://www.gstatic.com/generate_204',
    'https://cloudflare.com/cdn-cgi/trace',
]


def _leer_hora_internet_utc() -> Optional[datetime]:
    """Lee la hora GMT/UTC exacta del header 'Date' de un servidor confiable (o None)."""
    from email.utils import parsedate_to_datetime
    for url in _SERVIDORES_HORA:
        try:
            r = httpx.get(url, timeout=2.5, follow_redirects=True)
            date_header = r.headers.get('date')
            if date_header:
                dt = parsedate_to_datetime(date_header)  # aware (GMT)
                return dt.astimezone(timezone.utc).replace(tzinfo=None)  # naive UTC
        except Exception:
            continue
    return None


def _hora_confiable_utc() -> datetime:
    """Hora UTC actual confiable (internet con cache del offset; fallback reloj local)."""
    ahora_local = datetime.now()
    cache = _HORA_CACHE
    necesita_sync = (
        cache['offset'] is None
        or cache['ultima_sync'] is None
        or (ahora_local - cache['ultima_sync']).total_seconds() > 600
    )
    if necesita_sync:
        hora_internet = _leer_hora_internet_utc()
        if hora_internet is not None:
            cache['offset'] = hora_internet - datetime.now()
            cache['ultima_sync'] = datetime.now()
    if cache['offset'] is not None:
        return datetime.now() + cache['offset']
    return datetime.now()

# Conexión MongoDB
MONGO_URI = os.environ.get("MONGO_URI", "mongodb://localhost:27017")
client = MongoClient(MONGO_URI)
db = client["integra"]
coleccion_solicitudes = db["solicitud_veh_medical"]
coleccion_tramites = db["tramite_fmc"]
coleccion_tarifas = db["fletes_rutas_fmc"]
coleccion_divipolas = db["divipolas"]
coleccion_pedidos_medical = db["pedidos_medical"]
coleccion_historico = db["pedidos_medical_historico"]
coleccion_causales = db["causales"]
coleccion_baseusuarios = db["baseusuarios"]

# Índice sobre `consecutivo` en ambas colecciones. La generación de consecutivos
# consulta por prefijo (regional+fecha) y, tras Importar Vulcano, también consulta
# el histórico. SIN este índice, al crecer el histórico a miles/millones la consulta
# se vuelve lenta (collation scan). CON índice, el regex con ancla '^' usa el índice
# (barrido por prefijo) y queda rápido sin importar el tamaño. Es idempotente.
for _col_idx in (coleccion_pedidos_medical, coleccion_historico):
    try:
        _col_idx.create_index([("consecutivo", 1)])
    except Exception as _e:
        logger.warning(f"[STARTUP] No se pudo crear/verificar índice 'consecutivo': {_e}")

# Índice sobre fusion_info.datos_originales.consecutivo (solo docs fusionados).
# Importar Vulcano resuelve consecutivos que viven dentro de los originales de una
# fusión; sin este índice esa consulta es un collscan por cada fila del Excel.
# partialFilterExpression limita el índice a documentos fusionados (más chico). Idempotente.
try:
    coleccion_pedidos_medical.create_index(
        [("fusion_info.datos_originales.consecutivo", 1)],
        name="idx_datos_originales_consecutivo",
        partialFilterExpression={"fusion_info.es_fusionada": True},
    )
except Exception as _e:
    logger.warning(f"[STARTUP] No se pudo crear/verificar índice 'datos_originales.consecutivo': {_e}")

# Mapa regional -> municipio de la bodega de origen (para guardar en el campo
# `regional` de Mongo cuando el perfil es OPERATIVO).
REGIONAL_A_ORIGEN_BODEGA = {
    "BARRANQUILLA": "GALAPA",
    "CALI": "YUMBO",
    "MEDELLIN": "GIRARDOTA",
}


def regional_a_origen_bodega(regional: Optional[str]) -> Optional[str]:
    """Convierte la regional al municipio de la bodega de origen.
    BARRANQUILLA->GALAPA, CALI->YUMBO, MEDELLIN->GIRARDOTA. Las demás se conservan."""
    if not regional:
        return regional
    r = str(regional).upper().strip()
    return REGIONAL_A_ORIGEN_BODEGA.get(r, r)


# Mapeo inverso: bodega -> nombre de la regional (para filtros que reciben la bodega).
BODEGA_A_REGIONAL = {v: k for k, v in REGIONAL_A_ORIGEN_BODEGA.items()}


def _aplicar_filtro_regional_dropdown(filtro: dict, valor: str) -> None:
    """
    Filtra por una regional elegida en un dropdown de la UI (formato bodega:
    GALAPA/YUMBO/GIRARDOTA/BUCARAMANGA/FUNZA). Cubre todas las formas en que
    `regional`/`centro_costo` pudo quedar guardado (bodega, nombre de regional o código
    CEDI), porque conviven docs viejos y nuevos. También acepta el nombre de regional
    directamente (CALI/BARRANQUILLA/...).
    """
    v = (valor or "").upper().strip()
    if not v:
        return
    # Normalizar a nombre de regional; _aplicar_filtro_regional_operativo arma el $or.
    regional_nombre = BODEGA_A_REGIONAL.get(v, v)
    _aplicar_filtro_regional_operativo(filtro, regional_nombre)


def _aplicar_filtro_regional_operativo(filtro: dict, centro_distribucion: str) -> None:
    """
    Agrega a `filtro` (dict de consulta Mongo) las condiciones para que un perfil
    OPERATIVO vea solo las planillas de su regional.

    Cubre las distintas formas en que la regional queda almacenada, porque
    `centro_costo` suele guardarse como 'FMC' (no como código de bodega):
      - centro_costo = código de bodega (CO05)         -> cuando Siscore sí lo trae
      - regional = nombre de bodega (YUMBO/GALAPA/...) -> como se guarda para OPERATIVO
      - regional = nombre de la regional (CALI)         -> si se guardó sin conversión
    """
    cd = (centro_distribucion or "").upper().strip()
    if not cd:
        return
    regional_map = {
        "BARRANQUILLA": "CO04", "CALI": "CO05", "BUCARAMANGA": "CO06",
        "FUNZA": "CO07", "MEDELLIN": "CO09",
    }
    bodega_codigo = regional_map.get(cd, "")
    bodega_nombre = regional_a_origen_bodega(cd) or cd

    condiciones = []
    if bodega_codigo:
        condiciones.append({"centro_costo": bodega_codigo})
    if bodega_nombre and bodega_nombre != cd:
        condiciones.append({"regional": bodega_nombre})
    condiciones.append({"regional": cd})

    if len(condiciones) == 1:
        filtro.update(condiciones[0])
    else:
        filtro["$or"] = condiciones


# ----------------------------------------------------------------------------
# Notificación WhatsApp: aviso al operativo de que se creó su pedido.
# Plantilla "confirmacion_pedido_creado" (Meta, es_CO), texto:
#   Hola {{1}}, te confirmamos que el pedido {{2}} fue creado exitosamente
#   para la planilla {{3}}. Ya puedes continuar con el trámite en Integrapp.
# Mapeo de variables:
#   {{1}} = nombre del operativo que montó la planilla
#   {{2}} = número de pedido (pedido_vulcano)
#   {{3}} = número de planilla
# ----------------------------------------------------------------------------
PLANTILLA_PEDIDO_CREADO_NOMBRE = "confirmacion_pedido_creado"
PLANTILLA_PEDIDO_CREADO_IDIOMA = "es_CO"

# Notificación WhatsApp: aviso a analistas cuando planilla cambia de estado CREADO a visible.
# Plantilla "notificacion_analistas_pedidos" (Meta, es_CO), texto:
#   Hola {{1}}, la planilla {{2}} ha sido creada con el consecutivo {{3}}.
#   Regional: {{4}}
#   Creado por: {{5}}
#   Fecha: {{6}}
#   Por favor tramitala en *integrApp*.
# Mapeo de variables:
#   {{1}} = nombre del analista (o "Equipo de Analistas")
#   {{2}} = número de planilla
#   {{3}} = consecutivo completo
#   {{4}} = regional
#   {{5}} = usuario que creó la planilla
#   {{6}} = fecha de creación (formato dd/MM/yyyy)
# ----------------------------------------------------------------------------
PLANTILLA_ANALISTAS_PEDIDOS_NOMBRE = "notificacion_analistas_pedidos"
PLANTILLA_ANALISTAS_PEDIDOS_IDIOMA = "es_CO"

# Notificación WhatsApp: solicitud de autorización a coordinadores/control.
# Plantilla "solicitud_autorizacion" (Meta, es_CO), texto:
#   Hola {{1}}
#   La planilla {{2}} requiere tu autorización, por favor tramitala en *integrApp*.
# Mapeo de variables:
#   {{1}} = nombre del destinatario (coordinador/control)
#   {{2}} = número de planilla o consecutivo
# ----------------------------------------------------------------------------
PLANTILLA_SOLICITUD_AUTORIZACION_NOMBRE = "solicitud_autorizacion"
PLANTILLA_SOLICITUD_AUTORIZACION_IDIOMA = "es_CO"


def _normalizar_celular_co(celular: Optional[str]) -> Optional[str]:
    """Convierte un celular a formato internacional Colombia: solo dígitos con prefijo 57."""
    if not celular:
        return None
    limpio = "".join(c for c in str(celular) if c.isdigit())
    if not limpio:
        return None
    if not limpio.startswith("57"):
        limpio = "57" + limpio
    return limpio


def _notificar_pedido_creado_whatsapp(doc: dict, pedido: str):
    """
    Envía WhatsApp al operativo que registró la planilla (usuario_registro) para
    avisar que se creó el pedido. Resuelve celular y nombre en baseusuarios.
    Es a prueba de fallos: si algo falla solo queda en log (no rompe la asignación).
    """
    try:
        usuario_registro = (doc.get("usuario_registro") or "").strip()
        if not usuario_registro:
            logger.info("[NOTIF PEDIDO] Doc sin usuario_registro; no se notifica.")
            return

        usuario_doc = coleccion_baseusuarios.find_one({"usuario": usuario_registro.upper()})
        if not usuario_doc:
            logger.info(f"[NOTIF PEDIDO] Usuario '{usuario_registro}' no está en baseusuarios; no se notifica.")
            return

        celular = _normalizar_celular_co(usuario_doc.get("celular"))
        if not celular:
            logger.info(f"[NOTIF PEDIDO] Usuario '{usuario_registro}' sin celular válido; no se notifica.")
            return

        nombre = (usuario_doc.get("nombre") or usuario_registro).strip()
        planilla = doc.get("planilla") or doc.get("consecutivo") or ""

        res = enviar_template_sync(
            to=celular,
            template_name=PLANTILLA_PEDIDO_CREADO_NOMBRE,
            language_code=PLANTILLA_PEDIDO_CREADO_IDIOMA,
            body_params=[nombre, str(pedido), str(planilla)],
        )
        if res:
            logger.info(f"[NOTIF PEDIDO] WhatsApp OK -> {celular} ({nombre}) | pedido={pedido} planilla={planilla}")
        else:
            logger.warning(f"[NOTIF PEDIDO] WhatsApp NO enviado a {celular} ({nombre}) — revisar tokens o idioma de la plantilla '{PLANTILLA_PEDIDO_CREADO_NOMBRE}'.")
    except Exception as e:
        logger.error(f"[NOTIF PEDIDO] Error inesperado: {e}")


def _notificar_analistas_cambio_estado(doc: dict, estado_anterior: str, estado_nuevo: str):
    """
    Envía WhatsApp a todos los usuarios con perfil ANALISTA para avisar que una planilla
    cambió de estado CREADO a PREAPROBADO (u otro estado visible).

    Usa la plantilla "notificacion_analistas_pedidos" (Meta, es_CO).
    Es a prueba de fallos: si algo falla solo queda en log (no rompe el cambio de estado).

    Variables de la plantilla "notificacion_analistas_pedidos":
      {{1}} = nombre del analista
      {{2}} = número de planilla
      {{3}} = consecutivo completo
      {{4}} = regional
      {{5}} = usuario que creó la planilla
      {{6}} = fecha de creación (dd/MM/yyyy)
    """
    try:
        # Solo notificar si el cambio fue de CREADO a otro estado visible
        if estado_anterior != "CREADO":
            logger.info(f"[NOTIF ANALISTAS] No se notifica: cambio de {estado_anterior} a {estado_nuevo} (no es CREADO→visible)")
            return

        logger.info(f"[NOTIF ANALISTAS] Planilla pasó de CREADO a {estado_nuevo}, notificando analistas...")

        # Buscar todos los usuarios con perfil ANALISTA
        analistas = list(coleccion_baseusuarios.find({"perfil": "ANALISTA"}))

        if not analistas:
            logger.info("[NOTIF ANALISTAS] No hay usuarios con perfil ANALISTA en baseusuarios; no se notifica.")
            return

        planilla = doc.get("planilla") or ""
        consecutivo = doc.get("consecutivo") or planilla
        estado_legible = estado_nuevo.replace("_", " ")
        regional = doc.get("regional") or doc.get("centro_distribucion") or "N/A"
        usuario_registro = doc.get("usuario_registro") or "N/A"

        # Formatear fecha actual para la notificación
        from datetime import datetime
        fecha_actual = datetime.now().strftime("%d/%m/%Y")

        notificaciones_enviadas = 0
        errores = 0

        for analista in analistas:
            try:
                celular = _normalizar_celular_co(analista.get("celular"))
                if not celular:
                    logger.info(f"[NOTIF ANALISTAS] Analista {analista.get('usuario')} sin celular válido; se saltea.")
                    continue

                # Obtener nombre del analista
                nombre_analista = (analista.get("nombre") or analista.get("usuario") or "Analista").strip()

                # Usar la nueva plantilla específica para analistas
                # {{1}} = nombre analista, {{2}} = planilla, {{3}} = consecutivo
                # {{4}} = regional, {{5}} = usuario registro, {{6}} = fecha
                res = enviar_template_sync(
                    to=celular,
                    template_name=PLANTILLA_ANALISTAS_PEDIDOS_NOMBRE,
                    language_code=PLANTILLA_ANALISTAS_PEDIDOS_IDIOMA,
                    body_params=[
                        nombre_analista,
                        str(planilla),
                        str(consecutivo),
                        regional,
                        usuario_registro,
                        fecha_actual
                    ],
                )

                if res:
                    notificaciones_enviadas += 1
                    logger.info(f"[NOTIF ANALISTAS] WhatsApp OK -> {celular} ({nombre_analista}) | planilla={consecutivo} estado={estado_legible}")
                else:
                    logger.warning(f"[NOTIF ANALISTAS] WhatsApp NO enviado a {celular} ({nombre_analista})")
                    errores += 1

            except Exception as e:
                logger.error(f"[NOTIF ANALISTAS] Error al notificar a {analista.get('usuario')}: {e}")
                errores += 1

        logger.info(f"[NOTIF ANALISTAS] Resumen: {notificaciones_enviadas} enviadas, {errores} errores, {len(analistas)} analistas totales")

    except Exception as e:
        logger.error(f"[NOTIF ANALISTAS] Error en la función principal: {e}")


def _notificar_solicitud_autorizacion(doc: dict, estado_nuevo: str):
    """
    Envía WhatsApp a los perfiles correspondientes (COORDINADOR o CONTROL) cuando una planilla
    requiere su autorización después de pasar de CREADO a estados que requieren aprobación.

    Lógica de notificación:
    - Si estado_nuevo == "REQUIERE_APROBACION_COORDINADOR" → notificar solo a COORDINADOR
    - Si estado_nuevo == "REQUIERE_APROBACION_CONTROL" → notificar solo a CONTROL

    Usa la plantilla "solicitud_autorizacion" (Meta, es_CO).
    Es a prueba de fallos: si algo falla solo queda en log (no rompe el cambio de estado).

    Variables de la plantilla "solicitud_autorizacion":
      {{1}} = nombre del destinatario (coordinador/control)
      {{2}} = número de planilla o consecutivo
    """
    try:
        # Determinar perfil objetivo según el estado
        if estado_nuevo == "REQUIERE_APROBACION_COORDINADOR":
            perfil_objetivo = "COORDINADOR"
            logger.info(f"[NOTIF AUTORIZACIÓN] Planilla requiere aprobación COORDINADOR, notificando...")
        elif estado_nuevo == "REQUIERE_APROBACION_CONTROL":
            perfil_objetivo = "CONTROL"
            logger.info(f"[NOTIF AUTORIZACIÓN] Planilla requiere aprobación CONTROL, notificando...")
        else:
            logger.info(f"[NOTIF AUTORIZACIÓN] No se notifica: estado {estado_nuevo} no requiere autorización específica")
            return

        # Buscar usuarios con el perfil objetivo
        usuarios_perfil = list(coleccion_baseusuarios.find({"perfil": perfil_objetivo}))

        if not usuarios_perfil:
            logger.info(f"[NOTIF AUTORIZACIÓN] No hay usuarios con perfil {perfil_objetivo} en baseusuarios; no se notifica.")
            return

        planilla = doc.get("planilla") or ""
        consecutivo = doc.get("consecutivo") or planilla

        notificaciones_enviadas = 0
        errores = 0

        for usuario in usuarios_perfil:
            try:
                celular = _normalizar_celular_co(usuario.get("celular"))
                if not celular:
                    logger.info(f"[NOTIF AUTORIZACIÓN] {perfil_objetivo} {usuario.get('usuario')} sin celular válido; se saltea.")
                    continue

                # Obtener nombre del usuario
                nombre_usuario = (usuario.get("nombre") or usuario.get("usuario") or "Usuario").strip()

                # Usar la plantilla de solicitud de autorización
                # {{1}} = nombre usuario, {{2}} = planilla/consecutivo
                res = enviar_template_sync(
                    to=celular,
                    template_name=PLANTILLA_SOLICITUD_AUTORIZACION_NOMBRE,
                    language_code=PLANTILLA_SOLICITUD_AUTORIZACION_IDIOMA,
                    body_params=[nombre_usuario, str(consecutivo)],
                )

                if res:
                    notificaciones_enviadas += 1
                    logger.info(f"[NOTIF AUTORIZACIÓN] WhatsApp OK -> {celular} ({nombre_usuario}) | planilla={consecutivo} perfil={perfil_objetivo}")
                else:
                    logger.warning(f"[NOTIF AUTORIZACIÓN] WhatsApp NO enviado a {celular} ({nombre_usuario})")
                    errores += 1

            except Exception as e:
                logger.error(f"[NOTIF AUTORIZACIÓN] Error al notificar a {usuario.get('usuario')}: {e}")
                errores += 1

        logger.info(f"[NOTIF AUTORIZACIÓN] Resumen: {notificaciones_enviadas} enviadas, {errores} errores, {len(usuarios_perfil)} {perfil_objetivo} totales")

    except Exception as e:
        logger.error(f"[NOTIF AUTORIZACIÓN] Error en la función principal: {e}")


# NIT por nombre de cliente para el Excel de autorizados (campo "Cliente").
CLIENTE_A_NIT = {
    "FRESENIUS MEDICAL CARE": "901689684",
    "FRESENIUS KABI": "900402080",
    "DAVITA": "901689684",
    "CONGRUPO": "800146643",
}

# Configuración WS Siscore V3 (misma que en pedidos_v3)
SISCORE_V3_ENDPOINT = "https://integra-wms.appsiscore.com/app/ws/informe_v3.php"
SISCORE_V3_TOKEN = "n0ML0cFGhJwtq4lsAeUcMzrqkn94gX4TDaPuFbbXpoA"


class ConsultaPlanillasRequest(BaseModel):
    planillas: List[str]
    fecha_inicio: str
    fecha_fin: str
    perfil: Optional[str] = None
    centro_distribucion: Optional[str] = None


class VerificarFusionadasRequest(BaseModel):
    """Modelo para verificar planillas fusionadas (no requiere fechas)"""
    planillas: List[str]
    perfil: Optional[str] = None
    centro_distribucion: Optional[str] = None


class GuardarSolicitudRequest(BaseModel):
    usuario: str
    perfil: str
    centro_distribucion: str
    planilla: str
    piezas: int
    peso_real: float
    ruta: str
    codigos_pedido: str
    cantidad_pedidos: int
    cliente_origen: str
    municipio_destino: str
    departamento_destino: str
    regional: Optional[str] = None
    tarifa_calculada: float
    tipo_vehiculo: str
    total_solicitado: float
    tarifa_base: Optional[float] = None
    requiere_descargue: str = "NO"
    punto_adicional: bool = False
    desvio: bool = False
    aforo: Optional[float] = None
    placa: Optional[str] = None
    tipo_veh_sicetac: Optional[str] = None
    solicitud_id: Optional[str] = None


class ActualizarSolicitudRequest(BaseModel):
    usuario: str
    perfil: str
    centro_distribucion: str
    planilla: str
    piezas: int
    peso_real: float
    ruta: str
    codigos_pedido: str
    cantidad_pedidos: int
    cliente_origen: str
    municipio_destino: str
    departamento_destino: str
    regional: Optional[str] = None
    tarifa_calculada: float
    tipo_vehiculo: str
    total_solicitado: float
    tarifa_base: Optional[float] = None
    requiere_descargue: str = "NO"
    punto_adicional: bool = False
    desvio: bool = False
    aforo: Optional[float] = None
    placa: Optional[str] = None
    tipo_veh_sicetac: Optional[str] = None
    solicitud_id: str


class EnviarTramiteRequest(BaseModel):
    solicitud_id: str
    usuario: str


class ActualizarPlanillaPedidosRequest(BaseModel):
    """Modelo para actualizar una planilla en pedidos_medical"""
    planilla: str
    tarifa_base: Optional[float] = None
    tarifa_calculada: Optional[float] = None
    requiere_descargue: Optional[float] = 0  # Valor numérico del descargue
    punto_adicional: Optional[float] = 0     # Valor numérico del punto adicional
    desvio: Optional[float] = 0              # Valor numérico del desvío
    aforo: Optional[float] = None            # Valor numérico del aforo
    placa: Optional[str] = None
    tipo_veh_sicetac: Optional[str] = None
    peso_sicetac: Optional[float] = None  # Peso operacional editable (default = peso_real); se usa para trámites
    ruta: Optional[str] = None  # Ruta asignada/editada manualmente
    total_solicitado: float
    causal: Optional[str] = None
    estado: Optional[str] = None  # 'PREAPROBADO', 'REQUIERE_APROBACION_COORDINADOR', 'REQUIERE_APROBACION_CONTROL' o 'APROBADO'
    aprobado_por: Optional[str] = None
    fecha_aprobacion: Optional[str] = None
    municipio_destino: Optional[str] = None  # Municipio principal elegido manualmente
    usuario_modificacion: str  # Usuario que está editando (trazabilidad)


class ActualizarEstadoPlanillaRequest(BaseModel):
    """Modelo para actualizar el estado de aprobación de una planilla"""
    planilla: str
    estado: str  # 'PREAPROBADO', 'REQUIERE_APROBACION_COORDINADOR', 'REQUIERE_APROBACION_CONTROL' o 'APROBADO'
    aprobado_por: str  # Usuario que aprueba


# ============= ENDPOINT IMPORTAR VULCANO =============

@router.post("/importar-vulcano")
async def importar_vulcano(archivo: UploadFile = File(...)):
    """
    Importa pedidos Vulcano desde un Excel con columnas CONSECUTIVO y PEDIDO.
    Busca planillas en pedidos_medical por consecutivo, agrega pedido_vulcano,
    mueve el documento a pedidos_medical_historico y lo elimina de pedidos_medical.
    Solo accesible para ADMIN y ANALISTA (validar en frontend).
    """
    try:
        logger.info(f"=== IMPORTAR VULCANO ===")
        logger.info(f"Archivo recibido: {archivo.filename}")

        nombre_archivo = archivo.filename.lower()

        if nombre_archivo.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(
                archivo.file,
                engine='openpyxl' if nombre_archivo.endswith('.xlsx') else 'xlrd',
                dtype=str
            )
        else:
            raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx, .xls)")

        logger.info(f"Filas leídas: {len(df)}")

        # Normalizar columnas
        df.columns = [col.strip().upper().replace(" ", "_") for col in df.columns]
        logger.info(f"Columnas normalizadas: {list(df.columns)}")

        # Mapear columnas alternativas
        if "NO._PEDIDO" in df.columns and "PEDIDO" not in df.columns:
            df.rename(columns={"NO._PEDIDO": "PEDIDO"}, inplace=True)

        # Validar columnas requeridas
        columnas_requeridas = {"CONSECUTIVO", "PEDIDO"}
        if not columnas_requeridas.issubset(df.columns):
            faltantes = columnas_requeridas - set(df.columns)
            raise HTTPException(
                status_code=400,
                detail=f"El archivo debe tener las columnas: CONSECUTIVO, PEDIDO (o No. Pedido). Faltan: {', '.join(sorted(faltantes))}"
            )

        # Limpiar datos
        df["CONSECUTIVO"] = df["CONSECUTIVO"].astype(str).str.strip()
        df["PEDIDO"] = df["PEDIDO"].astype(str).str.strip()

        exitosos = 0
        no_encontrados = 0
        errores = 0
        detalles_no_encontrados = []
        # Soporte para fusiones: un pedido puede corresponder a una planilla original
        # embebida en fusion_info.datos_originales. La fusión pasa al histórico SOLO
        # cuando TODOS sus originales tienen pedido_vulcano.
        asignados_parciales = 0    # pedidos guardados en originales cuya fusión sigue activa
        fusiones_movidas = 0       # fusiones que pasaron al histórico
        parciales_por_fusion = {}  # {fusion_id: n} para descontar al completar una fusión

        for _, row in df.iterrows():
            consecutivo = row["CONSECUTIVO"]
            pedido = row["PEDIDO"]

            if not consecutivo or consecutivo == 'nan' or consecutivo == '':
                continue

            try:
                # --- Búsqueda en cascada ---
                # 1) Doc por consecutivo raíz (planilla normal, carro dividido o la fusión raíz).
                doc = coleccion_pedidos_medical.find_one({"consecutivo": consecutivo})
                es_no_fusion = bool(doc) and not (doc.get("fusion_info") or {}).get("es_fusionada")

                # 2) Si no es un doc no-fusionado, buscar dentro de los originales de una fusión.
                doc_fusion = None
                if not es_no_fusion:
                    doc_fusion = coleccion_pedidos_medical.find_one({
                        "fusion_info.es_fusionada": True,
                        "fusion_info.datos_originales.consecutivo": consecutivo,
                    })

                if doc_fusion:
                    # --- Ruta de fusión: asignar el pedido a un original ---
                    originales = doc_fusion.get("fusion_info", {}).get("datos_originales", []) or []
                    idx = next((i for i, o in enumerate(originales) if o.get("consecutivo") == consecutivo), None)
                    if idx is None:
                        no_encontrados += 1
                        detalles_no_encontrados.append(consecutivo)
                        logger.warning(f"[VULCANO-FUSION] Original no encontrado tras match (C={consecutivo})")
                        continue

                    # Asignación atómica al original (idempotente: sobreescribe, no duplica).
                    coleccion_pedidos_medical.update_one(
                        {"_id": doc_fusion["_id"], "fusion_info.datos_originales.consecutivo": consecutivo},
                        {"$set": {
                            "fusion_info.datos_originales.$[elem].pedido_vulcano": pedido,
                            "fusion_info.datos_originales.$[elem].fecha_pedido_vulcano": datetime.now(),
                        }},
                        array_filters=[{"elem.consecutivo": consecutivo}],
                    )

                    # Re-leer y evaluar completitud.
                    doc_actualizado = coleccion_pedidos_medical.find_one({"_id": doc_fusion["_id"]})
                    if doc_actualizado is None:
                        # Race condition: la fusión ya no existe (procesada por otro lado).
                        exitosos += 1
                        continue

                    originales_act = doc_actualizado.get("fusion_info", {}).get("datos_originales", []) or []
                    todos_con_pedido = bool(originales_act) and all(
                        o.get("consecutivo") and o.get("pedido_vulcano") for o in originales_act
                    )

                    if not todos_con_pedido:
                        # Éxito parcial: la fusión sigue activa esperando el resto de pedidos.
                        exitosos += 1
                        asignados_parciales += 1
                        parciales_por_fusion[doc_actualizado["_id"]] = \
                            parciales_por_fusion.get(doc_actualizado["_id"], 0) + 1
                        logger.info(
                            f"[VULCANO-FUSION] Pedido {pedido} asignado al original {consecutivo} "
                            f"(fusión {doc_actualizado.get('consecutivo')}) — pendiente completar"
                        )
                        continue

                    # Fusión completa: mover al histórico con todos sus pedidos.
                    doc_actualizado["fecha_movimiento_historico"] = datetime.now()
                    coleccion_historico.insert_one(doc_actualizado)
                    coleccion_pedidos_medical.delete_one({"_id": doc_actualizado["_id"]})

                    # Notificar por cada original (usuario_registro viene de la fusión, no del original).
                    usuario_registro_fusion = (doc_actualizado.get("usuario_registro") or "").strip()
                    for o in originales_act:
                        _notificar_pedido_creado_whatsapp(
                            {
                                "usuario_registro": usuario_registro_fusion,
                                "planilla": o.get("planilla"),
                                "consecutivo": o.get("consecutivo"),
                            },
                            o.get("pedido_vulcano"),
                        )

                    exitosos += 1
                    fusiones_movidas += 1
                    # Estos parciales ya completaron su fusión: dejar de contarlos como pendientes.
                    asignados_parciales -= parciales_por_fusion.pop(doc_actualizado["_id"], 0)
                    logger.info(
                        f"[VULCANO-FUSION] Fusión {doc_actualizado.get('consecutivo')} completada "
                        f"({len(originales_act)} originales) → histórico"
                    )

                elif doc:
                    # --- Doc no-fusionado: flujo histórico (mover al histórico) ---
                    doc["pedido_vulcano"] = pedido
                    doc["fecha_movimiento_historico"] = datetime.now()
                    coleccion_historico.insert_one(doc)
                    coleccion_pedidos_medical.delete_one({"_id": doc["_id"]})
                    exitosos += 1
                    _notificar_pedido_creado_whatsapp(doc, pedido)
                    logger.info(f"Planilla movida a historico: consecutivo={consecutivo}, pedido_vulcano={pedido}")

                else:
                    no_encontrados += 1
                    detalles_no_encontrados.append(consecutivo)
                    logger.warning(f"Consecutivo no encontrado: {consecutivo}")
                    continue

            except Exception as e:
                errores += 1
                logger.error(f"Error procesando consecutivo {consecutivo}: {str(e)}")

        logger.info(
            f"Importación Vulcano finalizada: {exitosos} exitosos ({asignados_parciales} parciales), "
            f"{fusiones_movidas} fusiones movidas, {no_encontrados} no encontrados, {errores} errores"
        )

        resultado = {
            "mensaje": f"Importación completada. {exitosos} pedido(s) asignado(s), {fusiones_movidas} fusión(es) movida(s) a histórico.",
            "exitosos": exitosos,
            "asignados_parciales": asignados_parciales,
            "fusiones_movidas": fusiones_movidas,
            "no_encontrados": no_encontrados,
            "errores": errores,
        }

        if detalles_no_encontrados:
            resultado["consecutivos_no_encontrados"] = detalles_no_encontrados[:20]

        return resultado

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al importar Vulcano: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo: {str(e)}")


class AsignarPedidoManualRequest(BaseModel):
    consecutivo: str
    pedido: str
    usuario: Optional[str] = None


@router.post("/asignar-pedido-manual")
async def asignar_pedido_manual(request: AsignarPedidoManualRequest):
    """
    Asigna manualmente el número de pedido Vulcano a una planilla (por consecutivo).
    Replica exactamente lo que hace 'importar-vulcano' para una sola fila:
    agrega pedido_vulcano, mueve el documento a pedidos_medical_historico y lo
    elimina de pedidos_medical. Pensado para el botón por planilla (ADMIN/ANALISTA).
    """
    try:
        consecutivo = (request.consecutivo or "").strip()
        pedido = (request.pedido or "").strip()

        if not consecutivo or not pedido:
            raise HTTPException(status_code=400, detail="consecutivo y pedido son obligatorios")

        logger.info(f"=== ASIGNAR PEDIDO MANUAL === consecutivo={consecutivo}, pedido={pedido}, usuario={request.usuario}")

        # Buscar planilla por consecutivo (igual que el import del Excel)
        doc = coleccion_pedidos_medical.find_one({"consecutivo": consecutivo})
        if not doc:
            raise HTTPException(status_code=404, detail=f"No se encontró planilla con consecutivo {consecutivo}")

        # Solo se puede asignar pedido a planillas APROBADAS
        if doc.get("estado") != "APROBADO":
            raise HTTPException(
                status_code=400,
                detail=f"La planilla con consecutivo {consecutivo} no está APROBADA (estado actual: {doc.get('estado')}). Solo se puede asignar el pedido a planillas aprobadas."
            )

        # Replicar lógica de una fila de importar-vulcano
        doc["pedido_vulcano"] = pedido
        doc["fecha_movimiento_historico"] = datetime.now()
        if request.usuario:
            doc["usuario_pedido_vulcano"] = request.usuario

        # Mover a histórico (delete-first idempotente por si el _id ya existiera ahí)
        coleccion_historico.delete_one({"_id": doc["_id"]})
        coleccion_historico.insert_one(doc)

        # Eliminar de pedidos_medical
        coleccion_pedidos_medical.delete_one({"_id": doc["_id"]})

        # Avisar por WhatsApp al operativo que montó la planilla
        _notificar_pedido_creado_whatsapp(doc, pedido)

        logger.info(f"Pedido manual asignado: consecutivo={consecutivo}, pedido_vulcano={pedido}")

        return {
            "mensaje": f"Pedido {pedido} asignado a consecutivo {consecutivo} y movido a histórico.",
            "exitoso": True,
            "consecutivo": consecutivo,
            "pedido": pedido,
            "planilla": doc.get("planilla")
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al asignar pedido manual: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al asignar pedido: {str(e)}")


class RetrocederASolicitudRequest(BaseModel):
    """Modelo para devolver una planilla del histórico a SolicitudVehiculos."""
    planilla: str
    usuario: Optional[str] = None  # Para trazabilidad
    motivo: Optional[str] = None   # Razón del retroceso


@router.post("/retroceder-a-solicitud")
async def retroceder_a_solicitud(request: RetrocederASolicitudRequest):
    """
    Operación inversa de 'importar-vulcano' / 'asignar-pedido-manual':
    devuelve una planilla de pedidos_medical_historico a pedidos_medical para que
    reaparezca en SolicitudVehiculos. Quita el pedido_vulcano y la fecha de
    movimiento, y deja el estado en APROBADO (lista para reasignar el pedido).
    Solo ADMIN (validar en frontend).
    """
    try:
        planilla = (request.planilla or "").strip()
        if not planilla:
            raise HTTPException(status_code=400, detail="planilla es obligatoria")

        logger.info(
            f"=== RETROCEDER A SOLICITUD === planilla={planilla}, "
            f"usuario={request.usuario}, motivo={request.motivo}"
        )

        # 1. Buscar en el histórico por planilla.
        doc = coleccion_historico.find_one({"planilla": planilla})
        if not doc:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró la planilla {planilla} en el histórico"
            )

        # 2. Anti-duplicado: si ya existe en pedidos_medical, abortar.
        existente = coleccion_pedidos_medical.find_one({
            "$or": [
                {"_id": doc["_id"]},
                {"consecutivo": doc.get("consecutivo")},
                {"planilla": doc.get("planilla")},
            ]
        })
        if existente:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"La planilla {planilla} ya existe en SolicitudVehiculos "
                    f"(pedidos_medical). No se puede retroceder para evitar duplicados."
                )
            )

        # 3. Trazabilidad del retroceso en historial_cambios.
        pedido_vulcano_previo = doc.get("pedido_vulcano")
        fecha_actual = datetime.now()
        historial_cambios = doc.get("historial_cambios", []) or []
        historial_cambios.append({
            "fecha": fecha_actual,
            "usuario": request.usuario,
            "accion": "retroceso_historico",
            "campos_modificados": [
                {"campo": "pedido_vulcano", "valor_anterior": pedido_vulcano_previo, "valor_nuevo": None},
                {"campo": "fecha_movimiento_historico", "valor_anterior": doc.get("fecha_movimiento_historico"), "valor_nuevo": None},
                {"campo": "estado", "valor_anterior": doc.get("estado"), "valor_nuevo": "APROBADO"},
            ],
            "motivo": request.motivo,
        })

        # 4. Limpiar campos del movimiento al histórico y dejar APROBADO.
        doc["estado"] = "APROBADO"
        doc["fecha_aprobacion"] = fecha_actual
        doc["fecha_modificacion"] = fecha_actual
        if request.usuario:
            doc["usuario_modificacion"] = request.usuario
        doc["historial_cambios"] = historial_cambios

        doc.pop("pedido_vulcano", None)
        doc.pop("fecha_movimiento_historico", None)
        doc.pop("usuario_pedido_vulcano", None)

        # En fusiones, limpiar también los pedidos embebidos en cada original.
        fusion_info = doc.get("fusion_info") or {}
        if fusion_info.get("es_fusionada") and fusion_info.get("datos_originales"):
            for original in fusion_info["datos_originales"]:
                original.pop("pedido_vulcano", None)
                original.pop("fecha_pedido_vulcano", None)

        # 5. Mover: delete-first idempotente + insert en activo + delete del histórico.
        coleccion_pedidos_medical.delete_one({"_id": doc["_id"]})
        coleccion_pedidos_medical.insert_one(doc)
        coleccion_historico.delete_one({"_id": doc["_id"]})

        logger.info(
            f"Planilla retrocedida del histórico a SolicitudVehiculos: "
            f"planilla={planilla}, consecutivo={doc.get('consecutivo')}, "
            f"pedido_vulcano_previo={pedido_vulcano_previo}"
        )

        return {
            "mensaje": (
                f"Planilla {planilla} devuelta a SolicitudVehiculos "
                f"(estado APROBADO, sin pedido Vulcano)."
            ),
            "exitoso": True,
            "planilla": planilla,
            "consecutivo": doc.get("consecutivo"),
            "pedido_vulcano_previo": pedido_vulcano_previo,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al retroceder planilla del histórico: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al retroceder la planilla: {str(e)}")


# Modelos para gestión de causales
class CausalRequest(BaseModel):
    """Modelo para crear/actualizar una causal"""
    nombre: str
    activo: bool = True


class CausalResponse(BaseModel):
    """Modelo de respuesta para causales"""
    _id: str
    nombre: str
    activo: bool
    fecha_creacion: Optional[datetime] = None


class ConsultarTarifaRequest(BaseModel):
    centro_costo: str
    ruta: str
    peso_real: float
    tipo_vehiculo: Optional[str] = None  # Si viene, la tarifa se busca por este tipo (p.ej. tipo_veh_sicetac);
                                         # si no, se deriva del peso (comportamiento histórico).


class GuardarBusquedaRequest(BaseModel):
    usuario: str
    perfil: str
    centro_distribucion: str
    planillas_buscadas: List[str]
    resultados_consolidados: List[dict]
    fecha_inicio: str
    fecha_fin: str
    planillas_a_eliminar: Optional[List[str]] = None  # Planillas a eliminar (para fusión)


def _numero_de_consecutivo(cons: str) -> Optional[int]:
    """Extrae la parte numérica de un consecutivo con formato REGIONAL-FECHA-N[L].
    Ej: 'FUNZA-20260701-1' -> 1 ; 'FUNZA-20260701-2B' -> 2. None si no parsea."""
    if not cons:
        return None
    parts = cons.split("-")
    if len(parts) < 3:
        return None
    numero_letra = parts[2]
    for i, char in enumerate(numero_letra):
        if char.isalpha():
            return int(numero_letra[:i]) if i > 0 else None
    return int(numero_letra) if numero_letra.isdigit() else None


def _generar_consecutivo(regional: str, fecha: datetime, es_fusion: bool = False, num_fusion: int = 1, fusion_id: Optional[str] = None, numeros_planillas_a_fusionar: Optional[List[int]] = None) -> dict:
    """
    Genera consecutivos únicos para planillas.

    Args:
        regional: Nombre de la regional (ej: "FUNZA")
        fecha: Fecha de la consulta
        es_fusion: Si es una fusión de planillas
        num_fusion: Número de planillas en la fusión
        fusion_id: ID del grupo de fusión (para reutilizar huecos)

    Returns:
        Dict con los consecutivos generados:
        - Para planilla individual: {"consecutivo": "FUNZA-20260527-1", "consecutivo_base": "FUNZA-20260527-1", "numero": 1, "letra": None}
        - Para fusión: [{"consecutivo": "FUNZA-20260527-1A", ...}, {"consecutivo": "FUNZA-20260527-1B", ...}]
    """
    fecha_str = fecha.strftime("%Y%m%d")
    prefijo = f"{regional}-{fecha_str}"

    # Buscar todos los consecutivos existentes para esta regional y fecha
    # EXCLUYENDO planillas fusionadas (marcadas como fusionada: true).
    # IMPORTANTE: se consulta TAMBIÉN el histórico (pedidos_medical_historico), porque
    # "Importar Vulcano" mueve las planillas allí y las borra de pedidos_medical.
    # Si no se considerara el histórico, al volver a cargar planillas el mismo día se
    # reutilizarían consecutivos ya usados (colisión).
    regex_pattern = f"^{prefijo}-\\d+[A-Z]?$"
    _query = {"consecutivo": {"$regex": regex_pattern}, "fusionada": {"$ne": True}}
    _projection = {"consecutivo": 1, "consecutivo_base": 1, "numero_consecutivo": 1, "letra_consecutivo": 1}
    existentes = list(coleccion_pedidos_medical.find(_query, _projection)) \
        + list(coleccion_historico.find(_query, _projection))

    # Extraer números y letras usadas
    numeros_usados = set()
    fusiones_activas = {}  # {numero_base: [letras_usadas]}

    for doc in existentes:
        cons = doc.get("consecutivo", "")
        if not cons:
            continue

        # Intentar obtener numero y letra directamente de los campos
        numero = doc.get("numero_consecutivo")
        letra = doc.get("letra_consecutivo")

        # Si no están disponibles, parsear del consecutivo
        if numero is None:
            parts = cons.split("-")
            if len(parts) >= 3:
                numero_letra = parts[2]

                # Encontrar dónde empieza la letra
                for i, char in enumerate(numero_letra):
                    if char.isalpha():
                        numero = int(numero_letra[:i]) if i > 0 else None
                        letra = numero_letra[i:]
                        break
                else:
                    # No hay letra, es todo el número
                    numero = int(numero_letra) if numero_letra.isdigit() else None

        if numero is not None:
            if letra:
                # Es una planilla fusionada
                if numero not in fusiones_activas:
                    fusiones_activas[numero] = []
                fusiones_activas[numero].append(letra)
            else:
                # Es una planilla individual
                numeros_usados.add(numero)

    # Considerar también los consecutivos EMBEBIDOS en fusiones (datos_originales).
    # Al fusionar p.ej. FUNZA-...-1 y FUNZA-...-2 en FUNZA-...-1A, los originales
    # dejan de ser documentos top-level (quedan dentro de fusion_info.datos_originales)
    # pero sus números SIGUEN ocupados. Sin este bloque, la próxima planilla individual
    # recibiría FUNZA-...-2 y colisionaría con el original embebido (bug de duplicado).
    # Hay índice idx_datos_originales_consecutivo que sostiene esta consulta.
    for _col in (coleccion_pedidos_medical, coleccion_historico):
        for _doc in _col.find(
            {"fusion_info.datos_originales.consecutivo": {"$regex": regex_pattern}},
            {"fusion_info.datos_originales.consecutivo": 1},
        ):
            for _original in (_doc.get("fusion_info") or {}).get("datos_originales", []) or []:
                _n = _numero_de_consecutivo(_original.get("consecutivo", ""))
                if _n is not None:
                    numeros_usados.add(_n)

    logger.info(f"[CONSECUTIVO] Regional: {regional}, Fecha: {fecha_str}")
    logger.info(f"[CONSECUTIVO] Números usados (individuales): {sorted(numeros_usados)}")
    logger.info(f"[CONSECUTIVO] Fusiones activas: {fusiones_activas}")

    if es_fusion:
        # Lógica para fusiones: usar el mismo número base con letras
        if fusion_id:
            # Reutilizar un número base de fusión existente
            # Buscar el fusion_id en documentos existentes
            fusion_existente = coleccion_pedidos_medical.find_one({
                "fusion_info.fusion_id": fusion_id
            })
            if fusion_existente:
                cons_base = fusion_existente.get("consecutivo_base", "")
                if cons_base:
                    parts = cons_base.split("-")
                    if len(parts) >= 3:
                        numero_base = int(''.join([c for c in parts[2] if c.isdigit()]))
                        # Obtener letras ya usadas en esta fusión
                        letras_usadas = fusiones_activas.get(numero_base, [])
                        # Generar letras para las planillas
                        letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                        consecutivos = []
                        letra_idx = 0

                        for i in range(num_fusion):
                            # Encontrar la siguiente letra disponible
                            while letra_idx < len(letras) and letras[letra_idx] in letras_usadas:
                                letra_idx += 1

                            if letra_idx >= len(letras):
                                raise HTTPException(
                                    status_code=500,
                                    detail=f"No hay más letras disponibles para el consecutivo base {prefijo}-{numero_base}"
                                )

                            letra = letras[letra_idx]
                            cons_completo = f"{prefijo}-{numero_base}{letra}"

                            consecutivos.append({
                                "consecutivo": cons_completo,
                                "consecutivo_base": f"{prefijo}-{numero_base}",
                                "numero": numero_base,
                                "letra": letra
                            })

                            letras_usadas.append(letra)
                            letra_idx += 1

                        return {"consecutivos": consecutivos, "numero_base": numero_base}

        # Nueva fusión: usar el número más pequeño de las planillas que se van a fusionar
        # Esto permite recuperar los números originales al dividir
        if numeros_planillas_a_fusionar and len(numeros_planillas_a_fusionar) > 0:
            # Usar el número más pequeño de las planillas que se van a fusionar
            numero_disponible = min(numeros_planillas_a_fusionar)
            logger.info(f"[CONSECUTIVO] Fusión usando número base {numero_disponible} de las planillas originales {numeros_planillas_a_fusionar}")
        else:
            # Si no se proporcionan números, usar el MÁXIMO número existente + 1 (no reutilizar huecos)
            max_numero = 0
            if numeros_usados:
                max_numero = max(numeros_usados)
            if fusiones_activas:
                max_numero = max(max_numero, max(fusiones_activas.keys()))

            numero_disponible = max_numero + 1

        # Generar letras para la fusión
        letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        consecutivos = []

        for i in range(num_fusion):
            if i >= len(letras):
                raise HTTPException(
                    status_code=500,
                    detail=f"No hay suficientes letras para fusionar {num_fusion} planillas"
                )

            letra = letras[i]
            cons_completo = f"{prefijo}-{numero_disponible}{letra}"

            consecutivos.append({
                "consecutivo": cons_completo,
                "consecutivo_base": f"{prefijo}-{numero_disponible}",
                "numero": numero_disponible,
                "letra": letra
            })

        fusiones_activas[numero_disponible] = [letras[i] for i in range(num_fusion)]

        return {
            "consecutivos": consecutivos,
            "numero_base": numero_disponible,
            "fusiones_activas": fusiones_activas
        }

    else:
        # Lógica para planillas individuales: usar el MÁXIMO número existente + 1
        # Esto asegura que los consecutivos no se reutilicen
        max_numero = 0
        if numeros_usados:
            max_numero = max(numeros_usados)
        if fusiones_activas:
            max_numero = max(max_numero, max(fusiones_activas.keys()))

        numero_disponible = max_numero + 1
        cons_completo = f"{prefijo}-{numero_disponible}"

        return {
            "consecutivo": cons_completo,
            "consecutivo_base": cons_completo,
            "numero": numero_disponible,
            "letra": None
        }


def _obtener_festivos_colombia(anio: int) -> List[str]:
    """
    Retorna lista de festivos de Colombia para un año dado (formato YYYY-MM-DD).
    Incluye festivos fijos y móviles (basados en Pascua).
    """
    from datetime import date

    festivos = []

    # Festivos fijos
    festivos_fijos = [
        (1, 1),   # 1 de enero
        (1, 6),   # 6 de enero
        (5, 1),   # 1 de mayo
        (7, 20),  # 20 de julio
        (8, 7),   # 7 de agosto
        (12, 8),  # 8 de diciembre
        (12, 25), # 25 de diciembre
    ]

    def _format_fecha(fecha: date) -> str:
        return fecha.strftime('%Y-%m-%d')

    def _mover_al_lunes(fecha: date) -> date:
        dia_sem = fecha.weekday()
        if dia_sem != 0:  # Si no es lunes
            dias_hasta_lunes = (7 - dia_sem) % 7
            if dias_hasta_lunes == 0:
                dias_hasta_lunes = 7
            return fecha + timedelta(days=dias_hasta_lunes)
        return fecha

    for mes, dia in festivos_fijos:
        festivos.append(_format_fecha(date(anio, mes, dia)))

    # Calcular Pascua
    a = anio % 19
    b = anio // 100
    c = anio % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes_pascua = (h + l - 7 * m + 114) // 31
    dia_pascua = ((h + l - 7 * m + 114) % 31) + 1
    pascua = date(anio, mes_pascua, dia_pascua)

    # Festivos móviles
    jueves_santo = pascua - timedelta(days=3)
    viernes_santo = pascua - timedelta(days=2)
    ascension = _mover_al_lunes(pascua + timedelta(days=39))
    corpus_christi = _mover_al_lunes(pascua + timedelta(days=60))
    sagrada_eucaristia = _mover_al_lunes(pascua + timedelta(days=68))

    festivos.extend([
        _format_fecha(jueves_santo),
        _format_fecha(viernes_santo),
        _format_fecha(ascension),
        _format_fecha(corpus_christi),
        _format_fecha(sagrada_eucaristia)
    ])

    return festivos


def _calcular_rango_3_dias_habiles() -> tuple[str, str]:
    """
    Calcula el rango de fechas: 40 días hábiles anteriores hasta hoy.
    No cuenta fines de semana ni festivos de Colombia.

    Returns:
        Tupla (fecha_inicial, fecha_final) en formato YYYY-MM-DD
    """
    hoy = datetime.now()
    festivos = _obtener_festivos_colombia(hoy.year)

    def es_festivo_o_fin_de_semana(fecha: datetime) -> bool:
        """Verifica si una fecha es festivo o fin de semana"""
        if fecha.weekday() >= 5:  # Sábado (5) o Domingo (6)
            return True
        fecha_str = fecha.strftime('%Y-%m-%d')
        return fecha_str in festivos

    # Restar 40 días hábiles
    dias_habiles_restar = 0
    fecha_actual = hoy

    while dias_habiles_restar < 40:
        fecha_actual -= timedelta(days=1)
        if not es_festivo_o_fin_de_semana(fecha_actual):
            dias_habiles_restar += 1

    fecha_final = hoy.strftime('%Y-%m-%d')
    fecha_inicial = fecha_actual.strftime('%Y-%m-%d')

    logger.info(f"Rango de fechas calculado (40 días hábiles): {fecha_inicial} a {fecha_final}")
    return fecha_inicial, fecha_final


def _get_proxy_url() -> Optional[str]:
    """Obtiene la configuración de proxy desde variables de entorno"""
    proxy_url = os.getenv('VULCANO_PROXY_URL')
    if proxy_url:
        logger.info(f"Proxy configurado: {proxy_url.split('@')[-1]}")
    return proxy_url


def _determinar_tipo_vehiculo(peso_real: float) -> str:
    """
    Determina el tipo de vehículo según el peso real.

    Rangos:
    - Hasta 1.000 kg → CARRY
    - 1.001 a 2.300 kg → NHR
    - 2.301 a 4.500 kg → TURBO
    - 4.501 a 6.100 kg → NIES
    - 6.101 a 9.000 kg → SENCILLO
    - 9.001 a 17.000 kg → PATINETA
    - Más de 17.000 kg → TRACTOMULA
    """
    if peso_real <= 1000:
        return "CARRY"
    elif peso_real <= 2300:
        return "NHR"
    elif peso_real <= 4500:
        return "TURBO"
    elif peso_real <= 6100:
        return "NIES"
    elif peso_real <= 9000:
        return "SENCILLO"
    elif peso_real <= 17000:
        return "PATINETA"
    else:
        return "TRACTOMULA"


def _obtener_tarifa_ruta(centro_costo: str, ruta: str, tipo_vehiculo: str) -> Optional[float]:
    try:
        tarifa = coleccion_tarifas.find_one({
            "ruta": ruta.upper()
        })

        if not tarifa:
            logger.warning(f"No se encontró tarifa para ruta={ruta}")
            return None

        # Mapear tipo_vehiculo al campo en BD
        tipo_map = {
            "CARRY": "carry",
            "NHR": "nhr",
            "TURBO": "turbo",
            "NIES": "nies",
            "SENCILLO": "sencillo",
            "PATINETA": "patineta",
            "TRACTOMULA": "tractomula"
        }

        campo = tipo_map.get(tipo_vehiculo.upper())
        if not campo:
            logger.error(f"Tipo de vehículo no válido: {tipo_vehiculo}")
            return None

        valor = tarifa.get(campo, 0)
        logger.info(f"Tarifa encontrada: ruta={ruta}, tipo={tipo_vehiculo}, valor={valor}")
        return float(valor)

    except Exception as e:
        logger.error(f"Error al obtener tarifa: {str(e)}")
        return None


async def _consultar_api_siscore_planillas(
    fecha_inicial: str,
    fecha_final: str,
    centro_distribucion: str = "TODOS"
) -> dict:
    """
    Consulta el API de Siscore V3 con rango de fechas (igual que pedidos_v3).

    Args:
        fecha_inicial: Fecha inicial en formato YYYY-MM-DD
        fecha_final: Fecha final en formato YYYY-MM-DD
        centro_distribucion: Centro de distribución ("TODOS" o específico)

    Returns:
        Diccionario con la respuesta del API
    """
    payload = {
        "token": SISCORE_V3_TOKEN,
        "fecha_inicial": fecha_inicial,
        "fecha_final": fecha_final,
        "centro_distribucion": centro_distribucion if centro_distribucion else "TODOS",
        "incluir_pedidos_manuales": "SI",
        "pedido_especifico": ""  # Vacío para traer todos los pedidos del rango
    }

    timeout = httpx.Timeout(600.0, connect=120.0)  # 10 minutos total, 2 minutos para conectar
    proxy_url = _get_proxy_url()

    logger.info(f"[API Siscore] Consultando rango: {fecha_inicial} a {fecha_final}")
    logger.info(f"[API Siscore] Centro distribución: {centro_distribucion if centro_distribucion else 'TODOS'}")
    logger.info(f"[API Siscore] Proxy: {'HABILITADO' if proxy_url else 'NO CONFIGURADO'}")

    try:
        async with httpx.AsyncClient(
            timeout=timeout,
            proxy=proxy_url,
            trust_env=False,
        ) as client:
            logger.info(f"[API Siscore] Enviando payload a: {SISCORE_V3_ENDPOINT}")
            logger.info(f"[API Siscore] Payload: {payload}")

            response = await client.post(
                SISCORE_V3_ENDPOINT,
                json=payload,
                headers={"Content-Type": "application/json"}
            )

            logger.info(f"[API Siscore] Status code: {response.status_code}")
            logger.info(f"[API Siscore] Headers: {dict(response.headers)}")

            response.raise_for_status()

            result = response.json()
            logger.info(f"[API Siscore] Respuesta recibida, tipo: {type(result)}")

            # Log específico para ver la estructura de la respuesta
            if isinstance(result, dict):
                logger.info(f"[API Siscore] Claves en respuesta: {list(result.keys())}")
                if 'registros' in result:
                    logger.info(f"[API Siscore] Cantidad de registros: {len(result.get('registros', []))}")
                else:
                    logger.warning(f"[API Siscore] No hay clave 'registros' en la respuesta")
            elif isinstance(result, list):
                logger.info(f"[API Siscore] La respuesta es una lista con {len(result)} elementos")

            return result

    except httpx.HTTPStatusError as e:
        error_msg = f"Error HTTP Siscore: {e.response.status_code}"
        try:
            error_detail = e.response.text[:500]
            error_msg += f" - {error_detail}"
        except:
            pass
        raise RuntimeError(error_msg)
    except httpx.ConnectTimeout:
        raise RuntimeError("Timeout conectando a Siscore. El servidor no respondió.")
    except httpx.ReadTimeout:
        raise RuntimeError("Timeout leyendo respuesta de Siscore. El endpoint tardó demasiado.")
    except Exception as e:
        raise RuntimeError(f"Error consultando Siscore: {type(e).__name__}: {str(e)}")


@router.post("/verificar-planillas-fusionadas")
async def verificar_planillas_fusionadas(request: VerificarFusionadasRequest):
    """
    Verifica si alguna de las planillas está fusionada ANTES de consultar Siscore.
    Evita perder tiempo consultando la API si las planillas ya están fusionadas.
    """
    try:
        logger.info(f"=== VERIFICANDO PLANILLAS FUSIONADAS ===")
        logger.info(f"Planillas a verificar: {request.planillas}")

        planillas_fusionadas = []
        for planilla_num in request.planillas:
            # Al fusionar, los originales se ELIMINAN como docs propios y se embeben en
            # fusion_info (planillas_originales / datos_originales) del documento fusionado.
            # Por eso se detecta buscando si esta planilla es ORIGINAL de alguna fusión,
            # en activos o en el histórico (por si la fusión ya fue despachada a Vulcano).
            query_fusion = {
                "fusion_info.es_fusionada": True,
                "$or": [
                    {"fusion_info.planillas_originales": planilla_num},
                    {"fusion_info.datos_originales.planilla": planilla_num},
                ],
            }
            doc_fusion = coleccion_pedidos_medical.find_one(query_fusion) \
                or coleccion_historico.find_one(query_fusion)
            if doc_fusion:
                planilla_fusionada = doc_fusion.get("planilla") or ""
                consecutivo_fusionada = doc_fusion.get("consecutivo") or ""
                planillas_fusionadas.append({
                    "planilla": planilla_num,
                    "fusionada_en": planilla_fusionada,
                    "consecutivo_fusionada": consecutivo_fusionada
                })
                logger.warning(f"⚠️ Planilla {planilla_num} está fusionada en {planilla_fusionada} (consecutivo: {consecutivo_fusionada})")

        return {
            "planillas_fusionadas": planillas_fusionadas,
            "total_fusionadas": len(planillas_fusionadas)
        }

    except Exception as e:
        logger.error(f"Error al verificar planillas fusionadas: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al verificar planillas fusionadas: {str(e)}"
        )


@router.post("/consultar-planillas")
async def consultar_planillas(request: ConsultaPlanillasRequest):
    """
    Consulta planillas en la API de Siscore V3 filtrando por rango de fechas.
    Calcula automáticamente 3 días hábiles hacia atrás.
    Devuelve todos los registros del rango.

    Args:
        request: Objeto con planillas, fecha_inicio y fecha_fin (puede estar vacío)

    Returns:
        Diccionario con todos los registros del rango de fechas
    """
    try:
        logger.info(f"=== INICIO CONSULTA DE PLANILLAS ===")
        logger.info(f"Planillas solicitadas: {request.planillas}")
        logger.info(f"Perfil: {request.perfil}")
        logger.info(f"Centro distribución solicitado: {request.centro_distribucion}")
        logger.info(f"Rango recibido: {request.fecha_inicio} a {request.fecha_fin}")

        # Calcular rango de 30 días hábiles si no se proporciona
        if not request.fecha_inicio or not request.fecha_fin:
            fecha_inicio, fecha_fin = _calcular_rango_3_dias_habiles()
            logger.info(f"Rango calculado automáticamente: {fecha_inicio} a {fecha_fin}")
        else:
            fecha_inicio = request.fecha_inicio
            fecha_fin = request.fecha_fin

        # Determinar centro de distribución según perfil
        # Si es operativo y tiene centro_distribucion, usarlo
        # Si es perfil global o no tiene centro_distribucion, usar TODOS
        if request.centro_distribucion and request.perfil not in ['ADMIN', 'COORDINADOR', 'CONTROL', 'ANALISTA']:
            # Convertir CO07 o FUNZA al formato especial para Siscore
            if request.centro_distribucion in ['CO07', 'FUNZA']:
                centro_dist = "FUNZA - SAN DIEGO 7G"
            else:
                centro_dist = request.centro_distribucion
            logger.info(f"Filtro aplicado: Operativo con centro_distribucion={centro_dist}")
        else:
            centro_dist = "TODOS"
            logger.info(f"Filtro aplicado: Perfil global o sin centro_distribucion, usando TODOS")

        # Consultar API de Siscore (con filtro de centro de distribución si aplica)
        respuesta_api = await _consultar_api_siscore_planillas(
            fecha_inicial=fecha_inicio,
            fecha_final=fecha_fin,
            centro_distribucion=centro_dist
        )

        # LOG: Mostrar respuesta completa de Siscore
        logger.info(f"=== RESPUESTA COMPLETA DE SISCORE ===")
        logger.info(f"Tipo: {type(respuesta_api)}")
        logger.info(f"Claves: {respuesta_api.keys() if isinstance(respuesta_api, dict) else 'No es dict'}")

        # La respuesta de Siscore tiene la estructura: {ok, total, filtros, data}
        # Los registros están en 'data'
        todos_registros = respuesta_api.get('data', [])

        logger.info(f"Total registros recibidos de Siscore: {len(todos_registros)}")

        # PRIMERO: Filtrar por las planillas que el usuario solicitó
        planillas_set = set(p.strip() for p in request.planillas)
        logger.info(f"Filtrando por planillas solicitadas: {planillas_set}")

        registros_filtrados = [
            reg for reg in todos_registros
            if (reg.get('Planilla') or '').strip() in planillas_set
        ]

        logger.info(f"Registros después de filtrar por planilla: {len(registros_filtrados)}")

        # SEGUNDO: Enriquecer SOLO los registros filtrados: buscar rutas faltantes por Divipola
        for reg in registros_filtrados:
            ruta_siscore = (reg.get('Ruta') or '').strip()
            divipola = (reg.get('Divipola') or '').strip()

            # Solo buscar si la ruta está vacía
            if not ruta_siscore or ruta_siscore == '' or ruta_siscore == '-':
                if divipola:
                    logger.info(f"Planilla {reg.get('Planilla')}: Ruta vacía, buscando por Divipola '{divipola}'")

                    # Buscar en colección de divipolas
                    divipola_doc = coleccion_divipolas.find_one({"divipola": divipola})

                    if divipola_doc:
                        ruta_encontrada = divipola_doc.get('ruta', '')
                        reg['Ruta'] = ruta_encontrada
                        logger.info(f"  ✅ Ruta encontrada en divipolas: {ruta_encontrada}")
                    else:
                        logger.warning(f"  ❌ No se encontró ruta para Divipola '{divipola}'")
                        reg['Ruta'] = '-'
                else:
                    logger.warning(f"  ⚠️ Ruta vacía pero sin Divipola para buscar")
                    reg['Ruta'] = '-'

        # LOG: Mostrar registros filtrados para depurar
        if registros_filtrados:
            logger.info(f"=== REGISTROS FILTRADOS Y ENRIQUECIDOS ===")
            for i, reg in enumerate(registros_filtrados[:5]):
                logger.info(f"Registro {i}: Planilla={reg.get('Planilla', 'N/A')}, Ruta={reg.get('Ruta', 'N/A')}, Divipola={reg.get('Divipola', 'N/A')}")

        # Devolver solo los registros filtrados y enriquecidos
        return {
            "registros": registros_filtrados,
            "total_registros": len(registros_filtrados),
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "planillas_buscadas": request.planillas
        }

    except Exception as e:
        logger.error(f"Error en consulta de planillas: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al consultar planillas: {str(e)}"
        )


@router.post("/consultar-planillas-bot")
def consultar_planillas_bot(request: ConsultaPlanillasRequest):
    """
    Consulta planillas vía el BOT de scraping del portal Siscore (Playwright).
    Reemplaza la fuente del WS de planillas (que dejó de funcionar para este fin)
    y devuelve la MISMA estructura que /siscore/consultar-planillas, de modo que
    el frontend no cambia su lógica de parseo/agrupación/tarifa/guardado.
    El endpoint viejo (/siscore/consultar-planillas) se conserva intacto.

    Es `def` (NO async) a propósito: FastAPI lo ejecuta en el threadpool y así el
    bot corre en su propio ProactorEventLoop (vía el wrapper síncrono), evitando
    el NotImplementedError de subprocess bajo el loop Selector de uvicorn/Windows.
    """
    try:
        # Import diferido: Playwright solo se carga si se invoca este endpoint,
        # así el servidor arranca aunque la dependencia aún no esté instalada.
        from Funciones import bot_siscore
        from Funciones.siscore_excel_mapper import construir_lookup_divipolas

        logger.info(f"=== CONSULTA PLANILLAS (BOT) ===")
        logger.info(f"Planillas solicitadas: {request.planillas}")

        planillas = [p.strip() for p in request.planillas if p and p.strip()]
        if not planillas:
            raise HTTPException(status_code=400, detail="No se recibieron planillas para consultar")

        # Lookup de divipolas para enriquecer Ruta/Departamento desde el Destino
        lookup_divipolas = construir_lookup_divipolas(coleccion_divipolas)

        # Reuso de sesión: mantiene un navegador logueado vivo entre requests
        # (evita bloqueos del TMS por exceso de logins).
        resultado = bot_siscore.session_manager.consultar(
            planillas,
            lookup_divipolas=lookup_divipolas,
        )

        logger.info(
            f"[BOT] Respuesta: {resultado.get('total_registros', 0)} registros, "
            f"{len(resultado.get('errores', []))} errores"
        )
        return resultado

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en consulta de planillas (bot): {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al consultar planillas (bot): {str(e)}"
        )


@router.get("/test-connection")
async def test_connection():
    """
    Endpoint para probar la conexión con Siscore
    """
    try:
        fecha_inicial, fecha_final = _calcular_rango_3_dias_habiles()

        respuesta = await _consultar_api_siscore_planillas(
            fecha_inicial=fecha_inicial,
            fecha_final=fecha_final,
            planillas=["TEST"]  # Planilla de prueba
        )

        return {
            "status": "connected",
            "message": "Conexión exitosa con Siscore",
            "timestamp": datetime.now().isoformat()
        }

    except Exception as e:
        return {
            "status": "error",
            "message": f"No se pudo conectar con Siscore: {str(e)}",
            "timestamp": datetime.now().isoformat()
        }


@router.get("/obtener-solicitudes-pendientes")
async def obtener_solicitudes_pendientes(usuario: str, perfil: str = "", centro_distribucion: str = ""):
    """
    Obtiene las solicitudes pendientes del usuario en solicitud_veh_medical.
    Para operativos, filtra por centro_distribucion (regional).
    Para perfiles globales (ADMIN, ANALISTA, COORDINADOR, CONTROL), muestra todas.
    """
    try:
        logger.info(f"Obteniendo solicitudes pendientes: usuario={usuario}, perfil={perfil}, centro_distribucion={centro_distribucion}")

        # Perfiles globales que ven todas las regionales
        perfiles_globales = ['ADMIN', 'ANALISTA', 'COORDINADOR', 'CONTROL']

        # Construir filtro
        filtro = {"estado": "pendiente"}

        # Si es operativo y tiene centro_distribucion, filtrar por su regional
        if perfil and perfil not in perfiles_globales and centro_distribucion:
            # Filtrar por centro_distribucion
            filtro["centro_distribucion"] = centro_distribucion
            logger.info(f"Filtro aplicado: Operativo con centro_distribucion={centro_distribucion}")
        else:
            logger.info(f"Filtro aplicado: Perfil global o sin centro_distribucion, mostrando todas las solicitudes")

        solicitudes = list(coleccion_solicitudes.find(filtro).sort("fecha_creacion", -1))

        # Convertir ObjectId a string
        for sol in solicitudes:
            sol["_id"] = str(sol["_id"])

        logger.info(f"Solicitudes encontradas: {len(solicitudes)}")

        return {
            "solicitudes": solicitudes,
            "total": len(solicitudes)
        }

    except Exception as e:
        logger.error(f"Error al obtener solicitudes pendientes: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener solicitudes: {str(e)}"
        )


@router.post("/guardar-solicitud")
async def guardar_solicitud(request: GuardarSolicitudRequest):
    """
    Guarda una solicitud de vehículo en la colección solicitud_veh_medical.
    """
    try:
        logger.info(f"[GUARDAR SOLICITUD] Planilla: {request.planilla}, Placa: {request.placa}, Aforo: {request.aforo}")

        nueva_solicitud = {
            "usuario": request.usuario,
            "perfil": request.perfil,
            "centro_distribucion": request.centro_distribucion,
            "planilla": request.planilla,
            "piezas": request.piezas,
            "peso_real": request.peso_real,
            "ruta": request.ruta,
            "codigos_pedido": request.codigos_pedido,
            "cantidad_pedidos": request.cantidad_pedidos,
            "cliente_origen": request.cliente_origen,
            "municipio_destino": request.municipio_destino,
            "departamento_destino": request.departamento_destino,
            "regional": request.regional,
            "tarifa_calculada": request.tarifa_calculada,
            "tipo_vehiculo": request.tipo_vehiculo,
            "total_solicitado": request.total_solicitado,
            "diferencia": request.total_solicitado - request.tarifa_calculada,
            "tarifa_base": request.tarifa_base,
            "requiere_descargue": request.requiere_descargue,
            "punto_adicional": request.punto_adicional,
            "desvio": request.desvio,
            "aforo": request.aforo,
            "placa": request.placa,
            "tipo_veh_sicetac": request.tipo_veh_sicetac,
            "estado": "pendiente",
            "fecha_creacion": datetime.now(),
            "fecha_actualizacion": datetime.now()
        }

        result = coleccion_solicitudes.insert_one(nueva_solicitud)
        nueva_solicitud["_id"] = str(result.inserted_id)

        logger.info(f"Solicitud guardada: {nueva_solicitud['_id']} para usuario {request.usuario}")

        return {
            "mensaje": "Solicitud guardada exitosamente",
            "solicitud": nueva_solicitud
        }

    except Exception as e:
        logger.error(f"Error al guardar solicitud: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al guardar solicitud: {str(e)}"
        )


@router.put("/actualizar-solicitud")
async def actualizar_solicitud(request: ActualizarSolicitudRequest):
    """
    Actualiza una solicitud existente en la colección solicitud_veh_medical.
    """
    try:
        from bson import ObjectId

        logger.info(f"[ACTUALIZAR SOLICITUD] ID: {request.solicitud_id}, Placa: {request.placa}, Aforo: {request.aforo}")

        if not ObjectId.is_valid(request.solicitud_id):
            raise HTTPException(status_code=400, detail="ID de solicitud inválido")

        # Campos a actualizar
        campos_actualizar = {
            "planilla": request.planilla,
            "piezas": request.piezas,
            "peso_real": request.peso_real,
            "ruta": request.ruta,
            "codigos_pedido": request.codigos_pedido,
            "cantidad_pedidos": request.cantidad_pedidos,
            "cliente_origen": request.cliente_origen,
            "municipio_destino": request.municipio_destino,
            "departamento_destino": request.departamento_destino,
            "regional": request.regional,
            "tarifa_calculada": request.tarifa_calculada,
            "tipo_vehiculo": request.tipo_vehiculo,
            "total_solicitado": request.total_solicitado,
            "diferencia": request.total_solicitado - request.tarifa_calculada,
            "tarifa_base": request.tarifa_base,
            "requiere_descargue": request.requiere_descargue,
            "punto_adicional": request.punto_adicional,
            "desvio": request.desvio,
            "aforo": request.aforo,
            "placa": request.placa,
            "tipo_veh_sicetac": request.tipo_veh_sicetac,
            "fecha_actualizacion": datetime.now()
        }

        # Actualizar el documento
        resultado = coleccion_solicitudes.update_one(
            {"_id": ObjectId(request.solicitud_id)},
            {"$set": campos_actualizar}
        )

        if resultado.matched_count == 0:
            raise HTTPException(status_code=404, detail="Solicitud no encontrada")

        logger.info(f"Solicitud actualizada: {request.solicitud_id} por usuario {request.usuario}")

        # Obtener el documento actualizado para retornarlo
        solicitud_actualizada = coleccion_solicitudes.find_one({"_id": ObjectId(request.solicitud_id)})
        solicitud_actualizada["_id"] = str(solicitud_actualizada["_id"])

        return {
            "mensaje": "Solicitud actualizada exitosamente",
            "solicitud": solicitud_actualizada
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al actualizar solicitud: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al actualizar solicitud: {str(e)}"
        )


@router.post("/enviar-tramite")
async def enviar_tramite(request: EnviarTramiteRequest):
    """
    Envía una solicitud a trámite en la colección tramite_fmc.
    """
    try:
        from bson import ObjectId

        if not ObjectId.is_valid(request.solicitud_id):
            raise HTTPException(status_code=400, detail="ID de solicitud inválido")

        # Obtener la solicitud original
        solicitud_orig = coleccion_solicitudes.find_one({"_id": ObjectId(request.solicitud_id)})

        if not solicitud_orig:
            raise HTTPException(status_code=404, detail="Solicitud no encontrada")

        # Crear el trámite
        tramite = {
            **solicitud_orig,
            "_id_orig": solicitud_orig["_id"],
            "estado": "en_revision",
            "usuario_envio": request.usuario,
            "fecha_envio": datetime.now(),
            "fecha_creacion": datetime.now()
        }

        # Eliminar el _id original para que MongoDB genere uno nuevo
        del tramite["_id"]

        result = coleccion_tramites.insert_one(tramite)
        tramite["_id"] = str(result.inserted_id)

        # Actualizar estado de la solicitud original
        coleccion_solicitudes.update_one(
            {"_id": ObjectId(request.solicitud_id)},
            {"$set": {"estado": "en_tramite", "fecha_actualizacion": datetime.now()}}
        )

        logger.info(f"Trámite enviado: {tramite['_id']} desde solicitud {request.solicitud_id}")

        return {
            "mensaje": "Solicitud enviada a trámite exitosamente",
            "tramite": tramite
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al enviar trámite: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al enviar trámite: {str(e)}"
        )


@router.get("/rutas")
async def listar_rutas():
    """Lista las rutas que tienen tarifa (para el autocompletar al asignar ruta a una planilla)."""
    try:
        rutas = sorted(set(r for r in coleccion_tarifas.distinct("ruta") if r))
        return {"rutas": rutas}
    except Exception as e:
        logger.error(f"Error al listar rutas: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al listar rutas: {str(e)}")


@router.post("/consultar-tarifa")
async def consultar_tarifa(request: ConsultarTarifaRequest):

    try:
        # Tipo de vehículo: si viene explícito (tipo_veh_sicetac elegido por el usuario),
        # se usa ESE para tarifar; si no, se deriva del peso (comportamiento histórico).
        tipo_vehiculo = (request.tipo_vehiculo or "").strip().upper() or _determinar_tipo_vehiculo(request.peso_real)

        # Obtener tarifa de fletes_rutas_fmc
        tarifa_calculada = _obtener_tarifa_ruta(request.centro_costo, request.ruta, tipo_vehiculo)

        logger.info(f"Tarifa consultada: ruta={request.ruta}, peso={request.peso_real}kg, tipo={tipo_vehiculo} (override={bool(request.tipo_vehiculo)}), tarifa={tarifa_calculada}")

        return {
            "tipo_vehiculo": tipo_vehiculo,
            "tarifa_calculada": tarifa_calculada if tarifa_calculada else 0
        }

    except Exception as e:
        logger.error(f"Error al consultar tarifa: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al consultar tarifa: {str(e)}"
        )


@router.post("/guardar-busqueda")
async def guardar_busqueda(request: GuardarBusquedaRequest):
    """
    Guarda cada planilla como un documento independiente en pedidos_medical (como libros).
    """
    try:
        logger.info(f"=== GUARDAR BÚSQUEDA ===")
        logger.info(f"Usuario: {request.usuario}")
        logger.info(f"Perfil: {request.perfil}")
        logger.info(f"Planillas buscadas: {request.planillas_buscadas}")
        logger.info(f"Cantidad de resultados: {len(request.resultados_consolidados)}")

        # Filtrar planillas NO encontradas: no se guardan en base porque son registros
        # vacíos (piezas=0, peso=0, ruta="-", etc.). Así tampoco se desperdician
        # números de consecutivo para planillas sin información real.
        resultados_a_guardar = [r for r in request.resultados_consolidados if r.get("encontrada")]
        omitidas = len(request.resultados_consolidados) - len(resultados_a_guardar)
        if omitidas:
            planillas_omitidas = [r.get("planilla") for r in request.resultados_consolidados if not r.get("encontrada")]
            logger.info(f"Omitiendo {omitidas} planilla(s) no encontrada(s) — no se guardan en base: {planillas_omitidas}")

        # VERIFICAR si alguna de las planillas buscadas está fusionada.
        # Los originales fusionados se eliminan y se embeben en fusion_info del doc fusionado,
        # así que se detecta buscando si la planilla es original de una fusión (activos o histórico).
        planillas_fusionadas_detectadas = []
        for planilla_num in request.planillas_buscadas:
            query_fusion = {
                "fusion_info.es_fusionada": True,
                "$or": [
                    {"fusion_info.planillas_originales": planilla_num},
                    {"fusion_info.datos_originales.planilla": planilla_num},
                ],
            }
            doc_fusion = coleccion_pedidos_medical.find_one(query_fusion) \
                or coleccion_historico.find_one(query_fusion)
            if doc_fusion:
                planilla_fusionada = doc_fusion.get("planilla") or ""
                consecutivo_fusionada = doc_fusion.get("consecutivo") or ""
                planillas_fusionadas_detectadas.append({
                    "planilla": planilla_num,
                    "fusionada_en": planilla_fusionada,
                    "consecutivo_fusionada": consecutivo_fusionada
                })
                logger.warning(f"⚠️ Planilla {planilla_num} está fusionada en {planilla_fusionada} (consecutivo: {consecutivo_fusionada})")

        fecha_creacion = datetime.now()

        # Agrupar resultados por regional y identificar fusiones
        resultados_por_regional = {}
        fusiones_por_regional = {}  # {regional: [resultados_fusionados]}

        for resultado in resultados_a_guardar:
            # Obtener regional prioritizando:
            # 1. Regional del resultado (que no sea '-' ni 'TODOS' ni vacía)
            # 2. Centro de distribución del resultado (que no sea '-' ni 'TODOS' ni vacío)
            # 3. Centro de costo del resultado
            # 4. Centro de distribución del USUARIO (fallback)
            regional_resultado = resultado.get("regional") or resultado.get("centro_distribucion") or resultado.get("centro_costo")

            # Si la regional del resultado es inválida ('-', 'TODOS', vacía), usar la del usuario
            if not regional_resultado or regional_resultado in ['-', 'TODOS', '']:
                regional_calculada = request.centro_distribucion or "TODOS"
            else:
                regional_calculada = regional_resultado

            # Limpiar y normalizar el nombre de la regional
            regional_calculada = regional_calculada.upper().strip()

            # Si aún queda en inválida, usar TODOS
            if regional_calculada in ['-', 'TODOS', '']:
                regional_calculada = "TODOS"

            # Guardar la regional calculada en el resultado para usarla después
            resultado["regional_calculada"] = regional_calculada

            fusion_info = resultado.get("fusion_info")

            if fusion_info and fusion_info.get("es_fusionada"):
                # Es una planilla fusionada
                if regional_calculada not in fusiones_por_regional:
                    fusiones_por_regional[regional_calculada] = []
                fusiones_por_regional[regional_calculada].append(resultado)
            else:
                # Es una planilla individual
                if regional_calculada not in resultados_por_regional:
                    resultados_por_regional[regional_calculada] = []
                resultados_por_regional[regional_calculada].append(resultado)

        # Procesar fusiones primero (para asignar un solo número base con letras)
        fusiones_procesadas = []  # [(resultado, consecutivo_info), ...]

        for regional, fusionados in fusiones_por_regional.items():
            # Obtener el fusion_id si existe (para reutilizar huecos)
            fusion_id = fusionados[0].get("fusion_info", {}).get("fusion_id") if fusionados else None

            # Extraer los números de las planillas que se van a fusionar
            # Buscar los consecutivos de las planillas originales en MongoDB
            numeros_planillas_a_fusionar = []
            for resultado in fusionados:
                fusion_info = resultado.get("fusion_info", {})
                planillas_originales = fusion_info.get("planillas_originales", [])

                # Buscar cada planilla original en MongoDB para obtener su consecutivo
                for planilla_num in planillas_originales:
                    doc_original = coleccion_pedidos_medical.find_one({"planilla": planilla_num})
                    if doc_original and doc_original.get("consecutivo"):
                        cons = doc_original.get("consecutivo")
                        parts = cons.split("-")
                        if len(parts) >= 3:
                            numero_letra = parts[2]
                            # Extraer número (sin letra)
                            numero = None
                            for i, char in enumerate(numero_letra):
                                if char.isalpha():
                                    numero = int(numero_letra[:i]) if i > 0 else None
                                    break
                            else:
                                # No hay letra, es todo el número
                                numero = int(numero_letra) if numero_letra.isdigit() else None

                            if numero is not None:
                                numeros_planillas_a_fusionar.append(numero)
                                logger.info(f"[CONSECUTIVO] Planilla original {planilla_num} tiene número {numero}")

            logger.info(f"[CONSECUTIVO] Números de planillas a fusionar: {numeros_planillas_a_fusionar}")

            # Generar consecutivos para la fusión
            try:
                consecutivo_info = _generar_consecutivo(
                    regional=regional,
                    fecha=fecha_creacion,
                    es_fusion=True,
                    num_fusion=len(fusionados),
                    fusion_id=fusion_id,
                    numeros_planillas_a_fusionar=numeros_planillas_a_fusionar if numeros_planillas_a_fusionar else None
                )

                logger.info(f"[CONSECUTIVO] Fusión {regional}: {len(fusionados)} planillas → número base {consecutivo_info.get('numero_base')}")

                # Asignar consecutivos a cada planilla fusionada
                letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                for i, resultado in enumerate(fusionados):
                    cons_info = {
                        "consecutivo": f"{regional}-{fecha_creacion.strftime('%Y%m%d')}-{consecutivo_info.get('numero_base')}{letras[i]}",
                        "consecutivo_base": f"{regional}-{fecha_creacion.strftime('%Y%m%d')}-{consecutivo_info.get('numero_base')}",
                        "numero": consecutivo_info.get('numero_base'),
                        "letra": letras[i],
                        "es_fusionada": True
                    }
                    fusiones_procesadas.append((resultado, cons_info))

            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"Error generando consecutivo para fusión: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Error generando consecutivo: {str(e)}")

        # Procesar planillas individuales por lotes (por regional)
        individuales_procesadas = []  # [(resultado, consecutivo_info), ...]

        for regional, resultados in resultados_por_regional.items():
            # Generar consecutivos para todas las planillas individuales de esta regional de una vez
            try:
                # Obtener el próximo número base disponible
                consecutivo_base = _generar_consecutivo(
                    regional=regional,
                    fecha=fecha_creacion,
                    es_fusion=False
                )
                numero_inicial = consecutivo_base["numero"]

                logger.info(f"[CONSECUTIVO] Individual {regional}: {len(resultados)} planillas comenzando desde {numero_inicial}")

                # Asignar consecutivos secuenciales
                for i, resultado in enumerate(resultados):
                    # Verificar si el resultado ya tiene un consecutivo (por ejemplo, al dividir una fusión)
                    if resultado.get("consecutivo"):
                        # Usar el consecutivo existente
                        cons_completo = resultado.get("consecutivo")
                        cons_base = resultado.get("consecutivo_base", cons_completo)

                        # Extraer número del consecutivo existente
                        numero = None
                        parts = cons_completo.split("-")
                        if len(parts) >= 3:
                            numero_letra = parts[2]
                            for j, char in enumerate(numero_letra):
                                if char.isalpha():
                                    numero = int(numero_letra[:j]) if j > 0 else None
                                    break
                            else:
                                numero = int(numero_letra) if numero_letra.isdigit() else None

                        cons_info_completo = {
                            "consecutivo": cons_completo,
                            "consecutivo_base": cons_base,
                            "numero": numero,
                            "letra": None,
                            "es_fusionada": False
                        }
                        individuales_procesadas.append((resultado, cons_info_completo))
                        logger.info(f"[CONSECUTIVO] Planilla {resultado.get('planilla')}: usando consecutivo existente {cons_completo}")
                    else:
                        # Generar nuevo consecutivo
                        numero = numero_inicial + i
                        cons_completo = f"{regional}-{fecha_creacion.strftime('%Y%m%d')}-{numero}"

                        cons_info_completo = {
                            "consecutivo": cons_completo,
                            "consecutivo_base": cons_completo,
                            "numero": numero,
                            "letra": None,
                            "es_fusionada": False
                        }
                        individuales_procesadas.append((resultado, cons_info_completo))
                        logger.info(f"[CONSECUTIVO] Planilla {resultado.get('planilla')}: {cons_completo}")

            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"Error generando consecutivos para planillas individuales: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Error generando consecutivos: {str(e)}")

        # Combinar todos los resultados procesados
        todos_procesados = fusiones_procesadas + individuales_procesadas

        # OPERATIVO: el campo `regional` se guarda como la bodega de origen
        # (CALI->YUMBO, BARRANQUILLA->GALAPA, MEDELLIN->GIRARDOTA). El consecutivo
        # NO se transforma (sigue con el nombre de la regional, ej. CALI-...).
        es_operativo = str(request.perfil or "").upper() == "OPERATIVO"

        # Estado inicial: las planillas creadas por OPERATIVO nacen en CREADO (borrador
        # visible solo para el creador y ADMIN hasta pasarlas a PREAPROBADO). Los demás
        # perfiles siguen naciendo en PREAPROBADO.
        estado_inicial = "CREADO" if es_operativo else "PREAPROBADO"

        # Guardar cada resultado como un documento independiente
        for resultado, cons_info in todos_procesados:
            logger.info(f"Procesando planilla: {resultado.get('planilla')} con consecutivo: {cons_info['consecutivo']}")

            planilla_doc = {
                "usuario_registro": request.usuario,  # Quien guardó (auditoría)
                "perfil": request.perfil,
                "centro_distribucion": request.centro_distribucion,
                "planilla": resultado.get("planilla"),
                "encontrada": resultado.get("encontrada", False),
                "piezas": resultado.get("piezas", 0),
                "peso_real": resultado.get("peso_real", 0),
                "peso_sicetac": resultado.get("peso_sicetac", resultado.get("peso_real", 0)),
                "ruta": resultado.get("ruta", "-"),
                "codigo_pedido": resultado.get("codigo_pedido", "-"),
                "cantidad_pedidos": resultado.get("cantidad_pedidos", 0),
                "cliente_origen": resultado.get("cliente_origen", "-"),
                "municipio_destino": resultado.get("municipio_destino", "-"),
                "departamento_destino": resultado.get("departamento_destino", "-"),
                "regional": (regional_a_origen_bodega(resultado.get("regional_calculada"))
                             if es_operativo else resultado.get("regional_calculada")),  # Usar la regional calculada con fallback
                "centro_costo": resultado.get("centro_costo"),
                "tarifa_calculada": resultado.get("tarifa_calculada", 0),
                "tipo_vehiculo": resultado.get("tipo_vehiculo"),
                "total_solicitado": resultado.get("total_solicitado", 0),
                "diferencia": resultado.get("total_solicitado", 0) - resultado.get("tarifa_calculada", 0),
                "flete_cobrado_fmc": resultado.get("flete_cobrado_fmc", 0),
                "cantidad_destinos": resultado.get("cantidad_destinos", 0),
                "municipios_destino_lista": resultado.get("municipios_destino_lista", "-"),
                "municipios_con_pedidos": resultado.get("municipios_con_pedidos", {}),
                "fusion_info": resultado.get("fusion_info"),  # Historial de fusión
                "tarifa_base": resultado.get("tarifa_base"),
                "requiere_descargue": resultado.get("requiere_descargue", "NO"),
                "punto_adicional": resultado.get("punto_adicional", False),
                "desvio": resultado.get("desvio", False),
                "aforo": resultado.get("aforo"),
                "placa": resultado.get("placa"),
                "tipo_veh_sicetac": resultado.get("tipo_veh_sicetac"),
                "fecha_creacion": fecha_creacion,
                "fecha_preaprobado": fecha_creacion if estado_inicial == "PREAPROBADO" else None,  # Visible para todos desde la creación (no operativo)
                "estado": resultado.get("estado", estado_inicial),  # CREADO (operativo) / PREAPROBADO (otros); puede ser REQUIERE_APROBACION_* o APROBADO
                "aprobado_por": resultado.get("aprobado_por"),
                "fecha_aprobacion": resultado.get("fecha_aprobacion"),
                # Campos de consecutivo
                "consecutivo": cons_info["consecutivo"],
                "consecutivo_base": cons_info["consecutivo_base"],
                "numero_consecutivo": cons_info["numero"],
                "letra_consecutivo": cons_info["letra"],
                "es_fusionada_consecutivo": cons_info["es_fusionada"],
                "registros_detalle": resultado.get("registros_detalle", []),
            }

            # Verificar si ya existe un documento con esta planilla
            existente = coleccion_pedidos_medical.find_one({"planilla": resultado.get("planilla")})

            if existente:
                # Actualizar el existente (conservar el consecutivo si ya tiene uno)
                # Preservar el estado y la fecha de preaprobado: una re-consulta no debe
                # rebajar una planilla que ya pasó a PREAPROBADO/APROBADO ni pisar su
                # fecha de visibilidad.
                planilla_doc["estado"] = existente.get("estado") or estado_inicial
                planilla_doc["fecha_preaprobado"] = existente.get("fecha_preaprobado") or planilla_doc.get("fecha_preaprobado")
                if existente.get("consecutivo"):
                    # Mantener el consecutivo existente y ACTUALIZAR cons_info para devolver al frontend
                    planilla_doc["consecutivo"] = existente.get("consecutivo")
                    planilla_doc["consecutivo_base"] = existente.get("consecutivo_base")
                    planilla_doc["numero_consecutivo"] = existente.get("numero_consecutivo")
                    planilla_doc["letra_consecutivo"] = existente.get("letra_consecutivo")
                    planilla_doc["es_fusionada_consecutivo"] = existente.get("es_fusionada_consecutivo", False)

                    # Actualizar cons_info con el consecutivo existente para devolver al frontend
                    cons_info["consecutivo"] = planilla_doc["consecutivo"]
                    cons_info["consecutivo_base"] = planilla_doc["consecutivo_base"]
                    cons_info["numero"] = planilla_doc["numero_consecutivo"]

                coleccion_pedidos_medical.update_one(
                    {"_id": existente["_id"]},
                    {"$set": planilla_doc}
                )
                logger.info(f"Planilla {resultado.get('planilla')}: actualizada con consecutivo {planilla_doc['consecutivo']}")
            else:
                # Insertar nuevo
                coleccion_pedidos_medical.insert_one(planilla_doc)
                logger.info(f"Planilla {resultado.get('planilla')}: guardada con consecutivo {planilla_doc['consecutivo']}")

        logger.info(f"Total guardado: {len(resultados_a_guardar)} planillas en pedidos_medical ({omitidas} omitida(s) por no encontrada(s))")

        # ELIMINAR las planillas originales que fueron fusionadas
        # Los datos originales se preservan en fusion_info de la planilla fusionada
        if request.planillas_a_eliminar and len(request.planillas_a_eliminar) > 0:
            logger.info(f"Planillas a eliminar por fusión: {request.planillas_a_eliminar}")

            # 1. Eliminar de pedidos_medical (si aún están ahí)
            resultado_delete = coleccion_pedidos_medical.delete_many(
                {"planilla": {"$in": request.planillas_a_eliminar}}
            )
            logger.info(f"Eliminadas {resultado_delete.deleted_count} planillas de pedidos_medical por fusión")

            # 2. Eliminar de pedidos_medical_historico (si ya fueron movidas por Vulcano)
            resultado_historico_delete = coleccion_historico.delete_many(
                {"planilla": {"$in": request.planillas_a_eliminar}}
            )
            logger.info(f"Eliminadas {resultado_historico_delete.deleted_count} planillas de historico por fusión")

        # Crear mapeo de planilla → consecutivo para devolver al frontend
        planillas_consecutivos = {}
        for resultado, cons_info in todos_procesados:
            planillas_consecutivos[resultado.get('planilla')] = {
                "consecutivo": cons_info['consecutivo'],
                "consecutivo_base": cons_info['consecutivo_base'],
                "numero": cons_info['numero'],
                "letra": cons_info['letra'],
                "es_fusionada": cons_info['es_fusionada']
            }

        return {
            "mensaje": f"Se guardaron/actualizaron {len(resultados_a_guardar)} planillas" + (f" ({omitidas} omitida(s) por no encontrada(s))" if omitidas else ""),
            "total": len(resultados_a_guardar),
            "omitidas_no_encontradas": omitidas,
            "consecutivos": planillas_consecutivos,
            "planillas_fusionadas_detectadas": planillas_fusionadas_detectadas
        }

    except Exception as e:
        logger.error(f"Error al guardar búsqueda: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al guardar búsqueda: {str(e)}"
        )


class DividirFusionRequest(BaseModel):
    planilla_fusionada: str
    usuario: str


@router.post("/dividir-fusion")
async def dividir_fusion(request: DividirFusionRequest):
    """
    Divide una planilla fusionada reconstruyendo las planillas originales
    desde fusion_info.datos_originales.
    """
    try:
        logger.info(f"=== DIVIDIR FUSION ===")
        logger.info(f"Planilla fusionada: {request.planilla_fusionada}")
        logger.info(f"Usuario: {request.usuario}")

        # Buscar la planilla fusionada
        fusionada = coleccion_pedidos_medical.find_one({"planilla": request.planilla_fusionada})
        if not fusionada:
            raise HTTPException(status_code=404, detail="Planilla fusionada no encontrada")

        fusion_info = fusionada.get("fusion_info", {})
        if not fusion_info.get("es_fusionada"):
            raise HTTPException(status_code=400, detail="Esta planilla no es una fusion")

        datos_originales = fusion_info.get("datos_originales", [])
        if not datos_originales:
            raise HTTPException(status_code=400, detail="No hay datos originales para reconstruir")

        logger.info(f"Reconstruyendo {len(datos_originales)} planillas originales desde datos_originales")

        # Reconstruir cada planilla original insertandola en pedidos_medical
        resultados_frontend = []
        planillas_consecutivos = {}

        for datos in datos_originales:
            planilla_num = datos.get("planilla")

            # Crear documento para insertar
            doc_insertar = {
                "planilla": planilla_num,
                "encontrada": datos.get("encontrada", True),
                "piezas": datos.get("piezas", 0),
                "peso_real": datos.get("peso_real", 0),
                "ruta": datos.get("ruta", "-"),
                "codigo_pedido": datos.get("codigo_pedido", "-"),
                "cantidad_pedidos": datos.get("cantidad_pedidos", 0),
                "cliente_origen": datos.get("cliente_origen", "-"),
                "municipio_destino": datos.get("municipio_destino", "-"),
                "departamento_destino": datos.get("departamento_destino", "-"),
                "regional": datos.get("regional"),
                "centro_costo": datos.get("centro_costo"),
                "tarifa_calculada": datos.get("tarifa_calculada", 0),
                "tipo_vehiculo": datos.get("tipo_vehiculo", "-"),
                "total_solicitado": datos.get("total_solicitado", 0),
                "tarifa_base": datos.get("tarifa_base"),
                "requiere_descargue": datos.get("requiere_descargue", 0),
                "punto_adicional": datos.get("punto_adicional", 0),
                "desvio": datos.get("desvio", 0),
                "aforo": datos.get("aforo", 0),
                "placa": datos.get("placa", ""),
                "tipo_veh_sicetac": datos.get("tipo_veh_sicetac"),
                "causal": datos.get("causal", ""),
                "cantidad_destinos": datos.get("cantidad_destinos", 0),
                "municipios_destino_lista": datos.get("municipios_destino_lista", "-"),
                "municipios_con_pedidos": datos.get("municipios_con_pedidos", {}),
                "consecutivo": datos.get("consecutivo"),
                "consecutivo_base": datos.get("consecutivo_base"),
                "flete_cobrado_fmc": datos.get("flete_cobrado_fmc", 0),
                "estado": datos.get("estado", "PREAPROBADO"),
                "fecha_preaprobado": (datos.get("fecha_preaprobado") or datetime.now())
                    if (datos.get("estado", "PREAPROBADO") == "PREAPROBADO") else datos.get("fecha_preaprobado"),
                "aprobado_por": datos.get("aprobado_por"),
                "fecha_aprobacion": datos.get("fecha_aprobacion"),
                "usuario": request.usuario,
                "perfil": fusionada.get("perfil", ""),
                "centro_distribucion": fusionada.get("centro_distribucion", ""),
                "fecha_creacion": datetime.now(),
                "registros_detalle": datos.get("registros_detalle", []),
            }

            # Insertar en pedidos_medical
            coleccion_pedidos_medical.insert_one(doc_insertar)
            logger.info(f"Reconstruida planilla original: {planilla_num} (consecutivo: {datos.get('consecutivo')})")

            # Formato para el frontend
            resultado = {
                "planilla": planilla_num,
                "encontrada": datos.get("encontrada", True),
                "piezas": datos.get("piezas", 0),
                "peso_real": datos.get("peso_real", 0),
                "ruta": datos.get("ruta", "-"),
                "codigo_pedido": datos.get("codigo_pedido", "-"),
                "cantidad_pedidos": datos.get("cantidad_pedidos", 0),
                "cliente_origen": datos.get("cliente_origen", "-"),
                "municipio_destino": datos.get("municipio_destino", "-"),
                "departamento_destino": datos.get("departamento_destino", "-"),
                "regional": datos.get("regional"),
                "centro_costo": datos.get("centro_costo"),
                "tarifa_calculada": datos.get("tarifa_calculada", 0),
                "tipo_vehiculo": datos.get("tipo_vehiculo", "-"),
                "total_solicitado": datos.get("total_solicitado", 0),
                "tarifa_base": datos.get("tarifa_base"),
                "requiere_descargue": datos.get("requiere_descargue", 0),
                "punto_adicional": datos.get("punto_adicional", 0),
                "desvio": datos.get("desvio", 0),
                "aforo": datos.get("aforo", 0),
                "placa": datos.get("placa", ""),
                "tipo_veh_sicetac": datos.get("tipo_veh_sicetac"),
                "causal": datos.get("causal", ""),
                "cantidad_destinos": datos.get("cantidad_destinos", 0),
                "municipios_destino_lista": datos.get("municipios_destino_lista", "-"),
                "municipios_con_pedidos": datos.get("municipios_con_pedidos", {}),
                "fusion_info": None,
                "estado": datos.get("estado", "PREAPROBADO"),
                "aprobado_por": datos.get("aprobado_por"),
                "fecha_aprobacion": datos.get("fecha_aprobacion"),
                "consecutivo": datos.get("consecutivo"),
                "consecutivo_base": datos.get("consecutivo_base"),
                "guardado": True
            }
            resultados_frontend.append(resultado)

            # Mapeo de consecutivos
            if datos.get("consecutivo"):
                planillas_consecutivos[planilla_num] = {
                    "consecutivo": datos["consecutivo"],
                    "consecutivo_base": datos["consecutivo_base"],
                    "numero": int(datos["consecutivo"].split("-")[-1]) if datos["consecutivo"] else None,
                    "letra": None,
                    "es_fusionada": False
                }

        # Eliminar la planilla fusionada
        coleccion_pedidos_medical.delete_one({"planilla": request.planilla_fusionada})
        logger.info(f"Eliminada planilla fusionada: {request.planilla_fusionada}")

        logger.info(f"Division completada: {len(resultados_frontend)} planillas reconstruidas")

        return {
            "mensaje": f"Se han restaurado {len(resultados_frontend)} planillas originales",
            "planillas": resultados_frontend,
            "consecutivos": planillas_consecutivos
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al dividir fusion: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al dividir fusion: {str(e)}")


# ============= DIVIDIR CONSECUTIVO (una planilla -> varios carros) =============

def _generar_consecutivo_division(consecutivo_original: str, num_carros: int) -> list:
    """
    Genera los consecutivos con letra para dividir una planilla en varios carros.
    Reutiliza el número base de la original y asigna las primeras letras libres.
    Ej: "BARRANQUILLA-20260619-3", 2 carros ->
        [{"consecutivo": "...-3A", "consecutivo_base": "...-3", "numero": 3, "letra": "A"},
         {"consecutivo": "...-3B", ...}]
    """
    if not consecutivo_original:
        raise HTTPException(status_code=400, detail="La planilla no tiene consecutivo para dividir")

    parts = consecutivo_original.split("-")
    if len(parts) < 3:
        raise HTTPException(status_code=400, detail=f"Consecutivo inválido para dividir: {consecutivo_original}")

    solo_digitos = "".join(ch for ch in parts[-1] if ch.isdigit())
    if not solo_digitos:
        raise HTTPException(status_code=400, detail=f"Consecutivo sin número base: {consecutivo_original}")
    numero_base = int(solo_digitos)
    prefijo = "-".join(parts[:-1])  # ej: "BARRANQUILLA-20260619"

    # Letras ya usadas para este número base (en activos e histórico)
    regex_pattern = f"^{prefijo}-{numero_base}[A-Z]?$"
    _query = {"consecutivo": {"$regex": regex_pattern}}
    _projection = {"letra_consecutivo": 1, "consecutivo": 1}
    existentes = list(coleccion_pedidos_medical.find(_query, _projection)) \
        + list(coleccion_historico.find(_query, _projection))

    letras_usadas = set()
    for doc in existentes:
        letra = doc.get("letra_consecutivo")
        if not letra:
            ult = doc.get("consecutivo", "").split("-")[-1]
            letra = "".join(ch for ch in ult if ch.isalpha())
        if letra:
            letras_usadas.add(str(letra).upper())

    letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    consecutivos = []
    letra_idx = 0
    for _ in range(num_carros):
        while letra_idx < len(letras) and letras[letra_idx] in letras_usadas:
            letra_idx += 1
        if letra_idx >= len(letras):
            raise HTTPException(status_code=500, detail="No hay más letras disponibles para dividir")
        letra = letras[letra_idx]
        consecutivos.append({
            "consecutivo": f"{prefijo}-{numero_base}{letra}",
            "consecutivo_base": f"{prefijo}-{numero_base}",
            "numero": numero_base,
            "letra": letra,
        })
        letras_usadas.add(letra)
        letra_idx += 1
    return consecutivos


class DividirConsecutivoRequest(BaseModel):
    """Divide una planilla en varios carros. El frontend ya calculó tipo_vehiculo y tarifa por carro."""
    planilla: str
    usuario: Optional[str] = None
    carros: List[dict]  # [{peso, tipo_vehiculo, tarifa_calculada, tarifa_base, total_solicitado}, ...]


class UnirCarrosRequest(BaseModel):
    """Revierte una división: elimina los carros y reconstruye la original."""
    planilla: str
    usuario: Optional[str] = None


@router.post("/dividir-consecutivo")
async def dividir_consecutivo(request: DividirConsecutivoRequest):
    """
    Divide una planilla en N 'carros' (mismo consecutivo base con letra A/B/C/D).
    Valida que la suma de pesos == peso total de la original, duplica los datos de la
    original en cada carro (solo cambian peso/tipo/flete) y elimina la original.
    """
    try:
        original = coleccion_pedidos_medical.find_one({"planilla": request.planilla})
        if not original:
            raise HTTPException(status_code=404, detail=f"Planilla {request.planilla} no encontrada")

        carros = request.carros or []
        if len(carros) < 2 or len(carros) > 4:
            raise HTTPException(status_code=400, detail="Debe dividir entre 2 y 4 carros")

        # Validar suma de pesos == peso total (tolerancia ±1 kg)
        try:
            peso_total_original = float(original.get("peso_real", 0) or 0)
        except (TypeError, ValueError):
            peso_total_original = 0
        suma_pesos = 0.0
        for c in carros:
            try:
                suma_pesos += float(c.get("peso", 0) or 0)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="Los pesos deben ser numéricos")
        if abs(suma_pesos - peso_total_original) > 1:
            raise HTTPException(
                status_code=400,
                detail=f"La suma de pesos ({suma_pesos}) no coincide con el peso total de la planilla ({peso_total_original})"
            )

        # Generar consecutivos con letra (3A, 3B, ...)
        consecutivos = _generar_consecutivo_division(original.get("consecutivo", ""), len(carros))

        # Snapshot de la original para poder revertir (unir)
        snapshot_campos = [
            "planilla", "piezas", "peso_real", "ruta", "codigo_pedido", "cantidad_pedidos",
            "cliente_origen", "municipio_destino", "departamento_destino", "regional",
            "centro_costo", "tarifa_calculada", "tipo_vehiculo", "total_solicitado",
            "tarifa_base", "requiere_descargue", "punto_adicional", "desvio", "aforo",
            "placa", "tipo_veh_sicetac", "cantidad_destinos", "municipios_destino_lista",
            "municipios_con_pedidos", "consecutivo", "consecutivo_base", "numero_consecutivo",
            "flete_cobrado_fmc", "estado", "aprobado_por", "fecha_aprobacion", "registros_detalle",
        ]
        datos_original = {k: original.get(k) for k in snapshot_campos if k in original}

        fecha_now = datetime.now()
        carros_creados = []   # [{planilla, letra, peso}] para division_info
        docs_a_insertar = []

        for i, carro in enumerate(carros):
            cons = consecutivos[i]
            peso_carro = float(carro.get("peso", 0) or 0)
            tarifa = float(carro.get("tarifa_calculada", 0) or 0)
            tarifa_base = carro.get("tarifa_base")
            if tarifa_base is None:
                tarifa_base = tarifa
            tarifa_base = float(tarifa_base or 0)
            total_solicitado = float(carro.get("total_solicitado", 0) or 0)
            tipo_veh = carro.get("tipo_vehiculo") or original.get("tipo_vehiculo", "")

            planilla_carro = f"{request.planilla}-{cons['letra']}"

            doc = {
                "usuario_registro": request.usuario or original.get("usuario_registro"),
                "perfil": original.get("perfil"),
                "centro_distribucion": original.get("centro_distribucion"),
                "planilla": planilla_carro,
                "encontrada": True,
                "piezas": original.get("piezas", 0),
                "peso_real": peso_carro,
                "ruta": original.get("ruta", "-"),
                "codigo_pedido": original.get("codigo_pedido", "-"),
                "cantidad_pedidos": original.get("cantidad_pedidos", 0),
                "cliente_origen": original.get("cliente_origen", "-"),
                "municipio_destino": original.get("municipio_destino", "-"),
                "departamento_destino": original.get("departamento_destino", "-"),
                "regional": original.get("regional"),
                "centro_costo": original.get("centro_costo"),
                "tarifa_calculada": tarifa,
                "tipo_vehiculo": tipo_veh,
                "total_solicitado": total_solicitado,
                "diferencia": total_solicitado - tarifa,
                "flete_cobrado_fmc": original.get("flete_cobrado_fmc", 0),
                "cantidad_destinos": original.get("cantidad_destinos", 0),
                "municipios_destino_lista": original.get("municipios_destino_lista", "-"),
                "municipios_con_pedidos": original.get("municipios_con_pedidos", {}),
                "fusion_info": None,
                "tarifa_base": tarifa_base,
                "requiere_descargue": original.get("requiere_descargue", 0),
                "punto_adicional": original.get("punto_adicional", 0),
                "desvio": original.get("desvio", 0),
                "aforo": original.get("aforo"),
                "placa": original.get("placa"),
                "tipo_veh_sicetac": original.get("tipo_veh_sicetac"),
                "fecha_creacion": fecha_now,
                "estado": "PREAPROBADO",
                "aprobado_por": None,
                "fecha_aprobacion": None,
                "consecutivo": cons["consecutivo"],
                "consecutivo_base": cons["consecutivo_base"],
                "numero_consecutivo": cons["numero"],
                "letra_consecutivo": cons["letra"],
                "es_fusionada_consecutivo": True,
                "registros_detalle": original.get("registros_detalle", []),
                "division_info": {
                    "es_dividida": True,
                    "planilla_original": request.planilla,
                    "consecutivo_original": original.get("consecutivo"),
                    "datos_original": datos_original,
                    "carros": [],  # se rellena abajo con todos los carros
                    "fecha_division": fecha_now,
                    "usuario": request.usuario,
                },
            }
            carros_creados.append({"planilla": planilla_carro, "letra": cons["letra"], "peso": peso_carro})
            docs_a_insertar.append(doc)

        # Rellenar la lista de carros en el division_info de cada uno
        for doc in docs_a_insertar:
            doc["division_info"]["carros"] = carros_creados

        # Insertar carros y eliminar la original
        for doc in docs_a_insertar:
            coleccion_pedidos_medical.insert_one(doc)
        coleccion_pedidos_medical.delete_one({"planilla": request.planilla})

        consecutivos_map = {
            d["planilla"]: {"consecutivo": d["consecutivo"], "consecutivo_base": d["consecutivo_base"]}
            for d in docs_a_insertar
        }
        logger.info(
            f"[DIVIDIR CONSECUTIVO] {request.planilla} -> {len(docs_a_insertar)} carros: "
            f"{[c['planilla'] for c in carros_creados]}"
        )

        return {
            "mensaje": f"Planilla dividida en {len(docs_a_insertar)} carros",
            "carros": [dict({k: v for k, v in d.items() if k != "_id"}, guardado=True, encontrada=True) for d in docs_a_insertar],
            "consecutivos": consecutivos_map,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al dividir consecutivo: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al dividir consecutivo: {str(e)}")


@router.post("/unir-carros")
async def unir_carros(request: UnirCarrosRequest):
    """
    Revierte una división de consecutivo: elimina todos los carros y reconstruye
    la planilla original desde division_info.datos_original.
    """
    try:
        carro = coleccion_pedidos_medical.find_one({"planilla": request.planilla})
        if not carro:
            raise HTTPException(status_code=404, detail=f"Planilla {request.planilla} no encontrada")

        division_info = carro.get("division_info") or {}
        if not division_info.get("es_dividida"):
            raise HTTPException(status_code=400, detail="Esta planilla no es un carro dividido")

        planillas_carros = [c.get("planilla") for c in (division_info.get("carros") or []) if c.get("planilla")]
        if request.planilla not in planillas_carros:
            planillas_carros.append(request.planilla)

        # Eliminar todos los carros (activos e histórico)
        if planillas_carros:
            coleccion_pedidos_medical.delete_many({"planilla": {"$in": planillas_carros}})
            coleccion_historico.delete_many({"planilla": {"$in": planillas_carros}})

        # Reconstruir la original desde el snapshot
        d = division_info.get("datos_original") or {}
        original_restaurada = {
            "usuario_registro": request.usuario or carro.get("usuario_registro"),
            "perfil": carro.get("perfil"),
            "centro_distribucion": carro.get("centro_distribucion"),
            "planilla": division_info.get("planilla_original") or d.get("planilla"),
            "encontrada": True,
            "piezas": d.get("piezas", 0),
            "peso_real": d.get("peso_real", 0),
            "ruta": d.get("ruta", "-"),
            "codigo_pedido": d.get("codigo_pedido", "-"),
            "cantidad_pedidos": d.get("cantidad_pedidos", 0),
            "cliente_origen": d.get("cliente_origen", "-"),
            "municipio_destino": d.get("municipio_destino", "-"),
            "departamento_destino": d.get("departamento_destino", "-"),
            "regional": d.get("regional"),
            "centro_costo": d.get("centro_costo"),
            "tarifa_calculada": d.get("tarifa_calculada", 0),
            "tipo_vehiculo": d.get("tipo_vehiculo", "-"),
            "total_solicitado": d.get("total_solicitado", 0),
            "diferencia": (d.get("total_solicitado", 0) or 0) - (d.get("tarifa_calculada", 0) or 0),
            "flete_cobrado_fmc": d.get("flete_cobrado_fmc", 0),
            "cantidad_destinos": d.get("cantidad_destinos", 0),
            "municipios_destino_lista": d.get("municipios_destino_lista", "-"),
            "municipios_con_pedidos": d.get("municipios_con_pedidos", {}),
            "fusion_info": None,
            "tarifa_base": d.get("tarifa_base"),
            "requiere_descargue": d.get("requiere_descargue", 0),
            "punto_adicional": d.get("punto_adicional", 0),
            "desvio": d.get("desvio", 0),
            "aforo": d.get("aforo"),
            "placa": d.get("placa"),
            "tipo_veh_sicetac": d.get("tipo_veh_sicetac"),
            "fecha_creacion": datetime.now(),
            "estado": d.get("estado", "PREAPROBADO"),
            "aprobado_por": d.get("aprobado_por"),
            "fecha_aprobacion": d.get("fecha_aprobacion"),
            "consecutivo": d.get("consecutivo"),
            "consecutivo_base": d.get("consecutivo_base"),
            "numero_consecutivo": d.get("numero_consecutivo"),
            "letra_consecutivo": None,
            "es_fusionada_consecutivo": False,
            "registros_detalle": d.get("registros_detalle", []),
            "division_info": None,
        }
        coleccion_pedidos_medical.insert_one(original_restaurada)
        logger.info(f"[UNIR CARROS] Restaurada original {original_restaurada['planilla']} desde división")

        return {
            "mensaje": f"Carros unidos; restaurada planilla {original_restaurada['planilla']}",
            "planilla": dict({k: v for k, v in original_restaurada.items() if k != "_id"}, guardado=True, encontrada=True),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al unir carros: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al unir carros: {str(e)}")


@router.put("/actualizar-planilla-pedidos")
async def actualizar_planilla_pedidos(request: ActualizarPlanillaPedidosRequest):
    """
    Actualiza una planilla específica en la colección pedidos_medical.
    Incluye trazabilidad completa de modificaciones.
    """
    try:
        logger.info(f"[ACTUALIZAR PLANILLA PEDIDOS] Planilla: {request.planilla}, Usuario: {request.usuario_modificacion}")

        # Obtener el documento actual antes de actualizar (para trazabilidad)
        doc_actual = coleccion_pedidos_medical.find_one({"planilla": request.planilla})
        if not doc_actual:
            logger.warning(f"[ACTUALIZAR PLANILLA PEDIDOS] No se encontró planilla: {request.planilla}")
            raise HTTPException(status_code=404, detail=f"Planilla {request.planilla} no encontrada en pedidos_medical")

        fecha_actual = datetime.now()

        # Crear registro de historial de cambios
        historial_cambios = doc_actual.get("historial_cambios", [])

        # Detectar qué campos cambiaron y registrar en historial
        campos_modificados = []
        if request.tarifa_base is not None and request.tarifa_base != doc_actual.get("tarifa_base"):
            campos_modificados.append({
                "campo": "tarifa_base",
                "valor_anterior": doc_actual.get("tarifa_base"),
                "valor_nuevo": request.tarifa_base
            })
        if request.tarifa_calculada is not None and request.tarifa_calculada != doc_actual.get("tarifa_calculada"):
            campos_modificados.append({
                "campo": "tarifa_calculada",
                "valor_anterior": doc_actual.get("tarifa_calculada"),
                "valor_nuevo": request.tarifa_calculada
            })
        if request.requiere_descargue != doc_actual.get("requiere_descargue"):
            campos_modificados.append({
                "campo": "requiere_descargue",
                "valor_anterior": doc_actual.get("requiere_descargue"),
                "valor_nuevo": request.requiere_descargue
            })
        if request.punto_adicional != doc_actual.get("punto_adicional"):
            campos_modificados.append({
                "campo": "punto_adicional",
                "valor_anterior": doc_actual.get("punto_adicional"),
                "valor_nuevo": request.punto_adicional
            })
        if request.desvio != doc_actual.get("desvio"):
            campos_modificados.append({
                "campo": "desvio",
                "valor_anterior": doc_actual.get("desvio"),
                "valor_nuevo": request.desvio
            })
        if request.aforo is not None and request.aforo != doc_actual.get("aforo"):
            campos_modificados.append({
                "campo": "aforo",
                "valor_anterior": doc_actual.get("aforo"),
                "valor_nuevo": request.aforo
            })
        if request.placa is not None and request.placa != doc_actual.get("placa"):
            campos_modificados.append({
                "campo": "placa",
                "valor_anterior": doc_actual.get("placa"),
                "valor_nuevo": request.placa
            })
        if request.tipo_veh_sicetac is not None and request.tipo_veh_sicetac != doc_actual.get("tipo_veh_sicetac"):
            campos_modificados.append({
                "campo": "tipo_veh_sicetac",
                "valor_anterior": doc_actual.get("tipo_veh_sicetac"),
                "valor_nuevo": request.tipo_veh_sicetac
            })
        if request.causal != doc_actual.get("causal"):
            campos_modificados.append({
                "campo": "causal",
                "valor_anterior": doc_actual.get("causal"),
                "valor_nuevo": request.causal
            })

        # Registrar cambio de estado
        estado_anterior = doc_actual.get("estado", "PREAPROBADO")
        if request.estado is not None and request.estado != estado_anterior:
            campos_modificados.append({
                "campo": "estado",
                "valor_anterior": estado_anterior,
                "valor_nuevo": request.estado
            })

        # Si hay cambios, agregar al historial
        if campos_modificados:
            nuevo_historial = {
                "fecha": fecha_actual,
                "usuario": request.usuario_modificacion,
                "accion": "edicion",
                "campos_modificados": campos_modificados,
                "causal": request.causal
            }
            historial_cambios.append(nuevo_historial)

        # Calcular diferencia contra la tarifa teórica que queda guardada.
        tarifa_calculada_final = (
            request.tarifa_calculada
            if request.tarifa_calculada is not None
            else (doc_actual.get("tarifa_calculada", 0) or 0)
        )
        diferencia = (request.total_solicitado or 0) - tarifa_calculada_final

        # Campos a actualizar
        campos_actualizar = {
            "tarifa_base": request.tarifa_base,
            "tarifa_calculada": tarifa_calculada_final,
            "requiere_descargue": request.requiere_descargue,
            "punto_adicional": request.punto_adicional,
            "desvio": request.desvio,
            "aforo": request.aforo,
            "placa": request.placa,
            "tipo_veh_sicetac": request.tipo_veh_sicetac,
            "peso_sicetac": request.peso_sicetac,
            "total_solicitado": request.total_solicitado,
            "diferencia": diferencia,
            "causal": request.causal,
            # Trazabilidad de modificación
            "usuario_modificacion": request.usuario_modificacion,
            "fecha_modificacion": fecha_actual,
            "historial_cambios": historial_cambios
        }

        # Si se envía ruta (edición de ruta), actualizarla
        if request.ruta is not None:
            campos_actualizar["ruta"] = request.ruta

        # Si se envía municipio_destino (cambio manual del municipio principal), actualizarlo.
        # Es un cambio puramente de etiqueta: NO afecta tarifa, total, estado ni la lista de municipios.
        if request.municipio_destino is not None:
            campos_actualizar["municipio_destino"] = request.municipio_destino
            if request.municipio_destino != doc_actual.get("municipio_destino"):
                campos_modificados.append({
                    "campo": "municipio_destino",
                    "valor_anterior": doc_actual.get("municipio_destino"),
                    "valor_nuevo": request.municipio_destino
                })

        # Si se envía estado, actualizarlo
        if request.estado is not None:
            campos_actualizar["estado"] = request.estado
            # Registrar la fecha en que dejó de ser CREADO (visible para todos) la primera
            # vez que ocurre; no sobreescribir si ya existía.
            if doc_actual.get("estado") == "CREADO" and request.estado != "CREADO" and not doc_actual.get("fecha_preaprobado"):
                campos_actualizar["fecha_preaprobado"] = fecha_actual
            if request.estado == "REQUIERE_APROBACION" and estado_anterior != "REQUIERE_APROBACION":
                # Si cambia a REQUIERE_APROBACION, registrar quién solicitó autorización
                campos_actualizar["usuario_solicitud_autorizacion"] = request.usuario_modificacion
                campos_actualizar["fecha_solicitud_autorizacion"] = fecha_actual
            if request.estado != "APROBADO":
                # Limpiar campos de aprobación si no está aprobada
                campos_actualizar["aprobado_por"] = None
                campos_actualizar["fecha_aprobacion"] = None
            else:
                campos_actualizar["aprobado_por"] = request.aprobado_por
                campos_actualizar["fecha_aprobacion"] = request.fecha_aprobacion

        # Actualizar el documento
        resultado = coleccion_pedidos_medical.update_one(
            {"planilla": request.planilla},
            {"$set": campos_actualizar}
        )

        if resultado.matched_count == 0:
            logger.warning(f"[ACTUALIZAR PLANILLA PEDIDOS] No se encontró planilla: {request.planilla}")
            raise HTTPException(status_code=404, detail=f"Planilla {request.planilla} no encontrada en pedidos_medical")

        logger.info(f"[ACTUALIZAR PLANILLA PEDIDOS] Planilla {request.planilla} actualizada - Modified: {resultado.modified_count}")
        logger.info(f"[TRAZABILIDAD] Usuario: {request.usuario_modificacion}, Fecha: {fecha_actual}")

        return {
            "mensaje": "Planilla actualizada exitosamente en pedidos_medical",
            "planilla": request.planilla,
            "modified_count": resultado.modified_count,
            "trazabilidad": {
                "usuario_modificacion": request.usuario_modificacion,
                "fecha_modificacion": fecha_actual.isoformat(),
                "estado_anterior": estado_anterior,
                "estado_nuevo": request.estado
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al actualizar planilla en pedidos_medical: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al actualizar planilla: {str(e)}"
        )


@router.put("/actualizar-estado-planilla")
async def actualizar_estado_planilla(request: ActualizarEstadoPlanillaRequest):
    """
    Actualiza el estado de aprobación de una planilla en pedidos_medical.
    Incluye trazabilidad de aprobación en el historial.
    """
    try:
        logger.info(f"=== ACTUALIZAR ESTADO PLANILLA ===")
        logger.info(f"Planilla: {request.planilla}")
        logger.info(f"Estado: {request.estado}")
        logger.info(f"Aprobado por: {request.aprobado_por}")

        # Obtener el documento actual antes de actualizar (para trazabilidad)
        doc_actual = coleccion_pedidos_medical.find_one({"planilla": request.planilla})
        if not doc_actual:
            logger.warning(f"[ACTUALIZAR ESTADO PLANILLA] No se encontró planilla: {request.planilla}")
            raise HTTPException(status_code=404, detail=f"Planilla {request.planilla} no encontrada en pedidos_medical")

        fecha_actual = _hora_confiable_utc()
        estado_anterior = doc_actual.get("estado", "PREAPROBADO")

        # Obtener historial existente
        historial_cambios = doc_actual.get("historial_cambios", [])

        # Si el estado cambió, agregar al historial
        if request.estado != estado_anterior:
            nuevo_historial = {
                "fecha": fecha_actual,
                "usuario": request.aprobado_por,
                "accion": "cambio_estado",
                "campos_modificados": [
                    {
                        "campo": "estado",
                        "valor_anterior": estado_anterior,
                        "valor_nuevo": request.estado
                    }
                ]
            }
            historial_cambios.append(nuevo_historial)

            logger.info(f"[TRAZABILIDAD] Cambio de estado: {estado_anterior} → {request.estado} por {request.aprobado_por}")

            # Volver a CREADO es una reapertura de edición (analista/admin): no avisa a
            # analistas ni solicita autorización (esos avisos son solo al salir de CREADO).
            if request.estado != "CREADO":
                # Notificar a analistas si el cambio fue de CREADO a estado visible
                _notificar_analistas_cambio_estado(doc_actual, estado_anterior, request.estado)

                # Notificar a coordinadores/control si requiere su autorización
                _notificar_solicitud_autorizacion(doc_actual, request.estado)

        # Campos a actualizar
        campos_actualizar = {
            "estado": request.estado,
            "aprobado_por": request.aprobado_por,
            "fecha_aprobacion": fecha_actual if request.estado == "APROBADO" else None,
            "historial_cambios": historial_cambios
        }

        # Registrar la fecha en que dejó de ser CREADO (visible para todos), sea cual sea
        # el estado destino (PREAPROBADO, REQUIERE_APROBACION_* o APROBADO). En otros
        # movimientos (p.ej. entre estados ya visibles) se conserva la existente.
        if doc_actual.get("estado") == "CREADO" and request.estado != "CREADO":
            campos_actualizar["fecha_preaprobado"] = fecha_actual
        elif request.estado == "CREADO":
            # Reapertura: la planilla vuelve a ser borrador del operativo (deja de ser
            # visible para todos), así que se limpia su fecha de visibilidad.
            campos_actualizar["fecha_preaprobado"] = None

        # Actualizar el documento
        resultado = coleccion_pedidos_medical.update_one(
            {"planilla": request.planilla},
            {"$set": campos_actualizar}
        )

        if resultado.matched_count == 0:
            logger.warning(f"No se encontró planilla: {request.planilla}")
            raise HTTPException(status_code=404, detail=f"Planilla {request.planilla} no encontrada en pedidos_medical")

        logger.info(f"Planilla {request.planilla} actualizada - Estado: {request.estado}")

        return {
            "mensaje": f"Planilla actualizada a estado {request.estado}",
            "planilla": request.planilla,
            "estado": request.estado,
            "estado_anterior": estado_anterior,
            "modified_count": resultado.modified_count
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al actualizar estado de planilla: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al actualizar estado: {str(e)}"
        )


class EliminarPlanillaRequest(BaseModel):
    """Modelo para eliminar una planilla"""
    planilla: str
    usuario: str  # Usuario que elimina (trazabilidad)


@router.delete("/eliminar-planilla")
async def eliminar_planilla(request: EliminarPlanillaRequest):
    """
    Elimina una planilla de la colección pedidos_medical.
    Si es una planilla fusionada, también elimina las planillas originales marcadas como fusionada: true.
    Incluye trazabilidad de quién eliminó.
    """
    try:
        logger.info(f"=== ELIMINAR PLANILLA ===")
        logger.info(f"Planilla: {request.planilla}")
        logger.info(f"Usuario: {request.usuario}")

        # Verificar si existe la planilla
        doc_actual = coleccion_pedidos_medical.find_one({"planilla": request.planilla})
        if not doc_actual:
            logger.warning(f"[ELIMINAR PLANILLA] No se encontró planilla: {request.planilla}")
            raise HTTPException(status_code=404, detail=f"Planilla {request.planilla} no encontrada en pedidos_medical")

        # Verificar si es una planilla fusionada
        fusion_info = doc_actual.get("fusion_info")
        es_fusionada = fusion_info and fusion_info.get("es_fusionada") == True

        planillas_eliminadas = [request.planilla]

        # Si es una planilla fusionada, eliminar también las planillas originales marcadas como fusionada: true
        if es_fusionada:
            logger.info(f"[ELIMINAR PLANILLA] La planilla {request.planilla} es una fusión, eliminando originales también")

            # Buscar planillas originales marcadas como fusionada: true que apuntan a esta planilla
            planillas_originales_fusionadas = list(coleccion_pedidos_medical.find({
                "fusionada": True,
                "fusionada_en.planilla_fusionada": request.planilla
            }))

            logger.info(f"[ELIMINAR PLANILLA] Encontradas {len(planillas_originales_fusionadas)} planillas originales fusionadas")

            # Eliminar cada planilla original
            for doc_original in planillas_originales_fusionadas:
                planilla_original = doc_original.get("planilla")
                logger.info(f"[ELIMINAR PLANILLA] Eliminando planilla original fusionada: {planilla_original}")
                coleccion_pedidos_medical.delete_one({"planilla": planilla_original})
                planillas_eliminadas.append(planilla_original)

        # Eliminar la planilla principal
        resultado = coleccion_pedidos_medical.delete_one({"planilla": request.planilla})

        if resultado.deleted_count == 0:
            logger.warning(f"[ELIMINAR PLANILLA] No se pudo eliminar planilla: {request.planilla}")
            raise HTTPException(status_code=500, detail=f"No se pudo eliminar la planilla {request.planilla}")

        logger.info(f"[ELIMINAR PLANILLA] Planilla {request.planilla} eliminada por {request.usuario}")
        if es_fusionada:
            logger.info(f"[ELIMINAR PLANILLA] Total planillas eliminadas: {len(planillas_eliminadas)} - {planillas_eliminadas}")

        return {
            "mensaje": f"Planilla {request.planilla} eliminada exitosamente",
            "planilla": request.planilla,
            "deleted_count": resultado.deleted_count,
            "eliminado_por": request.usuario,
            "es_fusionada": es_fusionada,
            "planillas_eliminadas": planillas_eliminadas,
            "total_eliminadas": len(planillas_eliminadas)
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al eliminar planilla: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al eliminar planilla: {str(e)}"
        )


def _mapear_tipo_vehiculo(tipo_vehiculo: str) -> str:
    """Mapea el tipo de vehículo interno al formato esperado por el sistema de transporte."""
    tipo_upper = tipo_vehiculo.upper() if tipo_vehiculo else ""
    if tipo_upper == "CARRY":
        return "CARRY"
    elif tipo_upper == "NHR":
        return "CAMIONETA"
    elif tipo_upper == "TURBO":
        return "TURBO"
    elif tipo_upper in {"NIES", "SENCILLO"}:
        return "SENCILLO"
    elif tipo_upper == "PATINETA":
        return "TRACTOCAMION"
    return tipo_vehiculo


def _repartir_flete(total_solicitado, piezas_list):
    """
    Reparte 'total_solicitado' entre las filas en proporción a sus piezas (cajas),
    usando división entera. La última fila absorbe el residuo de redondeo para que
    la suma de las partes sea EXACTAMENTE igual al total.
    Devuelve una lista de enteros del mismo len que piezas_list.
    """
    n = len(piezas_list)
    if n == 0:
        return []

    try:
        total = int(round(float(total_solicitado or 0)))
    except (TypeError, ValueError):
        total = 0

    piezas_int = []
    for p in piezas_list:
        try:
            piezas_int.append(int(round(float(p or 0))))
        except (TypeError, ValueError):
            piezas_int.append(0)

    total_piezas = sum(piezas_int)

    # Si no hay piezas para repartir, distribución equitativa con residuo en la última
    if total_piezas <= 0:
        base = total // n
        reparto = [base] * n
        reparto[-1] = total - base * (n - 1)
        return reparto

    reparto = []
    acumulado = 0
    for i in range(n - 1):
        parte = total * piezas_int[i] // total_piezas
        reparto.append(parte)
        acumulado += parte
    # La última fila absorbe el residuo para garantizar suma exacta
    reparto.append(total - acumulado)
    return reparto


def _consecutivo_original(datos_original, indice, consecutivo_fusionado):
    """
    Resuelve el consecutivo de una planilla original para la fila del Excel.
    1) datos_original['consecutivo'] si existe y no es vacío.
    2) Si no, deriva del consecutivo fusionado incrementando el sufijo alfabético
       por el índice (FUNZA-20260618-1A -> A=0, B=1, C=2...).
    3) Si el fusionado no termina en letra, lo usa tal cual y loggea warning.
    """
    cons = (datos_original or {}).get("consecutivo")
    if cons and str(cons).strip():
        return cons

    import re
    if consecutivo_fusionado:
        match = re.match(r"^(.*?)([A-Za-z]+)$", consecutivo_fusionado)
        if match:
            base = match.group(1)
            letra_base = match.group(2)[-1].upper()
            nueva_letra = chr(ord(letra_base) + indice)
            return f"{base}{nueva_letra}"
        logger.warning(
            f"Consecutivo fusionado '{consecutivo_fusionado}' sin sufijo alfabético; "
            f"se usará el mismo consecutivo para todas las filas de la fusión."
        )
    return consecutivo_fusionado or ""


def _normalizar_texto_simple(s):
    """Mayúsculas, sin tildes y espacios colapsados (para comparar nombres de cliente)."""
    import unicodedata
    if not s:
        return ""
    s = str(s).strip().upper()
    s = "".join(
        c for c in unicodedata.normalize("NFKD", s)
        if not unicodedata.combining(c)
    )
    return " ".join(s.split())


def _es_cliente_kabi(cliente_origen):
    """True si el cliente origen corresponde a FRESENIUS KABI (robusto a mayúsculas/tildes)."""
    return _normalizar_texto_simple(cliente_origen) == "FRESENIUS KABI"


def _expandir_fila_kabi(fila):
    """
    Para FRESENIUS KABI con registros_detalle: devuelve una SOLA fila (la original), con
    'Ubicación Descargue' = 'FKC_<Nombre>_<Cedula>' del PRIMER destinatario del detalle.
    Antes se generaba una fila por destinatario; ahora se colapsa a una sola fila para no
    repetir líneas en el Excel.
    Si ningún destinatario trae Nombre, la fila queda con Ubicación Descargue por divipolas.
    Para otros clientes (p.ej. FRESENIUS MEDICAL CARE) o sin detalle: devuelve [fila] sin cambios.
    Consume (pop) 'registros_detalle' de la fila para que no llegue al writer de Excel.
    """
    detalle = fila.pop("registros_detalle", None)
    if not _es_cliente_kabi(fila.get("cliente_origen")) or not detalle:
        return [fila]

    # Tomar el primer destinatario con Nombre no vacío.
    primer = None
    for item in detalle:
        nombre = (item.get("Nombre") or "").strip() if isinstance(item, dict) else ""
        if nombre:
            primer = item
            break

    if primer is not None:
        nombre = (primer.get("Nombre") or "").strip()
        cedula = (primer.get("Cedula") or "").strip()
        fila["ubicacion_descargue_override"] = f"FKC_{nombre}_{cedula}"
        logger.info(
            f"[EXPORTAR] KABI {fila.get('consecutivo', '')}: 1 fila con "
            f"Ubicación Descargue = FKC_{nombre}_{cedula} (primer destinatario)"
        )
    else:
        logger.info(
            f"[EXPORTAR] KABI {fila.get('consecutivo', '')}: 1 fila sin destinatario "
            f"con Nombre; Ubicación Descargue por divipolas."
        )

    return [fila]


def _expandir_doc_a_filas(doc):
    """
    Convierte un documento de pedidos_medical en una lista de 'filas lógicas' para el Excel.
    - Planilla NO fusionada (o sin datos_originales): devuelve 1 fila con los campos del doc.
    - Planilla fusionada: devuelve N filas (una por planilla original), con el Flete unidad
      (total_solicitado del doc fusionado) repartido proporcionalmente por piezas.
    Cada fila es un dict con los campos que _escribir_fila_planilla espera (kwargs).
    """
    fusion_info = doc.get("fusion_info") or {}
    es_fusionada = fusion_info.get("es_fusionada") is True
    datos_originales = fusion_info.get("datos_originales") or []

    # Caso normal: una sola fila con los datos del documento (comportamiento histórico)
    if not es_fusionada or not datos_originales:
        if es_fusionada and not datos_originales:
            logger.warning(
                f"Planilla {doc.get('planilla')} marcada como fusionada pero sin "
                f"datos_originales; se exporta como una sola fila."
            )
        return [{
            "consecutivo": doc.get("consecutivo", ""),
            "regional_doc": doc.get("regional"),
            "municipio_destino": doc.get("municipio_destino", ""),
            "codigo_pedido": doc.get("codigo_pedido", ""),
            "cliente_origen": doc.get("cliente_origen", ""),
            "tipo_vehiculo": doc.get("tipo_veh_sicetac") or doc.get("tipo_vehiculo", ""),
            "piezas": doc.get("piezas", 0),
            "peso_real": doc.get("peso_real", 0),
            "peso_sicetac": doc.get("peso_sicetac", doc.get("peso_real", 0)),
            "flete_unidad": doc.get("total_solicitado", 0),
            "punto_adicional_val": doc.get("punto_adicional", 0),
            "requiere_descargue_val": doc.get("requiere_descargue", 0),
            "registros_detalle": doc.get("registros_detalle", []),
        }]

    # Caso fusionado: repartir el flete total proporcionalmente por piezas (cajas)
    consecutivo_fusionado = doc.get("consecutivo", "")
    total_flete = doc.get("total_solicitado", 0)

    piezas_list = []
    for d in datos_originales:
        try:
            piezas_list.append(int(round(float(d.get("piezas", 0) or 0))))
        except (TypeError, ValueError):
            piezas_list.append(0)

    reparto = _repartir_flete(total_flete, piezas_list)

    logger.info(
        f"[EXPORTAR] Planilla fusionada {doc.get('planilla')} -> {len(datos_originales)} filas, "
        f"flete total {total_flete} repartido {reparto} según piezas {piezas_list}"
    )

    filas = []
    for i, d in enumerate(datos_originales):
        # tipo_vehiculo: TODAS las filas del fusionado usan el tipo SICETAC del vehículo
        # fusionado (el real del camión); cae al original y luego al tipo interno si faltan.
        # placa: se replica del doc fusionado si el original no la trae.
        tipo_veh = (doc.get("tipo_veh_sicetac") or d.get("tipo_veh_sicetac")
                    or doc.get("tipo_vehiculo") or d.get("tipo_vehiculo", ""))
        placa_val = d.get("placa")
        if placa_val is None:
            placa_val = doc.get("placa", "")
        filas.append({
            "consecutivo": _consecutivo_original(d, i, consecutivo_fusionado),
            "regional_doc": d.get("regional") or doc.get("regional"),
            "municipio_destino": d.get("municipio_destino", ""),
            "codigo_pedido": d.get("codigo_pedido", ""),
            "cliente_origen": d.get("cliente_origen", ""),
            "tipo_vehiculo": tipo_veh,
            "piezas": d.get("piezas", 0),
            "peso_real": d.get("peso_real", 0),
            "peso_sicetac": d.get("peso_sicetac", d.get("peso_real", 0)),
            "flete_unidad": reparto[i] if i < len(reparto) else 0,
            "punto_adicional_val": d.get("punto_adicional", 0),
            "requiere_descargue_val": d.get("requiere_descargue", 0),
            "registros_detalle": d.get("registros_detalle", []),
        })
    return filas


# Homologación de destinos para el Excel de aprobados (hoja "plantilla").
# Si el municipio de destino coincide con una clave (comparación en MAYÚSCULAS y sin
# espacios de más) se reemplaza por el valor al escribir la columna "Destino".
# Agregar aquí los renombrados que se vayan requiriendo homologar.
DESTINOS_RENOMBRAR_EXCEL = {
    "SANTIAGO DE CALI": "CALI",
    "LA UNION": "LA UNION ANT.",
}


def _renombrar_destino_excel(destino):
    """Reemplaza el nombre del destino por su forma homologada para el Excel de aprobados."""
    if not destino:
        return destino
    return DESTINOS_RENOMBRAR_EXCEL.get(str(destino).strip().upper(), destino)


def _escribir_fila_planilla(
    ws, row_num, *,
    consecutivo, regional_doc, municipio_destino, codigo_pedido,
    cliente_origen, tipo_vehiculo, piezas, peso_real,
    flete_unidad, punto_adicional_val, requiere_descargue_val,
    regional_usuario, divipolas_lookup, divipolas_por_poblacion,
    mapear_tipo_vehiculo, thin_border,
    ubicacion_descargue_override=None, peso_sicetac=None,
):
    """
    Escribe una fila de planilla en la hoja Excel y devuelve row_num + 1.
    Recibe los campos ya resueltos (de un doc normal o de un datos_original expandido).
    """
    # Regional del documento con fallback a la regional del usuario
    regional_doc = regional_doc or regional_usuario

    # Origen: reemplazar la regional por el municipio real de la bodega de origen
    _origen_map = {
        "BARRANQUILLA": "GALAPA",
        "CALI": "YUMBO",
    }
    origen = _origen_map.get(str(regional_doc).upper().strip(), regional_doc)

    # Tipo de viaje: Urbano si origen == destino, Nacional si son diferentes
    tipo_viaje = "URBANO" if str(regional_doc).upper() == str(municipio_destino).upper() else "NACIONAL"

    # Observación: DN + código pedido
    observacion = f"DN {codigo_pedido}" if codigo_pedido else "DN"

    # CENTRO COSTO: Regional + "CARGA MASIVA OPERACIONES CARGA" + Cliente Origen
    centro_costo = f"{regional_doc} CARGA MASIVA OPERACIONES CARGA {cliente_origen}"

    # Toneladas: Peso SICETAC / 1000 con 2 decimales (con fallback a peso real)
    peso_para_toneladas = peso_sicetac if peso_sicetac else peso_real
    try:
        peso_num = float(peso_para_toneladas) if peso_para_toneladas else 0
        toneladas = round(peso_num / 1000, 2)
    except (TypeError, ValueError):
        toneladas = 0

    # Valor unitario:
    # - FRESENIUS MEDICAL CARE: Piezas * 17000
    # - Otros clientes: Flete unidad / 0.7 (el flete representa el 70% del valor unitario)
    es_fmc = _normalizar_texto_simple(cliente_origen) == "FRESENIUS MEDICAL CARE"
    try:
        if es_fmc:
            piezas_num = int(piezas) if piezas else 0
            valor_unitario = piezas_num * 17000
        else:
            flete_num = float(flete_unidad) if flete_unidad else 0
            valor_unitario = round(flete_num / 0.7) if flete_num else 0
    except (TypeError, ValueError):
        valor_unitario = 0

    # Producto: MEDICAMENTOS... para FRESENIUS MEDICAL CARE y FRESENIUS KABI; VARIOS para los demás.
    producto = (
        "MEDICAMENTOS (CON EXCLUSION DE LOS PRODUCTOS DE LAS PARTIDAS 3002;  30"
        if (es_fmc or _es_cliente_kabi(cliente_origen))
        else "VARIOS"
    )

    # PUNTO ADICIONAL / CARGUE-DESCARGUE (banderas informativas)
    # Ambos van siempre en 0 en el Excel (por definición del formato de carga).
    punto_adicional = 0
    cargue_descargue = 0

    # Ubicación y dirección de descargue desde la colección divipolas.
    # Si hay override (filas duplicadas de FRESENIUS KABI), 'Ubicación Descargue' toma ese valor;
    # 'Direccion Descargue' siempre es el lookup divipolas (igual en todas las filas).
    ubicacion_lookup = ""
    direccion_lookup = ""
    if municipio_destino:
        lookup = divipolas_lookup.get(str(municipio_destino).strip())
        if not lookup:
            lookup = divipolas_por_poblacion.get(str(municipio_destino).strip().upper())
        if lookup:
            ubicacion_lookup = lookup.get("ubicacion_descargue", "")
            direccion_lookup = lookup.get("direccion_descargue", "")
    ubicacion_descargue = ubicacion_descargue_override if ubicacion_descargue_override else ubicacion_lookup
    direccion_descargue = direccion_lookup

    # Ubicación y dirección de cargue según regional
    _cargue_map = {
        "FUNZA":        ("BODEGA INTEGRA EL ROSAL",    "EL ROSAL"),
        "GIRARDOTA":    ("BODEGA INTEGRA GIRARDOTA",   "parque industrial del norte bodega 119"),
        "MEDELLIN":     ("BODEGA INTEGRA GIRARDOTA",   "parque industrial del norte bodega 119"),  # alias bodega GIRARDOTA (analista guarda 'MEDELLIN')
        "BARRANQUILLA": ("BODEGA INTEGRA GALAPA",      "GALAPA"),
        "GALAPA":       ("BODEGA INTEGRA GALAPA",      "GALAPA"),      # alias bodega (operativo guarda 'GALAPA')
        "CALI":         ("BODEGA INTEGRA YUMBO",       "Carrera 31 a #15-320"),
        "YUMBO":        ("BODEGA INTEGRA YUMBO",       "Carrera 31 a #15-320"),  # alias bodega (operativo guarda 'YUMBO')
        "BUCARAMANGA":  ("BODEGA INTEGRA BUCARAMANGA", "Parque industrial provincia de soto 1"),
    }
    _cargue = _cargue_map.get(
        str(regional_doc).upper().strip(),
        ("FME_BODEGA_INTEGRA_FUNZA", "PARQUE INDUSTRIAL SAN DIEGO")
    )
    ubicacion_cargue = _cargue[0]
    direccion_cargue = _cargue[1]

    # SEGURO: solo FRESENIUS KABI lleva 6000; el resto de clientes va en 0.
    seguro = 6000 if _es_cliente_kabi(cliente_origen) else 0

    datos = [
        consecutivo,                                      # Consecutivo
        tipo_viaje,                                       # Tipo de viaje
        "MASIVO",                                         # Linea de negocio
        "PENDIENTE",                                      # Estado
        observacion,                                      # Observación
        CLIENTE_A_NIT.get(str(cliente_origen).upper().strip(), cliente_origen),  # Cliente
        origen,                                           # Origen
        _renombrar_destino_excel(municipio_destino),      # Destino (homologado, ej. SANTIAGO DE CALI -> CALI)
        codigo_pedido,                                    # Pedido cliente
        codigo_pedido,                                    # Guía
        centro_costo,                                     # CENTRO COSTO
        ubicacion_cargue,                                 # Ubicacion Cargue
        direccion_cargue,                                 # Direccion cargue
        ubicacion_descargue,                              # Ubicacion Descargue
        direccion_descargue,                              # Direccion Descargue
        producto,                                        # Producto
        "NORMAL",                                         # Naturaleza
        mapear_tipo_vehiculo(tipo_vehiculo),              # Tipo de vehiculo
        "VEHICULOS",                                      # unidad
        1,                                                # Cantidad (siempre 1)
        "PAQUETES",                                       # Tipo embalaje
        toneladas,                                        # Toneladas
        flete_unidad,                                     # Flete unidad
        punto_adicional,                                  # PUNTO ADICIONAL
        cargue_descargue,                                 # CARGUE-DESCARGUE PER JURIDICA
        seguro,                                           # SEGURO
        "CUPO",                                           # Tipo pago
        0,                                                # Tolerancia
        0,                                                # Vlr hora STBY
        0,                                                # Vlr Declar Mercancia
        1,                                                # Aprobar Poliza
        "CUPO",                                           # Flete por
        valor_unitario,                                   # Valor unitario
        1,                                                # Aprobar cupo credito
        1,                                                # Aprobar rentabilidad
        "FURGON",                                         # Otras caracteristicas
        1,                                                # REMESAS
        1,                                                # REMISION DEL CLIENTE
        1,                                                # GUIA DE TRANSPORTE
        1                                                 # MANIFIESTO
    ]

    # Formatos numéricos por columna (índice 1-based dentro de `datos`).
    # 22 = Toneladas: siempre 2 decimales (ej. 5.00, 12.30). Si se reordena `datos`,
    # ajustar este índice (coincide con la posición de "Toneladas" en el header `columnas`).
    formatos_columna = {22: '0.00'}
    for col_idx, valor in enumerate(datos, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=valor)
        cell.border = thin_border
        if col_idx in formatos_columna:
            cell.number_format = formatos_columna[col_idx]

    return row_num + 1


class ExportarPlanillasExcelRequest(BaseModel):
    """Modelo para exportar planillas a Excel"""
    planillas: List[str]
    perfil: str
    centro_distribucion: Optional[str] = None


@router.post("/exportar-planillas-excel")
async def exportar_planillas_excel(request: ExportarPlanillasExcelRequest):
    """
    Exporta planillas a Excel con formato para sistema de transporte.
    Filtra por regional si el perfil es OPERATIVO.
    """
    try:
        from fastapi.responses import Response
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        import io

        logger.info(f"=== EXPORTAR PLANILLAS EXCEL ===")
        logger.info(f"Perfil: {request.perfil}")
        logger.info(f"Centro distribución: {request.centro_distribucion}")
        logger.info(f"Planillas solicitadas: {len(request.planillas)}")

        # Consultar planillas de MongoDB - SOLO APROBADAS
        consulta = {
            "planilla": {"$in": request.planillas},
            "estado": "APROBADO"
        }

        # Si es OPERATIVO, filtrar por regional
        if request.perfil == "OPERATIVO" and request.centro_distribucion:
            _aplicar_filtro_regional_operativo(consulta, request.centro_distribucion)

        planillas_db = list(coleccion_pedidos_medical.find(consulta))

        logger.info(f"Planillas encontradas en BD: {len(planillas_db)}")

        if not planillas_db:
            raise HTTPException(status_code=404, detail="No se encontraron planillas para exportar")

        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "plantilla"

        # Definir columnas NUEVO FORMATO
        columnas = [
            "Consecutivo", "Tipo de viaje", "Linea de negocio", "Estado", "Observación",
            "Cliente", "Origen", "Destino", "Pedido cliente", "Guia", "CENTRO COSTO",
            "Ubicación Cargue", "Direccion cargue", "Ubicación Descargue", "Direccion Descargue",
            "Producto", "Naturaleza", "Tipo de vehiculo", "unidad", "Cantidad", "Tipo embalaje",
            "Toneladas", "Flete unidad", "PUNTO ADICIONAL", "CARGUE-DESCARGUE PER JURIDICA",
            "SEGURO", "Tipo pago", "Tolerancia", "Vlr hora STBY", "Vlr Declar Mercancia",
            "Aprobar Poliza", "Flete por", "Valor unitario", "Aprobar cupo credito",
            "Aprobar rentabilidad", "Otras caracteristicas", "REMESAS", "REMISION DEL CLIENTE",
            "GUIA DE TRANSPORTE", "MANIFIESTO"
        ]

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004d40", end_color="004d40", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Escribir cabeceras
        for col_idx, columna in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_idx, value=columna)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Ajustar ancho de columnas NUEVO FORMATO
        column_widths = {
            'A': 20, 'B': 15, 'C': 15, 'D': 12, 'E': 40, 'F': 15, 'G': 20,
            'H': 20, 'I': 25, 'J': 20, 'K': 25, 'L': 25, 'M': 15, 'N': 12,
            'O': 15, 'P': 12, 'Q': 15, 'R': 18, 'S': 12, 'T': 12, 'U': 15,
            'V': 15, 'W': 12, 'X': 20, 'Y': 25, 'Z': 15, 'AA': 12, 'AB': 12,
            'AC': 12, 'AD': 15, 'AE': 15, 'AF': 15, 'AG': 12, 'AH': 12, 'AI': 15,
            'AJ': 12, 'AK': 20, 'AL': 20, 'AM': 12, 'AN': 20, 'AO': 20, 'AP': 12,
            'AQ': 20, 'AR': 18
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Regional del usuario
        regional_usuario = request.centro_distribucion or "FUNZA"

        # Pre-cargar divipolas para lookup de ubicación y dirección de descargue
        divipolas_lookup = {}
        for div_doc in coleccion_divipolas.find():
            divipolas_lookup[div_doc.get("divipola", "")] = {
                "ubicacion_descargue": div_doc.get("ubicacion_descargue", ""),
                "direccion_descargue": div_doc.get("direccion_descargue", ""),
            }
        # También indexar por nombre de población (normalizado)
        divipolas_por_poblacion = {}
        for div_doc in coleccion_divipolas.find():
            pob = div_doc.get("poblacion", "").strip().upper()
            if pob and pob not in divipolas_por_poblacion:
                divipolas_por_poblacion[pob] = {
                    "ubicacion_descargue": div_doc.get("ubicacion_descargue", ""),
                    "direccion_descargue": div_doc.get("direccion_descargue", ""),
                }

        # Escribir datos NUEVO FORMATO
        row_num = 2
        for doc in planillas_db:
            try:
                # Una planilla FUSIONADA se "expande" en N filas (una por planilla original),
                # repartiendo el Flete unidad (total_solicitado) proporcionalmente por piezas.
                # Una planilla normal produce 1 fila (comportamiento histórico).
                for fila_base in _expandir_doc_a_filas(doc):
                    # FRESENIUS KABI: además, duplica filas por destinatario (FKC_Nombre_Cedula)
                    for fila in _expandir_fila_kabi(fila_base):
                        row_num = _escribir_fila_planilla(
                            ws, row_num,
                            regional_usuario=regional_usuario,
                            divipolas_lookup=divipolas_lookup,
                            divipolas_por_poblacion=divipolas_por_poblacion,
                            mapear_tipo_vehiculo=_mapear_tipo_vehiculo,
                            thin_border=thin_border,
                            **fila,
                        )
            except Exception as e:
                logger.error(f"Error procesando planilla {doc.get('planilla', 'desconocido')}: {str(e)}")
                continue  # Saltar esta fila y continuar con la siguiente

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel generado con {row_num - 1} filas de datos")

        # Retornar archivo
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=planillas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al exportar Excel: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al exportar Excel: {str(e)}"
        )


class ExportarDetalleRequest(BaseModel):
    """Modelo para exportar el DETALLE completo de planillas a Excel (una fila por planilla)."""
    planillas: List[str]
    perfil: str
    centro_distribucion: Optional[str] = None


def _fmt_fecha_col(fecha) -> str:
    """Formatea una fecha Mongo (datetime UTC naive) a hora Colombia (UTC-5) 'YYYY-MM-DD HH:MM'."""
    if not fecha:
        return ""
    try:
        if isinstance(fecha, datetime):
            return (fecha - timedelta(hours=5)).strftime("%Y-%m-%d %H:%M")
        return str(fecha)
    except Exception:
        return str(fecha)


def _val_recargo_detalle(v, fallback: int):
    """Normaliza un recargo: número tal cual; booleano/string legacy -> fallback; resto 0."""
    if isinstance(v, bool):
        return fallback if v else 0
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str) and v.strip().upper() in ("SI", "TRUE", "1", "YES"):
        return fallback
    return 0


def _valores_detalle(obj: dict, fusion_consecutivo: Optional[str] = None, total_override: Optional[float] = None) -> list:
    """Arma la lista de valores (en orden de cabeceras) para una fila de detalle.
    Funciona igual para un doc top-level o para un original embebido en una fusión
    (los campos que el embebido no tenga quedarán vacíos). Si se pasa total_override,
    se usa ese valor para 'Total solicitado' (prorrateo de la fusión por cajas)."""
    total_sol = total_override if total_override is not None else (obj.get("total_solicitado") or 0)

    def _num(campo):
        v = obj.get(campo)
        return v if isinstance(v, (int, float)) else ("" if v is None else v)

    return [
        fusion_consecutivo or "",
        obj.get("consecutivo") or "",
        obj.get("consecutivo_base") or "",
        obj.get("planilla") or "",
        obj.get("regional") or "",
        obj.get("cliente_origen") or "",
        obj.get("ruta") or "",
        obj.get("municipio_destino") or "",
        obj.get("departamento_destino") or "",
        obj.get("municipios_destino_lista") or "",
        obj.get("codigo_pedido") or "",
        _num("cantidad_pedidos"),
        _num("piezas"),
        _num("peso_real"),
        _num("peso_sicetac"),
        obj.get("tipo_veh_sicetac") or "",
        obj.get("placa") or "",
        total_sol,
    ]


@router.post("/exportar-planillas-detalle-excel")
async def exportar_planillas_detalle_excel(request: ExportarDetalleRequest):
    """
    Exporta un Excel de DETALLE completo de las planillas indicadas: una fila por
    planilla/consecutivo con TODOS los campos disponibles (operación, carga, tarifas,
    recargos, fechas, trazabilidad, fusión/división). Al igual que exportar-planillas-excel,
    SOLO exporta las planillas con estado APROBADO. Solo ADMIN/ANALISTA (validar en frontend).
    """
    try:
        if not request.planillas:
            raise HTTPException(status_code=400, detail="planillas es obligatorio")

        # Consultar planillas de MongoDB - SOLO APROBADAS (igual que exportar-planillas-excel)
        consulta = {
            "planilla": {"$in": request.planillas},
            "estado": "APROBADO"
        }
        if request.perfil == "OPERATIVO" and request.centro_distribucion:
            _aplicar_filtro_regional_operativo(consulta, request.centro_distribucion)

        docs = list(coleccion_pedidos_medical.find(consulta).sort("fecha_creacion", -1))
        if not docs:
            raise HTTPException(
                status_code=404,
                detail="No se encontraron planillas con los criterios indicados"
            )

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from fastapi.responses import Response
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Detalle Planillas"

        cabeceras = [
            "Fusión (consecutivo)",
            "Consecutivo", "Consecutivo base", "Planilla",
            "Regional", "Cliente origen", "Ruta",
            "Municipio principal", "Departamento destino",
            "Todos los municipios", "Códigos pedido", "Cant. pedidos",
            "Piezas", "Peso real (kg)", "Peso SICETAC (kg)",
            "Vehículo SICETAC", "Placa",
            "Total solicitado",
        ]

        header_fill = PatternFill(start_color="004d40", end_color="004d40", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, header in enumerate(cabeceras, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        fila = 2
        for doc in docs:
            fusion_info = doc.get("fusion_info") or {}
            originales = fusion_info.get("datos_originales") or []
            if fusion_info.get("es_fusionada") and originales:
                # DIVIDIR la fusión: una fila por cada planilla original, indicando a
                # qué consecutivo de fusión pertenece (para no perder la trazabilidad).
                consecutivo_fusion = doc.get("consecutivo") or doc.get("planilla") or ""
                # Prorratear el Total solicitado de la fusión entre los originales según
                # sus cajas (piezas). La última parte absorbe el remanente para que la
                # suma de las partes coincida exactamente con el total de la fusión.
                total_fusion = doc.get("total_solicitado") or 0
                piezas_orig = [ (o.get("piezas") or 0) for o in originales ]
                suma_piezas = sum(piezas_orig)
                objetos = []
                acumulado = 0
                for i, o in enumerate(originales):
                    if suma_piezas > 0 and i < len(originales) - 1:
                        parte = round(total_fusion * piezas_orig[i] / suma_piezas)
                        acumulado += parte
                    elif suma_piezas > 0:
                        parte = total_fusion - acumulado  # última parte: remanente exacto
                    else:
                        parte = 0
                    objetos.append((o, consecutivo_fusion, parte))
            else:
                objetos = [(doc, None, None)]  # None = usar el total del propio documento

            for obj, consecutivo_fusion, total_override in objetos:
                valores = _valores_detalle(obj, fusion_consecutivo=consecutivo_fusion, total_override=total_override)
                for col_idx, valor in enumerate(valores, 1):
                    cell = ws.cell(row=fila, column=col_idx, value=valor)
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
                fila += 1

        # Anchos de columna auto-ajustados (topeados) para legibilidad.
        for col_idx in range(1, len(cabeceras) + 1):
            max_len = len(str(cabeceras[col_idx - 1]))
            for r in range(2, fila):
                v = ws.cell(row=r, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, min(len(str(v)), 55))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 52)

        ws.freeze_panes = "A2"  # Congelar la fila de cabeceras.

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"[DETALLE EXCEL] Generado con {fila - 2} planillas")

        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=detalle_planillas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al exportar Excel de detalle: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al exportar Excel de detalle: {str(e)}")


# ============= ENDPOINTS HISTÓRICO PEDIDOS =============

@router.get("/historico")
async def obtener_historico(
    fecha_inicio: str = "",
    fecha_fin: str = "",
    perfil: str = "",
    centro_distribucion: str = "",
    regional: str = ""
):
    """
    Obtiene planillas del historico (pedidos_medical_historico).
    Por defecto muestra las de hoy. Filtra por rango de fechas si se proporcionan.
    """
    try:
        # Filtro de fechas (día Colombia). El servidor corre en UTC, así que "hoy" y los
        # límites se calculan en zona America/Bogota (UTC-5).
        hoy = (datetime.now(timezone.utc) - _OFFSET_COLOMBIA).strftime("%Y-%m-%d")
        f_inicio = fecha_inicio if fecha_inicio else hoy
        f_fin = fecha_fin if fecha_fin else hoy

        # Ventana del día Colombia expresada como instantes UTC (Mongo guarda
        # fecha_movimiento_historico en UTC). Las +5 h alinean 00:00 Colombia con 05:00 UTC.
        fecha_inicio_dt = datetime.strptime(f_inicio, "%Y-%m-%d") + _OFFSET_COLOMBIA
        fecha_fin_dt = datetime.strptime(f_fin, "%Y-%m-%d") + timedelta(days=1) + _OFFSET_COLOMBIA

        filtro = {
            "fecha_movimiento_historico": {
                "$gte": fecha_inicio_dt,
                "$lt": fecha_fin_dt
            }
        }

        # Filtrar por regional para operativos
        perfiles_globales = ['ADMIN', 'ANALISTA', 'COORDINADOR', 'CONTROL']
        if perfil and perfil not in perfiles_globales and centro_distribucion:
            _aplicar_filtro_regional_operativo(filtro, centro_distribucion)

        # Filtro de regional elegido manualmente en el dropdown (perfiles globales).
        if regional:
            _aplicar_filtro_regional_dropdown(filtro, regional)

        logger.info(f"[HISTORICO] Filtro: {filtro}")

        docs = list(coleccion_historico.find(filtro).sort("fecha_movimiento_historico", -1))

        for doc in docs:
            doc["_id"] = str(doc["_id"])

        logger.info(f"[HISTORICO] Documentos encontrados: {len(docs)}")

        return {
            "planillas": docs,
            "total": len(docs),
            "fecha_inicio": f_inicio,
            "fecha_fin": f_fin
        }

    except Exception as e:
        logger.error(f"Error al obtener historico: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al obtener historico: {str(e)}")


class ExportarHistoricoExcelRequest(BaseModel):
    """Modelo para exportar historico a Excel con filtros"""
    fecha_inicio: str
    fecha_fin: str
    perfil: str
    centro_distribucion: Optional[str] = None
    regional: Optional[str] = None
    busqueda: Optional[str] = None


@router.post("/historico/exportar-excel")
async def exportar_historico_excel(request: ExportarHistoricoExcelRequest):
    """
    Exporta planillas del historico a Excel con los datos directos de MongoDB.
    Aplica los mismos filtros de fecha, perfil y regional que la vista.
    """
    try:
        from fastapi.responses import Response
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io

        logger.info(f"=== EXPORTAR HISTORICO EXCEL ===")
        logger.info(f"Filtros: {request.fecha_inicio} a {request.fecha_fin}, perfil={request.perfil}, centro={request.centro_distribucion}")

        # Construir filtro de fechas (día Colombia, igual que GET /historico)
        hoy = (datetime.now(timezone.utc) - _OFFSET_COLOMBIA).strftime("%Y-%m-%d")
        f_inicio = request.fecha_inicio if request.fecha_inicio else hoy
        f_fin = request.fecha_fin if request.fecha_fin else hoy

        fecha_inicio_dt = datetime.strptime(f_inicio, "%Y-%m-%d") + _OFFSET_COLOMBIA
        fecha_fin_dt = datetime.strptime(f_fin, "%Y-%m-%d") + timedelta(days=1) + _OFFSET_COLOMBIA

        filtro = {
            "fecha_movimiento_historico": {
                "$gte": fecha_inicio_dt,
                "$lt": fecha_fin_dt
            }
        }

        # Filtrar por regional para operativos
        perfiles_globales = ['ADMIN', 'ANALISTA', 'COORDINADOR', 'CONTROL']
        if request.perfil and request.perfil not in perfiles_globales and request.centro_distribucion:
            _aplicar_filtro_regional_operativo(filtro, request.centro_distribucion)

        # Filtro de regional elegido manualmente en el dropdown (perfiles globales).
        if request.regional:
            _aplicar_filtro_regional_dropdown(filtro, request.regional)

        planillas_db = list(coleccion_historico.find(filtro).sort("fecha_movimiento_historico", -1))
        logger.info(f"Planillas historico encontradas con filtros: {len(planillas_db)}")

        # Filtro de búsqueda textual si viene
        if request.busqueda and request.busqueda.strip():
            termino = request.busqueda.strip().lower()
            planillas_db = [
                doc for doc in planillas_db
                if termino in (doc.get("consecutivo") or "").lower()
                or termino in (doc.get("planilla") or "").lower()
                or termino in (doc.get("pedido_vulcano") or "").lower()
                or termino in (doc.get("ruta") or "").lower()
                or termino in (doc.get("municipio_destino") or "").lower()
                or termino in (doc.get("regional") or "").lower()
            ]
            logger.info(f"Después de filtro búsqueda '{request.busqueda}': {len(planillas_db)} registros")

        if not planillas_db:
            raise HTTPException(status_code=404, detail="No se encontraron planillas en historico con los filtros indicados")

        # --- Generar Excel limpio con datos directos de MongoDB ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Historico Pedidos"

        columnas = [
            "Consecutivo", "Planilla", "Pedido Vulcano", "Fecha Preaprobado", "Estado",
            "Total Solicitado", "Diferencia", "Regional", "Placa", "Piezas",
            "Peso Real", "Peso SICETAC", "Cant. Pedidos", "Ruta", "Tipo Vehículo",
            "Vehículo SICETAC", "Flete Teórico", "Flete Solicitado",
            "Descargue", "Punto Adic.", "Desvío", "Aforo",
            "Municipio Principal", "Cliente Origen", "Cant. Destinos",
            "Código Pedido", "Observaciones"
        ]

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="004d40", end_color="004d40", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_alignment = Alignment(horizontal="left", vertical="center")
        number_alignment = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Anchos por columna
        anchos = {
            "Consecutivo": 22, "Planilla": 14, "Pedido Vulcano": 16, "Fecha Preaprobado": 20,
            "Estado": 18, "Total Solicitado": 16, "Diferencia": 16, "Regional": 16,
            "Placa": 12, "Piezas": 10, "Peso Real": 12, "Peso SICETAC": 14,
            "Cant. Pedidos": 12, "Ruta": 18, "Tipo Vehículo": 14,
            "Vehículo SICETAC": 16, "Flete Teórico": 16, "Flete Solicitado": 16,
            "Descargue": 14, "Punto Adic.": 14, "Desvío": 14, "Aforo": 14,
            "Municipio Principal": 20, "Cliente Origen": 22,
            "Cant. Destinos": 13, "Código Pedido": 20, "Observaciones": 25
        }

        # Cabeceras
        for col_idx, columna in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_idx, value=columna)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)].width = anchos.get(columna, 16)

        # Colores para estados
        estado_fills = {
            "APROBADO": PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid"),
            "REQUIERE_APROBACION_COORDINADOR": PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid"),
            "REQUIERE_APROBACION_CONTROL": PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid"),
            "PREAPROBADO": PatternFill(start_color="e0f2fe", end_color="e0f2fe", fill_type="solid"),
        }
        estado_fonts = {
            "APROBADO": Font(bold=True, color="15803d", size=9),
            "REQUIERE_APROBACION_COORDINADOR": Font(bold=True, color="b45309", size=9),
            "REQUIERE_APROBACION_CONTROL": Font(bold=True, color="dc2626", size=9),
            "PREAPROBADO": Font(bold=True, color="0369a1", size=9),
        }
        money_font = Font(bold=True, color="005f56", size=10)
        diff_pos_font = Font(bold=True, color="b91c1c", size=10)
        diff_neg_font = Font(bold=True, color="15803d", size=10)
        even_fill = PatternFill(start_color="f8fffe", end_color="f8fffe", fill_type="solid")

        def fmt_recargo(val, fallback=0):
            if isinstance(val, (int, float)) and val != 0:
                return val
            if val is True or val == "SI":
                return fallback
            return 0

        def fmt_fecha(val):
            if not val:
                return ""
            if hasattr(val, "strftime"):
                return val.strftime("%Y-%m-%d %H:%M")
            return str(val)

        def num(val, default=0):
            try:
                if val is None or val == "":
                    return default
                return float(val)
            except (TypeError, ValueError):
                return default

        # Datos
        row_num = 2
        for i, doc in enumerate(planillas_db):
            try:
                estado = doc.get("estado", "PREAPROBADO")
                total_solicitado = num(doc.get("total_solicitado"))
                flete_teorico = num(doc.get("tarifa_calculada"))
                diferencia = num(doc.get("diferencia"), total_solicitado - flete_teorico)

                valores = [
                    doc.get("consecutivo", ""),
                    doc.get("planilla", ""),
                    doc.get("pedido_vulcano", ""),
                    fmt_fecha(doc.get("fecha_preaprobado") or doc.get("fecha_creacion")),
                    estado,
                    total_solicitado,
                    diferencia,
                    doc.get("regional", ""),
                    doc.get("placa", ""),
                    doc.get("piezas", 0),
                    doc.get("peso_real", 0),
                    doc.get("peso_sicetac", doc.get("peso_real", 0)),
                    doc.get("cantidad_pedidos", ""),
                    doc.get("ruta", ""),
                    doc.get("tipo_vehiculo", ""),
                    doc.get("tipo_veh_sicetac") or doc.get("tipo_vehiculo", ""),
                    flete_teorico,
                    doc.get("tarifa_base") or doc.get("tarifa_calculada", 0),
                    fmt_recargo(doc.get("requiere_descargue"), 50000),
                    fmt_recargo(doc.get("punto_adicional"), 80000),
                    fmt_recargo(doc.get("desvio"), 100000),
                    fmt_recargo(doc.get("aforo"), 0),
                    doc.get("municipio_destino", ""),
                    doc.get("cliente_origen", ""),
                    doc.get("cantidad_destinos", ""),
                    doc.get("codigo_pedido", ""),
                    doc.get("causal", ""),
                ]

                row_fill = even_fill if i % 2 == 0 else None

                for col_idx, valor in enumerate(valores, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=valor)
                    cell.border = thin_border

                    if row_fill:
                        cell.fill = row_fill

                    col_name = columnas[col_idx - 1]

                    # Columna Estado - badge
                    if col_name == "Estado":
                        cell.fill = estado_fills.get(estado, estado_fills["PREAPROBADO"])
                        cell.font = estado_fonts.get(estado, estado_fonts["PREAPROBADO"])
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        estado_label = {
                            "REQUIERE_APROBACION_COORDINADOR": "COORDINADOR",
                            "REQUIERE_APROBACION_CONTROL": "CONTROL",
                        }.get(estado, estado)
                        cell.value = estado_label

                    # Columnas monetarias
                    elif col_name in ("Flete Teórico", "Flete Solicitado", "Descargue", "Punto Adic.", "Desvío", "Aforo", "Total Solicitado"):
                        cell.font = money_font if col_name == "Total Solicitado" else Font(size=10)
                        cell.number_format = '$#,##0'
                        cell.alignment = number_alignment

                    # Columna Diferencia
                    elif col_name == "Diferencia":
                        if diferencia > 0:
                            cell.font = diff_pos_font
                        elif diferencia < 0:
                            cell.font = diff_neg_font
                        else:
                            cell.font = Font(color="666666", size=10)
                        cell.number_format = '$#,##0;[Red]-$#,##0;$0'
                        cell.alignment = number_alignment

                    # Columnas numéricas
                    elif col_name in ("Piezas", "Peso Real", "Peso SICETAC", "Cant. Pedidos", "Cant. Destinos"):
                        cell.alignment = number_alignment
                        if col_name in ("Peso Real", "Peso SICETAC"):
                            cell.number_format = '#,##0'

                    else:
                        cell.alignment = data_alignment

                row_num += 1

            except Exception as e:
                logger.error(f"Error procesando historico {doc.get('planilla', '?')}: {str(e)}")
                continue

        # Auto-filtro
        ws.auto_filter.ref = f"A1:{chr(64 + len(columnas)) if len(columnas) <= 26 else chr(64 + (len(columnas) - 1) // 26) + chr(65 + (len(columnas) - 1) % 26)}{row_num - 1}"

        # Congelar primera fila
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel historico generado: {row_num - 3} filas")

        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=historico_{f_inicio}_{f_fin}.xlsx"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al exportar historico Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al exportar historico: {str(e)}")


@router.get("/obtener-resultados-recientes")
async def obtener_resultados_recientes(limite: int = 100, perfil: str = "", centro_distribucion: str = "", usuario: str = ""):
    """
    Obtiene todos los planillas de pedidos_medical (cada documento es una planilla).
    Para operativos, filtra por su regional. Perfiles globales ven todas.
    Las planillas en estado CREADO (borrador del operativo) solo las ve el ADMIN y el
    propio operativo creador; el resto de perfiles no las ve hasta que pasen a PREAPROBADO.
    """
    try:
        logger.info(f"[OBTENER RESULTADOS] Limite: {limite}, Perfil: {perfil}, Centro: {centro_distribucion}, Usuario: {usuario}")

        # Construir filtro base
        filtro = {"fusionada": {"$ne": True}}

        # Perfiles globales ven todas las regionales
        perfiles_globales = ['ADMIN', 'ANALISTA', 'COORDINADOR', 'CONTROL']

        if perfil and perfil not in perfiles_globales and centro_distribucion:
            _aplicar_filtro_regional_operativo(filtro, centro_distribucion)
            logger.info(f"[OBTENER RESULTADOS] Filtro por regional (OPERATIVO): {filtro}")

        # Visibilidad del estado CREADO (borrador): ocultarlo salvo para ADMIN y para el
        # operativo creador. Se compone con el posible $or del filtro regional envolviéndolo
        # en un $and para evitar colisión de la clave "$or".
        condiciones_and = []
        if perfil != "ADMIN":
            if perfil == "OPERATIVO" and usuario:
                condiciones_and.append({"$or": [
                    {"estado": {"$ne": "CREADO"}},
                    {"estado": "CREADO", "usuario_registro": usuario},
                ]})
            else:
                condiciones_and.append({"estado": {"$ne": "CREADO"}})
        if condiciones_and:
            if "$or" in filtro:
                existente = dict(filtro)
                filtro.clear()
                filtro["$and"] = [existente] + condiciones_and
            else:
                for cond in condiciones_and:
                    filtro.update(cond)
            logger.info(f"[OBTENER RESULTADOS] Filtro visibilidad CREADO aplicado: {filtro}")

        # Contar total de documentos
        total_docs = coleccion_pedidos_medical.count_documents(filtro)
        logger.info(f"[OBTENER RESULTADOS] Total documentos con filtro: {total_docs}")

        # Traer planillas filtradas (excluye registros_detalle: pesado y no se usa en la tabla)
        planillas = list(coleccion_pedidos_medical.find(filtro, {"registros_detalle": 0}).sort("fecha_creacion", -1).limit(limite))

        # Convertir ObjectId a string
        for planilla in planillas:
            planilla["_id"] = str(planilla["_id"])

        logger.info(f"[OBTENER RESULTADOS] Planillas encontradas: {len(planillas)}")

        return {
            "planillas": planillas,
            "total": len(planillas)
        }

    except Exception as e:
        logger.error(f"Error al obtener resultados recientes: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener resultados: {str(e)}"
        )

        busquedas = list(coleccion_pedidos_medical.find(
            {"usuario": usuario}
        ).sort("fecha_creacion", -1).limit(limite))

        logger.info(f"[OBTENER BUSQUEDAS] Busquedas encontradas para usuario {usuario}: {len(busquedas)}")

        # Convertir ObjectId a string
        for bus in busquedas:
            bus["_id"] = str(bus["_id"])

        return {
            "busquedas": busquedas,
            "total": len(busquedas)
        }

    except Exception as e:
        logger.error(f"Error al obtener búsquedas recientes: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener búsquedas: {str(e)}"
        )


# ============= ENDPOINTS PARA GESTIÓN DE CAUSALES =============

@router.get("/causales")
async def obtener_causales():
    """
    Obtiene todas las causales activas para fusión de planillas.
    """
    try:
        causales = list(coleccion_causales.find({"activo": True}))
        for c in causales:
            c["_id"] = str(c["_id"])
        return {"causales": causales}
    except Exception as e:
        logger.error(f"Error al obtener causales: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al obtener causales: {str(e)}")


@router.get("/causales/todas")
async def obtener_todas_causales():
    """
    Obtiene todas las causales (activas e inactivas) - solo para admin.
    """
    try:
        causales = list(coleccion_causales.find({}).sort("fecha_creacion", -1))
        for c in causales:
            c["_id"] = str(c["_id"])
        return {"causales": causales}
    except Exception as e:
        logger.error(f"Error al obtener todas las causales: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al obtener causales: {str(e)}")


@router.post("/causales")
async def crear_causal(request: CausalRequest):
    """
    Crea una nueva causal de fusión.
    """
    try:
        # Verificar si ya existe una causal con ese nombre
        existente = coleccion_causales.find_one({"nombre": {"$regex": f"^{request.nombre}$", "$options": "i"}})
        if existente:
            raise HTTPException(status_code=400, detail="Ya existe una causal con ese nombre")

        nueva_causal = {
            "nombre": request.nombre,
            "activo": request.activo,
            "fecha_creacion": datetime.now()
        }
        result = coleccion_causales.insert_one(nueva_causal)
        nueva_causal["_id"] = str(result.inserted_id)
        logger.info(f"Causal creada: {nueva_causal['_id']} - {request.nombre}")
        return {"mensaje": "Causal creada exitosamente", "causal": nueva_causal}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al crear causal: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al crear causal: {str(e)}")


@router.put("/causales/{causal_id}")
async def actualizar_causal(causal_id: str, request: CausalRequest):
    """
    Actualiza una causal existente.
    """
    try:
        from bson import ObjectId
        if not ObjectId.is_valid(causal_id):
            raise HTTPException(status_code=400, detail="ID de causal inválido")

        campos_actualizar = {
            "nombre": request.nombre,
            "activo": request.activo
        }
        resultado = coleccion_causales.update_one(
            {"_id": ObjectId(causal_id)},
            {"$set": campos_actualizar}
        )

        if resultado.matched_count == 0:
            raise HTTPException(status_code=404, detail="Causal no encontrada")

        logger.info(f"Causal actualizada: {causal_id}")
        return {"mensaje": "Causal actualizada exitosamente"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al actualizar causal: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al actualizar causal: {str(e)}")


@router.delete("/causales/{causal_id}")
async def eliminar_causal(causal_id: str):
    """
    Elimina (desactiva) una causal.
    """
    try:
        from bson import ObjectId
        if not ObjectId.is_valid(causal_id):
            raise HTTPException(status_code=400, detail="ID de causal inválido")

        resultado = coleccion_causales.update_one(
            {"_id": ObjectId(causal_id)},
            {"$set": {"activo": False}}
        )

        if resultado.matched_count == 0:
            raise HTTPException(status_code=404, detail="Causal no encontrada")

        logger.info(f"Causal eliminada (desactivada): {causal_id}")
        return {"mensaje": "Causal eliminada exitosamente"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al eliminar causal: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al eliminar causal: {str(e)}")


@router.post("/causales/inicializar")
async def inicializar_causales():
    """
    Inicializa las causales por defecto si no existen.
    """
    try:
        causales_por_defecto = [
            {"nombre": "lleva paqueteo", "activo": True},
            {"nombre": "no se consiguio vehiculo", "activo": True}
        ]

        creadas = []
        for causal_def in causales_por_defecto:
            existente = coleccion_causales.find_one({"nombre": {"$regex": f"^{causal_def['nombre']}$", "$options": "i"}})
            if not existente:
                nueva_causal = {
                    "nombre": causal_def["nombre"],
                    "activo": causal_def["activo"],
                    "fecha_creacion": datetime.now()
                }
                result = coleccion_causales.insert_one(nueva_causal)
                nueva_causal["_id"] = str(result.inserted_id)
                creadas.append(nueva_causal)
                logger.info(f"Causal inicializada: {nueva_causal['_id']} - {causal_def['nombre']}")

        if creadas:
            return {"mensaje": f"Se inicializaron {len(creadas)} causales", "causales": creadas}
        else:
            return {"mensaje": "Las causales por defecto ya existen", "causales": []}
    except Exception as e:
        logger.error(f"Error al inicializar causales: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error al inicializar causales: {str(e)}")
