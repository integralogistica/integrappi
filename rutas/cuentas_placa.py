"""
Módulo «Cuentas por Placa» — Integrapp.

Catálogo de datos bancarios/conductor por placa, scoped por regional. Su objetivo
es autollenar la sección "Información bancaria" del formulario de Otros Costos
cuando se digita la placa (sugerencia: los campos siguen siendo editables).

Colección (base `integra`): `cuentas_por_placa`.
  - Una placa puede repetirse entre regionales distintas, pero NO dentro de la
    misma (índice unique (placa, regional)).
  - `regional` guarda el código CO canónico (CO05...) y `regional_info` los
    sinónimos {co, regional, bodega} para mostrar en la UI.

Perfiles: ADMIN, OPERATIVO y DESPACHADOR. Los regionales solo ven/editan SU
regional (derivada de baseusuarios, no confiable del frontend); ADMIN puede
operar cualquier regional y es el único que carga el Excel masivo.

Bancos y tipos de cuenta: los mismos del módulo Otros Costos
(BANCOS_OTROS_COSTOS_DEFAULT / TIPOS_CUENTA en rutas/otros_costos.py).
"""

import io
import logging
import re
import unicodedata
from datetime import datetime
from typing import Optional

from bson import ObjectId
from fastapi import APIRouter, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel
from pymongo.errors import DuplicateKeyError

from bd.bd_cliente import bd_cliente
from rutas.otros_costos import (
    BANCOS_OTROS_COSTOS_DEFAULT,
    CO_A_REGIONAL,
    REGIONAL_A_BODEGA,
    TIPOS_CUENTA,
    _ahora_utc,
    _hoy_colombia,
    _ip,
    _normalizar_regional,
    _requiere,
    _resolver_usuario,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/cuentas-placa", tags=["Cuentas por Placa"])

PERFILES_CUENTAS_PLACA = {"ADMIN", "OPERATIVO", "DESPACHADOR"}
PERFILES_REGIONALES_CP = {"OPERATIVO", "DESPACHADOR"}   # scope fijo a SU regional
# /por-placa es solo lectura de sugerencia: lo puede consultar cualquier perfil
# que entre al formulario de Otros Costos.
PERFILES_LOOKUP_CP = {"ADMIN", "OPERATIVO", "DESPACHADOR", "COORDINADOR", "CONTROL", "FINANCIERO", "ANALISTA"}

db = bd_cliente["integra"]
col_cuentas = db["cuentas_por_placa"]

# Índice unique compuesto — una placa por regional (idempotente al importar).
try:
    col_cuentas.create_index([("placa", 1), ("regional", 1)], unique=True, name="idx_cp_placa_regional")
except Exception as _e:
    logger.warning(f"[CUENTAS_PLACA] No se pudo crear índice unique placa+regional: {_e}")
try:
    col_cuentas.create_index("placa")
except Exception:
    pass

# Placa: 4-6 caracteres alfanuméricos, sin espacios/símbolos/guiones.
REGEX_PLACA = re.compile(r"^[A-Z0-9]{4,6}$")


def _sin_acentos(s) -> str:
    return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()


BANCOS_VALIDOS = {b["nombre"]: b["codigo"] for b in BANCOS_OTROS_COSTOS_DEFAULT}
# Comparación tolerante (sin acentos / mayúsculas) → valor canónico.
BANCOS_NORM = {_sin_acentos(n).upper().strip(): n for n in BANCOS_VALIDOS}
TIPOS_CUENTA_NORM = {_sin_acentos(str(t)).upper().strip(): t for t in TIPOS_CUENTA}

MEDIA_TYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CABECERAS_CUENTAS = ["PLACA", "NOMBRE CONDUCTOR", "TELEFONO", "NOMBRE BENEFICIARIO", "CEDULA", "BANCO", "TIPO DE CUENTA", "NUMERO CUENTA"]


# ── Helpers ───────────────────────────────────────────────────────────────────


def _norm_col(s) -> str:
    """Cabecera Excel → clave normalizada (sin acentos, lower, espacios colapsados)."""
    return re.sub(r"\s+", " ", _sin_acentos(s).strip().lower())


def _celda_texto(v) -> str:
    """Celda Excel → texto limpio (cero decimal de openpyxl: 1.23e+08 → '123000000')."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _validar_placa(placa: str) -> str:
    """Devuelve la placa normalizada (UPPER, solo alfanumérico) o lanza 422."""
    p = re.sub(r"[^A-Za-z0-9]", "", str(placa or "")).strip().upper()
    if not REGEX_PLACA.fullmatch(p):
        raise HTTPException(
            status_code=422,
            detail=f"Placa '{placa}' inválida: 4-6 caracteres alfanuméricos, sin espacios, guiones ni símbolos.",
        )
    return p


def _resolver_banco(valor: str) -> str:
    """Banco en cualquiera de sus escrituras → nombre canónico, o 422."""
    v = _sin_acentos(valor).upper().strip()
    if v in BANCOS_VALIDOS:
        return v
    if v in BANCOS_NORM:
        return BANCOS_NORM[v]
    raise HTTPException(
        status_code=422,
        detail=f"Banco '{valor}' no está en el listado permitido "
               f"(ej: {', '.join(list(BANCOS_VALIDOS)[:5])}...).",
    )


def _resolver_tipo_cuenta(valor: str) -> str:
    """Tipo de cuenta case-insensitive → valor canónico de TIPOS_CUENTA, o 422."""
    v = _sin_acentos(valor).upper().strip()
    if v in TIPOS_CUENTA_NORM:
        return TIPOS_CUENTA_NORM[v]
    raise HTTPException(
        status_code=422,
        detail=f"Tipo de cuenta '{valor}' inválido. Valores: {', '.join(TIPOS_CUENTA)}.",
    )


def _validar_digitos(valor: str, campo: str, minimo: int = 5, maximo: int = 20) -> str:
    v = re.sub(r"\D", "", str(valor or ""))
    if not re.fullmatch(rf"\d{{{minimo},{maximo}}}", v):
        raise HTTPException(status_code=422, detail=f"{campo} debe tener entre {minimo} y {maximo} dígitos.")
    return v


def _validar_telefono(valor: str) -> str:
    """Teléfono del conductor: opcional, pero si viene debe ser 7-15 dígitos."""
    v = re.sub(r"\D", "", str(valor or ""))
    if not v:
        return ""
    if not re.fullmatch(r"\d{7,15}", v):
        raise HTTPException(status_code=422, detail="El teléfono debe tener entre 7 y 15 dígitos.")
    return v


def _regional_scope(info: dict, regional_param: Optional[str]) -> str:
    """Código CO efectivo para filtrar/crear.

    Perfiles regionales: SIEMPRE su regional (ignora el parámetro). ADMIN: el
    parámetro normalizado, o '' (= todas). Lanza 422 si no reconoce."""
    if info["perfil"] in PERFILES_REGIONALES_CP:
        norm = _normalizar_regional(info.get("regional"))
        if not norm:
            raise HTTPException(
                status_code=422,
                detail=f"Su usuario no tiene una regional válida ({info.get('regional') or 'sin regional'}).",
            )
        return norm["co"]
    if regional_param:
        norm = _normalizar_regional(regional_param)
        if not norm:
            raise HTTPException(status_code=422, detail=f"Regional no reconocida: {regional_param}")
        return norm["co"]
    return ""


def _serializar(doc: dict) -> dict:
    out = dict(doc)
    out["id"] = str(out.pop("_id", ""))
    for k, v in out.items():
        if isinstance(v, datetime):
            out[k] = v.isoformat()
    return out


class CuentaPlacaRequest(BaseModel):
    usuario: str
    placa: str
    nombre_conductor: str
    telefono: str = ""
    nombre_beneficiario: str
    cedula: str
    banco: str
    tipo_cuenta: str
    numero_cuenta: str
    regional: str = ""   # solo lo usa ADMIN; los regionales quedan en su regional


# ── Endpoints ─────────────────────────────────────────────────────────────────
@router.get("/")
async def listar(
    usuario: str = Query(...),
    placa: str = Query(""),
    regional: str = Query(""),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
):
    """Catálogo paginado. Búsqueda por placa (substring); scope regional según perfil."""
    info = _resolver_usuario(usuario)
    _requiere(info, PERFILES_CUENTAS_PLACA, "ver el catálogo de cuentas por placa")
    co = _regional_scope(info, regional)

    filtro: dict = {}
    if placa:
        filtro["placa"] = {"$regex": re.escape(re.sub(r"[^A-Za-z0-9]", "", placa).upper()), "$options": "i"}
    if co:
        filtro["regional"] = co

    total = col_cuentas.count_documents(filtro)
    cursor = col_cuentas.find(filtro).sort("placa", 1).skip(skip).limit(limit)
    return {"total": total, "skip": skip, "limit": limit, "items": [_serializar(d) for d in cursor]}


@router.get("/catalogos")
async def catalogos(usuario: str = Query(...)):
    """Bancos / tipos de cuenta / regionales (bodega) para los selects del frontend."""
    info = _resolver_usuario(usuario)
    _requiere(info, PERFILES_CUENTAS_PLACA, "ver el catálogo de cuentas por placa")
    return {
        "bancos": BANCOS_OTROS_COSTOS_DEFAULT,
        "tipos_cuenta": TIPOS_CUENTA,
        # La regional se muestra por BODEGA (JUAN MINA, YUMBO, ...), igual que el
        # centro de distribución de Otros Costos; el código CO sigue como valor.
        "regionales": [
            {"co": co, "regional": regional, "bodega": REGIONAL_A_BODEGA.get(regional, regional)}
            for co, regional in CO_A_REGIONAL.items()
        ],
    }


@router.get("/por-placa")
async def por_placa(
    usuario: str = Query(...),
    placa: str = Query(...),
    regional: str = Query(""),
):
    """Lookup de auto-llenado para el formulario de Otros Costos.

    Scope: perfil regional → SU regional; perfil global → la regional del
    parámetro (el frontend envía el centro de distribución) o todas. Si hay
    varias coincidencias no sugiere: devuelve en qué regionales existe."""
    info = _resolver_usuario(usuario)
    _requiere(info, PERFILES_LOOKUP_CP, "consultar cuentas por placa")
    try:
        p = _validar_placa(placa)
    except HTTPException:
        return {"encontrada": False, "coincidencias": []}

    filtro: dict = {"placa": p}
    if info["perfil"] in PERFILES_REGIONALES_CP:
        filtro["regional"] = _regional_scope(info, None)
    elif regional:
        norm = _normalizar_regional(regional)
        if norm:
            filtro["regional"] = norm["co"]

    docs = list(col_cuentas.find(filtro))
    if len(docs) == 1:
        return {"encontrada": True, "cuenta": _serializar(docs[0])}
    return {
        "encontrada": False,
        "coincidencias": [
            {"placa": d.get("placa"), "regional": (d.get("regional_info") or {}).get("regional"),
             "bodega": (d.get("regional_info") or {}).get("bodega")}
            for d in docs
        ],
    }


@router.post("/crear")
async def crear(req: CuentaPlacaRequest, request: Request):
    info = _resolver_usuario(req.usuario)
    _requiere(info, PERFILES_CUENTAS_PLACA, "crear cuentas por placa")

    p = _validar_placa(req.placa)
    banco = _resolver_banco(req.banco)
    tipo = _resolver_tipo_cuenta(req.tipo_cuenta)
    cedula = _validar_digitos(req.cedula, "La cédula", 5, 15)
    numero = _validar_digitos(req.numero_cuenta, "El número de cuenta", 5, 20)
    telefono = _validar_telefono(req.telefono)
    if not (req.nombre_beneficiario or "").strip():
        raise HTTPException(status_code=422, detail="El nombre del beneficiario es obligatorio.")

    co = _regional_scope(info, req.regional)
    if not co:
        raise HTTPException(status_code=422, detail="Seleccione la regional de la cuenta.")
    norm = _normalizar_regional(co)
    ahora = _ahora_utc()
    doc = {
        "placa": p,
        "nombre_conductor": (req.nombre_conductor or "").strip().upper(),
        "telefono": telefono,
        "nombre_beneficiario": (req.nombre_beneficiario or "").strip().upper(),
        "cedula": cedula,
        "banco": banco,
        "tipo_cuenta": tipo,
        "numero_cuenta": numero,
        "regional": co,
        "regional_info": norm,
        "creado_por": {"usuario": info["usuario"], "nombre": info["nombre"], "rol": info["perfil"], "fecha": ahora},
        "actualizado_por": {"usuario": info["usuario"], "nombre": info["nombre"], "rol": info["perfil"], "fecha": ahora},
        "created_at": ahora,
        "updated_at": ahora,
    }
    try:
        col_cuentas.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail=f"La placa {p} ya existe en la regional {norm['regional']}.")
    return {"mensaje": "Cuenta creada", "cuenta": _serializar(doc)}


@router.put("/editar")
async def editar(req: CuentaPlacaRequest, request: Request, id: str = Query(...)):
    """Edita una cuenta. La regional NO es editable (borrar y crear si se equivocaron).
    Los perfiles regionales solo pueden editar cuentas de su regional."""
    info = _resolver_usuario(req.usuario)
    _requiere(info, PERFILES_CUENTAS_PLACA, "editar cuentas por placa")

    try:
        oid = ObjectId(id)
    except Exception:
        raise HTTPException(status_code=422, detail="Identificador inválido.")
    doc = col_cuentas.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Cuenta no encontrada.")
    if info["perfil"] != "ADMIN":
        co_usuario = _regional_scope(info, None)
        if doc.get("regional") != co_usuario:
            raise HTTPException(status_code=403, detail="Solo puede editar cuentas de su regional.")

    p = _validar_placa(req.placa)
    banco = _resolver_banco(req.banco)
    tipo = _resolver_tipo_cuenta(req.tipo_cuenta)
    cedula = _validar_digitos(req.cedula, "La cédula", 5, 15)
    numero = _validar_digitos(req.numero_cuenta, "El número de cuenta", 5, 20)
    telefono = _validar_telefono(req.telefono)
    if not (req.nombre_beneficiario or "").strip():
        raise HTTPException(status_code=422, detail="El nombre del beneficiario es obligatorio.")

    ahora = _ahora_utc()
    set_fields = {
        "placa": p,
        "nombre_conductor": (req.nombre_conductor or "").strip().upper(),
        "telefono": telefono,
        "nombre_beneficiario": (req.nombre_beneficiario or "").strip().upper(),
        "cedula": cedula,
        "banco": banco,
        "tipo_cuenta": tipo,
        "numero_cuenta": numero,
        "actualizado_por": {"usuario": info["usuario"], "nombre": info["nombre"], "rol": info["perfil"], "fecha": ahora},
        "updated_at": ahora,
    }
    try:
        col_cuentas.update_one({"_id": oid}, {"$set": set_fields})
    except DuplicateKeyError:
        raise HTTPException(
            status_code=409,
            detail=f"La placa {p} ya existe en la regional {(doc.get('regional_info') or {}).get('regional', doc.get('regional'))}.",
        )
    actualizado = col_cuentas.find_one({"_id": oid})
    return {"mensaje": "Cuenta actualizada", "cuenta": _serializar(actualizado)}


@router.delete("/eliminar")
async def eliminar(usuario: str = Query(...), id: str = Query(...), request: Request = None):
    info = _resolver_usuario(usuario)
    _requiere(info, PERFILES_CUENTAS_PLACA, "eliminar cuentas por placa")

    try:
        oid = ObjectId(id)
    except Exception:
        raise HTTPException(status_code=422, detail="Identificador inválido.")
    doc = col_cuentas.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Cuenta no encontrada.")
    if info["perfil"] != "ADMIN":
        co_usuario = _regional_scope(info, None)
        if doc.get("regional") != co_usuario:
            raise HTTPException(status_code=403, detail="Solo puede eliminar cuentas de su regional.")

    col_cuentas.delete_one({"_id": oid})
    return {"mensaje": f"Cuenta de la placa {doc.get('placa')} eliminada"}


# ── Carga masiva por Excel (solo ADMIN) ───────────────────────────────────────
@router.post("/importar-excel")
async def importar_excel(
    usuario: str = Query(...),
    regional: str = Form(...),
    archivo: UploadFile = File(...),
    request: Request = None,
):
    """Carga el catálogo desde un Excel con las columnas de CABECERAS_CUENTAS.

    Todas las filas quedan en la `regional` indicada (elegida por el ADMIN).
    Upsert por (placa, regional): si la placa ya existe ahí se ACTUALIZA. Los
    errores se acumulan por fila sin abortar el resto (patrón /importar-pago)."""
    info = _resolver_usuario(usuario)
    _requiere(info, {"ADMIN"}, "importar el catálogo de cuentas por placa")

    norm = _normalizar_regional(regional)
    if not norm:
        raise HTTPException(status_code=422, detail=f"Regional no reconocida: {regional}")
    co = norm["co"]

    contenido = await archivo.read()
    if not contenido:
        raise HTTPException(status_code=422, detail="El archivo está vacío.")
    try:
        wb = load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="El archivo no es un Excel válido (.xlsx).")
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if not filas:
        raise HTTPException(status_code=422, detail="El archivo no tiene filas.")

    # Localizar columnas por cabecera (tolerante a mayúsculas/acentos/espacios).
    cabeceras = {_norm_col(c): i for i, c in enumerate(filas[0]) if c is not None}

    def col_de(*claves: str) -> Optional[int]:
        return next((cabeceras[k] for k in cabeceras if k in claves), None)

    col_placa = col_de("placa")
    col_cond = col_de("nombre conductor", "conductor")
    col_tel = col_de("telefono", "telefono conductor", "celular")
    col_benef = col_de("nombre beneficiario", "beneficiario")
    col_ced = col_de("cedula")
    col_banco = col_de("banco")
    col_tipo = col_de("tipo de cuenta", "tipo cuenta", "tipo de cuenta bancaria")
    col_num = col_de("numero cuenta", "numero de cuenta", "n cuenta", "no cuenta")
    faltantes = [n for n, c in zip(CABECERAS_CUENTAS, [col_placa, col_cond, col_tel, col_benef, col_ced, col_banco, col_tipo, col_num]) if c is None]
    if faltantes:
        raise HTTPException(
            status_code=422,
            detail=f"El archivo debe tener las columnas: {', '.join(CABECERAS_CUENTAS)}. Faltan: {', '.join(faltantes)}.",
        )

    resultados, errores = [], []
    ahora = _ahora_utc()
    auditoria = {"usuario": info["usuario"], "nombre": info["nombre"], "rol": info["perfil"], "fecha": ahora}

    def celda(fila, idx):
        return _celda_texto(fila[idx] if idx < len(fila) else "")

    for n, fila in enumerate(filas[1:], start=2):
        if fila is None or all(v is None or str(v).strip() == "" for v in fila):
            continue
        try:
            p = _validar_placa(celda(fila, col_placa))
            banco = _resolver_banco(celda(fila, col_banco))
            tipo = _resolver_tipo_cuenta(celda(fila, col_tipo))
            cedula = _validar_digitos(celda(fila, col_ced), "La cédula", 5, 15)
            numero = _validar_digitos(celda(fila, col_num), "El número de cuenta", 5, 20)
            telefono = _validar_telefono(celda(fila, col_tel))
            benef = celda(fila, col_benef).strip().upper()
            if not benef:
                raise HTTPException(status_code=422, detail="El nombre del beneficiario es obligatorio.")
        except HTTPException as e:
            errores.append({"fila": n, "placa": celda(fila, col_placa), "detalle": str(e.detail)})
            continue

        res = col_cuentas.update_one(
            {"placa": p, "regional": co},
            {"$set": {
                "nombre_conductor": celda(fila, col_cond).strip().upper(),
                "telefono": telefono,
                "nombre_beneficiario": benef,
                "cedula": cedula,
                "banco": banco,
                "tipo_cuenta": tipo,
                "numero_cuenta": numero,
                "actualizado_por": auditoria,
                "updated_at": ahora,
             },
             "$setOnInsert": {
                "regional": co, "regional_info": norm,
                "creado_por": auditoria, "created_at": ahora,
             }},
            upsert=True,
        )
        resultados.append({"fila": n, "placa": p, "accion": "actualizada" if res.upserted_id is None else "creada"})

    return {
        "mensaje": f"{len(resultados)} procesadas, {len(errores)} con error",
        "procesadas": resultados,
        "errores": errores,
    }


# ── Plantilla y exportación Excel ─────────────────────────────────────────────
def _estilo_encabezados(ws, columnas):
    fill = PatternFill(start_color="004d40", end_color="004d40", fill_type="solid")
    for i, col in enumerate(columnas, 1):
        celda = ws.cell(row=1, column=i, value=col)
        celda.font = Font(color="FFFFFF", bold=True)
        celda.fill = fill
        ws.column_dimensions[chr(64 + i)].width = 24


@router.get("/plantilla")
async def plantilla():
    """Plantilla .xlsx con las 7 columnas + hoja de valores válidos (bancos/tipos)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuentas"
    ws.append(CABECERAS_CUENTAS)
    _estilo_encabezados(ws, CABECERAS_CUENTAS)

    ws_val = wb.create_sheet("Valores")
    ws_val.append(["BANCOS PERMITIDOS", "CODIGO"])
    for b in BANCOS_OTROS_COSTOS_DEFAULT:
        ws_val.append([b["nombre"], b["codigo"]])
    fila_tipos = len(BANCOS_OTROS_COSTOS_DEFAULT) + 4
    ws_val.cell(row=fila_tipos, column=1, value="TIPOS DE CUENTA PERMITIDOS").font = Font(bold=True)
    for i, t in enumerate(TIPOS_CUENTA, start=fila_tipos + 1):
        ws_val.cell(row=i, column=1, value=t)
    ws_val.column_dimensions["A"].width = 38
    ws_val.column_dimensions["B"].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content=buffer.getvalue(),
        media_type=MEDIA_TYPE_XLSX,
        headers={
            "Content-Disposition": f"attachment; filename=plantilla_cuentas_placa_{_hoy_colombia().strftime('%Y%m%d')}.xlsx",
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@router.get("/exportar-excel")
async def exportar_excel(usuario: str = Query(...), regional: str = Query("")):
    """Exporta el catálogo (con el scope/regional del filtro) a .xlsx."""
    info = _resolver_usuario(usuario)
    _requiere(info, PERFILES_CUENTAS_PLACA, "exportar el catálogo de cuentas por placa")
    co = _regional_scope(info, regional)

    filtro: dict = {}
    if co:
        filtro["regional"] = co

    wb = Workbook()
    ws = wb.active
    ws.title = "Cuentas"
    columnas = CABECERAS_CUENTAS + ["REGIONAL"]
    ws.append(columnas)
    _estilo_encabezados(ws, columnas)
    for d in col_cuentas.find(filtro).sort([("regional", 1), ("placa", 1)]):
        ws.append([
            d.get("placa", ""), d.get("nombre_conductor", ""), d.get("telefono", ""),
            d.get("nombre_beneficiario", ""), d.get("cedula", ""), d.get("banco", ""),
            d.get("tipo_cuenta", ""), d.get("numero_cuenta", ""),
            (d.get("regional_info") or {}).get("bodega", d.get("regional", "")),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return Response(
        content=buffer.getvalue(),
        media_type=MEDIA_TYPE_XLSX,
        headers={
            "Content-Disposition": f"attachment; filename=cuentas_por_placa_{_hoy_colombia().strftime('%Y%m%d')}.xlsx",
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )
