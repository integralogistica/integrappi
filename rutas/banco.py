from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import Response
import io
import logging
import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/banco", tags=["Banco"])

# Columnas definidas por posición X (basado en extractos de Bancolombia)
COLUMNS = [
    ("FECHA", 23, 71),
    ("DESCRIPCION", 72, 239),
    ("SUCURSAL_CANAL", 240, 328),
    ("REFERENCIA_1", 329, 397),
    ("REFERENCIA_2", 398, 466),
    ("DOCUMENTO", 467, 525),
    ("VALOR", 526, 600),
]

Y_TOLERANCE = 5


def _extract_header(words):
    """Extrae información del encabezado del PDF."""
    info = {
        "empresa": "",
        "numero_cuenta": "",
        "nit": "",
        "tipo_cuenta": "",
        "saldo_efectivo": "",
        "saldo_canje": "",
        "saldo_total": "",
    }

    full_text = " ".join(w["text"] for w in words)

    if words:
        min_top = min(w["top"] for w in words)
        empresa_words = [w for w in words if abs(w["top"] - min_top) < 2]
        info["empresa"] = " ".join(w["text"] for w in sorted(empresa_words, key=lambda w: w["x0"]))

    m = re.search(r"Cuenta:\s*(\d+)", full_text)
    if m:
        info["numero_cuenta"] = m.group(1)

    m = re.search(r"NIT:\s*(\d+)", full_text)
    if m:
        info["nit"] = m.group(1)

    m = re.search(r"Tipo de cuenta:\s*(\w+)", full_text)
    if m:
        info["tipo_cuenta"] = m.group(1)

    m = re.search(r"Saldo Efectivo Actual:\s*\$?([\d,]+\.\d{2})", full_text)
    if m:
        info["saldo_efectivo"] = m.group(1)

    m = re.search(r"Saldo en Canje Actual:\s*\$?([\d,]+\.\d{2})", full_text)
    if m:
        info["saldo_canje"] = m.group(1)

    m = re.search(r"Saldo Total Actual:\s*\$?([\d,]+\.\d{2})", full_text)
    if m:
        info["saldo_total"] = m.group(1)

    return info


def _group_words_by_row(words):
    """Agrupa palabras en filas basándose en su posición Y."""
    if not words:
        return []

    sorted_words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    rows = []
    current_row = [sorted_words[0]]

    for word in sorted_words[1:]:
        if abs(word["top"] - current_row[0]["top"]) <= Y_TOLERANCE:
            current_row.append(word)
        else:
            rows.append(current_row)
            current_row = [word]
    rows.append(current_row)

    return rows


def _parse_row(row_words):
    """Parsea una fila de palabras en una transacción usando posiciones de columna."""
    col_data = {col[0]: [] for col in COLUMNS}

    for word in row_words:
        assigned = False
        for col_name, x_start, x_end in COLUMNS:
            word_center = (word["x0"] + word["x1"]) / 2
            if x_start - 5 <= word_center <= x_end + 5:
                col_data[col_name].append(word["text"])
                assigned = True
                break
        if not assigned:
            word_center = (word["x0"] + word["x1"]) / 2
            best_col = min(COLUMNS, key=lambda c: abs(word_center - (c[1] + c[2]) / 2))
            col_data[best_col[0]].append(word["text"])

    fecha_str = " ".join(col_data["FECHA"])
    valor_str = " ".join(col_data["VALOR"])
    valor_str = valor_str.replace(",", "").replace(".", ",")
    desc_str = " ".join(col_data["DESCRIPCION"])
    suc_str = " ".join(col_data["SUCURSAL_CANAL"])
    ref1_str = " ".join(col_data["REFERENCIA_1"])
    ref2_str = " ".join(col_data["REFERENCIA_2"])
    doc_str = " ".join(col_data["DOCUMENTO"])

    # Fila de transacción válida: tiene fecha y valor
    if re.match(r"\d{4}/\d{2}/\d{2}", fecha_str) and re.match(r"-?[\d]+,\d{2}", valor_str):
        return {
            "tipo": "transaccion",
            "fecha": fecha_str,
            "descripcion": desc_str,
            "sucursal": suc_str,
            "referencia1": ref1_str,
            "referencia2": ref2_str,
            "documento": doc_str,
            "valor": valor_str,
        }

    # Fila huérfana: no tiene fecha ni valor, pero tiene datos en columnas de referencia/descripción
    orphan_data = ref1_str + ref2_str + doc_str + desc_str + suc_str
    if orphan_data.strip() and not fecha_str.strip() and not valor_str.strip():
        return {
            "tipo": "orfana",
            "descripcion": desc_str,
            "sucursal": suc_str,
            "referencia1": ref1_str,
            "referencia2": ref2_str,
            "documento": doc_str,
        }

    return None


def extract_transactions(pdf_bytes: bytes):
    """Extrae transacciones desde los bytes de un PDF."""
    transactions = []
    header_info = {}

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            words = page.extract_words(keep_blank_chars=True)

            if page_idx == 0:
                header_info = _extract_header(words)
                fecha_headers = [w for w in words if w["text"] == "FECHA"]
                if fecha_headers:
                    data_start_y = fecha_headers[0]["top"] + 10
                    data_words = [w for w in words if w["top"] > data_start_y]
                else:
                    data_words = words
            else:
                data_words = words

            data_words = [w for w in data_words
                          if not w["text"].startswith("Pagina")]

            rows = _group_words_by_row(data_words)

            for row_words in rows:
                parsed = _parse_row(row_words)
                if not parsed:
                    continue

                if parsed["tipo"] == "transaccion":
                    transactions.append(parsed)
                elif parsed["tipo"] == "orfana" and transactions:
                    prev = transactions[-1]
                    if parsed["referencia2"]:
                        if prev["referencia2"]:
                            prev["referencia2"] += " " + parsed["referencia2"]
                        else:
                            prev["referencia2"] = parsed["referencia2"]

    for txn in transactions:
        txn.pop("tipo", None)

    return header_info, transactions


def create_excel(header_info, transactions) -> bytes:
    """Genera Excel en memoria y retorna los bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracto Bancario"

    dark_blue = "003366"
    light_blue = "D6E4F0"
    green = "548235"
    red = "C00000"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    subtitle_font = Font(name="Calibri", bold=True, color=dark_blue, size=11)
    normal_font = Font(name="Calibri", size=10)
    header_fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    alt_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    columns = ["FECHA", "DESCRIPCION", "SUCURSAL/CANAL", "REFERENCIA 1", "REFERENCIA 2", "DOCUMENTO", "VALOR"]
    col_widths = [14, 42, 18, 18, 18, 18, 20]

    for col_idx, (col_name, width) in enumerate(zip(columns, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    row = 2
    for i, txn in enumerate(transactions):
        data = [
            txn["fecha"],
            txn["descripcion"],
            txn["sucursal"],
            txn["referencia1"],
            txn["referencia2"],
            txn["documento"],
            txn["valor"],
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border

            if col_idx == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                val_str = str(value).replace(",", "").replace("$", "")
                try:
                    num = float(val_str)
                    if num < 0:
                        cell.font = Font(name="Calibri", size=10, color=red)
                    else:
                        cell.font = Font(name="Calibri", size=10, color=green)
                except ValueError:
                    pass
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if i % 2 == 1:
                cell.fill = alt_fill

        row += 1

    row += 1
    ws.cell(row=row, column=1, value=f"Total transacciones: {len(transactions)}").font = subtitle_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{1 + len(transactions)}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@router.post("/pdf-a-excel")
async def pdf_a_excel(file: UploadFile = File(...)):
    """
    Recibe un PDF de extracto bancario y retorna un Excel con las transacciones.
    """
    try:
        if not file.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="El archivo debe ser un PDF")

        pdf_bytes = await file.read()

        if len(pdf_bytes) == 0:
            raise HTTPException(status_code=400, detail="El archivo esta vacio")

        logger.info(f"Procesando PDF: {file.filename} ({len(pdf_bytes)} bytes)")

        header_info, transactions = extract_transactions(pdf_bytes)

        if not transactions:
            raise HTTPException(status_code=404, detail="No se encontraron transacciones en el PDF")

        logger.info(f"Transacciones encontradas: {len(transactions)}")

        excel_bytes = create_excel(header_info, transactions)

        base_name = file.filename.rsplit(".", 1)[0]
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={base_name}_extracto.xlsx"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error procesando PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error procesando PDF: {str(e)}")
