"""
Módulo «Otros Costos» — Integrapp.

Gestiona los costos adicionales que algunos vehículos generan después de prestado
el servicio (parqueadero, peaje, cargue, horas adicionales, etc.). El equipo
operativo registra una solicitud asociada a uno o varios pedidos de Vulcano
(consultados en `pedidos_medical_historico`), y esta sigue un flujo de aprobación
por monto → pago → histórico, con trazabilidad completa.

Colecciones (base de datos `integra`):
  - `otros_costos`            ciclo activo (borrador, pendiente_aprobacion, devuelto, rechazado, aprobado)
  - `historico_otros_costos`  solicitudes pagadas
  - `anulados_otros_costos`   solicitudes anuladas

Seguridad: el perfil NO se confía del frontend. Cada endpoint recibe `usuario` y
lo resuelve contra `baseusuarios` para autorizar con el perfil real.
"""

import asyncio
import logging
import re
from datetime import datetime, timedelta, timezone
from typing import List, Literal, Optional

from bson import ObjectId
from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import Response
from pydantic import BaseModel, Field, field_validator
from pymongo.errors import DuplicateKeyError

from bd.bd_cliente import bd_cliente
from Funciones.whatsapp_utils_integra import enviar_template_sync

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/otros-costos", tags=["Otros Costos"])

# Colombia = UTC-5. El servidor corre en UTC y Mongo guarda los datetime como
# instantes UTC; los límites por "día Colombia" se alinean sumando 5 h.
_OFFSET_COLOMBIA = timedelta(hours=5)
LIMITE_COORDINADOR = 500000  # Coordinador aprueba hasta este valor inclusive
LIMITE_VALOR_SOLICITUD = 5_000_000  # Valor total máximo permitido por solicitud

# ── Plantillas de notificación WhatsApp (Meta, es_CO) ─────────────────────────
# Una plantilla por evento del flujo. Requieren crearse/aprobarse en Meta Business
# Manager; mientras no existan, el envío falla en silencio (solo log) y la acción
# del flujo igual se completa. El destinatario y el número se resuelven en backend
# desde `baseusuarios` (no se confía en el frontend).
#   (nombre, idioma)
PLANTILLA_OC_APROBACION = ("oc_solicitud_aprobacion", "es_CO")  # → COORDINADOR/CONTROL
PLANTILLA_OC_TRAMITE = ("oc_para_tramite", "es_CO")            # → ANALISTA
PLANTILLA_OC_PAGO = ("oc_para_pago", "es_CO")                  # → FINANCIERO
PLANTILLA_OC_PAGADA = ("oc_pago_realizado", "es_CO")           # → OPERATIVO creador
PLANTILLA_OC_DEVUELTA = ("oc_devuelta", "es_CO")               # → OPERATIVO creador
PLANTILLA_OC_RECH_ANUL = ("oc_rechazada_anulada", "es_CO")     # → OPERATIVO creador

# Catálogo por defecto de causales/tipos de costo (se siembra en `causales_otros_costos`).
# Editable directamente en Mongo: si la colección tiene documentos, se usa tal cual.
CAUSALES_OTROS_COSTOS_DEFAULT = [
    "AFORO", "CARGUE", "DESCARGUE", "DESVIO", "DEVOLUCIONES",
    "ENTREGA EN VEREDA", "OTROS", "PUNTO ADICIONAL", "RECOLECCIONES",
    "REQUERIMIENTO", "STAND BY", "TRASBORDO", "TRASLADO", "URGENCIA",
]
# Catálogo de bancos (nombre + código bancario). Se siembra en `bancos_otros_costos`
# con el mismo patrón que clientes/causales: editable directamente en Mongo.
# Nota: NEQUI/DAVIPLATA/BANCOLOMBIA A LA MANO son billeteras asociadas a un banco
# (comparten código con su banco matriz, según la lista oficial enviada).
BANCOS_OTROS_COSTOS_DEFAULT = [
    {"nombre": "BANCO DE LA REPUBLICA", "codigo": "0"},
    {"nombre": "BANCO DE BOGOTA", "codigo": "1"},
    {"nombre": "BANCO POPULAR", "codigo": "2"},
    {"nombre": "ITAU CORPBANCA COLOMBIA S.A.", "codigo": "6"},
    {"nombre": "BANCOLOMBIA", "codigo": "7"},
    {"nombre": "CITIBANK COLOMBIA", "codigo": "9"},
    {"nombre": "GNB SUDAMERIS S.A.", "codigo": "12"},
    {"nombre": "BBVA", "codigo": "13"},
    {"nombre": "COLPATRIA", "codigo": "19"},
    {"nombre": "OCCIDENTE", "codigo": "23"},
    {"nombre": "CAJA SOCIAL", "codigo": "32"},
    {"nombre": "BANCO AGRARIO DE COLOMBIA S.A.", "codigo": "40"},
    {"nombre": "DAVIVIENDA", "codigo": "51"},
    {"nombre": "AV VILLAS", "codigo": "52"},
    {"nombre": "BANCO W S.A.", "codigo": "53"},
    {"nombre": "BANCO CREDIFINANCIERA S.A.C.F", "codigo": "58"},
    {"nombre": "BANCAMIA", "codigo": "59"},
    {"nombre": "BANCO PICHINCHA S.A.", "codigo": "60"},
    {"nombre": "BANCOOMEVA", "codigo": "61"},
    {"nombre": "CMR FALABELLA S.A.", "codigo": "62"},
    {"nombre": "BANCO FINANDINA S.A.", "codigo": "63"},
    {"nombre": "BANCO SANTANDER DE NEGOCIOS COLOMBIA S.A.", "codigo": "65"},
    {"nombre": "BANCO COOPERATIVO COOPCENTRAL", "codigo": "66"},
    {"nombre": "BANCO COMPARTIR S.A", "codigo": "67"},
    {"nombre": "BANCO SERFINANZA S.A", "codigo": "69"},
    {"nombre": "NEQUI", "codigo": "507"},
    {"nombre": "DAVIPLATA", "codigo": "51"},
    {"nombre": "BANCOLOMBIA A LA MANO", "codigo": "7"},
    {"nombre": "LULO", "codigo": "70"},
]
TIPOS_CUENTA = ["Ahorros", "Corriente", "Depósito electrónico", "BILLETERA DIGITAL", "TARJETA PREPAGO"]
TIPOS_ID_TITULAR = ["CC", "NIT"]
# Catálogo por defecto de clientes para el formulario (se siembra en `clientes_otros_costos`).
# Editable directamente en Mongo: si la colección tiene documentos, se usa tal cual.
CLIENTES_OTROS_COSTOS_DEFAULT = [
    "FRESENIUS MEDICAL CARE",
    "CONGRUPO",
    "FRESENIUS KABI",
    "DAVITA",
    "FK SERVICIO TECNICO",
    "ORTOPEDICOS FUTURO COLOMBIA",
    "DISTRIBUIDORA COMTEK S.A.S.",
    "QUIMICA AVANZADA SAS",
    "MINISO COLOMBIA SAS",
]
ESTADOS_VALIDOS = [
    "borrador", "pendiente_aprobacion", "devuelto", "rechazado",
    "aprobado", "pagado", "anulado",
]

# ── Conexión MongoDB ──────────────────────────────────────────────────────────
client = bd_cliente
db = client["integra"]
col_activos = db["otros_costos"]
col_historico = db["historico_otros_costos"]
col_anulados = db["anulados_otros_costos"]
col_historico_pedidos = db["pedidos_medical_historico"]   # solo lectura (lookup)
col_usuarios = db["baseusuarios"]                          # resolución de identidad
col_clientes = db["clientes_otros_costos"]                 # catálogo de clientes (formulario)
col_causales = db["causales_otros_costos"]                 # catálogo de tipos de costo (formulario)
col_bancos = db["bancos_otros_costos"]                     # catálogo de bancos: nombre + código

# Índices (idempotentes al importar, igual patrón que siscore_consultas.py).
# `consecutivo` unique en activos → anti-colisión en la generación del consecutivo.
try:
    col_activos.create_index([("consecutivo", 1)], unique=True, name="idx_oc_consecutivo")
except Exception as _e:
    logger.warning(f"[OTROS_COSTOS] No se pudo crear índice unique 'consecutivo' en activos: {_e}")
for _col, _nombre in (
    (col_historico, "idx_oc_hist_consecutivo"),
    (col_anulados, "idx_oc_anul_consecutivo"),
):
    try:
        _col.create_index([("consecutivo", 1)], name=_nombre)
    except Exception as _e:
        logger.warning(f"[OTROS_COSTOS] No se pudo crear índice 'consecutivo': {_e}")
for _col in (col_activos, col_historico, col_anulados):
    for _fld in ("estado", "usuario_registro", "pedidos_normalizados"):
        try:
            _col.create_index(_fld)
        except Exception:
            pass
    try:
        _col.create_index([("created_at", -1)], name=f"idx_oc_created_{_col.name}")
    except Exception:
        pass


# ── Utilidades de tiempo ──────────────────────────────────────────────────────
def _ahora_utc() -> datetime:
    """UTC actual como datetime naive (igualconvención que el resto del sistema)."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _hoy_colombia() -> datetime:
    """Fecha/hora actual en zona Colombia (UTC-5) como datetime naive."""
    return _ahora_utc() - _OFFSET_COLOMBIA


# ── Regional: normalización de formatos ───────────────────────────────────────
# La regional vive en 3 formatos según el campo:
#   - baseusuarios.regional y otros_costos.regional_registro   -> CÓDIGO CO (CO04...)
#   - otros_costos.datos_servicio.centro_distribucion         -> NOMBRE DE BODEGA
#   - también aparece el nombre de ciudad (CALI/BARRANQUILLA/...)
# Estos maps permiten traducir entre los tres para filtrar de forma robusta
# (incluye docs viejos cuyo regional_registro pudo no quedar poblado).
CO_A_REGIONAL = {
    "CO04": "BARRANQUILLA", "CO05": "CALI", "CO06": "BUCARAMANGA",
    "CO07": "FUNZA", "CO09": "MEDELLIN",
}
REGIONAL_A_BODEGA = {
    "BARRANQUILLA": "JUAN MINA", "CALI": "YUMBO", "MEDELLIN": "GIRARDOTA",
    # BUCARAMANGA y FUNZA no tienen bodega de origen distinta: identidad.
    "BUCARAMANGA": "BUCARAMANGA", "FUNZA": "FUNZA",
}
BODEGA_A_REGIONAL = {b: r for r, b in REGIONAL_A_BODEGA.items()}
# Perfiles que ven TODAS las regionales (y a los que se les ofrece el dropdown).
PERFILES_GLOBALES_OC = {"ADMIN", "ANALISTA", "COORDINADOR", "CONTROL"}


def _normalizar_regional(valor: str) -> Optional[dict]:
    """Dada una regional en cualquiera de los 3 formatos (código CO, nombre de
    ciudad o nombre de bodega), devuelve los sinónimos canónicos, o None si no la
    reconoce. Ignora mayúsculas/espacios."""
    v = (valor or "").strip().upper()
    if not v:
        return None
    if v in CO_A_REGIONAL:                                   # por código CO
        regional = CO_A_REGIONAL[v]
        return {"co": v, "regional": regional,
                "bodega": REGIONAL_A_BODEGA.get(regional, regional)}
    if v in REGIONAL_A_BODEGA:                               # por nombre de ciudad
        co = next((c for c, r in CO_A_REGIONAL.items() if r == v), "")
        return {"co": co, "regional": v, "bodega": REGIONAL_A_BODEGA[v]}
    if v in BODEGA_A_REGIONAL:                               # por nombre de bodega
        regional = BODEGA_A_REGIONAL[v]
        co = next((c for c, r in CO_A_REGIONAL.items() if r == regional), "")
        return {"co": co, "regional": regional, "bodega": v}
    return None


def _sinonimos_regional(valor: str) -> set:
    """Set (en mayúsculas) de todas las formas de escribir la misma regional,
    para comparar contra cualquier campo del documento."""
    norm = _normalizar_regional(valor)
    if not norm:
        return set()
    return {s for s in (norm["co"], norm["regional"], norm["bodega"]) if s}


def _aplicar_filtro_regional(filtro: dict, valor: str) -> None:
    """Agrega al filtro un $or que casa la regional en los dos campos donde puede
    vivir: regional_registro (código CO) y datos_servicio.centro_distribucion
    (nombre de bodega). Cubre documentos viejos sin regional_registro. Si no
    reconoce el valor, deja el filtro intacto."""
    sinonimos = _sinonimos_regional(valor)
    if not sinonimos:
        return
    condiciones = []
    for s in sinonimos:
        condiciones.append({"regional_registro": s})
        condiciones.append({"datos_servicio.centro_distribucion": s})
    filtro["$or"] = condiciones


def _doc_coincide_regional(doc: dict, valor: str) -> bool:
    """True si la regional del documento (en cualquiera de sus campos/formatos)
    coincide con la regional indicada."""
    sinonimos = _sinonimos_regional(valor)
    if not sinonimos:
        return False
    ds = doc.get("datos_servicio") or {}
    valores_doc = {
        str(doc.get("regional_registro", "")).strip().upper(),
        str(ds.get("centro_distribucion", "")).strip().upper(),
    }
    return bool(valores_doc & sinonimos)


# ── Resolución de identidad / autorización ────────────────────────────────────
def _resolver_usuario(usuario: str) -> dict:
    """Devuelve el perfil/region/nombre REALES del usuario desde baseusuarios.

    El backend NO confía en el perfil que envía el frontend: lo deriva de la BD.
    Lanza 401 si no llega usuario / no existe, 403 si está inactivo.
    """
    if not usuario or not str(usuario).strip():
        raise HTTPException(status_code=401, detail="No autenticado")
    u = str(usuario).strip().upper()
    doc = col_usuarios.find_one({"usuario": u})
    if not doc:
        raise HTTPException(status_code=401, detail="Usuario no válido")
    if not doc.get("activo", True):
        raise HTTPException(status_code=403, detail="Usuario inactivo")
    return {
        "usuario": doc["usuario"],
        "perfil": (doc.get("perfil") or "").strip().upper(),
        "regional": doc.get("regional", ""),
        "nombre": doc.get("nombre", ""),
        "id": str(doc["_id"]),
    }


def _requiere(info: dict, perfiles: set[str], accion: str):
    if info["perfil"] not in perfiles:
        raise HTTPException(
            status_code=403,
            detail=f"Su perfil ({info['perfil']}) no tiene permiso para {accion}.",
        )


# ── Normalización de pedidos Vulcano (spec §4) ────────────────────────────────
def _normalizar_pedidos(texto: str) -> List[str]:
    """Normaliza una lista de pedidos (string) a una lista de strings sin ceros
    a la izquierda, sin duplicados, separando por coma/guion/puntoycoma/espacios."""
    if not texto:
        return []
    partes = re.split(r"[,\-;\s/]+", str(texto))
    out: List[str] = []
    seen: set[str] = set()
    for p in partes:
        digits = re.sub(r"\D", "", p)          # sólo dígitos (descarta caracteres inválidos)
        if not digits:
            continue
        norm = digits.lstrip("0") or "0"        # comparar sin ceros a la izquierda
        if norm not in seen:
            seen.add(norm)
            out.append(norm)
    return out


def _proyeccion_pedido(doc: dict) -> dict:
    """Proyecta un documento de pedidos_medical_historico al formato que necesita
    el módulo de Otros Costos (datos_servicio + referencia al origen)."""
    return {
        "_id_origen": str(doc["_id"]),
        "pedido_vulcano": str(doc.get("pedido_vulcano") or doc.get("codigo_pedido") or ""),
        "cliente": str(doc.get("cliente_origen") or ""),
        "centro_distribucion": str(doc.get("regional") or doc.get("centro_costo") or ""),
        "fecha_servicio": _fecha_a_str(doc.get("fecha_movimiento_historico"))
        or _fecha_a_str(doc.get("fecha_aprobacion"))
        or _fecha_a_str(doc.get("fecha_creacion")),
        "piezas": _a_numero(doc.get("piezas")),
        "peso_real": _a_numero(doc.get("peso_real")),
        "tipo_vehiculo": str(doc.get("tipo_vehiculo") or doc.get("tipo_veh_sicetac") or ""),
        "placa": str(doc.get("placa") or ""),
        "municipio_destino": str(doc.get("municipio_destino") or ""),
        "departamento_destino": str(doc.get("departamento_destino") or ""),
        "transportador": str(doc.get("transportador") or ""),
        "manifiesto": str(doc.get("manifiesto") or ""),
        "total_solicitado": _a_numero(doc.get("total_solicitado")),
        "regional": str(doc.get("regional") or ""),
        "estado_pedido": str(doc.get("estado") or ""),
    }


def _proyeccion_pedido_desde_original(original: dict, doc_fusion: dict) -> dict:
    """Proyecta un original embebido (fusion_info.datos_originales[i]) al formato de
    Otros Costos. Usa los datos INDIVIDUALES del pedido; para campos que solo existen
    a nivel del vehículo fusionado (manifiesto, transportador) usa el documento raíz."""
    oid = str(doc_fusion.get("_id", ""))
    sub = (original.get("consecutivo") or original.get("pedido_vulcano")
           or original.get("codigo_pedido") or "")
    return {
        "_id_origen": f"{oid}#{sub}" if sub else oid,
        "pedido_vulcano": str(original.get("pedido_vulcano") or original.get("codigo_pedido") or ""),
        "cliente": str(original.get("cliente_origen") or doc_fusion.get("cliente_origen") or ""),
        "centro_distribucion": str(original.get("regional") or original.get("centro_costo")
                                   or doc_fusion.get("regional") or ""),
        "fecha_servicio": (_fecha_a_str(original.get("fecha_preaprobado"))
                           or _fecha_a_str(original.get("fecha_pedido_vulcano"))
                           or _fecha_a_str(doc_fusion.get("fecha_movimiento_historico"))
                           or _fecha_a_str(doc_fusion.get("fecha_aprobacion"))
                           or _fecha_a_str(doc_fusion.get("fecha_creacion"))),
        "piezas": _a_numero(original.get("piezas")),
        "peso_real": _a_numero(original.get("peso_real")),
        "tipo_vehiculo": str(original.get("tipo_vehiculo") or original.get("tipo_veh_sicetac")
                             or doc_fusion.get("tipo_vehiculo") or ""),
        "placa": str(original.get("placa") or doc_fusion.get("placa") or ""),
        "municipio_destino": str(original.get("municipio_destino") or ""),
        "departamento_destino": str(original.get("departamento_destino")
                                    or doc_fusion.get("departamento_destino") or ""),
        "transportador": str(doc_fusion.get("transportador") or original.get("transportador") or ""),
        "manifiesto": str(doc_fusion.get("manifiesto") or original.get("manifiesto") or ""),
        "total_solicitado": _a_numero(original.get("total_solicitado")),
        "regional": str(original.get("regional") or doc_fusion.get("regional") or ""),
        "estado_pedido": str(original.get("estado") or doc_fusion.get("estado") or ""),
    }


def _buscar_pedidos_historico(pedidos_norm: List[str]) -> List[dict]:
    """Busca en pedidos_medical_historico los docs cuyo pedido_vulcano/codigo_pedido
    coincide (tolerante a ceros a la izquierda y separadores) con alguno de los
    pedidos normalizados. Regex armado solo con dígitos escapados (sin inyección).

    Si el documento es una planilla FUSIONADA, proyecta SOLO el/los original(es)
    embebido(s) en fusion_info.datos_originales que correspondan al pedido buscado,
    de modo que se traiga la información individual del pedido (sus propias piezas,
    peso, placa, destino, etc.) y no los totales agregados del vehículo fusionado."""
    if not pedidos_norm:
        return []
    condiciones = []
    for n in pedidos_norm:
        patron = rf"(^|[^0-9])0*{re.escape(n)}([^0-9]|$)"
        condiciones.append({"pedido_vulcano": {"$regex": patron}})
        condiciones.append({"codigo_pedido": {"$regex": patron}})
    encontrados: List[dict] = []
    vistos: set[str] = set()
    set_norm = set(pedidos_norm)
    for doc in col_historico_pedidos.find({"$or": condiciones}):
        fusion_info = doc.get("fusion_info") or {}
        datos_originales = (fusion_info.get("datos_originales")
                            if fusion_info.get("es_fusionada") else None)
        if datos_originales:
            # Planilla fusionada: proyectar solo el/los original(es) que coinciden.
            alguno = False
            for original in datos_originales:
                almacenados_orig = _normalizar_pedidos(
                    f"{original.get('pedido_vulcano','')},{original.get('codigo_pedido','')}"
                )
                if set_norm & set(almacenados_orig):
                    sub = (original.get("consecutivo") or original.get("pedido_vulcano")
                           or original.get("codigo_pedido") or "")
                    clave = f"{doc['_id']}#{sub}"
                    if clave in vistos:
                        continue
                    vistos.add(clave)
                    encontrados.append(_proyeccion_pedido_desde_original(original, doc))
                    alguno = True
            if alguno:
                continue
            # Fallback: la fusión coincidió por el pedido concatenado en la raíz, pero
            # ningún original individual coincide (p.ej. sin pedido asignado). Proyectar raíz.
        # No es fusión (o fallback): proyectar el documento top-level.
        almacenados = _normalizar_pedidos(
            f"{doc.get('pedido_vulcano','')},{doc.get('codigo_pedido','')}"
        )
        if set_norm & set(almacenados):                  # confirmar intersección real
            oid = str(doc["_id"])
            if oid in vistos:
                continue
            vistos.add(oid)
            encontrados.append(_proyeccion_pedido(doc))
    return encontrados


# ── Consecutivo {BODEGA}-OC-AAAAMMDD-NNNN ─────────────────────────────────────
# El prefijo es la bodega/regional del creador (YUMBO, FUNZA, ...), igual que en
# SolicitudVehiculos. Una secuencia independiente por regional y día.
def _generar_consecutivo_oc(fecha_col: datetime, regional: str = "") -> str:
    norm = _normalizar_regional(regional)
    bodega = (norm or {}).get("bodega") or "FUNZA"   # fallback FUNZA si no hay regional
    prefijo = f"{bodega}-OC-{fecha_col.strftime('%Y%m%d')}-"
    regex_prefijo = re.compile(rf"^{re.escape(prefijo)}(\d{{4}})$")
    max_n = 0
    for col in (col_activos, col_historico, col_anulados):
        for d in col.find({"consecutivo": {"$regex": f"^{re.escape(prefijo)}"}}, {"consecutivo": 1}):
            m = regex_prefijo.match(str(d.get("consecutivo", "")))
            if m:
                max_n = max(max_n, int(m.group(1)))
    return f"{prefijo}{max_n + 1:04d}"


def _insertar_con_reintento(doc: dict) -> str:
    """Inserta el doc generando consecutivo con reintento ante colisión (concurrencia)."""
    fecha_col = _hoy_colombia()
    ultimo_error = None
    for _ in range(3):
        doc["consecutivo"] = _generar_consecutivo_oc(fecha_col, doc.get("regional_registro", ""))
        try:
            col_activos.insert_one(doc)
            return doc["consecutivo"]
        except DuplicateKeyError as e:
            ultimo_error = e
            continue
    raise HTTPException(status_code=500, detail="No se pudo generar un consecutivo único")


# ── Trazabilidad (spec §11) ───────────────────────────────────────────────────
def _nuevo_movimiento(
    accion: str,
    estado_anterior: Optional[str],
    estado_nuevo: Optional[str],
    info: dict,
    observacion: str = "",
    ip: str = "",
) -> dict:
    return {
        "accion": accion,
        "estado_anterior": estado_anterior,
        "estado_nuevo": estado_nuevo,
        "usuario": info.get("usuario", ""),
        "nombre_usuario": info.get("nombre", ""),
        "rol": info.get("perfil", ""),
        "fecha": _ahora_utc(),
        "observacion": observacion or "",
        "ip": ip or "",
    }


# ── Notificaciones WhatsApp a los actores del flujo ───────────────────────────
# Fire-and-forget: si algo falla (sin celular, plantilla sin aprobar en Meta, error
# de red) solo queda en log y NO rompe la acción del flujo. El celular y el nombre
# se resuelven en `baseusuarios` (mismo patrón que siscore_consultas._notificar_*).
def _normalizar_celular_co(celular: Optional[str]) -> Optional[str]:
    """Convierte un celular a formato internacional Colombia: solo dígitos con 57."""
    if not celular:
        return None
    limpio = "".join(c for c in str(celular) if c.isdigit())
    if not limpio:
        return None
    if not limpio.startswith("57"):
        limpio = "57" + limpio
    return limpio


def _resolver_celular_nombre(usuario: str) -> Optional[tuple]:
    """Lookup único en baseusuarios → (celular_normalizado, nombre) o None."""
    if not usuario or not str(usuario).strip():
        return None
    doc = col_usuarios.find_one({"usuario": str(usuario).strip().upper()})
    if not doc:
        return None
    celular = _normalizar_celular_co(doc.get("celular"))
    if not celular:
        return None
    nombre = (doc.get("nombre") or doc.get("usuario") or "Usuario").strip()
    return (celular, nombre)


def _valor_cop(v) -> str:
    """Formatea un monto como pesos colombianos con punto de miles: 320000 → '320.000'."""
    try:
        return f"{int(round(float(v or 0))):,}".replace(",", ".")
    except (TypeError, ValueError):
        return "0"


def _enviar_a_perfil(perfil: str, plantilla: tuple, body_params_fn, *, consecutivo: str) -> None:
    """Itera los usuarios activos de `perfil` y les envía la plantilla. `body_params_fn`
    recibe (nombre) y devuelve la lista de parámetros del cuerpo ({{1}}, {{2}}, ...)."""
    try:
        usuarios = list(col_usuarios.find({
            "perfil": perfil,
            "$or": [{"activo": True}, {"activo": {"$exists": False}}],
        }))
        if not usuarios:
            logger.info(f"[NOTIF OC] No hay usuarios activos con perfil {perfil}; no se notifica ({consecutivo}).")
            return
        enviadas = 0
        for u in usuarios:
            celular = _normalizar_celular_co(u.get("celular"))
            if not celular:
                logger.info(f"[NOTIF OC] {perfil} {u.get('usuario')} sin celular válido; se saltea ({consecutivo}).")
                continue
            nombre = (u.get("nombre") or u.get("usuario") or perfil.title()).strip()
            res = enviar_template_sync(
                to=celular,
                template_name=plantilla[0],
                language_code=plantilla[1],
                body_params=body_params_fn(nombre),
            )
            if res:
                enviadas += 1
                logger.info(f"[NOTIF OC] WhatsApp OK -> {celular} ({nombre}) | {plantilla[0]} {consecutivo}")
            else:
                logger.warning(f"[NOTIF OC] WhatsApp NO enviado a {celular} ({nombre}) — revisar plantilla '{plantilla[0]}' ({consecutivo}).")
        logger.info(f"[NOTIF OC] {plantilla[0]} {consecutivo}: {enviadas}/{len(usuarios)} {perfil} notificados.")
    except Exception as e:
        logger.error(f"[NOTIF OC] Error en _enviar_a_perfil({perfil}, {plantilla[0]}): {e}")


def _enviar_a_creador(doc: dict, plantilla: tuple, body_params_fn) -> None:
    """Envía la plantilla al OPERATIVO que creó la solicitud (usuario_registro)."""
    try:
        resuelto = _resolver_celular_nombre(doc.get("usuario_registro", ""))
        if not resuelto:
            logger.info(f"[NOTIF OC] Creador sin celular válido; no se notifica ({doc.get('consecutivo', '')}).")
            return
        celular, nombre = resuelto
        res = enviar_template_sync(
            to=celular,
            template_name=plantilla[0],
            language_code=plantilla[1],
            body_params=body_params_fn(nombre),
        )
        if res:
            logger.info(f"[NOTIF OC] WhatsApp OK -> {celular} ({nombre}) | {plantilla[0]} {doc.get('consecutivo','')}")
        else:
            logger.warning(f"[NOTIF OC] WhatsApp NO enviado a {celular} ({nombre}) — revisar plantilla '{plantilla[0]}' ({doc.get('consecutivo','')}).")
    except Exception as e:
        logger.error(f"[NOTIF OC] Error en _enviar_a_creador({plantilla[0]}): {e}")


def _notificar_envio_aprobacion(doc: dict) -> None:
    """→ pendiente_aprobacion: avisa a COORDINADOR/CONTROL. Si el valor supera el
    límite del coordinador, el trámite es de Control; si no, lo ven ambos."""
    consec = doc.get("consecutivo", "")
    valor = _valor_cop(doc.get("valor_total", 0))
    if _a_numero(doc.get("valor_total")) > LIMITE_COORDINADOR:
        _enviar_a_perfil("CONTROL", PLANTILLA_OC_APROBACION, lambda n: [n, consec, valor], consecutivo=consec)
    else:
        _enviar_a_perfil("COORDINADOR", PLANTILLA_OC_APROBACION, lambda n: [n, consec, valor], consecutivo=consec)
        _enviar_a_perfil("CONTROL", PLANTILLA_OC_APROBACION, lambda n: [n, consec, valor], consecutivo=consec)


def _notificar_aprobacion(doc: dict) -> None:
    """→ aprobado: avisa a ANALISTA que debe tramitar en Vulcano."""
    consec = doc.get("consecutivo", "")
    valor = _valor_cop(doc.get("valor_total", 0))
    _enviar_a_perfil("ANALISTA", PLANTILLA_OC_TRAMITE, lambda n: [n, consec, valor], consecutivo=consec)


def _notificar_tramite_ok(doc: dict) -> None:
    """tramite_vulcano → ok: avisa a FINANCIERO que está listo para pagar."""
    consec = doc.get("consecutivo", "")
    valor = _valor_cop(doc.get("valor_total", 0))
    _enviar_a_perfil("FINANCIERO", PLANTILLA_OC_PAGO, lambda n: [n, consec, valor], consecutivo=consec)


def _notificar_pago(doc: dict) -> None:
    """→ pagado: avisa al creador (OPERATIVO) que se pagó su solicitud."""
    consec = doc.get("consecutivo", "")
    valor = _valor_cop(doc.get("valor_total", 0))
    _enviar_a_creador(doc, PLANTILLA_OC_PAGADA, lambda n: [n, consec, valor])


def _notificar_devolucion(doc: dict, motivo: str) -> None:
    """→ devuelto: avisa al creador con el motivo para que corrija."""
    consec = doc.get("consecutivo", "")
    motivo_texto = (motivo or "")[:200]
    _enviar_a_creador(doc, PLANTILLA_OC_DEVUELTA, lambda n: [n, consec, motivo_texto])


def _notificar_rechazo_anulacion(doc: dict, accion: str, motivo: str) -> None:
    """→ rechazado/anulado: avisa al creador. `accion` = 'rechazada' | 'anulada'."""
    consec = doc.get("consecutivo", "")
    motivo_texto = (motivo or "")[:200]
    _enviar_a_creador(doc, PLANTILLA_OC_RECH_ANUL, lambda n: [n, consec, accion, motivo_texto])


# ── Serialización / visibilidad de datos sensibles ────────────────────────────
def _fecha_a_str(v) -> Optional[str]:
    if isinstance(v, datetime):
        return v.isoformat()
    return v if v is not None else None


def _a_numero(v) -> float:
    try:
        if v is None or v == "":
            return 0.0
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _jsonable(v):
    if isinstance(v, ObjectId):
        return str(v)
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, list):
        return [_jsonable(x) for x in v]
    if isinstance(v, dict):
        return {k: _jsonable(x) for k, x in v.items()}
    return v


def _aplicar_visibilidad(doc: dict, perfil: str, usuario: Optional[str] = None) -> dict:
    """Enmascara datos bancarios sensibles salvo para ADMIN/FINANCIERO o el dueño
    de la solicitud (que necesita verlos para editarlos).

    El ANALISTA, que tramita en Vulcano las solicitudes aprobadas, ve la **cédula
    del titular** en claro (la necesita para el cruce), pero el **número de cuenta**
    sigue enmascarado para él."""
    if perfil in {"FINANCIERO", "ADMIN"}:
        return doc
    if usuario and doc.get("usuario_registro") == usuario:
        return doc  # el dueño ve sus propios datos bancarios
    db_ = doc.get("datos_bancarios")
    if isinstance(db_, dict):
        db_ = dict(db_)
        db_["numero_cuenta"] = _enmascarar(db_.get("numero_cuenta", ""))
        if perfil != "ANALISTA":
            db_["cedula_titular"] = _enmascarar(db_.get("cedula_titular", ""))
        doc["datos_bancarios"] = db_
    return doc


def _enmascarar(valor: str) -> str:
    s = str(valor or "")
    if len(s) <= 4:
        return "*" * len(s)
    return "*" * (len(s) - 4) + s[-4:]


def _serializar(doc: Optional[dict], perfil: str, usuario: Optional[str] = None) -> Optional[dict]:
    if doc is None:
        return None
    return _aplicar_visibilidad(_jsonable(doc), perfil, usuario)


# ── Validaciones (spec §8) ────────────────────────────────────────────────────
def _validar_solicitud(
    datos_servicio: "DatosServicio",
    costos: List["CostoConcepto"],
    datos_bancarios: "DatosBancarios",
    conductor: "Conductor",
    pedido_encontrado: bool,
) -> float:
    """Valida los campos obligatorios. Devuelve valor_total calculado."""
    if not datos_servicio.manifiesto or not str(datos_servicio.manifiesto).strip():
        raise HTTPException(status_code=422, detail="El manifiesto es obligatorio.")
    if not pedido_encontrado and not (datos_servicio.placa or "").strip():
        raise HTTPException(
            status_code=422,
            detail="La placa es obligatoria cuando el pedido no se encuentra en el histórico.",
        )

    if not costos:
        raise HTTPException(status_code=422, detail="Debe agregar al menos un concepto de costo.")
    valor_total = 0.0
    for c in costos:
        if not (c.tipo_costo or "").strip():
            raise HTTPException(status_code=422, detail="Cada concepto debe tener un tipo de costo.")
        if not (c.descripcion or "").strip():
            raise HTTPException(status_code=422, detail="La descripción del costo es obligatoria.")
        if _a_numero(c.valor) <= 0:
            raise HTTPException(status_code=422, detail="El valor de cada costo debe ser mayor que cero.")
        valor_total += _a_numero(c.valor)

    if valor_total <= 0:
        raise HTTPException(status_code=422, detail="El valor solicitado debe ser mayor que cero.")
    if valor_total > LIMITE_VALOR_SOLICITUD:
        raise HTTPException(
            status_code=422,
            detail=f"El valor total (${valor_total:,.0f}) supera el máximo permitido "
                   f"(${LIMITE_VALOR_SOLICITUD:,.0f}).",
        )

    # Bancarios
    if not (datos_bancarios.banco or "").strip():
        raise HTTPException(status_code=422, detail="El banco es obligatorio.")
    doc_banco = col_bancos.find_one({"nombre": {"$regex": f"^{re.escape(datos_bancarios.banco.strip())}$", "$options": "i"}})
    if not doc_banco:
        raise HTTPException(status_code=422, detail="El banco seleccionado no es válido.")
    # El código se resuelve desde el catálogo (fuente de verdad), no se confía del frontend.
    datos_bancarios.codigo_banco = str(doc_banco.get("codigo", ""))
    if not (datos_bancarios.numero_cuenta or "").strip():
        raise HTTPException(status_code=422, detail="El número de cuenta es obligatorio.")
    ced = re.sub(r"\D", "", datos_bancarios.cedula_titular or "")
    if not ced:
        raise HTTPException(status_code=422, detail="La cédula del titular es obligatoria.")
    if (datos_bancarios.tipo_id_titular or "").strip() not in TIPOS_ID_TITULAR:
        raise HTTPException(status_code=422, detail="El tipo de identificación del titular debe ser CC o NIT.")
    if not (datos_bancarios.nombre_titular or "").strip():
        raise HTTPException(status_code=422, detail="El nombre del titular de la cuenta es obligatorio.")
    if not (conductor.nombre or "").strip():
        raise HTTPException(status_code=422, detail="El nombre del conductor es obligatorio.")

    return round(valor_total, 2)


def _verificar_duplicado_interno(
    pedidos_norm: List[str],
    manifiesto: str,
    costos: List["CostoConcepto"],
    excluir_consecutivo: Optional[str] = None,
) -> List[dict]:
    """Advertencia (no bloqueante) de solicitudes que coincidan en pedido,
    manifiesto, tipo de costo y valor en los últimos 30 días."""
    if not pedidos_norm or not manifiesto:
        return []
    desde = _ahora_utc() - timedelta(days=30)
    tipos = {(c.tipo_costo or "").strip().upper() for c in costos}
    coincidencias: List[dict] = []
    set_norm = set(pedidos_norm)
    for col in (col_activos, col_historico):
        for d in col.find(
            {
                "manifiesto": str(manifiesto).strip(),
                "created_at": {"$gte": desde},
            }
        ):
            if excluir_consecutivo and d.get("consecutivo") == excluir_consecutivo:
                continue
            if not (set_norm & set(d.get("pedidos_normalizados") or [])):
                continue
            doc_tipos = {(c.get("tipo_costo") or "").strip().upper() for c in (d.get("costos") or [])}
            if tipos & doc_tipos:
                coincidencias.append({
                    "consecutivo": d.get("consecutivo"),
                    "estado": d.get("estado"),
                    "created_at": _fecha_a_str(d.get("created_at")),
                })
    return coincidencias


def _doc_por_consecutivo(consecutivo: str) -> tuple[Optional[dict], Optional[object]]:
    """Busca un documento por consecutivo en activos (y devuelve la colección).
    Retorna (doc, coleccion). Si no existe en activos, busca en histórico/anulados."""
    doc = col_activos.find_one({"consecutivo": consecutivo})
    if doc:
        return doc, col_activos
    for col in (col_historico, col_anulados):
        d = col.find_one({"consecutivo": consecutivo})
        if d:
            return d, col
    return None, None


def _scope_lectura(info: dict, historico: bool = False) -> dict:
    """
    Devuelve un sub-filtro Mongo que representa la BANDEJA del perfil: qué
    solicitudes puede ver/listar según el estado actual (el estado ES la etapa
    del flujo). _construir_filtro_listado compone este sub-filtro con los
    filtros del usuario vía $and, así cada perfil solo ve lo que está en 'su'
    paso del proceso.

    Modo histórico (solo lectura/auditoría): sin restricción de bandeja, salvo
    el OPERATIVO que solo ve sus propias solicitudes.
    """
    perfil = info["perfil"]
    usuario = info["usuario"]

    if historico:
        # Histórico = solo lectura/auditoría. Sin restricción de bandeja, salvo
        # el OPERATIVO/DESPACHADOR que ve las de su regional (comportamiento previo).
        if perfil in ("OPERATIVO", "DESPACHADOR"):
            base_hist: dict = {}
            if _normalizar_regional(info.get("regional", "")):
                _aplicar_filtro_regional(base_hist, info["regional"])
            else:
                base_hist = {"usuario_registro": usuario}
            return base_hist
        return {}

    # Modo activos: la bandeja se deriva del estado.
    if perfil == "ADMIN":
        return {}
    if perfil == "FINANCIERO":
        return {"estado": "aprobado", "tramite_vulcano": "ok"}   # listo para pagar
    if perfil == "ANALISTA":
        return {"estado": "aprobado"}                            # bandeja: trámite
    if perfil in ("CONTROL", "COORDINADOR"):
        return {"estado": "pendiente_aprobacion"}               # bandeja: aprobación
    if perfil in ("OPERATIVO", "DESPACHADOR"):
        # Bandeja (puede actuar): borrador/devuelto de su regional (o propias si
        # no tiene regional definida) + seguimiento en SOLO LECTURA de sus
        # propias solicitudes ya enviadas (otros estados activos).
        # DESPACHADOR tiene el mismo tratamiento que OPERATIVO (mismo rol operativo).
        base_regional: dict = {}
        if _normalizar_regional(info.get("regional", "")):
            _aplicar_filtro_regional(base_regional, info["regional"])  # deja {"$or": [...]}
        else:
            logger.warning(
                "[OTROS_COSTOS] %s %s sin regional definida; fallback a solicitudes propias.",
                perfil, usuario,
            )
            base_regional = {"usuario_registro": usuario}

        bandeja = {"$and": [base_regional, {"estado": {"$in": ["borrador", "devuelto"]}}]}
        seguimiento = [
            {"usuario_registro": usuario, "estado": e}
            for e in ("pendiente_aprobacion", "aprobado", "rechazado")
        ]
        return {"$or": [bandeja, *seguimiento]}

    # Perfil no contemplado: sin restricción (defensivo).
    return {}


# ════════════════════════════════════════════════════════════════════════════
# Modelos Pydantic
# ════════════════════════════════════════════════════════════════════════════
def _norm_upper(v):
    """Normaliza texto a MAYÚSCULAS sin espacios en los extremos (para Mongo)."""
    if v is None:
        return v
    return str(v).strip().upper()


def _norm_digitos(v):
    """Deja solo dígitos (para número de cuenta, cédula y teléfono)."""
    if v is None:
        return v
    return re.sub(r"\D", "", str(v))


class CostoConcepto(BaseModel):
    tipo_costo: str = ""
    descripcion: str = ""
    valor: float = 0

    @field_validator("descripcion", mode="before")
    @classmethod
    def _upper(cls, v):
        return _norm_upper(v)


class DatosServicio(BaseModel):
    cliente: str = ""
    centro_distribucion: str = ""
    fecha_servicio: Optional[str] = None
    piezas: float = 0
    peso_real: float = 0
    tipo_vehiculo: str = ""
    placa: str = ""
    municipio_destino: str = ""
    departamento_destino: str = ""
    transportador: str = ""
    manifiesto: str = ""

    @field_validator("placa", "municipio_destino", "departamento_destino", "transportador", "manifiesto", mode="before")
    @classmethod
    def _upper(cls, v):
        return _norm_upper(v)


class DatosBancarios(BaseModel):
    banco: str = ""
    codigo_banco: str = ""
    tipo_cuenta: str = ""
    numero_cuenta: str = ""
    tipo_id_titular: str = ""
    cedula_titular: str = ""
    nombre_titular: str = ""

    @field_validator("nombre_titular", "tipo_id_titular", mode="before")
    @classmethod
    def _upper(cls, v):
        return _norm_upper(v)

    @field_validator("numero_cuenta", "cedula_titular", mode="before")
    @classmethod
    def _digitos(cls, v):
        return _norm_digitos(v)


class Conductor(BaseModel):
    nombre: str = ""
    telefono: str = ""

    @field_validator("nombre", mode="before")
    @classmethod
    def _upper(cls, v):
        return _norm_upper(v)

    @field_validator("telefono", mode="before")
    @classmethod
    def _digitos(cls, v):
        return _norm_digitos(v)


class BuscarPedidosRequest(BaseModel):
    usuario: str
    pedido_vulcano: str


class CrearOtroCostoRequest(BaseModel):
    usuario: str
    enviar: bool = False
    pedido_vulcano_original: str
    pedidos_normalizados: List[str] = Field(default_factory=list)
    pedido_encontrado: bool = True
    motivo_no_encontrado: str = ""
    datos_servicio: DatosServicio
    costos: List[CostoConcepto]
    datos_bancarios: DatosBancarios
    conductor: Conductor = Field(default_factory=Conductor)


class EditarOtroCostoRequest(BaseModel):
    consecutivo: str
    usuario: str
    enviar: bool = False
    pedido_vulcano_original: Optional[str] = None
    pedidos_normalizados: Optional[List[str]] = None
    pedido_encontrado: Optional[bool] = None
    motivo_no_encontrado: Optional[str] = None
    datos_servicio: Optional[DatosServicio] = None
    costos: Optional[List[CostoConcepto]] = None
    datos_bancarios: Optional[DatosBancarios] = None
    conductor: Optional[Conductor] = None


class AccionConObservacionRequest(BaseModel):
    consecutivo: str
    usuario: str
    observacion: str = ""


class MarcarTramiteVulcanoRequest(BaseModel):
    consecutivo: str
    usuario: str
    tramite_vulcano: Literal["ok", "pendiente"]
    observacion: str = ""


class RegistrarPagoRequest(BaseModel):
    consecutivo: str
    usuario: str
    estado_pago: str = "PAGADO"
    fecha_pago: Optional[str] = None
    referencia: str = ""
    observaciones: str = ""


class AnularRequest(BaseModel):
    consecutivo: str
    usuario: str
    motivo: str = ""


class VerificarDuplicadoRequest(BaseModel):
    usuario: str
    pedido_vulcano: str
    manifiesto: str = ""
    tipo_costo: str = ""
    valor: float = 0


class ExportarExcelRequest(BaseModel):
    usuario: str
    estado: Optional[str] = None
    fecha_inicio: Optional[str] = None
    fecha_fin: Optional[str] = None
    pedido: Optional[str] = None
    placa: Optional[str] = None
    manifiesto: Optional[str] = None
    cliente: Optional[str] = None
    regional: Optional[str] = None
    origen: str = "historico"   # "historico" | "activos"


# ════════════════════════════════════════════════════════════════════════════
# Endpoints
# ════════════════════════════════════════════════════════════════════════════
@router.get("/tipos-costo")
async def tipos_costo():
    """Causales/tipos de costo para el formulario de Otros Costos.
    Auto-siembra la colección `causales_otros_costos` con el listado por defecto
    la primera vez (si está vacía); desde entonces es editable directamente en Mongo."""
    if col_causales.count_documents({}) == 0:
        col_causales.insert_many([{"nombre": c} for c in CAUSALES_OTROS_COSTOS_DEFAULT])
        logger.info(
            "[CAUSALES OTROS COSTOS] Colección sembrada con %d causales por defecto.",
            len(CAUSALES_OTROS_COSTOS_DEFAULT),
        )
    docs = list(col_causales.find({}, {"_id": 0, "nombre": 1}))
    return [d.get("nombre", "") for d in docs if d.get("nombre")]


@router.get("/bancos")
async def bancos():
    """Bancos para el formulario de Otros Costos (nombre + código bancario).
    Auto-siembra la colección `bancos_otros_costos` con el listado por defecto
    la primera vez (si está vacía); desde entonces es editable directamente en
    Mongo (documentos `{ "nombre": "...", "codigo": "..." }`)."""
    if col_bancos.count_documents({}) == 0:
        col_bancos.insert_many([dict(b) for b in BANCOS_OTROS_COSTOS_DEFAULT])
        logger.info(
            "[BANCOS OTROS COSTOS] Colección sembrada con %d bancos por defecto.",
            len(BANCOS_OTROS_COSTOS_DEFAULT),
        )
    docs = list(col_bancos.find({}, {"_id": 0, "nombre": 1, "codigo": 1}))
    return [
        {"nombre": d.get("nombre", ""), "codigo": str(d.get("codigo", ""))}
        for d in docs if d.get("nombre")
    ]


@router.get("/tipos-cuenta")
async def tipos_cuenta():
    return TIPOS_CUENTA


@router.get("/clientes")
async def clientes():
    """Clientes sugeridos para el campo Cliente del formulario de Otros Costos.
    Auto-siembra la colección `clientes_otros_costos` con el listado por defecto
    la primera vez (si está vacía); desde entonces es editable directamente en Mongo."""
    if col_clientes.count_documents({}) == 0:
        col_clientes.insert_many([{"nombre": c} for c in CLIENTES_OTROS_COSTOS_DEFAULT])
        logger.info(
            "[CLIENTES OTROS COSTOS] Colección sembrada con %d clientes por defecto.",
            len(CLIENTES_OTROS_COSTOS_DEFAULT),
        )
    docs = list(col_clientes.find({}, {"_id": 0, "nombre": 1}))
    return [d.get("nombre", "") for d in docs if d.get("nombre")]


@router.post("/buscar-pedidos")
async def buscar_pedidos(req: BuscarPedidosRequest):
    _resolver_usuario(req.usuario)  # valida sesión (cualquier perfil puede buscar)
    pedidos_norm = _normalizar_pedidos(req.pedido_vulcano)
    encontrados = _buscar_pedidos_historico(pedidos_norm)
    encontrados_norm = set()
    for e in encontrados:
        encontrados_norm |= set(_normalizar_pedidos(e.get("pedido_vulcano", "")))
    no_encontrados = [p for p in pedidos_norm if p not in encontrados_norm]

    totales = {
        "piezas": sum(_a_numero(e.get("piezas")) for e in encontrados),
        "peso_real": sum(_a_numero(e.get("peso_real")) for e in encontrados),
        "total_solicitado": sum(_a_numero(e.get("total_solicitado")) for e in encontrados),
    }

    # Diferencias entre los pedidos encontrados (cliente, destino, placa, vehículo)
    diferencias = {}
    advertencia = False
    if len(encontrados) > 1:
        for campo in ("cliente", "municipio_destino", "departamento_destino", "placa", "tipo_vehiculo"):
            valores = {str(e.get(campo, "")).strip() for e in encontrados if e.get(campo)}
            valores.discard("")
            if len(valores) > 1:
                diferencias[campo] = sorted(valores)
                advertencia = True

    return {
        "pedido_vulcano_original": req.pedido_vulcano,
        "pedidos_normalizados": pedidos_norm,
        "pedidos_encontrados": encontrados,
        "pedidos_no_encontrados": no_encontrados,
        "pedido_encontrado": len(encontrados) > 0,
        "totales": totales,
        "diferencias": diferencias,
        "advertencia_servicios_diferentes": advertencia,
    }


@router.post("/verificar-duplicado")
async def verificar_duplicado(req: VerificarDuplicadoRequest):
    _resolver_usuario(req.usuario)
    pedidos_norm = _normalizar_pedidos(req.pedido_vulcano)
    costo = CostoConcepto(tipo_costo=req.tipo_costo, valor=req.valor)
    coincidencias = _verificar_duplicado_interno(pedidos_norm, req.manifiesto, [costo])
    return {"posible_duplicado": len(coincidencias) > 0, "coincidencias": coincidencias}


@router.post("/crear")
async def crear_solicitud(req: CrearOtroCostoRequest, request: Request):
    info = _resolver_usuario(req.usuario)
    _requiere(info, {"OPERATIVO", "DESPACHADOR", "ADMIN"}, "crear solicitudes")

    if not str(req.pedido_vulcano_original or "").strip():
        raise HTTPException(status_code=422, detail="El pedido de Vulcano es obligatorio.")

    pedidos_norm = req.pedidos_normalizados or _normalizar_pedidos(req.pedido_vulcano_original)
    valor_total = _validar_solicitud(req.datos_servicio, req.costos, req.datos_bancarios, req.conductor, req.pedido_encontrado)
    requiere_control = valor_total > LIMITE_COORDINADOR
    ahora = _ahora_utc()
    estado = "pendiente_aprobacion" if req.enviar else "borrador"

    doc = {
        "pedido_vulcano_original": str(req.pedido_vulcano_original).strip(),
        "pedidos_normalizados": pedidos_norm,
        "pedido_encontrado": bool(req.pedido_encontrado),
        "motivo_no_encontrado": _norm_upper(req.motivo_no_encontrado) if req.motivo_no_encontrado else "",
        "datos_servicio": req.datos_servicio.model_dump(),
        "costos": [c.model_dump() for c in req.costos],
        "valor_total": valor_total,
        "requiere_aprobacion_control": requiere_control,
        "datos_bancarios": req.datos_bancarios.model_dump(),
        "conductor": req.conductor.model_dump(),
        "manifiesto": (req.datos_servicio.manifiesto or "").strip(),
        "estado": estado,
        "usuario_registro": info["usuario"],
        "perfil_registro": info["perfil"],
        "regional_registro": info["regional"],
        "creado_por": {
            "usuario": info["usuario"], "nombre": info["nombre"],
            "rol": info["perfil"], "fecha": ahora,
        },
        "aprobacion": {},
        "pago": {},
        "historial_movimientos": [
            _nuevo_movimiento("creacion", None, estado, info, "Solicitud creada", _ip(request))
        ],
        "created_at": ahora,
        "updated_at": ahora,
    }

    consecutivo = _insertar_con_reintento(doc)

    posible_dup = _verificar_duplicado_interno(pedidos_norm, doc["manifiesto"], req.costos, excluir_consecutivo=consecutivo)

    return {
        "mensaje": "Solicitud creada",
        "consecutivo": consecutivo,
        "estado": estado,
        "valor_total": valor_total,
        "requiere_aprobacion_control": requiere_control,
        "posible_duplicado": len(posible_dup) > 0,
        "duplicados": posible_dup,
        "solicitud": _serializar(doc, info["perfil"], info["usuario"]),
    }


@router.put("/editar")
async def editar_solicitud(req: EditarOtroCostoRequest, request: Request):
    info = _resolver_usuario(req.usuario)
    doc, col = _doc_por_consecutivo(req.consecutivo)
    if not doc or col is not col_activos:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada o ya no está activa.")

    estado_actual = doc.get("estado")
    if estado_actual in {"aprobado", "pagado", "anulado"}:
        raise HTTPException(
            status_code=422,
            detail="La solicitud está aprobada/pagada/anulada y no se puede editar.",
        )
    # Permiso: ADMIN, o OPERATIVO/DESPACHADOR dueño
    if info["perfil"] != "ADMIN":
        if info["perfil"] not in ("OPERATIVO", "DESPACHADOR") or doc.get("usuario_registro") != info["usuario"]:
            raise HTTPException(status_code=403, detail="Solo el creador o un administrador pueden editar.")
        # El operativo solo puede editar borrador o devuelto: una vez enviada a
        # aprobación (pendiente_aprobacion) la solicitud pasa al aprobador y ya
        # no le corresponde modificarla.
        if estado_actual not in {"borrador", "devuelto"}:
            raise HTTPException(
                status_code=422,
                detail="La solicitud ya fue enviada a aprobación y no se puede editar. "
                       "Debe ser devuelta para poder modificarla.",
            )

    datos_servicio = req.datos_servicio or DatosServicio(**doc.get("datos_servicio", {}))
    costos = req.costos if req.costos is not None else [CostoConcepto(**c) for c in doc.get("costos", [])]
    datos_bancarios = req.datos_bancarios or DatosBancarios(**doc.get("datos_bancarios", {}))
    conductor = req.conductor or Conductor(**doc.get("conductor", {}))

    valor_total = _validar_solicitud(datos_servicio, costos, datos_bancarios, conductor, doc.get("pedido_encontrado", True))
    requiere_control = valor_total > LIMITE_COORDINADOR

    set_fields = {
        "datos_servicio": datos_servicio.model_dump(),
        "costos": [c.model_dump() for c in costos],
        "valor_total": valor_total,
        "requiere_aprobacion_control": requiere_control,
        "datos_bancarios": datos_bancarios.model_dump(),
        "conductor": conductor.model_dump(),
        "manifiesto": (datos_servicio.manifiesto or "").strip(),
        "updated_at": _ahora_utc(),
    }
    if req.pedido_vulcano_original is not None:
        set_fields["pedido_vulcano_original"] = str(req.pedido_vulcano_original).strip()
        set_fields["pedidos_normalizados"] = req.pedidos_normalizados or _normalizar_pedidos(req.pedido_vulcano_original)
    if req.pedido_encontrado is not None:
        set_fields["pedido_encontrado"] = bool(req.pedido_encontrado)
    if req.motivo_no_encontrado is not None:
        set_fields["motivo_no_encontrado"] = _norm_upper(req.motivo_no_encontrado)

    # Si se solicita enviar a aprobación, transicionar el estado (igual que
    # /enviar-aprobacion). Solo borrador/devuelto pueden enviarse; la validación
    # de completitud ya se hizo arriba con _validar_solicitud().
    nuevo_estado = estado_actual
    if req.enviar:
        if estado_actual not in {"borrador", "devuelto"}:
            raise HTTPException(
                status_code=422,
                detail="Solo se pueden enviar a aprobación solicitudes en borrador o devueltas.",
            )
        nuevo_estado = "pendiente_aprobacion"
        set_fields["estado"] = nuevo_estado
        tipo_mov, motivo = "envio_aprobacion", "Edición y envío a aprobación"
    else:
        tipo_mov, motivo = "edicion", "Edición de la solicitud"

    mov = _nuevo_movimiento(tipo_mov, estado_actual, nuevo_estado, info, motivo, _ip(request))
    col_activos.update_one(
        {"consecutivo": req.consecutivo, "estado": estado_actual},
        {"$set": set_fields, "$push": {"historial_movimientos": mov}},
    )
    actualizado = col_activos.find_one({"consecutivo": req.consecutivo})
    # Si la edición envió a aprobación, avisar a COORDINADOR/CONTROL (fire-and-forget).
    if req.enviar and actualizado:
        await asyncio.to_thread(_notificar_envio_aprobacion, actualizado)
    return {"mensaje": "Solicitud actualizada", "estado": nuevo_estado, "solicitud": _serializar(actualizado, info["perfil"], info["usuario"])}


@router.post("/enviar-aprobacion")
async def enviar_aprobacion(req: AccionConObservacionRequest, request: Request):
    info = _resolver_usuario(req.usuario)
    _requiere(info, {"OPERATIVO", "DESPACHADOR", "ADMIN"}, "enviar solicitudes a aprobación")
    doc = col_activos.find_one({"consecutivo": req.consecutivo})
    if not doc:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")
    if doc.get("estado") not in {"borrador", "devuelto"}:
        raise HTTPException(status_code=422, detail="Solo se pueden enviar solicitudes en borrador o devueltas.")

    # Validar que esté completa antes de enviar
    _validar_solicitud(
        DatosServicio(**doc.get("datos_servicio", {})),
        [CostoConcepto(**c) for c in doc.get("costos", [])],
        DatosBancarios(**doc.get("datos_bancarios", {})),
        Conductor(**doc.get("conductor", {})),
        doc.get("pedido_encontrado", True),
    )

    estado_prev = doc.get("estado")
    mov = _nuevo_movimiento("envio_aprobacion", estado_prev, "pendiente_aprobacion", info, req.observacion, _ip(request))
    col_activos.update_one(
        {"consecutivo": req.consecutivo, "estado": estado_prev},
        {"$set": {"estado": "pendiente_aprobacion", "updated_at": _ahora_utc()},
         "$push": {"historial_movimientos": mov}},
    )
    # Avisar a COORDINADOR/CONTROL que tienen una solicitud pendiente (fire-and-forget).
    doc_actualizado = col_activos.find_one({"consecutivo": req.consecutivo})
    if doc_actualizado:
        await asyncio.to_thread(_notificar_envio_aprobacion, doc_actualizado)
    return {"mensaje": "Solicitud enviada a aprobación", "estado": "pendiente_aprobacion"}


def _obtener_y_validar_aprobacion(req, info) -> dict:
    doc = col_activos.find_one({"consecutivo": req.consecutivo})
    if not doc:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")
    if doc.get("estado") == "aprobado":
        raise HTTPException(status_code=409, detail="La solicitud ya está aprobada.")
    if doc.get("estado") != "pendiente_aprobacion":
        raise HTTPException(status_code=422, detail="La solicitud no está pendiente de aprobación.")
    valor_total = _a_numero(doc.get("valor_total"))
    if valor_total > LIMITE_COORDINADOR and info["perfil"] not in {"CONTROL", "ADMIN"}:
        raise HTTPException(
            status_code=403,
            detail=f"El valor (${valor_total:,.0f}) supera el límite del coordinador "
                   f"(${LIMITE_COORDINADOR:,.0f}). Requiere aprobación de Control.",
        )
    if info["perfil"] not in {"COORDINADOR", "CONTROL", "ADMIN"}:
        raise HTTPException(status_code=403, detail="Su perfil no puede aprobar solicitudes.")
    return doc


@router.post("/aprobar")
async def aprobar_solicitud(req: AccionConObservacionRequest, request: Request):
    info = _resolver_usuario(req.usuario)
    doc = _obtener_y_validar_aprobacion(req, info)
    ahora = _ahora_utc()
    aprobacion = {
        "usuario": info["usuario"], "nombre": info["nombre"], "rol": info["perfil"],
        "fecha": ahora, "observacion": req.observacion or "",
    }
    mov = _nuevo_movimiento("aprobacion", doc.get("estado"), "aprobado", info, req.observacion, _ip(request))
    res = col_activos.update_one(
        {"consecutivo": req.consecutivo, "estado": "pendiente_aprobacion"},
        {"$set": {"estado": "aprobado", "aprobacion": aprobacion, "tramite_vulcano": "pendiente", "updated_at": ahora},
         "$push": {"historial_movimientos": mov}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=409, detail="La solicitud cambió de estado (acción simultánea).")
    # Avisar a ANALISTA que debe tramitar en Vulcano (fire-and-forget). `doc` trae
    # valor_total y consecutivo de antes del update, que no cambian al aprobar.
    await asyncio.to_thread(_notificar_aprobacion, doc)
    return {"mensaje": "Solicitud aprobada", "estado": "aprobado"}


@router.post("/marcar-tramite-vulcano")
async def marcar_tramite_vulcano(req: MarcarTramiteVulcanoRequest, request: Request):
    """El ANALISTA confirma que el costo fue tramitado en Vulcano.
    Permite avanzar (pendiente→ok) y revertir (ok→pendiente). Solo sobre aprobados."""
    info = _resolver_usuario(req.usuario)
    _requiere(info, {"ANALISTA", "ADMIN"}, "marcar el trámite Vulcano")

    doc = col_activos.find_one({"consecutivo": req.consecutivo})
    if not doc:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")
    if doc.get("estado") != "aprobado":
        raise HTTPException(status_code=422, detail="Solo se tramita en Vulcano una solicitud aprobada.")

    ahora = _ahora_utc()
    tramite_info = {
        "usuario": info["usuario"], "nombre": info["nombre"], "rol": info["perfil"],
        "fecha": ahora, "observacion": req.observacion or "",
    }
    mov = _nuevo_movimiento("tramite_vulcano", None, req.tramite_vulcano, info, req.observacion, _ip(request))
    # Guardar atómicamente SOLO si sigue aprobado (anti-carrera).
    res = col_activos.update_one(
        {"consecutivo": req.consecutivo, "estado": "aprobado"},
        {"$set": {"tramite_vulcano": req.tramite_vulcano, "tramite_vulcano_info": tramite_info, "updated_at": ahora},
         "$push": {"historial_movimientos": mov}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=409, detail="La solicitud cambió de estado (acción simultánea).")
    # Al marcar OK, avisar a FINANCIERO que está listo para pagar. Al revertir a
    # pendiente no se notifica (lo ve en la app).
    if req.tramite_vulcano == "ok":
        await asyncio.to_thread(_notificar_tramite_ok, doc)
    return {"mensaje": f"Trámite Vulcano marcado como {req.tramite_vulcano}", "tramite_vulcano": req.tramite_vulcano}


@router.post("/devolver")
async def devolver_solicitud(req: AccionConObservacionRequest, request: Request):
    info = _resolver_usuario(req.usuario)
    _requiere(info, {"COORDINADOR", "CONTROL", "ADMIN", "ANALISTA", "FINANCIERO"}, "devolver solicitudes")
    if not (req.observacion or "").strip():
        raise HTTPException(status_code=422, detail="El motivo de devolución es obligatorio.")
    doc = col_activos.find_one({"consecutivo": req.consecutivo})
    if not doc:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")
    estado_prev = doc.get("estado")
    if estado_prev not in {"pendiente_aprobacion", "aprobado"}:
        raise HTTPException(
            status_code=422,
            detail="Solo se pueden devolver solicitudes pendientes o aprobadas.",
        )
    # Permiso según el estado desde el que se devuelve:
    # - Desde 'pendiente_aprobacion': Coordinador/Control/Admin (los que aprueban).
    # - Desde 'aprobado': Analista/Financiero/Admin (los que tramitan/pagan).
    perfil = info["perfil"]
    if estado_prev == "aprobado":
        permitidos, msg = {"ANALISTA", "FINANCIERO", "ADMIN"}, \
            "Desde 'aprobado' solo Analista/Financiero/Admin pueden devolver."
    else:
        permitidos, msg = {"COORDINADOR", "CONTROL", "ADMIN"}, \
            "Desde 'pendiente' solo Coordinador/Control/Admin pueden devolver."
    if perfil not in permitidos:
        raise HTTPException(status_code=403, detail=msg)
    # El Analista solo puede devolver mientras el trámite Vulcano esté pendiente
    # (su bandeja). Una vez marcó OK, la solicitud pasó a Financiero y ya no le toca.
    if perfil == "ANALISTA" and estado_prev == "aprobado" and doc.get("tramite_vulcano") == "ok":
        raise HTTPException(
            status_code=403,
            detail="El trámite ya fue marcado OK y la solicitud pasó a Financiero; no se puede devolver desde aquí.",
        )

    mov = _nuevo_movimiento("devolucion", estado_prev, "devuelto", info, req.observacion, _ip(request))
    set_fields = {"estado": "devuelto", "updated_at": _ahora_utc()}
    # Al devolver desde 'aprobado', resetear el trámite Vulcano para que el
    # analista lo revise de nuevo al reaprobar (/aprobar también lo resetea).
    if estado_prev == "aprobado":
        set_fields["tramite_vulcano"] = "pendiente"
    res = col_activos.update_one(
        {"consecutivo": req.consecutivo, "estado": estado_prev},
        {"$set": set_fields, "$push": {"historial_movimientos": mov}},
    )
    # Anti-carrera: si otro proceso cambió el estado entre el find y el update.
    if res.matched_count == 0:
        raise HTTPException(
            status_code=409,
            detail="La solicitud cambió de estado (acción simultánea). Recargue e intente de nuevo.",
        )
    # Avisar al creador con el motivo para que corrija (fire-and-forget).
    await asyncio.to_thread(_notificar_devolucion, doc, req.observacion)
    return {"mensaje": "Solicitud devuelta", "estado": "devuelto"}


@router.post("/rechazar")
async def rechazar_solicitud(req: AccionConObservacionRequest, request: Request):
    info = _resolver_usuario(req.usuario)
    _requiere(info, {"COORDINADOR", "CONTROL", "ADMIN"}, "rechazar solicitudes")
    if not (req.observacion or "").strip():
        raise HTTPException(status_code=422, detail="El motivo de rechazo es obligatorio.")
    doc = col_activos.find_one({"consecutivo": req.consecutivo})
    if not doc:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")
    if doc.get("estado") in {"rechazado", "pagado", "anulado"}:
        raise HTTPException(status_code=422, detail="La solicitud no se puede rechazar en su estado actual.")
    estado_prev = doc.get("estado")
    mov = _nuevo_movimiento("rechazo", estado_prev, "rechazado", info, req.observacion, _ip(request))
    col_activos.update_one(
        {"consecutivo": req.consecutivo, "estado": estado_prev},
        {"$set": {"estado": "rechazado", "updated_at": _ahora_utc()},
         "$push": {"historial_movimientos": mov}},
    )
    # Avisar al creador del rechazo con el motivo (fire-and-forget).
    await asyncio.to_thread(_notificar_rechazo_anulacion, doc, "rechazada", req.observacion)
    return {"mensaje": "Solicitud rechazada", "estado": "rechazado"}


def _mover_documento(doc: dict, col_destino) -> None:
    """Mueve el doc de col_activos a col_destino (patrón delete-first + insert idempotente)."""
    col_destino.delete_one({"_id": doc["_id"]})
    col_destino.insert_one(doc)
    col_activos.delete_one({"_id": doc["_id"]})


@router.post("/registrar-pago")
async def registrar_pago(req: RegistrarPagoRequest, request: Request):
    info = _resolver_usuario(req.usuario)
    _requiere(info, {"FINANCIERO", "ADMIN"}, "registrar pagos")

    doc = col_activos.find_one({"consecutivo": req.consecutivo})
    if not doc:
        # ¿ya está pagada en el histórico?
        if col_historico.find_one({"consecutivo": req.consecutivo, "estado": "pagado"}):
            raise HTTPException(status_code=409, detail="La solicitud ya está pagada.")
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")
    if doc.get("estado") == "pagado":
        raise HTTPException(status_code=409, detail="La solicitud ya está pagada.")
    if doc.get("estado") != "aprobado":
        raise HTTPException(status_code=422, detail="Solo se puede pagar una solicitud aprobada.")
    if doc.get("tramite_vulcano") != "ok":
        raise HTTPException(status_code=422, detail="La solicitud aún no ha sido tramitada en Vulcano.")

    ahora = _ahora_utc()
    estado_prev = doc.get("estado")
    pago = {
        "usuario": info["usuario"], "nombre": info["nombre"], "rol": info["perfil"],
        "estado_pago": (req.estado_pago or "PAGADO").strip().upper(),
        "fecha_pago": ahora,
        "fecha_pago_ingresada": req.fecha_pago or None,
        "referencia": req.referencia or "",
        "observaciones": req.observaciones or "",
    }
    mov_pago = _nuevo_movimiento("registro_pago", estado_prev, "pagado", info, req.observaciones, _ip(request))
    # Guardar atómicamente el pago SOLO si sigue aprobada (anti-doble-pago).
    res = col_activos.update_one(
        {"consecutivo": req.consecutivo, "estado": "aprobado", "tramite_vulcano": "ok"},
        {"$set": {"estado": "pagado", "pago": pago, "updated_at": ahora},
         "$push": {"historial_movimientos": mov_pago}},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=409, detail="La solicitud cambió de estado (acción simultánea).")

    doc_final = col_activos.find_one({"consecutivo": req.consecutivo})
    # Avisar al creador que se pagó su solicitud (fire-and-forget), antes de moverla.
    if doc_final:
        await asyncio.to_thread(_notificar_pago, doc_final)
    # Registrar el paso al histórico y mover
    mov_hist = _nuevo_movimiento("paso_historico", "pagado", "pagado", info, "Paso al histórico", _ip(request))
    col_activos.update_one({"consecutivo": req.consecutivo}, {"$push": {"historial_movimientos": mov_hist}})
    doc_final = col_activos.find_one({"consecutivo": req.consecutivo})
    _mover_documento(doc_final, col_historico)
    return {"mensaje": "Pago registrado y solicitud movida al histórico", "estado": "pagado"}


@router.post("/anular")
async def anular_solicitud(req: AnularRequest, request: Request):
    info = _resolver_usuario(req.usuario)
    _requiere(info, {"ADMIN"}, "anular solicitudes")
    doc, col = _doc_por_consecutivo(req.consecutivo)
    if not doc:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")
    if doc.get("estado") == "anulado":
        raise HTTPException(status_code=409, detail="La solicitud ya está anulada.")

    estado_prev = doc.get("estado")
    ahora = _ahora_utc()
    mov = _nuevo_movimiento("anulacion", estado_prev, "anulado", info, req.motivo, _ip(request))
    if col is col_activos:
        col_activos.update_one(
            {"consecutivo": req.consecutivo},
            {"$set": {"estado": "anulado", "motivo_anulacion": req.motivo or "",
                      "anulado_por": info["usuario"], "fecha_anulacion": ahora, "updated_at": ahora},
             "$push": {"historial_movimientos": mov}},
        )
        doc_final = col_activos.find_one({"consecutivo": req.consecutivo})
        _mover_documento(doc_final, col_anulados)
    else:
        # Ya estaba en histórico/anulados: sólo se marca (no se mueve de colección)
        col.update_one(
            {"consecutivo": req.consecutivo},
            {"$set": {"estado": "anulado", "motivo_anulacion": req.motivo or "",
                      "anulado_por": info["usuario"], "fecha_anulacion": ahora, "updated_at": ahora},
             "$push": {"historial_movimientos": mov}},
        )
    # Avisar al creador de la anulación con el motivo (fire-and-forget).
    await asyncio.to_thread(_notificar_rechazo_anulacion, doc, "anulada", req.motivo)
    return {"mensaje": "Solicitud anulada", "estado": "anulado"}


def _ip(request: Request) -> str:
    try:
        return request.client.host if request and request.client else ""
    except Exception:
        return ""


def _construir_filtro_listado(
    info: dict,
    estado: Optional[str],
    fecha_inicio: Optional[str],
    fecha_fin: Optional[str],
    pedido: Optional[str],
    placa: Optional[str],
    manifiesto: Optional[str],
    cliente: Optional[str],
    regional: Optional[str] = None,
    historico: bool = False,
) -> dict:
    # Filtros del usuario, acumulados APARTE para componerlos con el scope del
    # perfil vía $and. Así el $or de la bandeja de un perfil no se pisa con el
    # $or del filtro regional ni con un estado que elija el usuario.
    filtros: dict = {}
    # Descartar documentos huérfanos/incompletos (sin consecutivo) que pudieran existir.
    filtros["consecutivo"] = {"$exists": True}
    # Dropdown de regional: solo lo aplican los perfiles globales (evita que un
    # OPERATIVO/FINANCIERO bypassée su scope inyectando ?regional=...).
    if regional and info["perfil"] in PERFILES_GLOBALES_OC:
        _aplicar_filtro_regional(filtros, regional)

    if estado:
        filtros["estado"] = estado
    if pedido:
        pnorm = _normalizar_pedidos(pedido)
        if len(pnorm) == 1:
            filtros["pedidos_normalizados"] = pnorm[0]
        elif pnorm:
            filtros["pedidos_normalizados"] = {"$in": pnorm}
    if placa:
        filtros["datos_servicio.placa"] = str(placa).strip().upper()
    if manifiesto:
        filtros["manifiesto"] = str(manifiesto).strip()
    if cliente:
        filtros["datos_servicio.cliente"] = {"$regex": re.escape(str(cliente).strip()), "$options": "i"}
    if fecha_inicio or fecha_fin:
        rango = {}
        if fecha_inicio:
            rango["$gte"] = datetime.strptime(fecha_inicio, "%Y-%m-%d") + _OFFSET_COLOMBIA
        if fecha_fin:
            rango["$lt"] = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1) + _OFFSET_COLOMBIA
        filtros["created_at"] = rango

    scope = _scope_lectura(info, historico=historico)
    if not scope:
        return filtros                 # ADMIN (o histórico no operativo): sin $and
    return {"$and": [scope, filtros]}


@router.get("/")
async def listar_activos(
    request: Request,
    usuario: str = Query(...),
    estado: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    pedido: Optional[str] = Query(None),
    placa: Optional[str] = Query(None),
    manifiesto: Optional[str] = Query(None),
    cliente: Optional[str] = Query(None),
    regional: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
):
    info = _resolver_usuario(usuario)
    filtro = _construir_filtro_listado(info, estado, fecha_inicio, fecha_fin, pedido, placa, manifiesto, cliente, regional, historico=False)
    total = col_activos.count_documents(filtro)
    cursor = col_activos.find(filtro).sort("created_at", -1).skip(skip).limit(limit)
    items = [_serializar(d, info["perfil"], info["usuario"]) for d in cursor]
    return {"total": total, "skip": skip, "limit": limit, "items": items}


@router.get("/historico")
async def listar_historico(
    request: Request,
    usuario: str = Query(...),
    estado: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    pedido: Optional[str] = Query(None),
    placa: Optional[str] = Query(None),
    manifiesto: Optional[str] = Query(None),
    cliente: Optional[str] = Query(None),
    regional: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
):
    info = _resolver_usuario(usuario)
    # FINANCIERO ve todo el histórico (pagadas); OPERATIVO, las de su regional; el resto, todas.
    filtro = _construir_filtro_listado(info, estado, fecha_inicio, fecha_fin, pedido, placa, manifiesto, cliente, regional, historico=True)
    total = col_historico.count_documents(filtro)
    cursor = col_historico.find(filtro).sort("created_at", -1).skip(skip).limit(limit)
    items = [_serializar(d, info["perfil"], info["usuario"]) for d in cursor]
    return {"total": total, "skip": skip, "limit": limit, "items": items}


def _obtener_detalle(consecutivo: str, info: dict, col):
    doc = col.find_one({"consecutivo": consecutivo})
    if not doc:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")
    # Scope de detalle
    perfil = info["perfil"]
    if perfil in ("OPERATIVO", "DESPACHADOR"):
        # OPERATIVO/DESPACHADOR ve el detalle de cualquier solicitud de su
        # regional; si no tiene regional definida, solo las propias (mismo
        # fallback que el listado).
        if _normalizar_regional(info.get("regional", "")):
            if not _doc_coincide_regional(doc, info["regional"]):
                raise HTTPException(status_code=403, detail="No tiene acceso a esta solicitud.")
        elif doc.get("usuario_registro") != info["usuario"]:
            raise HTTPException(status_code=403, detail="No tiene acceso a esta solicitud.")
    if perfil == "FINANCIERO" and doc.get("estado") not in {"aprobado", "pagado"}:
        raise HTTPException(status_code=403, detail="No tiene acceso a esta solicitud.")
    return _serializar(doc, perfil, info.get("usuario"))


@router.get("/{consecutivo}")
async def obtener_detalle_activo(consecutivo: str, usuario: str = Query(...)):
    info = _resolver_usuario(usuario)
    return _obtener_detalle(consecutivo, info, col_activos)


@router.get("/historico/{consecutivo}")
async def obtener_detalle_historico(consecutivo: str, usuario: str = Query(...)):
    info = _resolver_usuario(usuario)
    return _obtener_detalle(consecutivo, info, col_historico)


@router.post("/exportar-excel")
async def exportar_excel(req: ExportarExcelRequest):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    info = _resolver_usuario(req.usuario)
    col = col_historico if req.origen != "activos" else col_activos
    filtro = _construir_filtro_listado(
        info, req.estado, req.fecha_inicio, req.fecha_fin, req.pedido, req.placa, req.manifiesto, req.cliente, req.regional,
        historico=(req.origen != "activos"),
    )

    docs = list(col.find(filtro).sort("created_at", -1).limit(5000))

    wb = Workbook()
    ws = wb.active
    ws.title = "Otros Costos"

    columnas = [
        "Consecutivo", "Pedido Vulcano", "Cliente", "Placa", "Manifiesto",
        "Tipo de Costo", "Valor Total", "Usuario Creación", "Usuario Aprobación",
        "Rol Aprobación", "Usuario Pago", "Estado Final", "Fecha Creación",
        "Fecha Aprobación", "Fecha Pago",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="004d40", end_color="004d40", fill_type="solid")
    thin = Border(*(Side(style="thin"),) * 4)
    for i, c in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")

    enmascarar = info["perfil"] not in {"FINANCIERO", "ADMIN"}
    for r, d in enumerate(docs, start=2):
        ds = d.get("datos_servicio", {}) or {}
        costos = d.get("costos", []) or []
        tipos = ", ".join((c.get("tipo_costo") or "") for c in costos)
        aprob = d.get("aprobacion", {}) or {}
        pago = d.get("pago", {}) or {}
        fila = [
            d.get("consecutivo", ""),
            d.get("pedido_vulcano_original", ""),
            ds.get("cliente", ""),
            ds.get("placa", ""),
            d.get("manifiesto", ""),
            tipos,
            d.get("valor_total", 0),
            (d.get("creado_por", {}) or {}).get("usuario", ""),
            aprob.get("usuario", ""),
            aprob.get("rol", ""),
            pago.get("usuario", ""),
            d.get("estado", ""),
            _fecha_a_str(d.get("created_at")),
            _fecha_a_str(aprob.get("fecha")),
            _fecha_a_str(pago.get("fecha_pago")),
        ]
        for i, v in enumerate(fila, 1):
            ws.cell(row=r, column=i, value=v).border = thin

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=otros_costos.xlsx"},
    )
