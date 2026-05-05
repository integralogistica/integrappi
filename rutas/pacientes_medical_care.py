from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import JSONResponse, StreamingResponse
import pandas as pd
import time
import json
from typing import List, Optional
from datetime import datetime, date, timedelta
from bd.bd_cliente import bd_cliente
from Funciones.normalizacion_medical_care import (
    fx_normalizar_paciente,
    fx_normalizar_direccion,
    fx_normalizar_celular,
    fx_separar_telefonos,
    fx_normalizar_municipio,
    fx_normalizar_cedula
)

router = APIRouter(prefix="/pacientes-medical-care", tags=["Medical Care"])

# Obtener base de datos y colección
bd = bd_cliente['integra']
coleccion = bd['pacientes_medical_care']
coleccion_cache = bd['cache_cruce_mc']
coleccion_cronograma = bd['cronograma_pacientes_mc']


def _get_cronograma_mes_actual() -> dict:
    """Retorna {cedula_normalizada: fecha_entrega} del mes actual (hora Colombia)."""
    import pytz
    from datetime import datetime as _dt
    anio_mes = _dt.now(pytz.timezone('America/Bogota')).strftime('%Y-%m')
    cursor = coleccion_cronograma.find(
        {'anio_mes': anio_mes},
        {'_id': 0, 'cedula': 1, 'fecha_entrega': 1}
    )
    return {doc['cedula']: doc.get('fecha_entrega', '') for doc in cursor}

# Mapeo código regional → nombre CEDI
_CEDI_MAPA = {
    'CO04': 'BARRANQUILLA', 'CO05': 'CALI', 'CO06': 'BUCARAMANGA',
    'CO07': 'FUNZA', 'CO09': 'MEDELLIN',
}

def _normalizar_cel(valor: str) -> str:
    """Devuelve solo dígitos de un número de celular, sin truncar."""
    return ''.join(filter(str.isdigit, valor or ''))


# ── Funciones para calcular días hábiles y estado del cruce ─────────────────────
def _obtener_festivos_colombia(anio: int) -> list:
    """Retorna lista de festivos de Colombia para un año dado (formato YYYY-MM-DD)."""
    from datetime import date, timedelta as _td
    from datetime import datetime as _dt

    festivos = []

    def _format_fecha(fecha):
        return fecha.strftime('%Y-%m-%d')

    def _mover_al_lunes(fecha):
        dia_sem = fecha.weekday()
        if dia_sem != 0:  # Si no es lunes (0)
            dias_hasta_lunes = (7 - dia_sem) % 7
            if dias_hasta_lunes == 0:
                dias_hasta_lunes = 7
            return fecha + _td(days=dias_hasta_lunes)
        return fecha

    # Festivos fijos
    festivos_fijos = [
        (1, 1),   # 1 de enero
        (1, 6),   # 6 de enero
        (5, 1),   # 1 de mayo
        (7, 20),  # 20 de julio
        (8, 7),   # 7 de agosto
        (12, 8),  # 8 de diciembre
        (12, 25), # 25 de diciembre
    ]
    for mes, dia in festivos_fijos:
        festivos.append(_format_fecha(date(anio, mes, dia)))

    # Festivos con Ley Emiliani (se mueven al lunes siguiente)
    festivos_emiliani = [
        (3, 19),  # San José
        (6, 29),  # San Pedro y San Pablo
        (8, 15),  # Asunción de la Virgen
        (10, 12), # Día de la Raza
        (11, 1),  # Todos los Santos
        (11, 11), # Independencia de Cartagena
    ]
    for mes, dia in festivos_emiliani:
        fecha = date(anio, mes, dia)
        festivos.append(_format_fecha(_mover_al_lunes(fecha)))

    # Calcular Semana Santa (algoritmo de Meeus/Jones/Butcher)
    a = anio % 19
    b = anio // 100
    c = anio % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes_pascua = (h + l - 7 * m + 114) // 31
    dia_pascua = ((h + l - 7 * m + 114) % 31) + 1
    pascua = date(anio, mes_pascua, dia_pascua)

    # Jueves Santo (3 días antes)
    jueves_santo = pascua - _td(days=3)
    festivos.append(_format_fecha(jueves_santo))

    # Viernes Santo (2 días antes)
    viernes_santo = pascua - _td(days=2)
    festivos.append(_format_fecha(viernes_santo))

    # Ascensión (39 días después, lunes siguiente = día 43)
    ascension = pascua + _td(days=43)
    festivos.append(_format_fecha(ascension))

    # Corpus Christi (60 días después, lunes siguiente = día 64)
    corpus_christi = pascua + _td(days=64)
    festivos.append(_format_fecha(corpus_christi))

    # Sagrado Corazón (68 días después, lunes siguiente = día 72)
    sagrado_corazon = pascua + _td(days=72)
    festivos.append(_format_fecha(sagrado_corazon))

    return sorted(festivos)


def _parsear_fecha_texto(fecha_str: str) -> Optional[datetime]:
    """Parsea fechas en formatos: YYYY-MM-DD, DD/MM/YYYY, DD MMM YYYY."""
    if not fecha_str:
        return None

    original = fecha_str
    fecha_str = fecha_str.strip()

    # Formato YYYY-MM-DD (primero, es el más confiable)
    if '/' not in fecha_str and '-' in fecha_str:
        try:
            return datetime.strptime(fecha_str[:10], '%Y-%m-%d')
        except ValueError:
            pass

    # Formato DD/MM/YYYY o DD/MM/YY (formato colombiano)
    if '/' in fecha_str:
        partes = fecha_str.split('/')
        if len(partes) == 3:
            p1, p2, p3 = partes[0].strip(), partes[1].strip(), partes[2].strip()

            # Validar que todos sean números
            try:
                v1, v2, v3 = int(p1), int(p2), int(p3)
            except ValueError:
                return None

            # Normalizar año a 4 dígitos
            if v3 < 100:
                v3 += 2000

            # Intentar DD/MM/YYYY primero (formato colombiano estándar)
            # v1 = día, v2 = mes, v3 = año
            if 1 <= v2 <= 12 and 1 <= v1 <= 31:  # mes y día válidos
                try:
                    return datetime(v3, v2, v1)
                except ValueError:
                    pass

            # Intentar MM/DD/YYYY (formato estadounidense) si DD/MM falló
            if 1 <= v1 <= 12 and 1 <= v2 <= 31:  # v1 como mes, v2 como día
                try:
                    return datetime(v3, v1, v2)
                except ValueError:
                    pass

            return None

    # Formato DD MMM YYYY (ej: "26 mar 2026")
    meses_map = {
        'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
    }
    partes = fecha_str.split()
    if len(partes) == 3:
        try:
            dia = int(partes[0])
            mes = meses_map.get(partes[1].lower())
            anio = int(partes[2])
            if mes:
                return datetime(anio, mes, dia)
        except (ValueError, IndexError):
            pass

    return None


def _calcular_dias_habiles(fecha_inicio_str: str, fecha_fin_str: str) -> int:
    """Calcula días hábiles entre dos fechas (excluye domingos y festivos de Colombia)."""
    import logging
    logger = logging.getLogger(__name__)

    if not fecha_inicio_str:
        return 0

    fecha_inicio = _parsear_fecha_texto(fecha_inicio_str)
    fecha_fin = _parsear_fecha_texto(fecha_fin_str)

    if not fecha_inicio or not fecha_fin:
        return 0

    # Validar que los años sean razonables
    if fecha_inicio.year < 2000 or fecha_inicio.year > 2100 or fecha_fin.year < 2000 or fecha_fin.year > 2100:
        logger.warning(f"[_calcular_dias_habiles] Años inválidos: inicio={fecha_inicio.year}, fin={fecha_fin.year}")
        # Si los años son inválidos, calcular días hábiles solo excluyendo domingos
        dias_habiles = 0
        fecha_actual = fecha_inicio
        while fecha_actual <= fecha_fin:
            if fecha_actual.weekday() != 6:  # No domingo
                dias_habiles += 1
            fecha_actual += timedelta(days=1)
        return dias_habiles

    # Normalizar a medianoche
    if isinstance(fecha_inicio, datetime):
        fecha_inicio = fecha_inicio.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(fecha_fin, datetime):
        fecha_fin = fecha_fin.replace(hour=0, minute=0, second=0, microsecond=0)

    # Obtener festivos para el rango de fechas
    anio_inicio = fecha_inicio.year
    anio_fin = fecha_fin.year
    festivos = []
    for anio in range(anio_inicio, anio_fin + 1):
        festivos.extend(_obtener_festivos_colombia(anio))

    dias_habiles = 0
    fecha_actual = fecha_inicio

    while fecha_actual <= fecha_fin:
        dia_semana = fecha_actual.weekday()  # 0 = lunes, 6 = domingo
        fecha_str = fecha_actual.strftime('%Y-%m-%d')

        # Excluir domingos (6) y festivos
        if dia_semana != 6 and fecha_str not in festivos:
            dias_habiles += 1

        fecha_actual += timedelta(days=1)

    return dias_habiles


def _determinar_estado_cruce(en_v3: bool, estado_pedido: str, f_pref_teorica: str, f_pedido: str = '') -> str:
    """Determina el estado según las reglas del cruce."""
    import logging
    import traceback
    logger = logging.getLogger(__name__)

    hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    # Normalizar estado_pedido para comparación robusta
    estado_pedido_norm = (estado_pedido or '').strip().upper()

    # Regla 1: Sin cruce
    if not en_v3:
        return 'sin cruce'

    # Regla 2: Retraso FMC (diferencia < 6 días hábiles entre F. Pedido y F. Pref. Integra)
    if en_v3 and f_pedido and f_pref_teorica:
        try:
            f_pedido_dt = _parsear_fecha_texto(f_pedido)
            if f_pedido_dt:
                dia_siguiente_pedido = (f_pedido_dt + timedelta(days=1)).strftime('%Y-%m-%d')
                dias_habiles_fmc = _calcular_dias_habiles(dia_siguiente_pedido, f_pref_teorica)
                if dias_habiles_fmc < 6:
                    return 'retraso FMC'
        except Exception as e:
            logger.warning(f"[_determinar_estado_cruce] Error FMC: {e}\n{traceback.format_exc()}")

    # Regla 3: Retraso operación (faltan 3 días hábiles o menos desde hoy)
    if en_v3 and estado_pedido_norm == 'POR PROGRAMAR' and f_pref_teorica:
        try:
            manana = (hoy + timedelta(days=1)).strftime('%Y-%m-%d')
            dias_habiles = _calcular_dias_habiles(manana, f_pref_teorica)
            if dias_habiles <= 3:
                return 'retraso operación'
        except Exception as e:
            logger.warning(f"[_determinar_estado_cruce] Error operación: {e}\n{traceback.format_exc()}")

    return '—'


# Columnas requeridas
COLUMNAS_REQUERIDAS = [
    'sede', 'paciente', 'cedula', 'direccion', 
    'departamento', 'municipio', 'ruta', 'cedi', 'celular'
]


@router.post("/cargar-masivo-stream")
async def cargar_pacientes_masivo_stream(
    archivo: UploadFile = File(...),
    usuario: str = Query(...)
):
    """
    Carga masiva de pacientes de Medical Care con progreso en tiempo real via SSE
    
    Args:
        archivo: Archivo Excel (.xlsx, .xls, .xlsm)
        usuario: Usuario que realiza la carga (parámetro de query string)
    
    Returns:
        StreamingResponse con eventos de progreso
    """
    import logging
    import asyncio
    import io
    
    # LEER EL CONTENIDO DEL ARCHIVO ANTES DE INICIAR EL STREAMING
    # Esto evita que el archivo se cierre antes de poder leerlo
    logger = logging.getLogger(__name__)
    logger.info(f"=== INICIO CARGA MASIVA STREAM ===")
    logger.info(f"Usuario: '{usuario}'")
    logger.info(f"Archivo: {archivo.filename if archivo else 'None'}")
    
    # Validar tipo de archivo
    if not archivo.filename.endswith(('.xlsx', '.xls', '.xlsm')):
        async def error_response():
            yield f"data: {json.dumps({'error': 'El archivo debe ser un Excel (.xlsx, .xls, .xlsm)'}, ensure_ascii=False)}\n\n"
        return StreamingResponse(error_response(), media_type="text/event-stream")
    
    logger.info("Leyendo contenido del archivo...")
    
    try:
        contenido = await archivo.read()
        logger.info(f"Archivo leído: {len(contenido)} bytes")
    except Exception as read_error:
        logger.error(f"Error al leer archivo: {str(read_error)}")
        async def error_response():
            yield f"data: {json.dumps({'error': f'Error al leer archivo: {str(read_error)}'}, ensure_ascii=False)}\n\n"
        return StreamingResponse(error_response(), media_type="text/event-stream")
    
    if not contenido or len(contenido) == 0:
        logger.error("El archivo está vacío")
        async def error_response():
            yield f"data: {json.dumps({'error': 'El archivo está vacío o no se pudo leer'}, ensure_ascii=False)}\n\n"
        return StreamingResponse(error_response(), media_type="text/event-stream")
    
    # Ahora procesamos el contenido del archivo dentro del generador
    async def generate_progress():
        """Generador de eventos SSE para el progreso"""
        tiempo_inicio = time.time()
        errores = []
        registros_exitosos = 0
        total_filas = 0
        
        try:
            # Leer archivo Excel desde el contenido en memoria
            buffer = io.BytesIO(contenido)
            
            logger.info("Intentando leer Excel...")
            
            # Determinar el tipo de archivo para pandas
            try:
                if archivo.filename.endswith('.xlsx'):
                    df = pd.read_excel(buffer, engine='openpyxl')
                elif archivo.filename.endswith('.xlsm'):
                    df = pd.read_excel(buffer, engine='openpyxl')
                else:
                    df = pd.read_excel(buffer, engine='xlrd')
                
                logger.info(f"Excel leído exitosamente: {len(df)} filas")
            except Exception as excel_error:
                logger.error(f"Error al leer Excel: {str(excel_error)}")
                yield f"data: {json.dumps({'error': f'Error al leer archivo Excel: {str(excel_error)}'}, ensure_ascii=False)}\n\n"
                return
            
            total_filas = len(df)
            yield f"data: {json.dumps({'stage': 'reading', 'progress': 0, 'message': 'Leyendo archivo Excel...'}, ensure_ascii=False)}\n\n"
            
            # Validar columnas requeridas
            columnas_archivo = [col.strip().lower() for col in df.columns]
            columnas_requeridas_lower = [col.lower() for col in COLUMNAS_REQUERIDAS]
            
            columnas_faltantes = set(columnas_requeridas_lower) - set(columnas_archivo)
            if columnas_faltantes:
                mensaje_error = f"Faltan las siguientes columnas: {', '.join(columnas_faltantes)}"
                yield f"data: {json.dumps({'error': mensaje_error}, ensure_ascii=False)}\n\n"
                return
            
            # Crear diccionario de mapeo de columnas (case-insensitive)
            mapeo_columnas = {}
            for col_req in COLUMNAS_REQUERIDAS:
                for col_archivo in df.columns:
                    if col_req.lower() == col_archivo.strip().lower():
                        mapeo_columnas[col_archivo] = col_req
                        break
            
            # Renombrar columnas a mayúsculas
            df = df.rename(columns=mapeo_columnas)
            
            yield f"data: {json.dumps({'stage': 'processing', 'progress': 0, 'total': total_filas, 'message': f'Procesando {total_filas} registros...'}, ensure_ascii=False)}\n\n"

            # Cargar todas las cédulas existentes en BD de una sola vez
            cedulas_en_bd = set(
                doc['cedula'] for doc in coleccion.find({}, {'cedula': 1, '_id': 0})
            )

            # Procesar cada fila con validación de duplicados
            documentos_a_insertar = []
            cedulas_ya_procesadas = set()

            for idx, fila in enumerate(df.to_dict('records')):
                try:
                    # Extraer valores originales con manejo de nulos
                    sede_original = str(fila.get('sede', '')).strip() if pd.notna(fila.get('sede')) else ''
                    paciente_original = str(fila.get('paciente', '')).strip() if pd.notna(fila.get('paciente')) else ''
                    cedula_original = str(fila.get('cedula', '')).strip() if pd.notna(fila.get('cedula')) else ''
                    direccion_original = str(fila.get('direccion', '')).strip() if pd.notna(fila.get('direccion')) else ''
                    departamento_original = str(fila.get('departamento', '')).strip() if pd.notna(fila.get('departamento')) else ''
                    municipio_original = str(fila.get('municipio', '')).strip() if pd.notna(fila.get('municipio')) else ''
                    ruta_original = str(fila.get('ruta', '')).strip() if pd.notna(fila.get('ruta')) else ''
                    cedi_original = str(fila.get('cedi', '')).strip() if pd.notna(fila.get('cedi')) else ''
                    celular_original = str(fila.get('celular', '')).strip() if pd.notna(fila.get('celular')) else ''
                    
                    # Validar campos obligatorios
                    if not cedula_original:
                        errores.append(f"Fila {idx + 2}: El campo 'cedula' es obligatorio")
                        continue

                    # Normalizar SOLO: paciente, dirección, cédula y celular
                    paciente_normalizado = fx_normalizar_paciente(paciente_original) if paciente_original else ''
                    cedula_normalizada = fx_normalizar_cedula(cedula_original)
                    direccion_normalizada = fx_normalizar_direccion(direccion_original)
                    telefono1, telefono2 = fx_separar_telefonos(celular_original)

                    # Validar resultados de normalización
                    if not cedula_normalizada:
                        errores.append(f"Fila {idx + 2}: Error al normalizar 'cedula'")
                        continue

                    # Validar duplicados por cédula en el archivo
                    if cedula_normalizada in cedulas_ya_procesadas:
                        errores.append(f"Fila {idx + 2}: La cédula {cedula_original} ya existe en el archivo cargado")
                        continue

                    # Validar duplicados en la base de datos (lookup O(1) contra el set precargado)
                    if cedula_normalizada in cedulas_en_bd:
                        errores.append(f"Fila {idx + 2}: La cédula {cedula_original} ya existe en la base de datos")
                        continue

                    direccion_final = direccion_normalizada if direccion_normalizada else direccion_original
                    llave = f"{paciente_normalizado} {direccion_final}".strip()

                    # Crear documento con campos originales y normalizados
                    documento = {
                        'sede': sede_original,
                        'paciente': paciente_normalizado,
                        'paciente_original': paciente_original,
                        'cedula': cedula_normalizada,
                        'cedula_original': cedula_original,
                        'direccion': direccion_final,
                        'direccion_original': direccion_original,
                        'departamento': departamento_original,
                        'municipio': municipio_original,
                        'ruta': ruta_original,
                        'cedi': cedi_original,
                        'celular': telefono1,
                        'celular_original': celular_original,
                        'telefono1': telefono1,
                        'telefono2': telefono2,
                        'llave': llave,
                        'estado': 'ACTIVO',
                        'usuario_carga': usuario,
                        'fecha_carga': time.strftime('%Y-%m-%d %H:%M:%S')
                    }

                    documentos_a_insertar.append(documento)
                    cedulas_ya_procesadas.add(cedula_normalizada)
                    registros_exitosos += 1

                    # Enviar progreso cada 10 filas o al final
                    if registros_exitosos % 10 == 0 or idx == total_filas - 1:
                        progreso = round((idx + 1) / total_filas * 100, 1)
                        yield f"data: {json.dumps({'stage': 'processing', 'progress': progreso, 'processed': registros_exitosos, 'total': total_filas, 'message': f'Procesando... {progreso}%'}, ensure_ascii=False)}\n\n"
                        await asyncio.sleep(0)  # Permitir que el event loop procese otros eventos
                
                except Exception as e:
                    errores.append(f"Fila {idx + 2}: Error al procesar - {str(e)}")
                    continue
            
            # Insertar documentos en MongoDB
            yield f"data: {json.dumps({'stage': 'saving', 'progress': 100, 'message': 'Guardando en base de datos...'}, ensure_ascii=False)}\n\n"
            
            if documentos_a_insertar:
                coleccion.insert_many(documentos_a_insertar)
            
            tiempo_fin = time.time()
            tiempo_segundos = round(tiempo_fin - tiempo_inicio, 2)
            
            # Enviar resultado final
            resultado = {
                'stage': 'complete',
                'progress': 100,
                'mensaje': f'Carga completada en {tiempo_segundos} segundos',
                'tiempo_segundos': tiempo_segundos,
                'registros_exitosos': registros_exitosos,
                'registros_con_errores': len(errores),
                'errores': errores[:50] if errores else []
            }
            yield f"data: {json.dumps(resultado, ensure_ascii=False)}\n\n"
            
        except Exception as e:
            tiempo_fin = time.time()
            tiempo_segundos = round(tiempo_fin - tiempo_inicio, 2)
            error_msg = {'error': f'Error al procesar el archivo: {str(e)}', 'tiempo_segundos': tiempo_segundos}
            yield f"data: {json.dumps(error_msg, ensure_ascii=False)}\n\n"
    
    return StreamingResponse(
        generate_progress(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


@router.post("/cargar-masivo")
async def cargar_pacientes_masivo(usuario: str, archivo: UploadFile = File(...)):
    """
    Carga masiva de pacientes de Medical Care desde un archivo Excel (versión sin streaming para compatibilidad)
    
    Args:
        usuario: Usuario que realiza la carga
        archivo: Archivo Excel (.xlsx, .xls, .xlsm)
    
    Returns:
        JSON con mensaje de éxito, tiempo de procesamiento y errores (si los hay)
    """
    import logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    
    tiempo_inicio = time.time()
    errores = []
    registros_exitosos = 0
    
    logger.info(f"=== INICIO CARGA MASIVA ===")
    logger.info(f"Usuario recibido: '{usuario}'")
    logger.info(f"Archivo recibido: {archivo.filename if archivo else 'None'}")
    logger.info(f"Tipo contenido: {archivo.content_type if archivo else 'None'}")
    
    # Validar tipo de archivo
    if not archivo.filename.endswith(('.xlsx', '.xls', '.xlsm')):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser un Excel (.xlsx, .xls, .xlsm)"
        )
    
    try:
        # Leer archivo Excel
        contenido = await archivo.read()
        
        # Determinar el tipo de archivo para pandas
        if archivo.filename.endswith('.xlsx'):
            df = pd.read_excel(contenido, engine='openpyxl')
        elif archivo.filename.endswith('.xlsm'):
            df = pd.read_excel(contenido, engine='openpyxl')
        else:
            df = pd.read_excel(contenido, engine='xlrd')
        
        # Validar columnas requeridas
        columnas_archivo = [col.strip().lower() for col in df.columns]
        columnas_requeridas_lower = [col.lower() for col in COLUMNAS_REQUERIDAS]
        
        columnas_faltantes = set(columnas_requeridas_lower) - set(columnas_archivo)
        if columnas_faltantes:
            raise HTTPException(
                status_code=400,
                detail=f"Faltan las siguientes columnas: {', '.join(columnas_faltantes)}"
            )
        
        # Crear diccionario de mapeo de columnas (case-insensitive)
        mapeo_columnas = {}
        for col_req in COLUMNAS_REQUERIDAS:
            for col_archivo in df.columns:
                if col_req.lower() == col_archivo.strip().lower():
                    mapeo_columnas[col_archivo] = col_req
                    break
        
        # Renombrar columnas a mayúsculas
        df = df.rename(columns=mapeo_columnas)
        
        # Cargar todas las cédulas existentes en BD de una sola vez
        cedulas_en_bd = set(
            doc['cedula'] for doc in coleccion.find({}, {'cedula': 1, '_id': 0})
        )

        # Procesar cada fila con validación de duplicados
        documentos_a_insertar = []
        cedulas_ya_procesadas = set()

        for idx, fila in enumerate(df.to_dict('records')):
            try:
                # Extraer valores originales con manejo de nulos
                sede_original = str(fila.get('sede', '')).strip() if pd.notna(fila.get('sede')) else ''
                paciente_original = str(fila.get('paciente', '')).strip() if pd.notna(fila.get('paciente')) else ''
                cedula_original = str(fila.get('cedula', '')).strip() if pd.notna(fila.get('cedula')) else ''
                direccion_original = str(fila.get('direccion', '')).strip() if pd.notna(fila.get('direccion')) else ''
                departamento_original = str(fila.get('departamento', '')).strip() if pd.notna(fila.get('departamento')) else ''
                municipio_original = str(fila.get('municipio', '')).strip() if pd.notna(fila.get('municipio')) else ''
                ruta_original = str(fila.get('ruta', '')).strip() if pd.notna(fila.get('ruta')) else ''
                cedi_original = str(fila.get('cedi', '')).strip() if pd.notna(fila.get('cedi')) else ''
                celular_original = str(fila.get('celular', '')).strip() if pd.notna(fila.get('celular')) else ''
                
                # Validar campos obligatorios
                if not cedula_original:
                    errores.append(f"Fila {idx + 2}: El campo 'cedula' es obligatorio")
                    continue

                # Normalizar SOLO: paciente, dirección, cédula y celular
                paciente_normalizado = fx_normalizar_paciente(paciente_original) if paciente_original else ''
                cedula_normalizada = fx_normalizar_cedula(cedula_original)
                direccion_normalizada = fx_normalizar_direccion(direccion_original)
                telefono1, telefono2 = fx_separar_telefonos(celular_original)

                # Validar resultados de normalización
                if not cedula_normalizada:
                    errores.append(f"Fila {idx + 2}: Error al normalizar 'cedula'")
                    continue

                # Validar duplicados por cédula en el archivo
                if cedula_normalizada in cedulas_ya_procesadas:
                    errores.append(f"Fila {idx + 2}: La cédula {cedula_original} ya existe en el archivo cargado")
                    continue

                # Validar duplicados en la base de datos (lookup O(1) contra el set precargado)
                if cedula_normalizada in cedulas_en_bd:
                    errores.append(f"Fila {idx + 2}: La cédula {cedula_original} ya existe en la base de datos")
                    continue

                direccion_final = direccion_normalizada if direccion_normalizada else direccion_original
                llave = f"{paciente_normalizado} {direccion_final}".strip()

                # Crear documento con campos originales y normalizados
                documento = {
                    'sede': sede_original,
                    'paciente': paciente_normalizado,
                    'paciente_original': paciente_original,
                    'cedula': cedula_normalizada,
                    'cedula_original': cedula_original,
                    'direccion': direccion_final,
                    'direccion_original': direccion_original,
                    'departamento': departamento_original,
                    'municipio': municipio_original,
                    'ruta': ruta_original,
                    'cedi': cedi_original,
                    'celular': telefono1,
                    'celular_original': celular_original,
                    'telefono1': telefono1,
                    'telefono2': telefono2,
                    'llave': llave,
                    'estado': 'ACTIVO',
                    'usuario_carga': usuario,
                    'fecha_carga': time.strftime('%Y-%m-%d %H:%M:%S')
                }

                documentos_a_insertar.append(documento)
                cedulas_ya_procesadas.add(cedula_normalizada)
                registros_exitosos += 1

            except Exception as e:
                errores.append(f"Fila {idx + 2}: Error al procesar - {str(e)}")
                continue
        
        # Insertar documentos en MongoDB
        if documentos_a_insertar:
            coleccion.insert_many(documentos_a_insertar)
        
        tiempo_fin = time.time()
        tiempo_segundos = round(tiempo_fin - tiempo_inicio, 2)
        
        # Construir respuesta
        if errores:
            return JSONResponse(
                status_code=207,  # Multi-Status
                content={
                    'mensaje': f'Carga completada con {len(errores)} errores',
                    'tiempo_segundos': tiempo_segundos,
                    'registros_exitosos': registros_exitosos,
                    'registros_con_errores': len(errores),
                    'errores': errores[:50]  # Limitar a primeros 50 errores
                }
            )
        else:
            return {
                'mensaje': f'Se importaron correctamente {registros_exitosos} registros',
                'tiempo_segundos': tiempo_segundos,
                'registros_exitosos': registros_exitosos,
                'registros_con_errores': 0
            }
    
    except HTTPException:
        raise
    except Exception as e:
        tiempo_fin = time.time()
        tiempo_segundos = round(tiempo_fin - tiempo_inicio, 2)
        raise HTTPException(
            status_code=500,
            detail={
                'mensaje': f'Error al procesar el archivo: {str(e)}',
                'tiempo_segundos': tiempo_segundos
            }
        )


@router.get("/")
async def obtener_pacientes(
    skip: int = 0,
    limit: int = 100,
    cedi: str = None
):
    """
    Obtiene la lista de pacientes de Medical Care con paginación.
    Si se pasa cedi, filtra solo los pacientes de ese CEDI (case-insensitive).
    """
    try:
        filtro = {}
        if cedi:
            filtro['cedi'] = {'$regex': f'^{cedi}$', '$options': 'i'}

        cursor = coleccion.find(filtro).skip(skip).limit(limit)
        pacientes = []
        for doc in cursor:
            doc['_id'] = str(doc['_id'])
            pacientes.append(doc)

        return {
            'pacientes': pacientes,
            'total': len(pacientes),
            'skip': skip,
            'limit': limit
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al obtener pacientes: {str(e)}'
        )


@router.get("/buscar")
async def buscar_paciente(cedula: str = None, paciente: str = None, cedi: str = None):
    """
    Busca pacientes por cédula o nombre de paciente.
    Si se pasa cedi, restringe la búsqueda a ese CEDI (case-insensitive).
    """
    try:
        filtro = {}

        if cedula:
            filtro['cedula'] = fx_normalizar_cedula(cedula)

        if paciente:
            filtro['paciente'] = {'$regex': paciente, '$options': 'i'}

        if cedi:
            filtro['cedi'] = {'$regex': f'^{cedi}$', '$options': 'i'}

        cursor = coleccion.find(filtro).limit(100)
        pacientes = []
        for doc in cursor:
            doc['_id'] = str(doc['_id'])
            pacientes.append(doc)

        return {
            'pacientes': pacientes,
            'total': len(pacientes)
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al buscar pacientes: {str(e)}'
        )


def _fmt_fecha(v) -> str:
    """Convierte datetime/date/str a string 'YYYY-MM-DD'. Retorna '' si None o '0000-00-00'."""
    if v is None:
        return ''
    if isinstance(v, str):
        s = v[:10]
        return '' if s.startswith('0000') or s == '' else s
    try:
        return v.strftime('%Y-%m-%d')
    except Exception:
        return ''


_MESES_ES = ['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic']

def _fmt_fecha_legible(v) -> str:
    """Convierte 'YYYY-MM-DD HH:MM:SS' o datetime a '9 abr 2026'."""
    try:
        if isinstance(v, str):
            from datetime import datetime as _dt
            d = _dt.strptime(v[:10], '%Y-%m-%d')
        else:
            d = v
        return f"{d.day} {_MESES_ES[d.month - 1]} {d.year}"
    except Exception:
        return str(v)


def _motor_cruce(pacientes: list, registros_v3: list, cronograma_dict: dict):
    """
    Motor central del cruce pacientes <-> V3.
    Generador: emite dicts de progreso y un dict final con stage='complete' y key 'result'.
    Criterio 1: nombre paciente vs cliente_destino >= 95%  (match_tipo 'nombre')
    Criterio 2: llave paciente vs llave V3 >= 73%          (match_tipo 'llave')
    Criterio 3: celular paciente vs telefono V3 exacto     (match_tipo 'celular')
    Tras la etapa de pacientes, clasifica cada V3 restante como:
      matched       : mejor similitud >= 0.75 con algún paciente (se agrega a llaves_v3_con_paciente)
      sin_paciente  : mejor similitud < 0.75
      llave_vacia   : V3 sin llave (no pudo participar en el cruce)

    NUEVA REGLA: Cuando un paciente hace match por nombre (>=95%), también reclama
    automáticamente los pedidos V3 que tengan la misma llave, aunque el nombre no
    alcance el 95%. Además, cualquier V3 con similitud de llave >= 75% contra
    cualquier paciente se considera emparejado (zona gris eliminada como categoría).
    """
    from rapidfuzz.fuzz import ratio as fuzz_ratio
    from rapidfuzz import process as _fuzz_process

    total_pacientes = len(pacientes)
    total_v3 = len(registros_v3)
    yield {'stage': 'loading', 'progress': 8,
           'message': f'{total_pacientes} pacientes y {total_v3} pedidos V3 cargados'}

    # ── Construir índices V3 ─────────────────────────────────────────────────
    llaves_v3 = [doc['llave'] for doc in registros_v3 if doc.get('llave')]

    docs_v3_por_llave: dict = {}
    contador_pedidos_por_llave: dict = {}
    for doc in registros_v3:
        llave = doc.get('llave') or ''
        if not llave:
            continue
        contador_pedidos_por_llave[llave] = contador_pedidos_por_llave.get(llave, 0) + 1
        if llave not in docs_v3_por_llave:
            docs_v3_por_llave[llave] = doc
        else:
            exist = docs_v3_por_llave[llave]
            fe_exist = exist.get('fecha_entrega', '')
            fe_curr  = doc.get('fecha_entrega', '')
            if fe_curr and not fe_exist:
                docs_v3_por_llave[llave] = doc
            elif fe_curr and fe_exist:
                if doc.get('estado_pedido') == 'ENTREGADO' and exist.get('estado_pedido') != 'ENTREGADO':
                    docs_v3_por_llave[llave] = doc

    nombres_v3 = [
        (doc.get('cliente_destino', ''), doc['llave'])
        for doc in registros_v3 if doc.get('llave') and doc.get('cliente_destino')
    ]
    nombres_v3_strs = [n for n, _ in nombres_v3]
    dict_telefonos_v3 = {
        _normalizar_cel(doc.get('telefono_original', '')): doc['llave']
        for doc in registros_v3
        if doc.get('llave') and len(_normalizar_cel(doc.get('telefono_original', ''))) >= 7
    }

    # ── Etapa 2: pacientes -> V3 ─────────────────────────────────────────────
    resultado_pacientes: list = []
    llaves_v3_con_paciente: set = set()
    paso_reporte = max(1, total_pacientes // 20)

    for idx, p in enumerate(pacientes):
        llave_paciente = p.get('llave', '') or ''
        if not llave_paciente:
            continue

        cedi_raw = p.get('cedi', '') or ''
        cedi = _CEDI_MAPA.get(cedi_raw.upper(), cedi_raw.upper())
        paciente_norm = p.get('paciente', '') or ''
        en_v3 = False
        llave_v3_match = ''
        match_tipo = None
        similitud = 0.0
        llaves_nombre_match: list = []  # todas las llaves V3 con nombre >= 95%

        # Criterio 1: nombre >= 95% — reclamar TODOS los registros con ese nombre
        if paciente_norm and nombres_v3_strs:
            res_todos = _fuzz_process.extract(
                paciente_norm, nombres_v3_strs, scorer=fuzz_ratio, score_cutoff=95
            )
            if res_todos:
                en_v3 = True
                match_tipo = 'nombre'
                similitud = round(res_todos[0][1], 1)          # mejor score
                llave_v3_match = nombres_v3[res_todos[0][2]][1]  # llave del mejor para display
                llaves_nombre_match = list({nombres_v3[r[2]][1] for r in res_todos})

        # Criterio 2: llave >= 73%
        if not en_v3 and llave_paciente and llaves_v3:
            res_l = _fuzz_process.extractOne(llave_paciente, llaves_v3, scorer=fuzz_ratio)
            if res_l:
                similitud = round(res_l[1], 1)
                llave_v3_match = res_l[0]
                if res_l[1] >= 73:
                    en_v3 = True
                    match_tipo = 'llave'

        # Criterio 3: celular exacto
        if not en_v3 and dict_telefonos_v3:
            tel1 = _normalizar_cel(p.get('telefono1', '') or '')
            tel2 = _normalizar_cel(p.get('telefono2', '') or '')
            celular_p = next(
                (t for t in (tel1, tel2) if len(t) >= 7 and t in dict_telefonos_v3), ''
            )
            if celular_p:
                en_v3 = True
                llave_v3_match = dict_telefonos_v3[celular_p]
                match_tipo = 'celular'

        if en_v3 and llave_v3_match:
            for lv in (llaves_nombre_match if llaves_nombre_match else [llave_v3_match]):
                llaves_v3_con_paciente.add(lv)

        doc_v3 = docs_v3_por_llave.get(llave_v3_match, {}) if (en_v3 and llave_v3_match) else {}
        paciente_result = {
            'paciente':           p.get('paciente_original', ''),
            'cedula':             p.get('cedula_original', ''),
            'direccion_original': p.get('direccion_original', ''),
            'ruta':               p.get('ruta', '') or 'SIN RUTA',
            'cedi':               cedi,
            'llave':              llave_paciente,
            'similitud':          similitud,
            'match_tipo':         match_tipo,
            'llave_v3':           llave_v3_match,
            'en_v3':              en_v3,
            'estado':             p.get('estado', 'ACTIVO'),
            'estado_pedido':      doc_v3.get('estado_pedido', ''),
            'fecha_pedido':       _fmt_fecha(doc_v3.get('fecha_pedido')),
            'fecha_preferente':   _fmt_fecha(doc_v3.get('fecha_preferente')),
            'fecha_entrega':      _fmt_fecha(doc_v3.get('fecha_entrega')),
            'planilla':           doc_v3.get('planilla', ''),
            'municipio_destino':  doc_v3.get('municipio_destino', ''),
            'divipola':           doc_v3.get('divipola', ''),
            'ruta_v3':            doc_v3.get('ruta', ''),
            'cliente_destino_v3': doc_v3.get('cliente_destino_original', ''),
            'celular_paciente':   ' / '.join(filter(None, [
                                      p.get('telefono1', '') or '',
                                      p.get('telefono2', '') or '',
                                  ])),
            'telefono_v3':        doc_v3.get('telefono_original', ''),
            'f_pref_teorica':     cronograma_dict.get(p.get('cedula', ''), ''),
            'cant_pedidos_v3':    sum(contador_pedidos_por_llave.get(lv, 0) for lv in llaves_nombre_match)
                                  if (en_v3 and llaves_nombre_match)
                                  else (contador_pedidos_por_llave.get(llave_v3_match, 0)
                                        if (en_v3 and llave_v3_match) else 0),
        }
        try:
            paciente_result['estado_cruce'] = _determinar_estado_cruce(
                en_v3=en_v3,
                estado_pedido=paciente_result['estado_pedido'],
                f_pref_teorica=paciente_result['f_pref_teorica'],
                f_pedido=paciente_result['fecha_pedido'],
            )
        except Exception:
            paciente_result['estado_cruce'] = '—'
        resultado_pacientes.append(paciente_result)

        if (idx + 1) % paso_reporte == 0 or (idx + 1) == total_pacientes:
            pct = round(10 + ((idx + 1) / total_pacientes) * 50)
            yield {
                'stage': 'comparing_patients', 'progress': pct,
                'processed': idx + 1, 'total': total_pacientes,
                'message': f'Paciente {idx + 1} de {total_pacientes}',
            }

    # ── Agrupar ocupación por ruta ───────────────────────────────────────────
    rutas_ocupacion: dict = {}
    for p in resultado_pacientes:
        ruta = p['ruta']
        if ruta not in rutas_ocupacion:
            rutas_ocupacion[ruta] = {
                'pacientes': [], 'total': 0, 'en_v3': 0,
                'entregados': 0, 'cedi': p['cedi'], 'planillas': set(),
            }
        rutas_ocupacion[ruta]['pacientes'].append(p)
        rutas_ocupacion[ruta]['total'] += 1
        if p['en_v3']:
            rutas_ocupacion[ruta]['en_v3'] += 1
        if p['en_v3'] and p.get('estado_pedido') == 'ENTREGADO':
            rutas_ocupacion[ruta]['entregados'] += 1
        if p.get('planilla'):
            rutas_ocupacion[ruta]['planillas'].add(p['planilla'])

    ocupacion_resultado = []
    for ruta, datos in sorted(rutas_ocupacion.items()):
        total = datos['total']
        en_v3 = datos['en_v3']
        entregados = datos['entregados']
        ocupacion_resultado.append({
            'ruta':                ruta,
            'cedi':                datos['cedi'],
            'total_pacientes':     total,
            'pacientes_en_v3':     en_v3,
            'ocupacion_pct':       round(en_v3 / total * 100, 1) if total else 0.0,
            'pacientes_entregados': entregados,
            'pct_entregados':      round(entregados / total * 100, 1) if total else 0.0,
            'vehiculos':           len(datos['planillas']),
            'pacientes':           sorted(datos['pacientes'],
                                          key=lambda x: (not x['en_v3'], -x['similitud'])),
        })

    # ── Etapa 3: V3 sin paciente / zona gris / llave vacía ──────────────────
    llaves_pacientes = [p['llave'] for p in resultado_pacientes if p.get('llave')]
    # Mapa llave → cedula para lookup del cronograma en registros sin paciente
    llave_a_cedula = {p.get('llave', ''): p.get('cedula', '') for p in pacientes if p.get('llave') and p.get('cedula')}
    sin_paciente: list = []
    zona_gris: list = []
    llave_vacia: list = []
    paso_reporte_v3 = max(1, total_v3 // 20)

    yield {'stage': 'comparing_v3', 'progress': 62,
           'message': f'Verificando {total_v3} pedidos V3...'}

    for idx, reg in enumerate(registros_v3):
        llave_v3 = reg.get('llave', '') or ''
        bodega   = reg.get('bodega_origen', '') or ''
        cedi_v3  = _CEDI_MAPA.get(bodega.upper(), bodega.upper())

        reg_base = {
            'codigo_pedido':    reg.get('codigo_pedido', ''),
            'cliente_destino':  reg.get('cliente_destino_original', ''),
            'direccion_destino': reg.get('direccion_destino_original', ''),
            'ruta':             reg.get('ruta', '') or 'SIN RUTA',
            'cedi':             cedi_v3,
            'estado_pedido':    reg.get('estado_pedido', ''),
            'telefono':         reg.get('telefono_original', ''),
            'fecha_preferente': _fmt_fecha(reg.get('fecha_preferente')),
            'llave':            llave_v3,
        }

        if not llave_v3:
            llave_vacia.append(reg_base)
        elif llave_v3 in llaves_v3_con_paciente:
            pass  # reclamado por un paciente (match directo nombre/llave/celular)
        else:
            res_sp = (
                _fuzz_process.extractOne(llave_v3, llaves_pacientes, scorer=fuzz_ratio)
                if llaves_pacientes else None
            )
            mejor_sim = res_sp[1] / 100.0 if res_sp else 0.0
            mejor_llave_p = res_sp[0] if res_sp else ''
            if mejor_sim >= 0.75:
                # Similitud suficiente con algún paciente → se considera emparejado
                # (aunque ese paciente haya sido reclamado por otro V3 como mejor match).
                llaves_v3_con_paciente.add(llave_v3)
            else:
                cedula_cercana = llave_a_cedula.get(mejor_llave_p, '')
                sin_paciente.append({
                    **reg_base,
                    'similitud': round(mejor_sim * 100, 1),
                    'llave_paciente_cercana': mejor_llave_p,
                    'f_pref_teorica': cronograma_dict.get(cedula_cercana, '') if cedula_cercana else '',
                })

        if (idx + 1) % paso_reporte_v3 == 0 or (idx + 1) == total_v3:
            pct = round(62 + ((idx + 1) / total_v3) * 28)
            yield {
                'stage': 'comparing_v3', 'progress': pct,
                'processed': idx + 1, 'total': total_v3,
                'message': f'V3 {idx + 1} de {total_v3}',
            }

    def _agrupar_por_ruta(lista: list) -> list:
        rutas: dict = {}
        for reg in lista:
            ruta = reg['ruta']
            if ruta not in rutas:
                rutas[ruta] = {'registros': [], 'cedi': reg['cedi']}
            rutas[ruta]['registros'].append(reg)
        return [
            {
                'ruta': ruta, 'cedi': datos['cedi'], 'total': len(datos['registros']),
                'registros': sorted(datos['registros'],
                                    key=lambda x: x.get('similitud', 0), reverse=True),
            }
            for ruta, datos in sorted(rutas.items())
        ]

    yield {'stage': 'saving', 'progress': 95, 'message': 'Preparando resultados...'}

    yield {
        'stage': 'complete',
        'result': {
            'ocupacion_resultado':  ocupacion_resultado,
            'v3_sin_paciente':      _agrupar_por_ruta(sin_paciente),
            'v3_zona_gris':         _agrupar_por_ruta(zona_gris),
            'v3_llave_vacia':       llave_vacia,
            'total_sin_paciente':   len(sin_paciente),
            'total_zona_gris':      len(zona_gris),
            'total_llave_vacia':    len(llave_vacia),
            'total_v3':             total_v3,
        },
    }


def ejecutar_cruce_automatico(usuario: str = 'sync_automatico') -> dict:
    """
    Ejecuta el cruce completo pacientes <-> V3 y guarda en cache_cruce_mc.
    Llamado automáticamente tras cada sync_v3 exitoso. No usa SSE.
    Usa _motor_cruce como motor central (misma lógica que el endpoint SSE).
    """
    import logging, threading
    logger = logging.getLogger(__name__)

    try:
        cronograma_dict = _get_cronograma_mes_actual()
        pacientes = list(coleccion.find(
            {},
            {'llave': 1, 'paciente': 1, 'paciente_original': 1, 'direccion_original': 1,
             'ruta': 1, 'estado': 1, 'cedula': 1, 'cedula_original': 1, 'cedi': 1,
             'telefono1': 1, 'telefono2': 1}
        ))
        coleccion_v3 = bd['v3']
        registros_v3 = list(coleccion_v3.find(
            {'llave': {'$exists': True}},
            {'llave': 1, 'cliente_destino': 1, 'cliente_destino_original': 1,
             'direccion_destino_original': 1, 'ruta': 1, 'estado_pedido': 1,
             'codigo_pedido': 1, 'bodega_origen': 1, 'telefono_original': 1,
             'fecha_pedido': 1, 'fecha_preferente': 1,
             'fecha_entrega': 1, 'planilla': 1,
             'divipola': 1, 'municipio_destino': 1}
        ))

        result = None
        for event in _motor_cruce(pacientes, registros_v3, cronograma_dict):
            if event.get('stage') == 'complete':
                result = event['result']

        if not result:
            raise RuntimeError('_motor_cruce no retornó resultado')

        fecha_calculo = time.strftime('%Y-%m-%d %H:%M:%S')
        coleccion_cache.update_one(
            {'tipo': 'cruce_completo'},
            {'$set': {
                'tipo':               'cruce_completo',
                'ocupacion_rutas':    result['ocupacion_resultado'],
                'v3_sin_paciente':    result['v3_sin_paciente'],
                'v3_zona_gris':       result['v3_zona_gris'],
                'v3_llave_vacia':     result['v3_llave_vacia'],
                'total_sin_paciente': result['total_sin_paciente'],
                'total_zona_gris':    result['total_zona_gris'],
                'total_llave_vacia':  result['total_llave_vacia'],
                'total_v3':           result['total_v3'],
                'calculado_por':      usuario,
                'fecha_calculo':      fecha_calculo,
            }},
            upsert=True
        )
        logger.info(
            f"[cruce_automatico] OK — {len(pacientes)} pacientes, "
            f"{result['total_v3']} V3, "
            f"{result['total_sin_paciente']} sin paciente, "
            f"{result['total_zona_gris']} zona gris, "
            f"{result['total_llave_vacia']} llave vacía"
        )
        threading.Thread(
            target=enviar_excel_cruce_por_correo,
            args=(usuario, fecha_calculo),
            daemon=True
        ).start()
        return {
            'ok':                 True,
            'total_pacientes':    len(pacientes),
            'total_sin_paciente': result['total_sin_paciente'],
            'total_zona_gris':    result['total_zona_gris'],
            'total_llave_vacia':  result['total_llave_vacia'],
            'fecha_calculo':      fecha_calculo,
        }

    except Exception as e:
        logger.error(f"[cruce_automatico] Error: {e}")
        return {'ok': False, 'error': str(e)}


@router.post("/recalcular-cruce")
async def recalcular_cruce(usuario: str, enviar_correo: bool = True):
    """
    Ejecuta el cruce pacientes <-> V3 con progreso en tiempo real via SSE.
    Usa _motor_cruce como motor central (misma lógica que ejecutar_cruce_automatico).
    """
    import logging, threading
    logger = logging.getLogger(__name__)
    logger.info(f"[/recalcular-cruce] usuario={usuario}, enviar_correo={enviar_correo}")

    def generar_eventos():
        try:
            yield f"data: {json.dumps({'stage': 'loading', 'progress': 0, 'message': 'Cargando pacientes y pedidos V3...'})}\n\n"

            cronograma_dict = _get_cronograma_mes_actual()
            pacientes = list(coleccion.find(
                {},
                {'llave': 1, 'paciente': 1, 'paciente_original': 1, 'direccion_original': 1,
                 'ruta': 1, 'estado': 1, 'cedula': 1, 'cedula_original': 1, 'cedi': 1,
                 'telefono1': 1, 'telefono2': 1}
            ))
            coleccion_v3 = bd['v3']
            registros_v3 = list(coleccion_v3.find(
                {'llave': {'$exists': True}},
                {'llave': 1, 'cliente_destino': 1, 'cliente_destino_original': 1,
                 'direccion_destino_original': 1, 'ruta': 1, 'estado_pedido': 1,
                 'codigo_pedido': 1, 'bodega_origen': 1, 'telefono_original': 1,
                 'fecha_pedido': 1, 'fecha_preferente': 1,
                 'fecha_entrega': 1, 'planilla': 1,
                 'divipola': 1, 'municipio_destino': 1}
            ))

            for event in _motor_cruce(pacientes, registros_v3, cronograma_dict):
                if event.get('stage') == 'complete':
                    result = event['result']
                    fecha_calculo = time.strftime('%Y-%m-%d %H:%M:%S')
                    coleccion_cache.update_one(
                        {'tipo': 'cruce_completo'},
                        {'$set': {
                            'tipo':               'cruce_completo',
                            'ocupacion_rutas':    result['ocupacion_resultado'],
                            'v3_sin_paciente':    result['v3_sin_paciente'],
                            'v3_zona_gris':       result['v3_zona_gris'],
                            'v3_llave_vacia':     result['v3_llave_vacia'],
                            'total_sin_paciente': result['total_sin_paciente'],
                            'total_zona_gris':    result['total_zona_gris'],
                            'total_llave_vacia':  result['total_llave_vacia'],
                            'total_v3':           result['total_v3'],
                            'calculado_por':      usuario,
                            'fecha_calculo':      fecha_calculo,
                        }},
                        upsert=True
                    )
                    if enviar_correo:
                        threading.Thread(
                            target=enviar_excel_cruce_por_correo,
                            args=(usuario, fecha_calculo),
                            daemon=True
                        ).start()

                    # Enviar notificaciones WhatsApp a usuarios con notificaciones_mc configuradas
                    threading.Thread(
                        target=_enviar_notificaciones_whatsapp_cruce,
                        daemon=True
                    ).start()

                    logger.info(
                        f"[recalcular-cruce] Completado. fecha={fecha_calculo}, "
                        f"pacientes={len(pacientes)}, total_v3={result['total_v3']}, "
                        f"sin_paciente={result['total_sin_paciente']}, "
                        f"zona_gris={result['total_zona_gris']}, "
                        f"llave_vacia={result['total_llave_vacia']}"
                    )
                    yield f"data: {json.dumps({'stage': 'complete', 'progress': 100, 'message': 'Cruce completado', 'rutas': result['ocupacion_resultado'], 'v3_sin_paciente': result['v3_sin_paciente'], 'v3_zona_gris': result['v3_zona_gris'], 'v3_llave_vacia': result['v3_llave_vacia'], 'total_sin_paciente': result['total_sin_paciente'], 'total_zona_gris': result['total_zona_gris'], 'total_llave_vacia': result['total_llave_vacia'], 'total_v3': result['total_v3'], 'fecha_calculo': fecha_calculo, 'calculado_por': usuario})}\n\n"
                else:
                    yield f"data: {json.dumps(event)}\n\n"

        except Exception as e:
            yield f"data: {json.dumps({'error': str(e)})}\n\n"

    return StreamingResponse(generar_eventos(), media_type="text/event-stream")


@router.get("/ocupacion-rutas")
async def ocupacion_rutas():
    """
    Retorna el último resultado de ocupación por rutas guardado en cache.
    Para recalcular usar POST /recalcular-cruce.
    """
    cache = coleccion_cache.find_one({'tipo': 'cruce_completo'}, {'_id': 0})
    if not cache:
        return {
            'rutas': [], 'fecha_calculo': None, 'calculado_por': None,
            'total_sin_paciente': 0, 'total_zona_gris': 0, 'total_llave_vacia': 0,
            'total_v3': 0,
        }
    return {
        'rutas':              cache.get('ocupacion_rutas', []),
        'fecha_calculo':      cache.get('fecha_calculo'),
        'calculado_por':      cache.get('calculado_por'),
        'total_sin_paciente': cache.get('total_sin_paciente', 0),
        'total_zona_gris':    cache.get('total_zona_gris', 0),
        'total_llave_vacia':  cache.get('total_llave_vacia', 0),
        'total_v3':           cache.get('total_v3', 0),
    }


@router.get("/v3-sin-paciente")
async def v3_sin_paciente():
    """
    Retorna el último resultado de V3 sin paciente guardado en cache.
    Incluye zona_gris (similitud >= 75% pero sin reclamar) y llave_vacia.
    Para recalcular usar POST /recalcular-cruce.
    """
    cache = coleccion_cache.find_one({'tipo': 'cruce_completo'}, {'_id': 0})
    if not cache:
        return {
            'total_sin_paciente': 0, 'rutas': [],
            'v3_zona_gris': [], 'total_zona_gris': 0,
            'v3_llave_vacia': [], 'total_llave_vacia': 0,
            'total_v3': 0,
            'fecha_calculo': None, 'calculado_por': None,
        }
    return {
        'total_sin_paciente': cache.get('total_sin_paciente', 0),
        'rutas':              cache.get('v3_sin_paciente', []),
        'v3_zona_gris':       cache.get('v3_zona_gris', []),
        'total_zona_gris':    cache.get('total_zona_gris', 0),
        'v3_llave_vacia':     cache.get('v3_llave_vacia', []),
        'total_llave_vacia':  cache.get('total_llave_vacia', 0),
        'total_v3':           cache.get('total_v3', 0),
        'fecha_calculo':      cache.get('fecha_calculo'),
        'calculado_por':      cache.get('calculado_por'),
    }


def _generar_excel_bytes(cache: dict, cedi: str = None, solo_sin_paciente: bool = False, nombre_hoja: str = 'Pacientes sin montar', sin_hoja_v3: bool = False) -> tuple:
    """
    Genera el Excel del cruce y retorna (bytes, nombre_archivo).
    Reutilizado por el endpoint de descarga y el envío por correo.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ocupacion_rutas_data = cache.get('ocupacion_rutas', [])
    v3_sin_paciente_data = cache.get('v3_sin_paciente', [])
    fecha_calculo        = cache.get('fecha_calculo', '')
    calculado_por        = cache.get('calculado_por', '')

    if cedi:
        ocupacion_rutas_data = [r for r in ocupacion_rutas_data if (r.get('cedi') or '').upper() == cedi.upper()]
        v3_sin_paciente_data = [r for r in v3_sin_paciente_data if (r.get('cedi') or '').upper() == cedi.upper()]

    from datetime import datetime as _dt, timedelta as _td
    _hoy = _dt.now().replace(hour=0, minute=0, second=0, microsecond=0)
    _limite = _hoy + _td(days=5)
    _mes_actual = _hoy.month
    _anio_actual = _hoy.year

    def _fecha_dt(s):
        """Parsea 'YYYY-MM-DD' a datetime, retorna None si falla."""
        try:
            return _dt.strptime(s[:10], '%Y-%m-%d') if s else None
        except Exception:
            return None

    wb = openpyxl.Workbook()
    header_fill    = PatternFill('solid', fgColor='004D40')
    header_font    = Font(bold=True, color='FFFFFF', size=10)
    title_font     = Font(bold=True, size=12, color='004D40')
    entregado_fill = PatternFill('solid', fgColor='F1F8F1')   # verde claro — en_v3 + ENTREGADO
    urgente_fill   = PatternFill('solid', fgColor='FFF3F3')   # rojo claro  — fecha_preferente próxima/vencida
    red_font       = Font(bold=True, color='C62828', size=10)
    center         = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left           = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_border    = Border(
        left=Side(style='thin', color='BDBDBD'), right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),  bottom=Side(style='thin', color='BDBDBD')
    )

    def set_header_row(ws, row, cols):
        for c, title in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=title)
            cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center, thin_border

    def style_cell(cell, fill=None, font=None):
        cell.border = thin_border
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        cell.alignment = left

    if not solo_sin_paciente:
        ws1 = wb.active
        ws1.title = 'Pacientes sin montar'
        ws1['A1'] = f'Pacientes sin montar (< 6 días hábiles, mes actual)  |  {_fmt_fecha_legible(fecha_calculo)}'
        ws1['A1'].font, ws1['A1'].alignment = title_font, center
        ws1.merge_cells('A1:F1')
        ws1.row_dimensions[1].height = 22
        set_header_row(ws1, 2, ['CEDI', 'Ruta', 'Paciente', 'Cédula', 'Dirección', 'F. Pref. Integra'])

        # Calcular mañana para filtro de días hábiles
        _manana = (_hoy + _td(days=1)).strftime('%Y-%m-%d')
        fila = 3
        import logging
        logger = logging.getLogger(__name__)
        for r in ocupacion_rutas_data:
            for p in r['pacientes']:
                # FILTRO 0: Si es reporte de retraso operación (sin_hoja_v3=True), incluir pacientes con estado_cruce = "retraso operación"
                # Si es reporte normal, solo pacientes SIN CRUCE (en_v3 = False)
                if sin_hoja_v3:
                    # Para retraso operación: incluir pacientes con en_v3=True y estado_cruce = "retraso operación"
                    if not p.get('en_v3', False) or p.get('estado_cruce', '').lower() not in ('retraso operación', 'retraso operacion'):
                        continue  # Excluir pacientes sin cruce o sin estado de retraso
                else:
                    # Para reporte normal: excluir pacientes ya cruzados
                    if p.get('en_v3', False):
                        continue

                f_pref_teorica = p.get('f_pref_teorica', '')
                paciente_nombre = p.get('paciente', '')
                paciente_cedula = p.get('cedula', '')
                ruta_nombre = r.get('ruta', '')

                # Log para diagnóstico: pacientes que se excluyen
                exclusion_reason = None

                if not f_pref_teorica:
                    exclusion_reason = 'sin f_pref_teorica'
                else:
                    # Usar _parsear_fecha_texto que acepta múltiples formatos (DD/MM/YYYY, YYYY-MM-DD, etc.)
                    f_pref_dt = _parsear_fecha_texto(f_pref_teorica)
                    if not f_pref_dt:
                        exclusion_reason = f'fecha inválida: {f_pref_teorica}'
                    elif f_pref_dt.month != _mes_actual or f_pref_dt.year != _anio_actual:
                        exclusion_reason = f'mes diferente: {f_pref_dt.month}/{f_pref_dt.year} vs {_mes_actual}/{_anio_actual}'
                    else:
                        # FILTRO 2: Solo incluir si F. Pref. Integra < 6 días hábiles
                        dias_habiles = _calcular_dias_habiles(_manana, f_pref_teorica)
                        if dias_habiles >= 6:
                            exclusion_reason = f'días hábiles >= 6: {dias_habiles}'

                if exclusion_reason:
                    logger.info(f"[Excel filtro] Excluido: {ruta_nombre} | {paciente_nombre} ({paciente_cedula}) | f_pref={f_pref_teorica} | {exclusion_reason}")
                    continue  # Excluir según el motivo

                # Determinar color de fondo según urgencia
                fecha_pref_dt_urgente = _parsear_fecha_texto(f_pref_teorica)
                es_rojo = fecha_pref_dt_urgente is not None and fecha_pref_dt_urgente <= _limite
                fill = urgente_fill if es_rojo else None

                vals = [r.get('cedi',''), r['ruta'], p['paciente'], p['cedula'],
                        p.get('direccion_original',''), f_pref_teorica]

                for c, val in enumerate(vals, 1):
                    font = red_font if es_rojo and c == 6 else None
                    style_cell(ws1.cell(row=fila, column=c, value=val), fill, font)
                fila += 1
        for i, w in enumerate([14,18,28,14,32,12], 1):
            ws1.column_dimensions[get_column_letter(i)].width = w
        ws1.freeze_panes = 'A3'
        if not sin_hoja_v3:
            ws2 = wb.create_sheet('Pedidos sin paciente asociado')
    else:
        ws2 = wb.active
        ws2.title = 'Pedidos sin paciente asociado'

    if not sin_hoja_v3:
        if solo_sin_paciente:
            # ── Formato cliente: "Pacientes no montados" ───────────────────────────
            ws2['A1'] = f'Pedidos sin paciente asociado  |  {_fmt_fecha_legible(fecha_calculo)}'
            ws2['A1'].font, ws2['A1'].alignment = title_font, center
            ws2.merge_cells('A1:F1')
            ws2.row_dimensions[1].height = 22
            set_header_row(ws2, 2, ['CEDI', 'Ruta', 'Código Pedido', 'Cliente Destino',
                                     'Dirección', 'F. Pref. Integra'])
            fila2 = 3
            # Calcular mañana para filtro de días hábiles
            _manana = (_hoy + _td(days=1)).strftime('%Y-%m-%d')
            total_sin_filtrado = 0
            for r in v3_sin_paciente_data:
                for reg in r['registros']:
                    # FILTRO 1: Solo incluir si F. Pref. Integra existe y es del mes actual
                    f_pref_integra = reg.get('f_pref_teorica', '')
                    if not f_pref_integra:
                        continue  # Excluir si no tiene fecha preferente teórica

                    # Usar _parsear_fecha_texto que acepta múltiples formatos
                    f_pref_dt = _parsear_fecha_texto(f_pref_integra)
                    if not f_pref_dt or f_pref_dt.month != _mes_actual or f_pref_dt.year != _anio_actual:
                        continue  # Excluir si no es del mes actual

                    # FILTRO 2: Solo incluir si F. Pref. Integra < 6 días hábiles
                    dias_habiles = _calcular_dias_habiles(_manana, f_pref_integra)
                    if dias_habiles >= 6:
                        continue  # Excluir si faltan 6 o más días hábiles

                    total_sin_filtrado += 1
                    fecha_pref_dt_urgente = _parsear_fecha_texto(f_pref_integra)
                    es_rojo = fecha_pref_dt_urgente is not None and fecha_pref_dt_urgente <= _limite
                    fill = urgente_fill if es_rojo else None

                    vals = [r.get('cedi',''), r['ruta'], reg.get('codigo_pedido',''),
                            reg.get('cliente_destino',''), reg.get('direccion_destino',''), f_pref_integra]

                    for c, val in enumerate(vals, 1):
                        font = red_font if es_rojo and c == 6 else None
                        style_cell(ws2.cell(row=fila2, column=c, value=val), fill, font)
                    fila2 += 1
            # Actualizar título con el conteo filtrado
            ws2['A1'] = f'{total_sin_filtrado} pedidos de V3 sin paciente (< 6 días hábiles, mes actual)  |  {_fmt_fecha_legible(fecha_calculo)}'
            for i, w in enumerate([14, 18, 16, 28, 32, 12], 1):
                ws2.column_dimensions[get_column_letter(i)].width = w
        else:
            # ── Formato interno: "V3 Sin Paciente" (original) ─────────────────────
            ws2['A1'] = f'V3 Sin Paciente (< 6 días hábiles, mes actual)  |  {_fmt_fecha_legible(fecha_calculo)}'
            ws2['A1'].font, ws2['A1'].alignment = title_font, center
            ws2.merge_cells('A1:F1')
            ws2.row_dimensions[1].height = 22
            set_header_row(ws2, 2, ['CEDI', 'Ruta', 'Código Pedido', 'Cliente Destino',
                                     'Dirección', 'F. Pref. Integra'])
            fila2 = 3
            # Calcular mañana para filtro de días hábiles
            _manana = (_hoy + _td(days=1)).strftime('%Y-%m-%d')
            for r in v3_sin_paciente_data:
                for reg in r['registros']:
                    # FILTRO 1: Solo incluir si F. Pref. Integra existe y es del mes actual
                    f_pref_integra = reg.get('f_pref_teorica', '')
                    if not f_pref_integra:
                        continue  # Excluir si no tiene fecha preferente teórica

                    # Usar _parsear_fecha_texto que acepta múltiples formatos
                    f_pref_dt = _parsear_fecha_texto(f_pref_integra)
                    if not f_pref_dt or f_pref_dt.month != _mes_actual or f_pref_dt.year != _anio_actual:
                        continue  # Excluir si no es del mes actual

                    # FILTRO 2: Solo incluir si F. Pref. Integra < 6 días hábiles
                    dias_habiles = _calcular_dias_habiles(_manana, f_pref_integra)
                    if dias_habiles >= 6:
                        continue  # Excluir si faltan 6 o más días hábiles

                    fecha_pref_dt_urgente = _parsear_fecha_texto(f_pref_integra)
                    es_rojo = fecha_pref_dt_urgente is not None and fecha_pref_dt_urgente <= _limite
                    fill = urgente_fill if es_rojo else None

                    vals = [r.get('cedi',''), r['ruta'], reg.get('codigo_pedido',''),
                            reg.get('cliente_destino',''), reg.get('direccion_destino',''), f_pref_integra]

                    for c, val in enumerate(vals, 1):
                        font = red_font if es_rojo and c == 6 else None
                        style_cell(ws2.cell(row=fila2, column=c, value=val), fill, font)
                    fila2 += 1
            for i, w in enumerate([14, 18, 16, 28, 32, 12], 1):
                ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = 'A3'

    buffer = io.BytesIO()
    wb.save(buffer)
    prefijo = 'v3_sin_paciente' if solo_sin_paciente else 'cruce_mc'
    nombre = f"{prefijo}{'_'+cedi if cedi else ''}_{fecha_calculo.replace(' ','_').replace(':','-')}.xlsx"
    return buffer.getvalue(), nombre


def enviar_excel_cruce_por_correo(calculado_por: str, fecha_calculo: str):
    """
    Envía el Excel del cruce segmentado por notificaciones_mc:
    - 'retraso_operacion': Excel completo (ambas hojas) → usuarios operacionales MC
    - 'sin_cruce': Solo hoja 'V3 Sin Paciente' → usuarios + contactos cliente (CLIENTE_FMC)
    """
    import os
    import logging
    import resend as _resend

    logger = logging.getLogger(__name__)
    try:
        api_key   = os.getenv('RESEND_API_KEY', '')
        mail_from = os.getenv('MAIL_FROM', 'no-reply@integralogistica.com')
        if not api_key:
            logger.warning('[cruce_email] RESEND_API_KEY no configurada')
            return
        _resend.api_key = api_key

        col_usuarios = bd['baseusuarios']
        cache = coleccion_cache.find_one({'tipo': 'cruce_completo'}, {'_id': 0})
        if not cache:
            return

        fecha_legible = _fmt_fecha_legible(fecha_calculo)
        calculado_str = f' por <strong>{calculado_por}</strong>' if calculado_por != 'sync_automatico' else ''

        # ── 1. Retraso operación: Solo pacientes con estado retraso, filtrado por CEDI del usuario ──
        usuarios_retraso = list(col_usuarios.find(
            {'notificaciones_mc': 'retraso_operacion', 'correo': {'$exists': True, '$nin': [None, '']}},
            {'correo': 1, 'regional': 1, 'perfil': 1, '_id': 0}
        ))
        for usr in usuarios_retraso:
            correo = usr.get('correo')
            if not correo:
                continue
            es_admin = (usr.get('perfil') or '').upper() == 'ADMIN'
            regional_usr = (usr.get('regional') or '').strip().upper()

            # Mapear código de regional a nombre de CEDI
            mapa_regional_cedi = {
                'CO04': 'BARRANQUILLA',
                'CO05': 'CALI',
                'CO06': 'BUCARAMANGA',
                'CO07': 'FUNZA',
                'CO09': 'MEDELLIN',
            }
            cedi_usr = mapa_regional_cedi.get(regional_usr, regional_usr).upper()

            cache_retraso = dict(cache)
            rutas_filtradas = []
            for ruta in cache.get('ocupacion_rutas', []):
                if not es_admin and cedi_usr and (ruta.get('cedi') or '').upper() != cedi_usr:
                    continue
                pacientes_retraso = [
                    p for p in ruta.get('pacientes', [])
                    if p.get('estado_cruce', '').lower() in ('retraso operación', 'retraso operacion')
                ]
                if pacientes_retraso:
                    rutas_filtradas.append({**ruta, 'pacientes': pacientes_retraso})
            if not rutas_filtradas:
                logger.info(f'[cruce_email] Retraso-op: sin registros para {correo} (CEDI={cedi_usr})')
                continue
            cache_retraso['ocupacion_rutas'] = rutas_filtradas
            cache_retraso['v3_sin_paciente'] = []
            total_retraso = sum(len(r['pacientes']) for r in rutas_filtradas)
            excel_bytes, nombre_archivo = _generar_excel_bytes(cache_retraso, nombre_hoja='Pacientes con Retraso', sin_hoja_v3=True)
            _resend.Emails.send({
                'from':    f'IntegrApp <{mail_from}>',
                'to':      [correo],
                'subject': f'Retrasos Operación — {fecha_legible} ({total_retraso})',
                'html': (
                    f'<p>Se adjunta el reporte de <strong>Retrasos de Operación</strong> '
                    f'generado el <strong>{fecha_legible}</strong>{calculado_str}.</p>'
                    f'<p>Total registros con retraso: <strong>{total_retraso}</strong>.</p>'
                    f'<p>Saludos,<br>IntegrApp</p>'
                ),
                'attachments': [{'filename': nombre_archivo, 'content': list(excel_bytes)}],
            })
            logger.info(f'[cruce_email] Retraso-op enviado a {correo} (CEDI={cedi_usr}, admin={es_admin}, {total_retraso} registros)')
        if not usuarios_retraso:
            logger.warning('[cruce_email] Sin destinatarios para retraso_operacion')

        # ── 2. Sin cruce: Pacientes sin cruce + V3 sin paciente, filtrado por CEDI del usuario ──
        # NUEVA REGLA: Solo incluir registros con F. Pref. Integra < 6 días hábiles desde hoy Y del mes actual
        hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        manana = (hoy + timedelta(days=1)).strftime('%Y-%m-%d')
        mes_actual = hoy.month
        anio_actual = hoy.year
        usuarios_sin_cruce = list(col_usuarios.find(
            {'notificaciones_mc': 'sin_cruce', 'correo': {'$exists': True, '$nin': [None, '']}},
            {'correo': 1, 'regional': 1, 'perfil': 1, '_id': 0}
        ))
        for usr in usuarios_sin_cruce:
            correo = usr.get('correo')
            if not correo:
                continue
            es_admin = (usr.get('perfil') or '').upper() == 'ADMIN'
            es_cliente_fmc = (usr.get('perfil') or '').upper() == 'CLIENTE_FMC'
            ver_todo = es_admin or es_cliente_fmc
            regional_usr = (usr.get('regional') or '').strip().upper()

            # Mapear código de regional a nombre de CEDI
            mapa_regional_cedi = {
                'CO04': 'BARRANQUILLA',
                'CO05': 'CALI',
                'CO06': 'BUCARAMANGA',
                'CO07': 'FUNZA',
                'CO09': 'MEDELLIN',
            }
            cedi_usr = mapa_regional_cedi.get(regional_usr, regional_usr).upper()

            cache_sc = dict(cache)
            rutas_sc = []
            for ruta in cache.get('ocupacion_rutas', []):
                # Filtrar por CEDI del usuario (si tiene una asignada)
                # ADMIN y CLIENTE_FMC ven todo, usuarios con regional ven solo su CEDI
                if not ver_todo and regional_usr:
                    ruta_cedi = (ruta.get('cedi') or '').upper()
                    if ruta_cedi != cedi_usr:
                        continue
                pacientes_sc = [
                    p for p in ruta.get('pacientes', [])
                    if not p.get('en_v3', False)
                ]
                # FILTRO POR FECHA: Solo pacientes con F. Pref. Integra < 6 días hábiles Y del mes actual
                pacientes_sc_filtrados = []
                for p in pacientes_sc:
                    f_pref_teorica = p.get('f_pref_teorica', '')
                    if not f_pref_teorica:
                        continue  # Excluir si no tiene fecha preferente teórica

                    # Usar _parsear_fecha_texto que acepta múltiples formatos
                    f_pref_dt = _parsear_fecha_texto(f_pref_teorica)
                    if not f_pref_dt or f_pref_dt.month != mes_actual or f_pref_dt.year != anio_actual:
                        continue  # Excluir si no es del mes actual

                    dias_habiles = _calcular_dias_habiles(manana, f_pref_teorica)
                    if dias_habiles < 6:  # Menos de 6 días hábiles desde mañana
                        pacientes_sc_filtrados.append(p)
                if pacientes_sc_filtrados:
                    rutas_sc.append({**ruta, 'pacientes': pacientes_sc_filtrados})
            if not rutas_sc:
                logger.info(f'[cruce_email] Sin-cruce: sin registros para {correo} (CEDI={cedi_usr})')
                continue
            cache_sc['ocupacion_rutas'] = rutas_sc
            # Filtrar v3_sin_paciente por CEDI y por fecha < 6 días hábiles Y del mes actual
            v3_sp_filtrado = []
            for r in cache.get('v3_sin_paciente', []):
                # Filtrar por CEDI del usuario (si tiene una asignada)
                if not ver_todo and regional_usr:
                    ruta_cedi = (r.get('cedi') or '').upper()
                    if ruta_cedi != cedi_usr:
                        continue
                registros_filtrados = []
                for reg in r.get('registros', []):
                    f_pref_teorica = reg.get('f_pref_teorica', '')
                    if not f_pref_teorica:
                        continue  # Excluir si no tiene fecha preferente teórica

                    # Usar _parsear_fecha_texto que acepta múltiples formatos
                    f_pref_dt = _parsear_fecha_texto(f_pref_teorica)
                    if not f_pref_dt or f_pref_dt.month != mes_actual or f_pref_dt.year != anio_actual:
                        continue  # Excluir si no es del mes actual

                    dias_habiles = _calcular_dias_habiles(manana, f_pref_teorica)
                    if dias_habiles < 6:  # Menos de 6 días hábiles desde mañana
                        registros_filtrados.append(reg)
                if registros_filtrados:
                    v3_sp_filtrado.append({**r, 'registros': registros_filtrados})
            cache_sc['v3_sin_paciente'] = v3_sp_filtrado
            total_sin_cruce = sum(len(r['pacientes']) for r in rutas_sc)
            total_v3_sp = sum(len(r.get('registros', [])) for r in v3_sp_filtrado)
            excel_sc, nombre_sc = _generar_excel_bytes(cache_sc)
            _resend.Emails.send({
                'from':    f'IntegrApp <{mail_from}>',
                'to':      [correo],
                'subject': f'Sin Cruce + V3 Sin Paciente — {fecha_legible}',
                'html': (
                    f'<p>Se adjunta el reporte generado el <strong>{fecha_legible}</strong>{calculado_str}.</p>'
                    f'<p>El archivo contiene dos hojas:</p>'
                    f'<ul>'
                    f'<li><strong>Pacientes sin cruce</strong>: {total_sin_cruce} pacientes no han sido tramitados por parte de FMC (F. Pref. Integra < 6 días hábiles, mes actual).</li>'
                    f'<li><strong>V3 Sin Paciente</strong>: {total_v3_sp} pedidos en la V3 de los cuales no tenemos información del paciente (F. Pref. Integra < 6 días hábiles, mes actual).</li>'
                    f'</ul>'
                    f'<p>Saludos,<br>IntegrApp</p>'
                ),
                'attachments': [{'filename': nombre_sc, 'content': list(excel_sc)}],
            })
            logger.info(f'[cruce_email] Sin-cruce enviado a {correo} (CEDI={cedi_usr}, admin={es_admin}, {total_sin_cruce} sin cruce, {total_v3_sp} v3 sin paciente)')
        if not usuarios_sin_cruce:
            logger.warning('[cruce_email] Sin destinatarios para sin_cruce')

    except Exception as e:
        logger.error(f'[cruce_email] Error: {e}')


def _enviar_notificaciones_whatsapp_cruce():
    """
    Envía notificaciones WhatsApp tras un recálculo manual del cruce.
    Reutiliza la misma lógica que sync_api_v3 pero sin el contexto del sync.
    """
    from Funciones.whatsapp_utils_integra import enviar_template_sync
    from bd.bd_cliente import bd_cliente as _bd_cli
    import logging

    logger_local = logging.getLogger(__name__)

    try:
        # Obtener cache completo del cruce
        cruce_cache = coleccion_cache.find_one({'tipo': 'cruce_completo'})
        if not cruce_cache:
            logger_local.warning("[whatsapp_cruce] No hay cache de cruce disponible")
            return

        # Obtener fecha/hora actual
        from datetime import datetime as _dt
        ahora = _dt.now()
        fecha_hora = ahora.strftime('%Y-%m-%d %H:%M:%S')

        # Mapeo de regional a CEDI
        def _mapear_regional_a_cedi(regional: str) -> str:
            mapa = {
                'CO04': 'BARRANQUILLA',
                'CO05': 'CALI',
                'CO06': 'BUCARAMANGA',
                'CO07': 'FUNZA',
                'CO09': 'MEDELLIN',
            }
            return mapa.get(regional, regional)

        def _obtener_estadisticas_por_regional(cruce_cache: dict, regional: str = None) -> dict:
            """Obtiene estadísticas del cruce filtradas por regional, con desglose por CEDI para admin."""
            cedi_filtro = _mapear_regional_a_cedi(regional) if regional else None

            # Obtener pacientes desde ocupacion_rutas
            ocupacion_rutas = cruce_cache.get('ocupacion_rutas', [])
            pacientes_por_ruta = {}

            # Diccionario para agrupar por CEDI
            stats_por_cedi = {}

            for ruta in ocupacion_rutas:
                ruta_cedi = ruta.get('cedi', '')
                if cedi_filtro and ruta_cedi != cedi_filtro:
                    continue

                # Inicializar contador para este CEDI si no existe
                if ruta_cedi not in stats_por_cedi:
                    stats_por_cedi[ruta_cedi] = {
                        'total_pacientes': 0,
                        'retraso_operacion': 0,
                        'sin_cruce': 0,
                    }

                for paciente in ruta.get('pacientes', []):
                    cedula = paciente.get('cedula')
                    pacientes_por_ruta[cedula] = paciente

                    # Contar por CEDI
                    stats_por_cedi[ruta_cedi]['total_pacientes'] += 1

                    if paciente.get('estado_cruce') == 'retraso operación':
                        stats_por_cedi[ruta_cedi]['retraso_operacion'] += 1

                    if not paciente.get('en_v3', False):
                        stats_por_cedi[ruta_cedi]['sin_cruce'] += 1

            # Si se filtra por regional, retornar solo ese CEDI
            if cedi_filtro:
                return {
                    'total_retraso_operacion': stats_por_cedi.get(cedi_filtro, {}).get('retraso_operacion', 0),
                    'total_sin_cruce': stats_por_cedi.get(cedi_filtro, {}).get('sin_cruce', 0),
                    'total_pacientes': stats_por_cedi.get(cedi_filtro, {}).get('total_pacientes', 0),
                }

            # Si no hay filtro (admin), retornar totales y desglose
            total_retraso_operacion = sum(s['retraso_operacion'] for s in stats_por_cedi.values())
            total_sin_cruce = sum(s['sin_cruce'] for s in stats_por_cedi.values())
            total_pacientes = sum(s['total_pacientes'] for s in stats_por_cedi.values())

            return {
                'total_retraso_operacion': total_retraso_operacion,
                'total_sin_cruce': total_sin_cruce,
                'total_pacientes': total_pacientes,
                'desglose_por_cedi': stats_por_cedi,  # Incluye desglose para admin
            }

        # Obtener usuarios con MEDICAL_CARE y notificaciones_mc activas
        col_usuarios = _bd_cli['integra']['baseusuarios']

        # Buscar usuarios que tienen MEDICAL_CARE en clientes y notificaciones_mc configuradas
        usuarios_notif = list(col_usuarios.find({
            'clientes': 'MEDICAL_CARE',
            '$or': [
                {'notificaciones_mc': {'$exists': True, '$ne': [], '$nin': [[None], ''], '$type': 'array'}},
                {'notificaciones_mc': {'$exists': True, '$ne': '', '$ne': None, '$type': 'string'}},
            ],
            'celular': {'$exists': True, '$ne': None, '$ne': ''}
        }))

        if not usuarios_notif:
            logger_local.info("[whatsapp_cruce] No hay usuarios con notificaciones MC configuradas")
            return

        logger_local.info(f"[whatsapp_cruce] Enviando notificaciones a {len(usuarios_notif)} usuarios")

        for usuario in usuarios_notif:
            celular = usuario.get('celular', '').strip()
            regional = usuario.get('regional', '')
            notificaciones_raw = usuario.get('notificaciones_mc', [])
            nombre_usuario = usuario.get('nombre', '')
            perfil_usuario = usuario.get('perfil', '')

            # Normalizar notificaciones a lista (maneja ambos formatos: string o array)
            if isinstance(notificaciones_raw, str):
                notificaciones = [notificaciones_raw]
            else:
                notificaciones = notificaciones_raw or []

            # Normalizar celular: eliminar espacios, guiones, paréntesis; anteponer 57 si no tiene
            celular_limpio = ''.join(c for c in celular if c.isdigit())
            if not celular_limpio.startswith('57'):
                celular_limpio = '57' + celular_limpio

            # Verificar si es ADMIN para enviar todas las regiones o solo la suya
            es_admin = (perfil_usuario or '').upper() == 'ADMIN'
            regional_para_stats = None if es_admin else regional

            # Obtener estadísticas (todas si es ADMIN, solo su regional si no)
            stats = _obtener_estadisticas_por_regional(cruce_cache, regional_para_stats)

            # Determinar el texto de regional para el mensaje
            if es_admin:
                texto_regional = "TODAS LAS REGIONALES"
            else:
                texto_regional = _mapear_regional_a_cedi(regional)

            # Función auxiliar para formatear desglose por CEDI
            def _formatear_desglose_cedi(stats_dict, tipo):
                """Genera string con desglose por CEDI para admin."""
                if 'desglose_por_cedi' not in stats_dict:
                    return ""

                # Orden de CEDIS
                orden_cedis = ['FUNZA', 'CALI', 'MEDELLIN', 'BARRANQUILLA', 'BUCARAMANGA']
                partes = []
                for cedi in orden_cedis:
                    if cedi in stats_dict['desglose_por_cedi']:
                        valor = stats_dict['desglose_por_cedi'][cedi][tipo]
                        partes.append(f"{cedi}: {valor}")

                return ' | '.join(partes) if partes else ""

            # Enviar notificaciones según tipo
            for tipo_notif in notificaciones:
                try:
                    if tipo_notif == 'retraso_operacion' and stats['total_retraso_operacion'] > 0:
                        if es_admin and 'desglose_por_cedi' in stats:
                            # Mensaje con desglose para admin (una sola línea)
                            desglose = _formatear_desglose_cedi(stats, 'retraso_operacion')
                            mensaje = (
                                f"🚨 Retraso Operación {texto_regional} | {desglose} | "
                                f"Total: {stats['total_retraso_operacion']} pedidos | El Excel con el detalle fue enviado a tu correo"
                            )
                        else:
                            # Mensaje simple para operativo
                            mensaje = (
                                f"🚨 Retraso Operación {texto_regional} | "
                                f"Tienes {stats['total_retraso_operacion']} pedidos con retraso operación que requieren montaje urgente | "
                                f"El Excel con el detalle fue enviado a tu correo"
                            )
                        res = enviar_template_sync(
                            to=celular_limpio,
                            template_name='confirmar_actualizacion',
                            language_code='es_CO',
                            body_params=[mensaje],
                        )
                        if res:
                            logger_local.info(f"[whatsapp_cruce] WS enviado a {celular_limpio} ({nombre_usuario}) - retraso_operacion")
                        else:
                            logger_local.warning(f"[whatsapp_cruce] WS no enviado a {celular_limpio} (tokens/error)")

                    elif tipo_notif == 'sin_cruce' and stats['total_sin_cruce'] > 0:
                        if es_admin and 'desglose_por_cedi' in stats:
                            # Mensaje con desglose para admin (una sola línea)
                            desglose = _formatear_desglose_cedi(stats, 'sin_cruce')
                            mensaje = (
                                f"⚠️ Pacientes Sin Montar {texto_regional} | {desglose} | "
                                f"Total: {stats['total_sin_cruce']} pacientes | El Excel con el detalle fue enviado a tu correo"
                            )
                        else:
                            # Mensaje simple para operativo
                            mensaje = (
                                f"⚠️ Pacientes Sin Montar {texto_regional} | "
                                f"Tienes {stats['total_sin_cruce']} pacientes que aún no han sido montados por parte del cliente | "
                                f"El Excel con el detalle fue enviado a tu correo"
                            )
                        res = enviar_template_sync(
                            to=celular_limpio,
                            template_name='confirmar_actualizacion',
                            language_code='es_CO',
                            body_params=[mensaje],
                        )
                        if res:
                            logger_local.info(f"[whatsapp_cruce] WS enviado a {celular_limpio} ({nombre_usuario}) - sin_cruce")
                        else:
                            logger_local.warning(f"[whatsapp_cruce] WS no enviado a {celular_limpio} (tokens/error)")
                except Exception as e:
                    logger_local.error(f"[whatsapp_cruce] Error enviando notificación a {celular_limpio}: {e}")

    except Exception as e:
        logger_local.error(f"[whatsapp_cruce] Error general: {e}")


@router.get("/exportar-cruce-excel")
async def exportar_cruce_excel(cedi: str = None):
    """
    Genera y descarga un Excel con los resultados del último cruce.
    Dos hojas: 'Ocupacion Rutas' y 'V3 Sin Paciente'.
    Si se pasa cedi, filtra solo esa regional.
    """
    from fastapi.responses import StreamingResponse as SR
    import io

    import io
    cache = coleccion_cache.find_one({'tipo': 'cruce_completo'}, {'_id': 0})
    if not cache:
        raise HTTPException(status_code=404, detail='No hay datos calculados. Ejecute el recálculo primero.')
    excel_bytes, nombre = _generar_excel_bytes(cache, cedi)
    return SR(
        io.BytesIO(excel_bytes),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nombre}"'}
    )


@router.delete("/eliminar-todos")
async def eliminar_todos_pacientes(usuario: str):
    """
    Elimina todos los pacientes de Medical Care (solo ADMIN)
    
    Args:
        usuario: Usuario que realiza la eliminación
    
    Returns:
        JSON con confirmación de eliminación
    """
    try:
        resultado = coleccion.delete_many({})
        return {
            'mensaje': f'Se eliminaron {resultado.deleted_count} registros',
            'usuario': usuario,
            'registros_eliminados': resultado.deleted_count
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al eliminar pacientes: {str(e)}'
        )


@router.post("/")
async def crear_paciente(usuario: str, paciente_data: dict):
    """
    Crea un nuevo paciente individual
    
    Args:
        usuario: Usuario que crea el paciente
        paciente_data: Datos del paciente (sede, paciente, cedula, direccion, departamento, municipio, ruta, cedi, celular)
    
    Returns:
        JSON con el paciente creado
    """
    try:
        # Extraer valores originales
        sede_original = paciente_data.get('sede', '').strip()
        paciente_original = paciente_data.get('paciente', '').strip()
        cedula_original = paciente_data.get('cedula', '').strip()
        direccion_original = paciente_data.get('direccion', '').strip()
        departamento_original = paciente_data.get('departamento', '').strip()
        municipio_original = paciente_data.get('municipio', '').strip()
        ruta_original = paciente_data.get('ruta', '').strip()
        cedi_original = paciente_data.get('cedi', '').strip()
        celular_original = paciente_data.get('celular', '').strip()
        estado = paciente_data.get('estado', 'ACTIVO').strip().upper()
        
        # Validar campos obligatorios
        if not cedula_original:
            raise HTTPException(
                status_code=400,
                detail="El campo 'cedula' es obligatorio"
            )

        # Normalizar SOLO: paciente, dirección, cédula y celular
        paciente_normalizado = fx_normalizar_paciente(paciente_original) if paciente_original else ''
        cedula_normalizada = fx_normalizar_cedula(cedula_original)
        direccion_normalizada = fx_normalizar_direccion(direccion_original)
        telefono1, telefono2 = fx_separar_telefonos(celular_original)

        if not cedula_normalizada:
            raise HTTPException(
                status_code=400,
                detail="Error al normalizar el campo 'cedula'"
            )

        # Validar duplicados
        existe_en_bd = coleccion.find_one({'cedula': cedula_normalizada})
        if existe_en_bd:
            raise HTTPException(
                status_code=409,
                detail=f"Ya existe un paciente con la cédula {cedula_original}"
            )

        direccion_final = direccion_normalizada if direccion_normalizada else direccion_original
        llave = f"{paciente_normalizado} {direccion_final}".strip()

        # Crear documento
        documento = {
            'sede': sede_original,
            'paciente': paciente_normalizado,
            'paciente_original': paciente_original,
            'cedula': cedula_normalizada,
            'cedula_original': cedula_original,
            'direccion': direccion_final,
            'direccion_original': direccion_original,
            'departamento': departamento_original,
            'municipio': municipio_original,
            'ruta': ruta_original,
            'cedi': cedi_original,
            'celular': telefono1,
            'celular_original': celular_original,
            'telefono1': telefono1,
            'telefono2': telefono2,
            'llave': llave,
            'estado': estado,
            'usuario_carga': usuario,
            'fecha_carga': time.strftime('%Y-%m-%d %H:%M:%S')
        }
        
        resultado = coleccion.insert_one(documento)
        documento['_id'] = str(resultado.inserted_id)
        
        return {
            'mensaje': 'Paciente creado exitosamente',
            'paciente': documento
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al crear paciente: {str(e)}'
        )


@router.put("/{paciente_id}")
async def actualizar_paciente(paciente_id: str, usuario: str, paciente_data: dict):
    """
    Actualiza un paciente existente
    
    Args:
        paciente_id: ID del paciente a actualizar
        usuario: Usuario que realiza la actualización
        paciente_data: Datos actualizados del paciente
    
    Returns:
        JSON con el paciente actualizado
    """
    try:
        from bson import ObjectId
        
        # Validar ObjectId
        if not ObjectId.is_valid(paciente_id):
            raise HTTPException(
                status_code=400,
                detail="ID de paciente inválido"
            )
        
        # Verificar que el paciente existe
        paciente_existente = coleccion.find_one({'_id': ObjectId(paciente_id)})
        if not paciente_existente:
            raise HTTPException(
                status_code=404,
                detail="Paciente no encontrado"
            )
        
        # Extraer valores originales
        sede_original = paciente_data.get('sede', paciente_existente.get('sede', '')).strip()
        paciente_original = paciente_data.get('paciente', paciente_existente.get('paciente_original', '')).strip()
        cedula_original = paciente_data.get('cedula', paciente_existente.get('cedula_original', '')).strip()
        direccion_original = paciente_data.get('direccion', paciente_existente.get('direccion_original', '')).strip()
        departamento_original = paciente_data.get('departamento', paciente_existente.get('departamento_original', '')).strip()
        municipio_original = paciente_data.get('municipio', paciente_existente.get('municipio_original', '')).strip()
        ruta_original = paciente_data.get('ruta', paciente_existente.get('ruta_original', '')).strip()
        cedi_original = paciente_data.get('cedi', paciente_existente.get('cedi_original', '')).strip()
        celular_original = paciente_data.get('celular', paciente_existente.get('celular_original', '')).strip()
        estado = paciente_data.get('estado', paciente_existente.get('estado', 'ACTIVO')).strip().upper()
        
        # Normalizar SOLO: paciente, dirección, cédula y celular
        paciente_normalizado = fx_normalizar_paciente(paciente_original) if paciente_original else ''
        cedula_normalizada = fx_normalizar_cedula(cedula_original)
        direccion_normalizada = fx_normalizar_direccion(direccion_original)
        telefono1, telefono2 = fx_separar_telefonos(celular_original)

        # Si la cédula cambia, validar duplicados
        cedula_actual = paciente_existente.get('cedula', '')
        if cedula_normalizada != cedula_actual:
            existe_en_bd = coleccion.find_one({
                'cedula': cedula_normalizada,
                '_id': {'$ne': ObjectId(paciente_id)}
            })
            if existe_en_bd:
                raise HTTPException(
                    status_code=409,
                    detail=f"Ya existe otro paciente con la cédula {cedula_original}"
                )

        direccion_final = direccion_normalizada if direccion_normalizada else direccion_original
        llave = f"{paciente_normalizado} {direccion_final}".strip()

        # Crear documento actualizado
        documento_actualizado = {
            'sede': sede_original,
            'paciente': paciente_normalizado,
            'paciente_original': paciente_original,
            'cedula': cedula_normalizada,
            'cedula_original': cedula_original,
            'direccion': direccion_final,
            'direccion_original': direccion_original,
            'departamento': departamento_original,
            'municipio': municipio_original,
            'ruta': ruta_original,
            'cedi': cedi_original,
            'celular': telefono1,
            'celular_original': celular_original,
            'telefono1': telefono1,
            'telefono2': telefono2,
            'llave': llave,
            'estado': estado,
            'usuario_actualizacion': usuario,
            'fecha_actualizacion': time.strftime('%Y-%m-%d %H:%M:%S')
        }
        
        coleccion.update_one(
            {'_id': ObjectId(paciente_id)},
            {'$set': documento_actualizado}
        )
        
        documento_actualizado['_id'] = paciente_id
        
        return {
            'mensaje': 'Paciente actualizado exitosamente',
            'paciente': documento_actualizado
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al actualizar paciente: {str(e)}'
        )


@router.delete("/{paciente_id}")
async def eliminar_paciente(paciente_id: str, usuario: str):
    """
    Elimina un paciente individual
    
    Args:
        paciente_id: ID del paciente a eliminar
        usuario: Usuario que realiza la eliminación
    
    Returns:
        JSON con confirmación de eliminación
    """
    try:
        from bson import ObjectId
        
        # Validar ObjectId
        if not ObjectId.is_valid(paciente_id):
            raise HTTPException(
                status_code=400,
                detail="ID de paciente inválido"
            )
        
        # Verificar que el paciente existe
        paciente_existente = coleccion.find_one({'_id': ObjectId(paciente_id)})
        if not paciente_existente:
            raise HTTPException(
                status_code=404,
                detail="Paciente no encontrado"
            )
        
        # Eliminar paciente
        coleccion.delete_one({'_id': ObjectId(paciente_id)})
        
        return {
            'mensaje': 'Paciente eliminado exitosamente',
            'paciente_id': paciente_id,
            'usuario': usuario
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al eliminar paciente: {str(e)}'
        )


@router.get("/{paciente_id}")
async def obtener_paciente(paciente_id: str):
    """
    Obtiene un paciente por ID
    
    Args:
        paciente_id: ID del paciente
    
    Returns:
        JSON con los datos del paciente
    """
    try:
        from bson import ObjectId
        
        # Validar ObjectId
        if not ObjectId.is_valid(paciente_id):
            raise HTTPException(
                status_code=400,
                detail="ID de paciente inválido"
            )
        
        paciente = coleccion.find_one({'_id': ObjectId(paciente_id)})
        if not paciente:
            raise HTTPException(
                status_code=404,
                detail="Paciente no encontrado"
            )
        
        paciente['_id'] = str(paciente['_id'])
        
        return {
            'paciente': paciente
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f'Error al obtener paciente: {str(e)}'
        )


# Función auxiliar para normalización base (reutilizada)
def fx_normalizar_base(txt: str) -> str:
    """Normalización básica"""
    if not txt:
        return txt
    
    # Convertir a mayúsculas y recortar
    t0 = txt.strip().upper()
    
    # Compactar espacios
    return ' '.join(t0.split())
